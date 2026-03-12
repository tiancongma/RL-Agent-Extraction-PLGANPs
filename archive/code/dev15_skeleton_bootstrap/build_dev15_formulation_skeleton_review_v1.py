#!/usr/bin/env python3
"""
Build DEV15 formulation skeleton candidate scaffold and review workbook.

Example:
python src/archive_methods/dev15_skeleton_bootstrap/build_dev15_formulation_skeleton_review_v1.py --overwrite
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from src.archive_methods.dev15_skeleton_bootstrap.formulation_skeleton_common import (
        BOUNDARY_CONFIDENCE_OPTIONS,
        CANDIDATE_COLUMNS,
        FORMULATION_EXISTS_OPTIONS,
        REVIEW_COLUMNS,
        REVIEW_STATUS_OPTIONS,
        SOURCE_TYPE_OPTIONS,
        build_candidate_rows_from_source,
        detect_candidate_source,
        discover_dev15_manifest,
        ensure_candidate_rows,
        load_manifest_rows,
        norm_text,
        project_root,
        slugify_doi,
        write_tsv,
    )
except ModuleNotFoundError:
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from src.archive_methods.dev15_skeleton_bootstrap.formulation_skeleton_common import (
        BOUNDARY_CONFIDENCE_OPTIONS,
        CANDIDATE_COLUMNS,
        FORMULATION_EXISTS_OPTIONS,
        REVIEW_COLUMNS,
        REVIEW_STATUS_OPTIONS,
        SOURCE_TYPE_OPTIONS,
        build_candidate_rows_from_source,
        detect_candidate_source,
        discover_dev15_manifest,
        ensure_candidate_rows,
        load_manifest_rows,
        norm_text,
        project_root,
        slugify_doi,
        write_tsv,
    )


def _write_candidate_scaffold_files(
    out_candidates_dir: Path,
    candidate_rows: List[Dict[str, str]],
) -> None:
    out_candidates_dir.mkdir(parents=True, exist_ok=True)
    out_all = out_candidates_dir / "dev15_formulation_candidates.tsv"
    write_tsv(out_all, CANDIDATE_COLUMNS, candidate_rows)

    by_key: Dict[str, List[Dict[str, str]]] = {}
    for row in candidate_rows:
        by_key.setdefault(row["paper_key"], []).append(row)

    for paper_key, rows in by_key.items():
        doi = rows[0].get("doi", "")
        out_jsonl = out_candidates_dir / f"{paper_key}__{slugify_doi(doi)}__candidates.jsonl"
        with out_jsonl.open("w", encoding="utf-8", newline="") as f:
            for row in rows:
                f.write(json.dumps(row, ensure_ascii=True) + "\n")


def _build_workbook(
    workbook_path: Path,
    manifest_rows: List[Dict[str, str]],
    candidate_rows: List[Dict[str, str]],
    source_path: Path | None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "review_formulations"
    ws_opts = wb.create_sheet("dropdown_options")
    ws_instr = wb.create_sheet("instructions")
    ws_summary = wb.create_sheet("source_summary")

    for j, col in enumerate(REVIEW_COLUMNS, start=1):
        ws.cell(row=1, column=j, value=col)

    manifest_by_key = {r["paper_key"]: r for r in manifest_rows}
    paper_keys_order = sorted({r["paper_key"] for r in candidate_rows})
    alt_fill_a = PatternFill("solid", fgColor="FFFFFF")
    alt_fill_b = PatternFill("solid", fgColor="F8FAFC")
    paper_fill_map = {
        key: (alt_fill_a if i % 2 == 0 else alt_fill_b)
        for i, key in enumerate(paper_keys_order)
    }

    for idx, row in enumerate(candidate_rows, start=2):
        paper_key = row["paper_key"]
        paper_meta = manifest_by_key[paper_key]
        label = norm_text(row.get("candidate_formulation_label", ""))
        source_type = norm_text(row.get("source_type_candidate", ""))
        source_locator = norm_text(row.get("evidence_pointer_candidate", ""))
        notes = norm_text(row.get("notes_candidate", ""))

        exists_default = "yes" if (label or source_locator or source_type) else "uncertain"
        confidence_default = "medium" if exists_default == "yes" else "low"

        values = {
            "paper_key": paper_key,
            "doi": paper_meta["doi"],
            "paper_title": paper_meta["paper_title"],
            "formulation_id": row["candidate_formulation_id"],
            "formulation_label_raw": label,
            "source_type": source_type,
            "source_locator": source_locator,
            "formulation_exists_gt": exists_default,
            "formulation_boundary_confidence": confidence_default,
            "review_status": "pending",
            "notes": notes,
            "helper_incomplete": '=IF(OR($F{r}="",$H{r}="",$I{r}="",$J{r}=""),"missing_required","")'.format(r=idx),
            "helper_duplicate_id": '=IF(COUNTIFS($A:$A,$A{r},$D:$D,$D{r})>1,"duplicate_id","")'.format(r=idx),
        }
        for j, col in enumerate(REVIEW_COLUMNS, start=1):
            ws.cell(row=idx, column=j, value=values[col])
            ws.cell(row=idx, column=j).fill = paper_fill_map[paper_key]

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 14,
        "B": 24,
        "C": 45,
        "D": 18,
        "E": 28,
        "F": 20,
        "G": 34,
        "H": 16,
        "I": 22,
        "J": 18,
        "K": 28,
        "L": 18,
        "M": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row_i in range(2, ws.max_row + 1):
        for col in ["C", "E", "G", "K"]:
            ws[f"{col}{row_i}"].alignment = Alignment(wrap_text=True, vertical="top")

    # Dropdown sheet
    ws_opts["A1"] = "source_type_options"
    for i, v in enumerate(SOURCE_TYPE_OPTIONS, start=2):
        ws_opts[f"A{i}"] = v
    ws_opts["B1"] = "formulation_exists_gt_options"
    for i, v in enumerate(FORMULATION_EXISTS_OPTIONS, start=2):
        ws_opts[f"B{i}"] = v
    ws_opts["C1"] = "formulation_boundary_confidence_options"
    for i, v in enumerate(BOUNDARY_CONFIDENCE_OPTIONS, start=2):
        ws_opts[f"C{i}"] = v
    ws_opts["D1"] = "review_status_options"
    for i, v in enumerate(REVIEW_STATUS_OPTIONS, start=2):
        ws_opts[f"D{i}"] = v
    ws_opts.sheet_state = "hidden"

    max_row = ws.max_row
    dv_source = DataValidation(type="list", formula1="=dropdown_options!$A$2:$A$6", allow_blank=True)
    dv_exists = DataValidation(type="list", formula1="=dropdown_options!$B$2:$B$4", allow_blank=True)
    dv_conf = DataValidation(type="list", formula1="=dropdown_options!$C$2:$C$4", allow_blank=True)
    dv_status = DataValidation(type="list", formula1="=dropdown_options!$D$2:$D$4", allow_blank=True)
    for dv in [dv_source, dv_exists, dv_conf, dv_status]:
        ws.add_data_validation(dv)
    dv_source.add(f"F2:F{max_row}")
    dv_exists.add(f"H2:H{max_row}")
    dv_conf.add(f"I2:I{max_row}")
    dv_status.add(f"J2:J{max_row}")

    # Conditional formats
    missing_fill = PatternFill("solid", fgColor="FDE68A")
    uncertain_fill = PatternFill("solid", fgColor="FECACA")
    duplicate_fill = PatternFill("solid", fgColor="FCA5A5")
    ws.conditional_formatting.add(
        f"A2:K{max_row}",
        FormulaRule(formula=['=OR($F2="",$H2="",$I2="",$J2="")'], fill=missing_fill),
    )
    ws.conditional_formatting.add(
        f"A2:K{max_row}",
        FormulaRule(formula=['=OR($H2="uncertain",$J2="needs_second_pass",$I2="low")'], fill=uncertain_fill),
    )
    ws.conditional_formatting.add(
        f"D2:D{max_row}",
        FormulaRule(formula=['=COUNTIFS($A:$A,$A2,$D:$D,$D2)>1'], fill=duplicate_fill),
    )

    # Instructions sheet
    instructions = [
        "This workbook is for DEV-only formulation skeleton curation (debug iteration, not final publication benchmark).",
        "Goal per row: confirm whether a candidate formulation exists in the paper and set review status.",
        "You mainly edit these columns: formulation_label_raw, source_type, source_locator, formulation_exists_gt, formulation_boundary_confidence, review_status, notes.",
        "Use dropdowns whenever possible to minimize typing.",
        "For real formulations: formulation_exists_gt=yes and review_status=reviewed.",
        "For false candidates: formulation_exists_gt=no and review_status=reviewed.",
        "For uncertain cases: formulation_exists_gt=uncertain and review_status=needs_second_pass.",
        "Export includes confirmed reviewed formulations only.",
        "Stable IDs are deterministic and prefilled as {paper_key}_F##.",
        "Conditional colors: yellow=missing required dropdown, red=uncertain/second-pass, dark red=duplicate formulation_id within paper.",
    ]
    ws_instr["A1"] = "DEV15 Formulation Skeleton Review Instructions"
    ws_instr["A1"].font = Font(bold=True)
    for i, line in enumerate(instructions, start=3):
        ws_instr[f"A{i}"] = line
    ws_instr.column_dimensions["A"].width = 130
    ws_instr.freeze_panes = "A3"

    # Source summary sheet
    ws_summary["A1"] = "paper_key"
    ws_summary["B1"] = "doi"
    ws_summary["C1"] = "paper_title"
    ws_summary["D1"] = "candidate_rows"
    counts: Dict[str, int] = {}
    for row in candidate_rows:
        counts[row["paper_key"]] = counts.get(row["paper_key"], 0) + 1
    for i, paper in enumerate(manifest_rows, start=2):
        ws_summary[f"A{i}"] = paper["paper_key"]
        ws_summary[f"B{i}"] = paper["doi"]
        ws_summary[f"C{i}"] = paper["paper_title"]
        ws_summary[f"D{i}"] = counts.get(paper["paper_key"], 0)
    ws_summary["F1"] = "candidate_source_path"
    ws_summary["F2"] = str(source_path) if source_path else "NONE (default blank scaffold)"
    ws_summary.column_dimensions["A"].width = 14
    ws_summary.column_dimensions["B"].width = 24
    ws_summary.column_dimensions["C"].width = 45
    ws_summary.column_dimensions["D"].width = 16
    ws_summary.column_dimensions["F"].width = 80

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)


def main() -> int:
    repo = project_root()
    default_out_dir = repo / "data/cleaned/labels/manual/dev15_formulation_skeleton"
    default_workbook = default_out_dir / "dev15_formulation_skeleton_review_v1.xlsx"
    default_candidates_dir = default_out_dir / "candidates"

    ap = argparse.ArgumentParser()
    ap.add_argument("--manifest-tsv", type=Path, default=None, help="DEV manifest TSV. Auto-detected if omitted.")
    ap.add_argument("--candidate-tsv", type=Path, default=None, help="Optional candidate TSV with scaffold schema.")
    ap.add_argument("--out-workbook", type=Path, default=default_workbook, help="Output review workbook path.")
    ap.add_argument("--out-candidates-dir", type=Path, default=default_candidates_dir, help="Output candidate scaffold directory.")
    ap.add_argument(
        "--default-rows-per-paper",
        type=int,
        default=8,
        help="Minimum rows per paper for manual review (existing candidates keep more rows).",
    )
    ap.add_argument("--overwrite", action="store_true", help="Allow overwriting output workbook.")
    args = ap.parse_args()

    manifest_path = discover_dev15_manifest(args.manifest_tsv)
    manifest_rows = load_manifest_rows(manifest_path)
    if len(manifest_rows) != 15:
        print(f"[WARN] Expected 15 papers in DEV set, got {len(manifest_rows)} from {manifest_path}")

    if args.out_workbook.exists() and not args.overwrite:
        raise FileExistsError(f"Workbook already exists (use --overwrite): {args.out_workbook}")

    source_path = None
    if args.candidate_tsv:
        source_path = args.candidate_tsv
        if not source_path.exists():
            raise FileNotFoundError(f"Candidate TSV not found: {source_path}")
        candidate_rows = build_candidate_rows_from_source(source_path, manifest_rows)
    else:
        source_path = detect_candidate_source(
            dev_keys=[r["paper_key"] for r in manifest_rows],
            dev_dois=[r["doi"] for r in manifest_rows],
        )
        candidate_rows = build_candidate_rows_from_source(source_path, manifest_rows) if source_path else []

    candidate_rows = ensure_candidate_rows(
        manifest_rows=manifest_rows,
        candidate_rows=candidate_rows,
        default_rows_per_paper=max(1, args.default_rows_per_paper),
    )
    _write_candidate_scaffold_files(args.out_candidates_dir, candidate_rows)
    _build_workbook(args.out_workbook, manifest_rows, candidate_rows, source_path)

    print(f"[OK] manifest_path\t{manifest_path}")
    print(f"[OK] candidate_source\t{source_path if source_path else 'NONE'}")
    print(f"[OK] candidate_rows\t{len(candidate_rows)}")
    print(f"[OK] candidates_dir\t{args.out_candidates_dir}")
    print(f"[OK] workbook\t{args.out_workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


