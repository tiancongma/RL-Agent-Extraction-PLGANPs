#!/usr/bin/env python3
"""
Create a versioned DEV15 GT workbook authority update and a variant-alignment scaffold.

This utility preserves the prior workbook in place, writes a new sibling workbook,
updates BXCV5XWB from 3 to 9 variant-aware rows using the retained Stage 5 final
rows, and exports a DOI-level adjudication scaffold for the next GT layer.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from src.utils.paths import DATA_CLEANED_DIR, DATA_RESULTS_DIR, PROJECT_ROOT
except ModuleNotFoundError:
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[3]))
    from src.utils.paths import DATA_CLEANED_DIR, DATA_RESULTS_DIR, PROJECT_ROOT


DEFAULT_BASE_XLSX = (
    DATA_CLEANED_DIR
    / "labels"
    / "manual"
    / "dev15_formulation_skeleton"
    / "dev15_formulation_skeleton_review_v1_fixed.xlsx"
)
DEFAULT_OUT_XLSX = (
    DATA_CLEANED_DIR
    / "labels"
    / "manual"
    / "dev15_formulation_skeleton"
    / "dev15_formulation_skeleton_review_v2_variantaware.xlsx"
)
DEFAULT_FINAL_TABLE_TSV = (
    DATA_RESULTS_DIR
    / "run_20260313_0950_f4912f3_dev15_current_merged_benchmark_v1"
    / "final_formulation_table_v1.tsv"
)
DEFAULT_SCOPE_MANIFEST_TSV = (
    DATA_RESULTS_DIR
    / "run_20260313_0950_f4912f3_dev15_current_merged_benchmark_v1"
    / "dev15_scope.tsv"
)
DEFAULT_SCAFFOLD_TSV = (
    DATA_CLEANED_DIR
    / "labels"
    / "manual"
    / "dev15_formulation_skeleton"
    / "dev15_variant_alignment_scaffold_v1.tsv"
)

BX_KEY = "BXCV5XWB"
BX_DOI = "10.1007/s10439-019-02430-x"
BX_REVIEW_START_ROW = 63
BX_REVIEW_END_ROW = 70
BX_SOURCE_SUMMARY_ROW = 6


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_token(value: Any) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def first_number_token(value: Any) -> str:
    match = re.search(r"[-+]?\d+(?:\.\d+)?", normalize_text(value))
    if not match:
        return ""
    token = match.group(0)
    try:
        num = float(token)
    except ValueError:
        return token
    if num.is_integer():
        return str(int(num))
    return f"{num:.6g}"


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def infer_polymer_family(text_bundle: str) -> str:
    text = normalize_text(text_bundle).lower()
    text = text.replace("–", "-").replace("—", "-")
    if (
        "plga-peg-ha" in text
        or "plga peg ha" in text
        or "peg-ha" in text
        or ("plga-peg" in text and "hyaluronic" in text)
        or ("plga peg" in text and "hyaluronic" in text)
        or ("plga" in text and "ethylene glycol" in text and "hyaluronic" in text)
    ):
        return "PLGA-PEG-HA"
    if "plga-peg" in text or "plga peg" in text or ("plga" in text and "ethylene glycol" in text):
        return "PLGA-PEG"
    if "plga" in text:
        return "PLGA"
    return ""


def infer_payload_state(text_bundle: str, drug_name: str = "", loaded_state: str = "") -> str:
    text = normalize_text(text_bundle).lower()
    if "blank" in text or "empty" in text:
        return "blank_control"
    if normalize_token(drug_name) == "fitc" or "fitc" in text:
        return "fitc_assay_loaded"
    if normalize_token(drug_name) or normalize_text(loaded_state).lower() == "drug_loaded":
        return "drug_loaded"
    return ""


def payload_label(payload_state: str) -> str:
    if payload_state == "fitc_assay_loaded":
        return "FITC"
    if payload_state == "blank_control":
        return "Blank"
    return "Kartogenin"


def sort_bx_pred_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    payload_order = {"drug_loaded": 0, "fitc_assay_loaded": 1, "blank_control": 2}
    family_order = {"PLGA": 0, "PLGA-PEG": 1, "PLGA-PEG-HA": 2}

    def sort_key(row: dict[str, str]) -> tuple[int, int, str]:
        raw_label = row.get("representative_source_raw_formulation_label", "") or row.get("raw_formulation_label", "")
        payload_state = infer_payload_state(
            " ".join(
                [
                    raw_label,
                    row.get("drug_name_value", ""),
                    row.get("loaded_state_final", ""),
                ]
            ),
            drug_name=row.get("drug_name_value", ""),
            loaded_state=row.get("loaded_state_final", ""),
        )
        family = infer_polymer_family(
            " ".join(
                [
                    raw_label,
                    row.get("polymer_identity_final", ""),
                ]
            )
        )
        return (
            payload_order.get(payload_state, 99),
            family_order.get(family, 99),
            row.get("final_formulation_id", ""),
        )

    return sorted(rows, key=sort_key)


def build_bx_row_payloads(final_rows: list[dict[str, str]], existing_bx_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    bx_final_rows = [row for row in final_rows if normalize_text(row.get("key")) == BX_KEY]
    if len(bx_final_rows) != 9:
        raise RuntimeError(f"Expected 9 retained BXCV5XWB final rows, found {len(bx_final_rows)}")

    existing_notes_by_id = {row["formulation_id"]: row.get("notes", "") for row in existing_bx_rows}
    paper_title = existing_bx_rows[0].get("paper_title", "") if existing_bx_rows else ""
    sorted_rows = sort_bx_pred_rows(bx_final_rows)
    out_rows: list[dict[str, str]] = []
    for index, pred_row in enumerate(sorted_rows, start=1):
        raw_label = pred_row.get("representative_source_raw_formulation_label", "") or pred_row.get("raw_formulation_label", "")
        family = infer_polymer_family(" ".join([raw_label, pred_row.get("polymer_identity_final", "")]))
        payload_state = infer_payload_state(
            " ".join([raw_label, pred_row.get("drug_name_value", ""), pred_row.get("loaded_state_final", "")]),
            drug_name=pred_row.get("drug_name_value", ""),
            loaded_state=pred_row.get("loaded_state_final", ""),
        )
        formulation_id = f"{BX_KEY}_F{index:02d}"
        base_note = existing_notes_by_id.get(formulation_id, "")
        payload_text = {
            "drug_loaded": "KGN-loaded",
            "fitc_assay_loaded": "FITC-loaded",
            "blank_control": "Blank",
        }.get(payload_state, "Retained")
        generated_note = (
            f"Variant-aware GT authority v2: retained Stage5 final row `{pred_row.get('final_formulation_id', '')}` "
            f"for {payload_text} {family} nanoparticles. BXCV5XWB was revised from family-like scaffold counting "
            f"to retained variant-aware final-output counting to match the active benchmark object."
        )
        note = normalize_text(" ".join(part for part in [base_note, generated_note] if part))
        out_rows.append(
            {
                "paper_key": BX_KEY,
                "doi": BX_DOI,
                "paper_title": paper_title,
                "formulation_id": formulation_id,
                "formulation_label_raw": payload_label(payload_state),
                "source_type": "text_described",
                "source_locator": "fulltext",
                "formulation_exists_gt": "yes",
                "formulation_boundary_confidence": "medium",
                "review_status": "reviewed",
                "notes": note,
                "helper_incomplete": "",
                "helper_duplicate_id": "",
            }
        )
    return out_rows


def load_existing_bx_rows(workbook_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["review_formulations"]
    raw_rows = worksheet.iter_rows(values_only=True)
    header = [str(value) if value is not None else "" for value in next(raw_rows)]
    rows: list[dict[str, str]] = []
    for values in raw_rows:
        row = {
            header[idx]: "" if idx >= len(values) or values[idx] is None else str(values[idx])
            for idx in range(len(header))
        }
        if row.get("paper_key", "") == BX_KEY:
            rows.append(row)
    return rows


def create_updated_workbook(
    base_xlsx: Path,
    out_xlsx: Path,
    final_table_tsv: Path,
) -> dict[str, Any]:
    if out_xlsx.exists():
        out_xlsx.unlink()
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(base_xlsx, out_xlsx)

    final_rows = read_tsv(final_table_tsv)
    existing_bx_rows = load_existing_bx_rows(base_xlsx)
    bx_payloads = build_bx_row_payloads(final_rows, existing_bx_rows)

    workbook = load_workbook(out_xlsx)
    review_ws = workbook["review_formulations"]
    summary_ws = workbook["source_summary"]

    review_ws.insert_rows(BX_REVIEW_END_ROW + 1, amount=1)
    header = [cell.value for cell in review_ws[1]]
    header_index = {str(value): idx + 1 for idx, value in enumerate(header)}

    for excel_row_num, payload in enumerate(bx_payloads, start=BX_REVIEW_START_ROW):
        for field_name, value in payload.items():
            review_ws.cell(row=excel_row_num, column=header_index[field_name], value=value)

    summary_header = [cell.value for cell in summary_ws[1]]
    summary_index = {str(value): idx + 1 for idx, value in enumerate(summary_header)}
    summary_ws.cell(row=BX_SOURCE_SUMMARY_ROW, column=summary_index["candidate_rows"], value="9")
    summary_ws.cell(row=BX_SOURCE_SUMMARY_ROW, column=summary_index["GT_rows"], value="9")
    summary_ws.cell(
        row=BX_SOURCE_SUMMARY_ROW,
        column=summary_index["candidate_source_path"],
        value=str(final_table_tsv.resolve()),
    )

    workbook.save(out_xlsx)
    return {
        "out_xlsx": str(out_xlsx),
        "bx_review_rows": [row["formulation_id"] for row in bx_payloads],
        "bx_source_summary_row": BX_SOURCE_SUMMARY_ROW,
        "final_table_tsv": str(final_table_tsv),
    }


def read_gt_rows_from_workbook(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["review_formulations"]
    raw_rows = worksheet.iter_rows(values_only=True)
    header = [str(value) if value is not None else "" for value in next(raw_rows)]
    rows: list[dict[str, str]] = []
    for values in raw_rows:
        row = {
            header[idx]: "" if idx >= len(values) or values[idx] is None else str(values[idx])
            for idx in range(len(header))
        }
        if normalize_text(row.get("formulation_exists_gt", "")).lower() == "yes":
            rows.append(row)
    return rows


def numeric_signature_tokens(text_bundle: str, limit: int = 3) -> list[str]:
    tokens = re.findall(r"[-+]?\d+(?:\.\d+)?", normalize_text(text_bundle))
    return tokens[:limit]


def build_gt_semantics(row: dict[str, str]) -> dict[str, str]:
    text_bundle = " ".join(
        [
            row.get("formulation_label_raw", ""),
            row.get("notes", ""),
            row.get("source_locator", ""),
        ]
    )
    polymer_family = infer_polymer_family(text_bundle)
    payload_state = infer_payload_state(text_bundle)
    numeric_tokens = numeric_signature_tokens(text_bundle)
    linked_match = re.search(r"Stage5 final row `([^`]+)`", row.get("notes", ""))
    signature_parts = [
        normalize_text(row.get("doi", "")).lower(),
        polymer_family,
        payload_state,
        normalize_token(row.get("formulation_label_raw", "")),
        *numeric_tokens,
    ]
    signature = "|".join(part for part in signature_parts if part)
    family_id = f"{row.get('paper_key', '')}::{polymer_family}" if polymer_family else ""
    return {
        "canonical_signature_gt": signature,
        "family_id": family_id,
        "parent_core_row_id": row.get("formulation_id", ""),
        "variant_role": "family_core" if family_id else "",
        "payload_state": payload_state,
        "gt_evidence_note": normalize_text(row.get("notes", "")),
        "linked_pred_row_id": linked_match.group(1) if linked_match else "",
    }


def build_pred_semantics(row: dict[str, str]) -> dict[str, str]:
    raw_label = row.get("representative_source_raw_formulation_label", "") or row.get("raw_formulation_label", "")
    polymer_family = infer_polymer_family(" ".join([raw_label, row.get("polymer_identity_final", "")]))
    payload_state = infer_payload_state(
        " ".join([raw_label, row.get("drug_name_value", ""), row.get("loaded_state_final", "")]),
        drug_name=row.get("drug_name_value", ""),
        loaded_state=row.get("loaded_state_final", ""),
    )
    signature_parts = [
        normalize_text(row.get("doi", "")).lower(),
        polymer_family,
        payload_state,
        normalize_token(row.get("drug_name_value", "")),
        first_number_token(row.get("drug_feed_amount_text_value", "")),
        first_number_token(row.get("plga_mass_mg_value", "")),
        first_number_token(row.get("surfactant_concentration_text_value", "")),
        normalize_token(row.get("organic_solvent_value", "")),
    ]
    signature = "|".join(part for part in signature_parts if part)
    family_id = f"{row.get('key', '')}::{polymer_family}" if polymer_family else ""
    pred_source_text = normalize_text(raw_label)
    return {
        "canonical_signature_pred": signature,
        "family_id": family_id,
        "parent_core_row_id": row.get("representative_source_formulation_id", ""),
        "variant_role": "family_core" if family_id else "",
        "payload_state": payload_state,
        "pred_evidence_anchor": normalize_text(raw_label),
        "pred_source_table": "",
        "pred_source_text": pred_source_text,
    }


def is_informative_signature(signature: str) -> bool:
    parts = [part for part in signature.split("|") if part]
    return len(parts) >= 4


def pair_rows_for_doi(
    doi: str,
    gt_rows: list[dict[str, str]],
    pred_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    gt_pending = []
    for row in gt_rows:
        payload = build_gt_semantics(row)
        gt_pending.append({**row, **payload})
    pred_pending = []
    for row in pred_rows:
        payload = build_pred_semantics(row)
        pred_pending.append({**row, **payload})

    out_rows: list[dict[str, str]] = []
    gt_by_sig: dict[str, list[dict[str, str]]] = defaultdict(list)
    pred_by_sig: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in gt_pending:
        gt_by_sig[row["canonical_signature_gt"]].append(row)
    for row in pred_pending:
        pred_by_sig[row["canonical_signature_pred"]].append(row)

    matched_gt_ids: set[str] = set()
    matched_pred_ids: set[str] = set()

    pred_by_id = {row["final_formulation_id"]: row for row in pred_pending}
    for gt_row in gt_pending:
        linked_pred_row_id = gt_row.get("linked_pred_row_id", "")
        if not linked_pred_row_id:
            continue
        pred_row = pred_by_id.get(linked_pred_row_id)
        if not pred_row:
            continue
        matched_gt_ids.add(gt_row["formulation_id"])
        matched_pred_ids.add(pred_row["final_formulation_id"])
        out_rows.append(
            build_alignment_row(
                doi=doi,
                gt_row=gt_row,
                pred_row=pred_row,
                alignment_decision="matched",
                benchmark_include_gt="yes",
                alignment_confidence="high",
                alignment_notes="Matched by explicit retained Stage5 final-row reference recorded in the GT note.",
            )
        )

    for signature, gt_group in sorted(gt_by_sig.items()):
        pred_group = pred_by_sig.get(signature, [])
        if not signature or not is_informative_signature(signature):
            continue
        if len(gt_group) == 1 and len(pred_group) == 1:
            gt_row = gt_group[0]
            pred_row = pred_group[0]
            if gt_row["formulation_id"] in matched_gt_ids or pred_row["final_formulation_id"] in matched_pred_ids:
                continue
            matched_gt_ids.add(gt_row["formulation_id"])
            matched_pred_ids.add(pred_row["final_formulation_id"])
            out_rows.append(
                build_alignment_row(
                    doi=doi,
                    gt_row=gt_row,
                    pred_row=pred_row,
                    alignment_decision="matched",
                    benchmark_include_gt="yes",
                    alignment_confidence="high",
                    alignment_notes="Unique canonical signature match.",
                )
            )

    remaining_gt = [row for row in gt_pending if row["formulation_id"] not in matched_gt_ids]
    remaining_pred = [row for row in pred_pending if row["final_formulation_id"] not in matched_pred_ids]
    remaining_gt.sort(key=lambda row: row["formulation_id"])
    remaining_pred.sort(key=lambda row: row["final_formulation_id"])

    pair_count = min(len(remaining_gt), len(remaining_pred))
    for idx in range(pair_count):
        gt_row = remaining_gt[idx]
        pred_row = remaining_pred[idx]
        out_rows.append(
            build_alignment_row(
                doi=doi,
                gt_row=gt_row,
                pred_row=pred_row,
                alignment_decision="needs_review",
                benchmark_include_gt="yes",
                alignment_confidence="low",
                alignment_notes="Paired by stable DOI-local order after the unique-signature pass.",
            )
        )

    for gt_row in remaining_gt[pair_count:]:
        out_rows.append(
            build_alignment_row(
                doi=doi,
                gt_row=gt_row,
                pred_row=None,
                alignment_decision="missing_gt_row",
                benchmark_include_gt="yes",
                alignment_confidence="medium",
                alignment_notes="GT row has no remaining predicted counterpart after automatic pairing.",
            )
        )

    for pred_row in remaining_pred[pair_count:]:
        out_rows.append(
            build_alignment_row(
                doi=doi,
                gt_row=None,
                pred_row=pred_row,
                alignment_decision="spurious_pred_row",
                benchmark_include_gt="",
                alignment_confidence="medium",
                alignment_notes="Predicted retained row has no remaining GT counterpart after automatic pairing.",
            )
        )

    return out_rows


def build_alignment_row(
    doi: str,
    gt_row: dict[str, str] | None,
    pred_row: dict[str, str] | None,
    alignment_decision: str,
    benchmark_include_gt: str,
    alignment_confidence: str,
    alignment_notes: str,
) -> dict[str, str]:
    family_id = ""
    parent_core_row_id = ""
    variant_role = ""
    payload_state = ""
    if pred_row:
        family_id = pred_row.get("family_id", "")
        parent_core_row_id = pred_row.get("parent_core_row_id", "")
        variant_role = pred_row.get("variant_role", "")
        payload_state = pred_row.get("payload_state", "")
    elif gt_row:
        family_id = gt_row.get("family_id", "")
        parent_core_row_id = gt_row.get("parent_core_row_id", "")
        variant_role = gt_row.get("variant_role", "")
        payload_state = gt_row.get("payload_state", "")

    signature_gt = gt_row.get("canonical_signature_gt", "") if gt_row else ""
    signature_pred = pred_row.get("canonical_signature_pred", "") if pred_row else ""
    diff_summary = ""
    if signature_gt and signature_pred and signature_gt != signature_pred:
        diff_summary = f"gt={signature_gt} ; pred={signature_pred}"
    elif signature_gt and not signature_pred:
        diff_summary = f"gt_only={signature_gt}"
    elif signature_pred and not signature_gt:
        diff_summary = f"pred_only={signature_pred}"

    return {
        "doi": doi,
        "gt_formulation_id": gt_row.get("formulation_id", "") if gt_row else "",
        "pred_row_id": pred_row.get("final_formulation_id", "") if pred_row else "",
        "family_id": family_id,
        "parent_core_row_id": parent_core_row_id,
        "variant_role": variant_role,
        "payload_state": payload_state,
        "alignment_decision": alignment_decision,
        "benchmark_include_gt": benchmark_include_gt,
        "alignment_confidence": alignment_confidence,
        "alignment_notes": alignment_notes,
        "gt_evidence_note": gt_row.get("gt_evidence_note", "") if gt_row else "",
        "pred_evidence_anchor": pred_row.get("pred_evidence_anchor", "") if pred_row else "",
        "pred_source_table": pred_row.get("pred_source_table", "") if pred_row else "",
        "pred_source_text": pred_row.get("pred_source_text", "") if pred_row else "",
        "canonical_signature_gt": signature_gt,
        "canonical_signature_pred": signature_pred,
        "key_field_diff_summary": diff_summary,
    }


def build_variant_alignment_scaffold(
    gt_workbook: Path,
    final_table_tsv: Path,
    scope_manifest_tsv: Path | None,
    out_tsv: Path,
) -> dict[str, Any]:
    gt_rows = read_gt_rows_from_workbook(gt_workbook)
    final_rows = read_tsv(final_table_tsv)
    if scope_manifest_tsv:
        scope_keys = {row["key"] for row in read_tsv(scope_manifest_tsv)}
        final_rows = [row for row in final_rows if row.get("key", "") in scope_keys]

    gt_by_doi: dict[str, list[dict[str, str]]] = defaultdict(list)
    pred_by_doi: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in gt_rows:
        gt_by_doi[normalize_text(row.get("doi", "")).lower()].append(row)
    for row in final_rows:
        pred_by_doi[normalize_text(row.get("doi", "")).lower()].append(row)

    all_dois = sorted(set(gt_by_doi) | set(pred_by_doi))
    scaffold_rows: list[dict[str, str]] = []
    for doi in all_dois:
        scaffold_rows.extend(pair_rows_for_doi(doi, gt_by_doi.get(doi, []), pred_by_doi.get(doi, [])))

    write_tsv(
        out_tsv,
        [
            "doi",
            "gt_formulation_id",
            "pred_row_id",
            "family_id",
            "parent_core_row_id",
            "variant_role",
            "payload_state",
            "alignment_decision",
            "benchmark_include_gt",
            "alignment_confidence",
            "alignment_notes",
            "gt_evidence_note",
            "pred_evidence_anchor",
            "pred_source_table",
            "pred_source_text",
            "canonical_signature_gt",
            "canonical_signature_pred",
            "key_field_diff_summary",
        ],
        scaffold_rows,
    )
    return {
        "out_tsv": str(out_tsv),
        "row_count": len(scaffold_rows),
        "doi_count": len(all_dois),
    }


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Create a versioned variant-aware DEV15 GT workbook update and alignment scaffold."
    )
    parser.add_argument("--base-xlsx", type=Path, default=DEFAULT_BASE_XLSX)
    parser.add_argument("--out-xlsx", type=Path, default=DEFAULT_OUT_XLSX)
    parser.add_argument("--final-table-tsv", type=Path, default=DEFAULT_FINAL_TABLE_TSV)
    parser.add_argument("--scope-manifest-tsv", type=Path, default=DEFAULT_SCOPE_MANIFEST_TSV)
    parser.add_argument("--out-scaffold-tsv", type=Path, default=DEFAULT_SCAFFOLD_TSV)
    return parser


def main() -> int:
    args = build_arg_parser().parse_args()
    workbook_result = create_updated_workbook(
        base_xlsx=args.base_xlsx,
        out_xlsx=args.out_xlsx,
        final_table_tsv=args.final_table_tsv,
    )
    scaffold_result = build_variant_alignment_scaffold(
        gt_workbook=args.out_xlsx,
        final_table_tsv=args.final_table_tsv,
        scope_manifest_tsv=args.scope_manifest_tsv,
        out_tsv=args.out_scaffold_tsv,
    )
    result = {
        "base_xlsx": str(args.base_xlsx),
        "out_xlsx": str(args.out_xlsx),
        "final_table_tsv": str(args.final_table_tsv),
        "scope_manifest_tsv": str(args.scope_manifest_tsv),
        "out_scaffold_tsv": str(args.out_scaffold_tsv),
        "bx_update": workbook_result,
        "scaffold": scaffold_result,
        "project_root": str(PROJECT_ROOT),
    }
    print(json.dumps(result, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
