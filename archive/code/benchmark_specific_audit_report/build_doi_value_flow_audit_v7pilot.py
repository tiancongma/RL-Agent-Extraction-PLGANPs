#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from src.archive_methods.older_weak_label_pilot_variants.auto_extract_weak_labels_v7pilot_r3_fixflat import (
    LLM_PROMPT_TEMPLATE,
    call_gemini,
    ensure_genai,
    normalize_doi,
    safe_json_load,
)


TARGET_FIELDS = [
    "la_ga_ratio",
    "plga_mw_kDa",
    "polymer_name_raw",
    "organic_solvent",
    "encapsulation_efficiency_percent",
    "surfactant_name",
    "surfactant_concentration_text",
]


def _norm(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    return " ".join(str(v).strip().split())


def _nonnull_value(v: Any) -> str:
    # value_raw is intentionally taken from `value` only (not value_text),
    # to diagnose null-value vs quote-presence gaps.
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    return _norm(v)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build DOI value-flow audit for v7 pilot (diagnostic only).")
    p.add_argument(
        "--run-dir",
        default="data/results/run_20260306_1552_v7pilot3r3fixflat_dev/weak_labels_v7pilot_r3_fixflat",
    )
    p.add_argument("--doi", default="10.1016/j.ejpb.2004.09.002")
    p.add_argument(
        "--manifest-tsv",
        default="data/cleaned/goren_2025/index/splits/dev_manifest_v7pilot3_2026-03-06.tsv",
    )
    p.add_argument("--model", default="gemini-2.5-flash")
    p.add_argument("--max-chars", type=int, default=50000)
    p.add_argument("--retries", type=int, default=1)
    p.add_argument("--sleep", type=float, default=0.4)
    p.add_argument(
        "--raw-out",
        default="data/diagnostics/raw_llm_response_10.1016.j.ejpb.2004.09.002.txt",
    )
    p.add_argument(
        "--out-xlsx",
        default="data/cleaned/labels/manual/doi_10.1016_j.ejpb.2004.09.002_value_flow_audit.xlsx",
    )
    return p.parse_args()


def resolve_io(run_dir: Path) -> Tuple[Path, Path]:
    jsonl_candidates = [
        run_dir / "weak_labels__v7pilot.jsonl",
        run_dir / "weak_labels__v7pilot_r3_fixparse.jsonl",
        run_dir / "weak_labels__v7pilot_r3_fixflat.jsonl",
        run_dir / "weak_labels__v7pilot_r3.jsonl",
    ]
    tsv_candidates = [
        run_dir / "weak_labels__v7pilot.tsv",
        run_dir / "weak_labels__v7pilot_r3_fixparse.tsv",
        run_dir / "weak_labels__v7pilot_r3_fixflat.tsv",
        run_dir / "weak_labels__v7pilot_r3.tsv",
    ]
    jsonl = next((p for p in jsonl_candidates if p.exists()), None)
    tsv = next((p for p in tsv_candidates if p.exists()), None)
    if not jsonl or not tsv:
        raise FileNotFoundError(f"Could not find jsonl/tsv under {run_dir}")
    return jsonl, tsv


def find_text_path(manifest_tsv: Path, target_doi: str) -> Path:
    df = pd.read_csv(manifest_tsv, sep="\t", dtype=str).fillna("")
    df["doi_norm"] = df["doi"].map(normalize_doi)
    m = df[df["doi_norm"] == target_doi]
    if m.empty:
        raise ValueError(f"DOI not found in manifest: {target_doi}")
    p = Path(str(m.iloc[0]["text_path"]).replace("\\", "/"))
    if not p.exists():
        raise FileNotFoundError(f"text_path missing: {p}")
    return p


def get_field_obj(form: Dict[str, Any], field_name: str) -> Dict[str, Any]:
    fields = form.get("fields", [])
    if field_name == "polymer_name_raw":
        pmw = get_field_obj(form, "plga_mw_kDa")
        if pmw:
            return {
                "value": pmw.get("value_text", ""),
                "scope": pmw.get("scope", ""),
                "membership_confidence": pmw.get("membership_confidence", ""),
                "evidence_region_type": pmw.get("evidence_region_type", ""),
                "evidence_quote": pmw.get("value_text", ""),
            }
        return {}
    if isinstance(fields, list):
        for it in fields:
            if isinstance(it, dict) and _norm(it.get("field_name")).lower() == field_name.lower():
                return {
                    "value": it.get("value"),
                    "scope": it.get("scope", ""),
                    "membership_confidence": it.get("membership_confidence", ""),
                    "evidence_region_type": it.get("evidence_region_type", ""),
                    "evidence_quote": it.get("value_text", ""),
                }
    elif isinstance(fields, dict):
        it = fields.get(field_name)
        if isinstance(it, dict):
            return {
                "value": it.get("value"),
                "scope": it.get("scope", ""),
                "membership_confidence": it.get("membership_confidence", ""),
                "evidence_region_type": it.get("evidence_region_type", ""),
                "evidence_quote": it.get("value_text", ""),
            }
        if it is not None:
            return {"value": it, "scope": "", "membership_confidence": "", "evidence_region_type": "", "evidence_quote": ""}
    return {}


def tsv_cols(field_name: str) -> Tuple[str, str]:
    if field_name == "polymer_name_raw":
        return "plga_mw_kDa_value_text", "plga_mw_kDa_scope"
    return f"{field_name}_value", f"{field_name}_scope"


def style_xlsx(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max(12, max_len + 2), 70)
    wb.save(path)


def main() -> None:
    args = parse_args()
    run_dir = Path(args.run_dir)
    out_xlsx = Path(args.out_xlsx)
    raw_out = Path(args.raw_out)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    raw_out.parent.mkdir(parents=True, exist_ok=True)
    target_doi = normalize_doi(args.doi)

    jsonl_path, tsv_path = resolve_io(run_dir)

    # Step 1: raw LLM response reconstruction (not stored in run artifacts).
    text_path = find_text_path(Path(args.manifest_tsv), target_doi)
    txt = text_path.read_text(encoding="utf-8", errors="ignore")
    if args.max_chars > 0 and len(txt) > args.max_chars:
        txt = txt[: args.max_chars]
    if raw_out.exists():
        raw_response = raw_out.read_text(encoding="utf-8", errors="ignore")
    else:
        ensure_genai(args.model)
        raw_response = call_gemini(args.model, LLM_PROMPT_TEMPLATE + txt, args.retries, args.sleep)
        raw_out.write_text(raw_response, encoding="utf-8")
    raw_obj = safe_json_load(raw_response)

    # Build replay map (raw-response detectable values) by (formulation_id, field_name).
    replay_map: Dict[Tuple[str, str], str] = {}
    replay_order_map: Dict[Tuple[int, str], str] = {}
    replay_forms = raw_obj.get("formulations", [])
    if not isinstance(replay_forms, list):
        replay_forms = []
    for ridx, fm in enumerate(replay_forms, start=1):
        if not isinstance(fm, dict):
            continue
        fid = _norm(fm.get("formulation_id", fm.get("id", "")))
        for field in TARGET_FIELDS:
            fo = get_field_obj(fm, field)
            detected = _nonnull_value(fo.get("value"))
            if detected == "":
                detected = _norm(fo.get("evidence_quote", ""))
            replay_map[(fid, field)] = detected
            replay_order_map[(ridx, field)] = detected

    # Step 2 + Step 3: parsed object and JSONL layers from run JSONL.
    # Parsed object here is the run-level pre-write equivalent because no mutation
    # occurs between parsed `data["formulations"]` and JSONL write in the script.
    base_rows: List[Dict[str, Any]] = []
    with jsonl_path.open("r", encoding="utf-8", errors="replace") as f:
        for line in f:
            if not line.strip():
                continue
            obj = json.loads(line)
            if normalize_doi(obj.get("doi", "")) != target_doi:
                continue
            forms = obj.get("formulations", [])
            if not isinstance(forms, list):
                forms = []
            for run_idx, fm in enumerate(forms, start=1):
                if not isinstance(fm, dict):
                    continue
                fid = _norm(fm.get("formulation_id", fm.get("id", "")))
                for field in TARGET_FIELDS:
                    fo = get_field_obj(fm, field)
                    json_v = _nonnull_value(fo.get("value"))
                    raw_val = replay_map.get((fid, field), "")
                    if raw_val == "":
                        raw_val = replay_order_map.get((run_idx, field), "")
                    base_rows.append(
                        {
                            "doi": target_doi,
                            "formulation_id": fid,
                            "field_name": field,
                            "raw_response_value": raw_val,
                            "parsed_object_value_raw": json_v,
                            "jsonl_value_raw": json_v,
                            "parsed_scope": _norm(fo.get("scope", "")),
                            "jsonl_scope": _norm(fo.get("scope", "")),
                            "evidence_quote": _norm(fo.get("evidence_quote", "")),
                            "evidence_region_type": _norm(fo.get("evidence_region_type", "")),
                        }
                    )
    value_flow = pd.DataFrame(base_rows)

    # Step 4: TSV layer
    tsv_df = pd.read_csv(tsv_path, sep="\t", dtype=str).fillna("")
    tsv_df["doi_norm"] = tsv_df["doi"].map(normalize_doi)
    tsv_df = tsv_df[tsv_df["doi_norm"] == target_doi].copy()
    tsv_map: Dict[Tuple[str, str], Dict[str, str]] = {}
    for _, r in tsv_df.iterrows():
        fid = _norm(r.get("formulation_id", ""))
        for field in TARGET_FIELDS:
            vcol, scol = tsv_cols(field)
            tsv_map[(fid, field)] = {"tsv_value": _norm(r.get(vcol, "")), "tsv_scope": _norm(r.get(scol, ""))}

    value_flow["tsv_value"] = value_flow.apply(
        lambda x: tsv_map.get((_norm(x["formulation_id"]), _norm(x["field_name"])), {}).get("tsv_value", ""),
        axis=1,
    )
    value_flow["tsv_scope"] = value_flow.apply(
        lambda x: tsv_map.get((_norm(x["formulation_id"]), _norm(x["field_name"])), {}).get("tsv_scope", ""),
        axis=1,
    )

    # Step 5: missing-stage classifier
    def missing_stage(row: pd.Series) -> str:
        if _norm(row["raw_response_value"]) == "":
            return "missing_in_raw_response"
        if _norm(row["parsed_object_value_raw"]) == "":
            return "missing_in_parsed_object"
        if _norm(row["jsonl_value_raw"]) == "":
            return "missing_in_jsonl"
        if _norm(row["tsv_value"]) == "":
            return "missing_in_tsv"
        return "present_everywhere"

    value_flow["value_missing_stage"] = value_flow.apply(missing_stage, axis=1)
    value_flow = value_flow[
        [
            "doi",
            "formulation_id",
            "field_name",
            "raw_response_value",
            "parsed_object_value_raw",
            "jsonl_value_raw",
            "tsv_value",
            "parsed_scope",
            "jsonl_scope",
            "tsv_scope",
            "evidence_quote",
            "evidence_region_type",
            "value_missing_stage",
        ]
    ].sort_values(["formulation_id", "field_name"])

    # Step 6: summary
    stage_counts = value_flow["value_missing_stage"].value_counts().to_dict()
    summary_rows = [
        {"metric": "total_field_objects_inspected", "value": int(len(value_flow))},
        {"metric": "missing_in_raw_response", "value": int(stage_counts.get("missing_in_raw_response", 0))},
        {"metric": "missing_in_parsed_object", "value": int(stage_counts.get("missing_in_parsed_object", 0))},
        {"metric": "missing_in_jsonl", "value": int(stage_counts.get("missing_in_jsonl", 0))},
        {"metric": "missing_in_tsv", "value": int(stage_counts.get("missing_in_tsv", 0))},
        {"metric": "present_everywhere", "value": int(stage_counts.get("present_everywhere", 0))},
    ]
    summary_df = pd.DataFrame(summary_rows)
    top_problem = (
        value_flow[value_flow["value_missing_stage"] != "present_everywhere"]["field_name"].value_counts().reset_index()
    )
    top_problem.columns = ["field_name", "problem_count"]

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xw:
        value_flow.to_excel(xw, sheet_name="VALUE_FLOW", index=False)
        summary_df.to_excel(xw, sheet_name="SUMMARY", index=False)
        if not top_problem.empty:
            top_problem.to_excel(xw, sheet_name="SUMMARY", index=False, startrow=len(summary_df) + 3)
            ws = xw.book["SUMMARY"]
            ws.cell(row=len(summary_df) + 3, column=1, value="TOP_PROBLEMATIC_FIELDS")

    style_xlsx(out_xlsx)

    first5 = value_flow[value_flow["value_missing_stage"] != "present_everywhere"].head(5)
    print(f"workbook_path={out_xlsx.resolve()}")
    print(f"total_field_objects_analyzed={len(value_flow)}")
    print(f"missing_in_raw_response={stage_counts.get('missing_in_raw_response', 0)}")
    print(f"missing_in_parsed_object={stage_counts.get('missing_in_parsed_object', 0)}")
    print(f"missing_in_jsonl={stage_counts.get('missing_in_jsonl', 0)}")
    print(f"missing_in_tsv={stage_counts.get('missing_in_tsv', 0)}")
    print(f"present_everywhere={stage_counts.get('present_everywhere', 0)}")
    print("first_5_problematic_rows=" + json.dumps(first5.to_dict(orient="records"), ensure_ascii=False))


if __name__ == "__main__":
    main()


