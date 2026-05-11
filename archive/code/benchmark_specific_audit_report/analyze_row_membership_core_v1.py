#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


DEFAULT_RUN_ID = "run_20260219_1623_780eb83_goren18_weaklabels_v1"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analyze DOI-level multiplicity differences after collapsing to formulation_core."
    )
    parser.add_argument("--run-id", default=DEFAULT_RUN_ID)
    parser.add_argument(
        "--core-projected-tsv",
        default="",
        help="Default: data/results/<run_id>/benchmark_goren_2025/core_eval_v1/core_projected_to_curated.tsv",
    )
    parser.add_argument(
        "--curated-tsv",
        default="data/benchmark/goren_2025/overlap_goren18_v1/goren18_curated_overlap_subset.tsv",
    )
    parser.add_argument(
        "--core-alignment-tsv",
        default="",
        help="Default: data/results/<run_id>/benchmark_goren_2025/core_eval_v1/core_alignment_rows.tsv (optional)",
    )
    parser.add_argument(
        "--out-dir",
        default="",
        help="Default: data/results/<run_id>/benchmark_goren_2025/core_eval_v1",
    )
    return parser.parse_args()


def normalize_doi(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"^doi\s*:\s*", "", s)
    s = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", s)
    s = re.sub(r"^doi\.org/", "", s)
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[)\],.;]+$", "", s)
    return s


def format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    for col in ws.columns:
        name = str(col[0].value or "")
        letter = col[0].column_letter
        if name in {"reference_normalized_doi", "doi_norm"}:
            ws.column_dimensions[letter].width = 42
        elif name in {"difference_flag"}:
            ws.column_dimensions[letter].width = 18
        elif name in {"notes"}:
            ws.column_dimensions[letter].width = 52
        else:
            ws.column_dimensions[letter].width = 18


def main() -> None:
    args = parse_args()
    run_id = args.run_id
    out_dir = Path(args.out_dir) if args.out_dir else Path(f"data/results/{run_id}/benchmark_goren_2025/core_eval_v1")
    core_projected = Path(args.core_projected_tsv) if args.core_projected_tsv else out_dir / "core_projected_to_curated.tsv"
    curated_tsv = Path(args.curated_tsv)
    core_alignment = Path(args.core_alignment_tsv) if args.core_alignment_tsv else out_dir / "core_alignment_rows.tsv"

    missing = [str(p) for p in [core_projected, curated_tsv] if not p.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required input file(s): {missing}")

    out_dir.mkdir(parents=True, exist_ok=True)

    core_df = pd.read_csv(core_projected, sep="\t", dtype=str).fillna("")
    curated_df = pd.read_csv(curated_tsv, sep="\t", dtype=str).fillna("")

    if "reference" not in core_df.columns:
        raise RuntimeError("core_projected_to_curated.tsv missing required DOI/reference column: reference")
    if "doi_norm" in curated_df.columns:
        curated_doi = curated_df["doi_norm"].map(normalize_doi)
    elif "reference" in curated_df.columns:
        curated_doi = curated_df["reference"].map(normalize_doi)
    else:
        raise RuntimeError("Curated overlap TSV missing DOI column (doi_norm/reference).")

    core_df["reference_normalized_doi"] = core_df["reference"].map(normalize_doi)
    curated_df["reference_normalized_doi"] = curated_doi

    core_counts = (
        core_df.groupby("reference_normalized_doi", dropna=False).size().reset_index(name="core_projected_row_count_per_doi")
    )
    curated_counts = (
        curated_df.groupby("reference_normalized_doi", dropna=False).size().reset_index(name="curated_row_count_per_doi")
    )
    doi_rows = core_counts.merge(curated_counts, on="reference_normalized_doi", how="outer").fillna(0)
    doi_rows["core_projected_row_count_per_doi"] = doi_rows["core_projected_row_count_per_doi"].astype(int)
    doi_rows["curated_row_count_per_doi"] = doi_rows["curated_row_count_per_doi"].astype(int)
    doi_rows["difference"] = doi_rows["core_projected_row_count_per_doi"] - doi_rows["curated_row_count_per_doi"]

    def _flag(v: int) -> str:
        if v > 0:
            return "core_gt_curated"
        if v < 0:
            return "core_lt_curated"
        return "equal"

    doi_rows["difference_flag"] = doi_rows["difference"].map(_flag)
    doi_rows["abs_difference"] = doi_rows["difference"].abs()
    doi_rows["notes"] = doi_rows.apply(
        lambda r: "collapsed core count greater than curated row multiplicity"
        if r["difference_flag"] == "core_gt_curated"
        else (
            "collapsed core count lower than curated row multiplicity"
            if r["difference_flag"] == "core_lt_curated"
            else ""
        ),
        axis=1,
    )
    doi_rows = doi_rows.sort_values(["abs_difference", "reference_normalized_doi"], ascending=[False, True]).reset_index(drop=True)

    summary = {
        "selected_join_key_name": "reference_normalized_doi",
        "core_rows_total": int(len(core_df)),
        "curated_rows_total": int(len(curated_df)),
        "doi_total": int(len(doi_rows)),
        "core_gt_curated_doi_count": int((doi_rows["difference_flag"] == "core_gt_curated").sum()),
        "core_lt_curated_doi_count": int((doi_rows["difference_flag"] == "core_lt_curated").sum()),
        "equal_doi_count": int((doi_rows["difference_flag"] == "equal").sum()),
        "top_abs_difference_dois": doi_rows.head(10)[
            ["reference_normalized_doi", "core_projected_row_count_per_doi", "curated_row_count_per_doi", "difference"]
        ].to_dict(orient="records"),
    }

    alignment_crosstab = pd.DataFrame()
    if core_alignment.exists():
        ar = pd.read_csv(core_alignment, sep="\t", dtype=str).fillna("")
        if "doi_norm" in ar.columns:
            ar["reference_normalized_doi"] = ar["doi_norm"].map(normalize_doi)
            flag_map = dict(zip(doi_rows["reference_normalized_doi"], doi_rows["difference_flag"]))
            ar["difference_flag"] = ar["reference_normalized_doi"].map(flag_map).fillna("unknown")
            if "mode" in ar.columns and "failure_type" in ar.columns:
                alignment_crosstab = (
                    ar.groupby(["difference_flag", "mode", "failure_type"], dropna=False)
                    .size()
                    .reset_index(name="count")
                    .sort_values("count", ascending=False)
                )
                summary["alignment_crosstab_rows"] = int(len(alignment_crosstab))

    summary_path = out_dir / "core_row_membership_summary.json"
    rows_path = out_dir / "core_row_membership_rows.tsv"
    xlsx_path = out_dir / "core_row_membership_analysis.xlsx"
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    doi_rows.to_csv(rows_path, sep="\t", index=False)

    summary_sheet = pd.DataFrame([{"metric": k, "value": json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v} for k, v in summary.items() if k != "top_abs_difference_dois"])
    top_sheet = pd.DataFrame(summary["top_abs_difference_dois"])
    core_gt = doi_rows[doi_rows["difference_flag"] == "core_gt_curated"].copy()
    core_lt = doi_rows[doi_rows["difference_flag"] == "core_lt_curated"].copy()
    equal = doi_rows[doi_rows["difference_flag"] == "equal"].copy()

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary_sheet.to_excel(writer, sheet_name="summary", index=False, startrow=0)
        top_sheet.to_excel(writer, sheet_name="summary", index=False, startrow=len(summary_sheet) + 3)
        doi_rows.to_excel(writer, sheet_name="doi_counts", index=False)
        core_gt.to_excel(writer, sheet_name="core_gt_curated", index=False)
        core_lt.to_excel(writer, sheet_name="core_lt_curated", index=False)
        equal.to_excel(writer, sheet_name="equal", index=False)
        if not alignment_crosstab.empty:
            alignment_crosstab.to_excel(writer, sheet_name="alignment_crosstab", index=False)

    wb = load_workbook(xlsx_path)
    for ws_name in wb.sheetnames:
        format_sheet(wb[ws_name])
    wb.save(xlsx_path)

    print(f"doi_total={summary['doi_total']}")
    print(f"core_gt_curated_doi_count={summary['core_gt_curated_doi_count']}")
    print(f"core_lt_curated_doi_count={summary['core_lt_curated_doi_count']}")
    print(f"equal_doi_count={summary['equal_doi_count']}")
    print(f"output_summary={summary_path}")
    print(f"output_rows={rows_path}")
    print(f"output_xlsx={xlsx_path}")


if __name__ == "__main__":
    main()
