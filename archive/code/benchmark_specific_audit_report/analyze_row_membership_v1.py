#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


DEFAULT_RUN_ID = "run_20260219_1623_780eb83_goren18_weaklabels_v1"
DEFAULT_CURATED = "data/benchmark/goren_2025/overlap_goren18_v1/goren18_curated_overlap_subset.tsv"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analyze projected-vs-curated row-set membership for benchmark diagnostics."
    )
    parser.add_argument("--run-id", default=DEFAULT_RUN_ID)
    parser.add_argument(
        "--projected-tsv",
        default="",
        help="Default: data/results/<run_id>/benchmark_goren_2025/projected_to_curated.tsv",
    )
    parser.add_argument("--curated-tsv", default=DEFAULT_CURATED)
    parser.add_argument(
        "--alignment-tsv",
        default="",
        help="Default: data/results/<run_id>/benchmark_goren_2025/alignment_rows.tsv",
    )
    parser.add_argument(
        "--out-dir",
        default="",
        help="Default: data/results/<run_id>/benchmark_goren_2025/",
    )
    return parser.parse_args()


def normalize_doi(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"^doi:\s*", "", s)
    s = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", s)
    s = re.sub(r"^doi\.org/", "", s)
    s = re.sub(r"[)\],.;]+$", "", s)
    return s


def looks_like_doi_series(series: pd.Series) -> bool:
    sample = series.astype(str).map(normalize_doi)
    sample = sample[sample != ""].head(50)
    if sample.empty:
        return False
    doi_like = sample.str.match(r"^10\.\d{4,9}/\S+$", na=False)
    return float(doi_like.mean()) >= 0.7


def select_join_key(projected: pd.DataFrame, curated: pd.DataFrame) -> tuple[str, pd.Series, pd.Series]:
    if "doi_norm" in projected.columns and "doi_norm" in curated.columns:
        return (
            "doi_norm",
            projected["doi_norm"].map(normalize_doi),
            curated["doi_norm"].map(normalize_doi),
        )
    if "reference" in projected.columns and "reference" in curated.columns:
        p_ref = projected["reference"].map(normalize_doi)
        c_ref = curated["reference"].map(normalize_doi)
        if looks_like_doi_series(p_ref) and looks_like_doi_series(c_ref):
            return ("reference_normalized_doi", p_ref, c_ref)
    common_cols = [c for c in projected.columns if c in curated.columns]
    for c in common_cols:
        if c in {"small_molecule_name", "solvent", "surfactant_name", "EE", "LC"}:
            continue
        p = projected[c].astype(str).str.strip()
        q = curated[c].astype(str).str.strip()
        if (p != "").any() and (q != "").any():
            return (c, p, q)
    raise RuntimeError(
        "No valid shared join key detected. Please provide explicit join columns."
    )


def format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column_letter in {"F", "G", "H", "I", "J"}:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in ws.columns:
        letter = col[0].column_letter
        name = str(col[0].value or "")
        if name in {"join_key_value", "notes", "failure_type"}:
            ws.column_dimensions[letter].width = 52
        elif name in {"membership_category", "mode"}:
            ws.column_dimensions[letter].width = 20
        else:
            ws.column_dimensions[letter].width = 18


def main() -> None:
    args = parse_args()
    run_id = args.run_id
    out_dir = Path(args.out_dir) if args.out_dir else Path(f"data/results/{run_id}/benchmark_goren_2025")
    projected_path = Path(args.projected_tsv) if args.projected_tsv else out_dir / "projected_to_curated.tsv"
    curated_path = Path(args.curated_tsv)
    alignment_path = Path(args.alignment_tsv) if args.alignment_tsv else out_dir / "alignment_rows.tsv"

    required = [projected_path, curated_path]
    missing = [str(p) for p in required if not p.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required input file(s): {missing}")

    projected = pd.read_csv(projected_path, sep="\t", dtype=str).fillna("")
    curated = pd.read_csv(curated_path, sep="\t", dtype=str).fillna("")

    join_key_name, projected_key, curated_key = select_join_key(projected, curated)
    projected = projected.copy()
    curated = curated.copy()
    projected["_join_key"] = projected_key
    curated["_join_key"] = curated_key

    p_counts = projected.groupby("_join_key", dropna=False).size().rename("projected_row_count")
    c_counts = curated.groupby("_join_key", dropna=False).size().rename("curated_row_count")
    rows = pd.concat([p_counts, c_counts], axis=1).fillna(0).reset_index().rename(columns={"_join_key": "join_key_value"})
    rows["projected_row_count"] = rows["projected_row_count"].astype(int)
    rows["curated_row_count"] = rows["curated_row_count"].astype(int)

    def classify(r: pd.Series) -> str:
        if r["projected_row_count"] > 0 and r["curated_row_count"] > 0:
            return "intersection"
        if r["projected_row_count"] > 0 and r["curated_row_count"] == 0:
            return "projected_only"
        return "curated_only"

    rows["membership_category"] = rows.apply(classify, axis=1)
    rows["example_key"] = ""
    rows["example_formulation_id"] = ""
    rows["doi_norm"] = rows["join_key_value"] if "doi" in join_key_name else ""
    rows["notes"] = rows.apply(
        lambda r: "multiple formulations per DOI"
        if (r["projected_row_count"] > 1 or r["curated_row_count"] > 1)
        else "",
        axis=1,
    )

    rows = rows[
        [
            "join_key_value",
            "membership_category",
            "projected_row_count",
            "curated_row_count",
            "example_key",
            "example_formulation_id",
            "doi_norm",
            "notes",
        ]
    ]

    projected_only = rows[rows["membership_category"] == "projected_only"].copy()
    curated_only = rows[rows["membership_category"] == "curated_only"].copy()
    intersection = rows[rows["membership_category"] == "intersection"].copy()

    summary = {
        "selected_join_key_name": join_key_name,
        "counts": {
            "projected_rows": int(len(projected)),
            "curated_rows": int(len(curated)),
            "intersection": int(intersection["projected_row_count"].sum()),
            "projected_only": int(projected_only["projected_row_count"].sum()),
            "curated_only": int(curated_only["curated_row_count"].sum()),
        },
        "unique_key_counts": {
            "projected_unique_keys": int((p_counts.index.to_series().astype(str) != "").sum()),
            "curated_unique_keys": int((c_counts.index.to_series().astype(str) != "").sum()),
            "intersection_unique_keys": int(len(intersection)),
            "projected_only_unique_keys": int(len(projected_only)),
            "curated_only_unique_keys": int(len(curated_only)),
        },
        "top_projected_only_join_keys": projected_only.sort_values(
            "projected_row_count", ascending=False
        )
        .head(20)[["join_key_value", "projected_row_count"]]
        .to_dict(orient="records"),
        "top_curated_only_join_keys": curated_only.sort_values(
            "curated_row_count", ascending=False
        )
        .head(20)[["join_key_value", "curated_row_count"]]
        .to_dict(orient="records"),
    }

    # Optional alignment crosstabs
    alignment_crosstab = pd.DataFrame()
    membership_by_failure = pd.DataFrame()
    membership_by_mode = pd.DataFrame()
    if alignment_path.exists():
        align = pd.read_csv(alignment_path, sep="\t", dtype=str).fillna("")
        if "doi_norm" in align.columns:
            align["_join_key"] = align["doi_norm"].map(normalize_doi)
        elif "reference" in align.columns:
            align["_join_key"] = align["reference"].map(normalize_doi)
        else:
            align["_join_key"] = ""
        map_cat = rows.set_index("join_key_value")["membership_category"].to_dict()
        align["membership_category"] = align["_join_key"].map(map_cat).fillna("unknown")

        if "failure_type" in align.columns:
            membership_by_failure = (
                align.groupby(["membership_category", "failure_type"], dropna=False)
                .size()
                .reset_index(name="count")
                .sort_values(["membership_category", "count"], ascending=[True, False])
            )
        if "mode" in align.columns:
            membership_by_mode = (
                align.groupby(["membership_category", "mode"], dropna=False)
                .size()
                .reset_index(name="count")
                .sort_values(["membership_category", "count"], ascending=[True, False])
            )

        parts = []
        if not membership_by_failure.empty:
            tmp = membership_by_failure.copy()
            tmp["crosstab_type"] = "membership_vs_failure_type"
            parts.append(tmp[["crosstab_type", "membership_category", "failure_type", "count"]])
        if not membership_by_mode.empty:
            tmp = membership_by_mode.copy()
            tmp["failure_type"] = tmp["mode"]
            tmp["crosstab_type"] = "membership_vs_mode"
            parts.append(tmp[["crosstab_type", "membership_category", "failure_type", "count"]])
        if parts:
            alignment_crosstab = pd.concat(parts, ignore_index=True)

        summary["alignment_crosstabs"] = {
            "membership_vs_failure_type": membership_by_failure.to_dict(orient="records"),
            "membership_vs_mode": membership_by_mode.to_dict(orient="records"),
        }

    out_dir.mkdir(parents=True, exist_ok=True)
    summary_json_path = out_dir / "row_membership_summary.json"
    rows_tsv_path = out_dir / "row_membership_rows.tsv"
    xlsx_path = out_dir / "row_membership_analysis.xlsx"

    summary_json_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    rows.to_csv(rows_tsv_path, sep="\t", index=False)

    # Build summary sheet data
    summary_rows = [
        {"metric": "selected_join_key_name", "value": join_key_name},
        {"metric": "projected_rows", "value": summary["counts"]["projected_rows"]},
        {"metric": "curated_rows", "value": summary["counts"]["curated_rows"]},
        {"metric": "intersection_rows", "value": summary["counts"]["intersection"]},
        {"metric": "projected_only_rows", "value": summary["counts"]["projected_only"]},
        {"metric": "curated_only_rows", "value": summary["counts"]["curated_only"]},
        {"metric": "projected_unique_keys", "value": summary["unique_key_counts"]["projected_unique_keys"]},
        {"metric": "curated_unique_keys", "value": summary["unique_key_counts"]["curated_unique_keys"]},
        {"metric": "intersection_unique_keys", "value": summary["unique_key_counts"]["intersection_unique_keys"]},
        {"metric": "projected_only_unique_keys", "value": summary["unique_key_counts"]["projected_only_unique_keys"]},
        {"metric": "curated_only_unique_keys", "value": summary["unique_key_counts"]["curated_only_unique_keys"]},
    ]
    top_proj = pd.DataFrame(summary["top_projected_only_join_keys"])
    if not top_proj.empty:
        top_proj = top_proj.rename(columns={"projected_row_count": "count"})
        top_proj["list_name"] = "top_projected_only"
    top_cur = pd.DataFrame(summary["top_curated_only_join_keys"])
    if not top_cur.empty:
        top_cur = top_cur.rename(columns={"curated_row_count": "count"})
        top_cur["list_name"] = "top_curated_only"
    top_df = pd.concat([top_proj, top_cur], ignore_index=True) if (not top_proj.empty or not top_cur.empty) else pd.DataFrame(columns=["list_name", "join_key_value", "count"])

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="summary", index=False, startrow=0)
        top_df.to_excel(writer, sheet_name="summary", index=False, startrow=len(summary_rows) + 3)
        projected_only.to_excel(writer, sheet_name="projected_only", index=False)
        curated_only.to_excel(writer, sheet_name="curated_only", index=False)
        intersection.to_excel(writer, sheet_name="intersection", index=False)
        if not alignment_crosstab.empty:
            alignment_crosstab.to_excel(writer, sheet_name="alignment_crosstab", index=False)

    wb = load_workbook(xlsx_path)
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        format_sheet(ws)
    wb.save(xlsx_path)

    print(f"selected_join_key={join_key_name}")
    print(
        "summary_counts="
        + json.dumps(
            {
                "projected_rows": summary["counts"]["projected_rows"],
                "curated_rows": summary["counts"]["curated_rows"],
                "intersection": summary["counts"]["intersection"],
                "projected_only": summary["counts"]["projected_only"],
                "curated_only": summary["counts"]["curated_only"],
            },
            ensure_ascii=False,
        )
    )
    print(f"output_summary_json={summary_json_path}")
    print(f"output_rows_tsv={rows_tsv_path}")
    print(f"output_xlsx={xlsx_path}")


if __name__ == "__main__":
    main()
