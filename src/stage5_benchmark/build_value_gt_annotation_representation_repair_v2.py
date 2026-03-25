#!/usr/bin/env python3
from __future__ import annotations

"""
Build a preservation-aware Layer3 candidate-value workbook refresh.

This tool treats an older compact value-annotation workbook as a reviewer
surface with possible human edits. It diffs the current workbook against its
machine baseline TSV, preserves human-edited cells, refreshes unchanged machine
cells, and remaps values into a wider representation-aware schema so that
concentration, ratio, mass, and phase-volume values are no longer forced into
the same columns.
"""

import argparse
import csv
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


LEGACY_MAIN_SHEET = "value_gt_annotation"
ALLOWED_PRESERVATION_STATUSES = {
    "preserved_human_edit",
    "refreshed_machine_cell",
    "corrected_projection_error",
    "left_blank",
    "unchanged",
}

OUTPUT_COLUMNS = [
    "paper_key",
    "doi",
    "gt_formulation_id",
    "family_id",
    "parent_core",
    "variant_role",
    "benchmark_default_include",
    "formulation_label",
    "seed_pred_representative_source_formulation_id",
    "gt_row_decision",
    "boundary_anchor_label",
    "polymer_name",
    "polymer_grade",
    "polymer_mw_raw",
    "polymer_mw_kDa",
    "la_ga_ratio_raw",
    "la_ga_ratio_normalized",
    "polymer_mass_mg",
    "polymer_concentration_value",
    "polymer_concentration_unit",
    "polymer_concentration_phase",
    "polymer_to_solvent_ratio_raw",
    "polymer_to_drug_ratio_raw",
    "drug_name",
    "drug_mass_mg",
    "drug_concentration_value",
    "drug_concentration_unit",
    "drug_to_polymer_ratio_raw",
    "surfactant_name",
    "surfactant_mass_mg",
    "surfactant_concentration_value",
    "surfactant_concentration_unit",
    "stabilizer_name",
    "helper_material_name",
    "method_type",
    "solvent_name",
    "co_solvent_name",
    "W1_volume_mL",
    "O_volume_mL",
    "W2_volume_mL",
    "external_aqueous_phase_volume_mL",
    "internal_aqueous_phase_volume_mL",
    "phase_ratio_raw",
    "sonication_time_s",
    "homogenization_time_min",
    "stirring_time_h",
    "evaporation_time_h",
    "centrifugation_g",
    "centrifugation_time_min",
    "ee_percent",
    "lc_percent",
    "dl_percent",
    "particle_size_nm",
    "pdi",
    "zeta_mV",
    "value_source_type",
    "candidate_notes",
]

SCHEMA_DIFF_ROWS = [
    ("paper_key", "existing_reused", "identity", "Paper key anchor", "preserve_human_edit_if_present_else_refresh"),
    ("doi", "existing_reused", "identity", "DOI anchor", "preserve_human_edit_if_present_else_refresh"),
    ("gt_formulation_id", "existing_reused", "identity", "Authoritative GT formulation row id", "preserve_human_edit_if_present_else_refresh"),
    ("family_id", "new_added", "identity", "Variant-aware family id from authoritative boundary review", "refresh_machine_only"),
    ("parent_core", "new_added", "identity", "Parent core formulation id from authoritative boundary review", "refresh_machine_only"),
    ("variant_role", "new_added", "identity", "Variant role from authoritative boundary review", "refresh_machine_only"),
    ("benchmark_default_include", "new_added", "identity", "GT benchmark include flag", "refresh_machine_only"),
    ("formulation_label", "existing_reinterpreted", "identity", "Reviewer-facing formulation label; prefers boundary anchor label", "preserve_human_edit_if_present_else_refresh"),
    ("seed_pred_representative_source_formulation_id", "existing_reused", "identity", "Representative source formulation id", "preserve_human_edit_if_present_else_refresh"),
    ("gt_row_decision", "existing_reused", "identity", "GT keep/drop decision anchor", "preserve_human_edit_if_present_else_refresh"),
    ("boundary_anchor_label", "existing_reused", "identity", "Boundary anchor label from GT review", "preserve_human_edit_if_present_else_refresh"),
    ("polymer_name", "new_added", "polymer_identity", "Polymer name candidate", "preserve_human_edit_if_present_else_refresh"),
    ("polymer_grade", "new_added", "polymer_identity", "Polymer grade candidate", "preserve_human_edit_if_present_else_refresh"),
    ("polymer_mw_raw", "existing_reinterpreted", "polymer_identity", "Raw polymer MW candidate text", "preserve_human_edit_if_present_else_refresh"),
    ("polymer_mw_kDa", "new_added", "polymer_identity", "Polymer MW candidate in kDa when directly parseable", "preserve_human_edit_if_present_else_refresh"),
    ("la_ga_ratio_raw", "existing_reinterpreted", "polymer_identity", "Raw LA:GA ratio candidate text", "preserve_human_edit_if_present_else_refresh"),
    ("la_ga_ratio_normalized", "new_added", "polymer_identity", "Normalized LA:GA ratio when directly parseable", "preserve_human_edit_if_present_else_refresh"),
    ("polymer_mass_mg", "existing_reinterpreted", "polymer_amount", "Polymer mass candidate text only; concentration never goes here", "preserve_human_edit_if_present_else_refresh"),
    ("polymer_concentration_value", "new_added", "polymer_amount", "Polymer concentration numeric/text value", "preserve_human_edit_if_present_else_refresh"),
    ("polymer_concentration_unit", "new_added", "polymer_amount", "Polymer concentration unit", "preserve_human_edit_if_present_else_refresh"),
    ("polymer_concentration_phase", "new_added", "polymer_amount", "Phase in which polymer concentration is reported", "preserve_human_edit_if_present_else_refresh"),
    ("polymer_to_solvent_ratio_raw", "new_added", "polymer_amount", "Raw polymer-to-solvent ratio", "preserve_human_edit_if_present_else_refresh"),
    ("polymer_to_drug_ratio_raw", "new_added", "polymer_amount", "Raw polymer-to-drug ratio", "preserve_human_edit_if_present_else_refresh"),
    ("drug_name", "existing_reinterpreted", "drug", "Drug identity candidate", "preserve_human_edit_if_present_else_refresh"),
    ("drug_mass_mg", "existing_reinterpreted", "drug", "Drug mass candidate text only; concentration never goes here", "preserve_human_edit_if_present_else_refresh"),
    ("drug_concentration_value", "new_added", "drug", "Drug concentration numeric/text value", "preserve_human_edit_if_present_else_refresh"),
    ("drug_concentration_unit", "new_added", "drug", "Drug concentration unit", "preserve_human_edit_if_present_else_refresh"),
    ("drug_to_polymer_ratio_raw", "existing_reinterpreted", "drug", "Raw drug-to-polymer ratio", "preserve_human_edit_if_present_else_refresh"),
    ("surfactant_name", "existing_reinterpreted", "surfactant", "Surfactant/stabilizer identity candidate", "preserve_human_edit_if_present_else_refresh"),
    ("surfactant_mass_mg", "existing_reinterpreted", "surfactant", "Surfactant mass candidate text only", "preserve_human_edit_if_present_else_refresh"),
    ("surfactant_concentration_value", "new_added", "surfactant", "Surfactant concentration value", "preserve_human_edit_if_present_else_refresh"),
    ("surfactant_concentration_unit", "new_added", "surfactant", "Surfactant concentration unit", "preserve_human_edit_if_present_else_refresh"),
    ("stabilizer_name", "new_added", "surfactant", "Stabilizer candidate identity", "preserve_human_edit_if_present_else_refresh"),
    ("helper_material_name", "new_added", "surfactant", "Helper material candidate identity", "preserve_human_edit_if_present_else_refresh"),
    ("method_type", "new_added", "phase_process", "Preparation method candidate", "refresh_machine_only"),
    ("solvent_name", "existing_reinterpreted", "phase_process", "Primary organic solvent name", "preserve_human_edit_if_present_else_refresh"),
    ("co_solvent_name", "new_added", "phase_process", "Secondary solvent/co-solvent name", "preserve_human_edit_if_present_else_refresh"),
    ("W1_volume_mL", "new_added", "phase_process", "Internal aqueous phase volume candidate", "preserve_human_edit_if_present_else_refresh"),
    ("O_volume_mL", "new_added", "phase_process", "Organic phase volume candidate", "preserve_human_edit_if_present_else_refresh"),
    ("W2_volume_mL", "new_added", "phase_process", "External aqueous phase volume candidate when explicitly W2", "preserve_human_edit_if_present_else_refresh"),
    ("external_aqueous_phase_volume_mL", "new_added", "phase_process", "External aqueous phase volume candidate", "preserve_human_edit_if_present_else_refresh"),
    ("internal_aqueous_phase_volume_mL", "new_added", "phase_process", "Internal aqueous phase volume candidate", "preserve_human_edit_if_present_else_refresh"),
    ("phase_ratio_raw", "new_added", "phase_process", "Raw phase ratio candidate", "preserve_human_edit_if_present_else_refresh"),
    ("sonication_time_s", "new_added", "process", "Sonication time candidate", "refresh_machine_only"),
    ("homogenization_time_min", "new_added", "process", "Homogenization time candidate", "refresh_machine_only"),
    ("stirring_time_h", "new_added", "process", "Stirring time candidate", "refresh_machine_only"),
    ("evaporation_time_h", "new_added", "process", "Evaporation time candidate", "refresh_machine_only"),
    ("centrifugation_g", "new_added", "process", "Centrifugation force candidate", "refresh_machine_only"),
    ("centrifugation_time_min", "new_added", "process", "Centrifugation time candidate", "refresh_machine_only"),
    ("ee_percent", "existing_reinterpreted", "outcome", "Encapsulation efficiency candidate", "preserve_human_edit_if_present_else_refresh"),
    ("lc_percent", "existing_reinterpreted", "outcome", "Loading content candidate", "preserve_human_edit_if_present_else_refresh"),
    ("dl_percent", "new_added", "outcome", "Drug loading candidate", "preserve_human_edit_if_present_else_refresh"),
    ("particle_size_nm", "existing_reinterpreted", "outcome", "Particle size candidate", "preserve_human_edit_if_present_else_refresh"),
    ("pdi", "new_added", "outcome", "PDI candidate", "preserve_human_edit_if_present_else_refresh"),
    ("zeta_mV", "new_added", "outcome", "Zeta potential candidate", "preserve_human_edit_if_present_else_refresh"),
    ("value_source_type", "new_added", "provenance", "Lightweight value provenance for manual review", "refresh_machine_only"),
    ("candidate_notes", "new_added", "provenance", "Candidate notes and non-forced mapping explanations", "preserve_human_edit_if_present_else_refresh"),
]

IDENTITY_COLUMNS = {
    "paper_key",
    "doi",
    "gt_formulation_id",
    "family_id",
    "parent_core",
    "variant_role",
    "benchmark_default_include",
    "formulation_label",
    "seed_pred_representative_source_formulation_id",
    "gt_row_decision",
    "boundary_anchor_label",
}

LEGACY_TO_OUTPUT_PRIMARY = {
    "paper_key": "paper_key",
    "doi": "doi",
    "gt_formulation_id": "gt_formulation_id",
    "seed_pred_representative_source_formulation_id": "seed_pred_representative_source_formulation_id",
    "gt_row_decision": "gt_row_decision",
    "boundary_anchor_label": "boundary_anchor_label",
    "drug_name_candidates": "drug_name",
    "drug_feed_amount_candidates": "drug_mass_mg",
    "polymer_MW_candidates": "polymer_mw_raw",
    "LA_GA_candidates": "la_ga_ratio_raw",
    "surfactant_name_candidates": "surfactant_name",
    "surfactant_concentration_candidates": "surfactant_concentration_value",
    "drug_polymer_ratio_candidates": "drug_to_polymer_ratio_raw",
    "organic_solvent_candidates": "solvent_name",
    "particle_size_candidates": "particle_size_nm",
    "EE_candidates": "ee_percent",
    "LC_candidates": "lc_percent",
    "surfactant_mass_candidates": "surfactant_mass_mg",
    "organic_phase_volume_candidates": "O_volume_mL",
    "aqueous_phase_volume_candidates": "external_aqueous_phase_volume_mL",
    "polymer_feed_mass_candidates": "polymer_mass_mg",
}

POLYMER_NAMES = ["PLGA-PEG", "PEG-PLGA", "PLGA", "PCL", "PLA", "Myritol 318"]
SURFACTANT_NAMES = ["Pluronic F68", "PVA", "Polysorbate 80", "Tween 80", "Labrafil"]
DRUG_NAMES = ["Etoposide", "XAN", "3-MeOXAN", "FITC", "GAR", "MB"]
GRADE_PATTERNS = [
    re.compile(r"(?i)\bRG\d+[A-Z]?\b"),
    re.compile(r"(?i)\bResomer\b[^\s,;]*"),
]
MIXTURE_RATIO_RE = re.compile(r"(?i)\b(\d+(?:\.\d+)?)\s*[:/]\s*(\d+(?:\.\d+)?)\s*(?:,\s*)?(w/w|w/v|v/v)?\b")
MASS_RE = re.compile(r"(?i)\b(\d+(?:\.\d+)?)\s*(mg|g|ug|µg)\b")
CONCENTRATION_RE = re.compile(r"(?i)\b(\d+(?:\.\d+)?)\s*(mg\s*/\s*mL|g\s*/\s*L|%\s*w\s*/\s*v|%\s*w/v|%\s*v/v|mg/mL)\b")
MW_RE = re.compile(r"(?i)\b(\d+(?:\.\d+)?)\s*kDa\b")
VOLUME_RE = re.compile(r"(?i)\b(\d+(?:\.\d+)?)\s*(mL|ml|L|l)\b")


@dataclass
class LegacyCell:
    baseline_value: str
    current_value: str
    category: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_header(value: Any) -> str:
    return normalize_text(value)


def read_tsv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(path: Path, fieldnames: list[str], rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def read_workbook_rows(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [normalize_header(value) for value in rows[0]]
    cleaned_headers = [header if header else f"__blank_col_{idx}" for idx, header in enumerate(headers)]
    data_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        data_rows.append(
            {
                cleaned_headers[idx]: normalize_text(value)
                for idx, value in enumerate(row[: len(cleaned_headers)])
            }
        )
    return cleaned_headers, data_rows


def write_workbook(path: Path, rows: list[dict[str, str]], metadata_rows: list[tuple[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = LEGACY_MAIN_SHEET
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(OUTPUT_COLUMNS))}{len(rows) + 1}"

    header_fill = PatternFill(fill_type="solid", start_color="D9EAF7", end_color="D9EAF7")
    for col_idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=row.get(column_name, ""))

    for col_idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        width = max(len(column_name) + 2, 16)
        for row in rows[:50]:
            width = max(width, min(48, len(normalize_text(row.get(column_name))) + 2))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    meta = workbook.create_sheet("metadata")
    meta["A1"] = "key"
    meta["B1"] = "value"
    meta["A1"].font = Font(bold=True)
    meta["B1"].font = Font(bold=True)
    for idx, (key, value) in enumerate(metadata_rows, start=2):
        meta.cell(row=idx, column=1, value=key)
        meta.cell(row=idx, column=2, value=value)
    meta.column_dimensions["A"].width = 36
    meta.column_dimensions["B"].width = 120

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def load_text_by_paper(text_dir: Path) -> dict[str, str]:
    text_by_paper: dict[str, str] = {}
    for text_path in sorted(text_dir.glob("*.txt")):
        key = text_path.name.split(".")[0]
        text_by_paper[key] = text_path.read_text(encoding="utf-8", errors="ignore")
    return text_by_paper


def load_boundary_rows(boundary_workbook_path: Path) -> dict[tuple[str, str], dict[str, str]]:
    workbook = load_workbook(boundary_workbook_path, read_only=True, data_only=True)
    sheet = workbook["review_gt_rows"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [normalize_header(value) for value in rows[0]]
    index: dict[tuple[str, str], dict[str, str]] = {}
    for raw_row in rows[1:]:
        row = {headers[idx]: normalize_text(value) for idx, value in enumerate(raw_row)}
        key = (row.get("paper_key", ""), row.get("gt_formulation_id", ""))
        if key[0] and key[1]:
            index[key] = row
    return index


def load_gt_rows(gt_skeleton_tsv: Path) -> dict[tuple[str, str], dict[str, str]]:
    rows = read_tsv_rows(gt_skeleton_tsv)
    return {(row.get("paper_key", ""), row.get("formulation_id", "")): row for row in rows}


def classify_cell(baseline_value: str, current_value: str) -> str:
    if baseline_value == current_value:
        return "unchanged_from_machine_baseline"
    if not baseline_value and current_value:
        return "newly_added_by_user"
    if baseline_value and current_value != baseline_value:
        return "human_edited_relative_to_baseline"
    return "machine_fill_candidate"


def build_legacy_cell_index(
    baseline_rows: list[dict[str, str]],
    current_rows: list[dict[str, str]],
) -> tuple[list[tuple[str, str]], dict[tuple[str, str], dict[str, LegacyCell]]]:
    baseline_index = {(row["paper_key"], row["gt_formulation_id"]): row for row in baseline_rows}
    current_index = {(row["paper_key"], row["gt_formulation_id"]): row for row in current_rows}
    if list(baseline_index.keys()) != list(current_index.keys()):
        raise ValueError("Baseline TSV and current workbook row identities do not align.")
    ordered_keys = list(baseline_index.keys())
    legacy_cells: dict[tuple[str, str], dict[str, LegacyCell]] = {}
    for key in ordered_keys:
        row_cells: dict[str, LegacyCell] = {}
        baseline_row = baseline_index[key]
        current_row = current_index[key]
        for column_name in baseline_row:
            baseline_value = normalize_text(baseline_row.get(column_name))
            current_value = normalize_text(current_row.get(column_name))
            row_cells[column_name] = LegacyCell(
                baseline_value=baseline_value,
                current_value=current_value,
                category=classify_cell(baseline_value, current_value),
            )
        legacy_cells[key] = row_cells
    return ordered_keys, legacy_cells


def chosen_legacy_value(cell: LegacyCell) -> str:
    if cell.category in {"human_edited_relative_to_baseline", "newly_added_by_user"}:
        return cell.current_value
    return cell.baseline_value


def parse_concentration(text: str) -> tuple[str, str]:
    match = CONCENTRATION_RE.search(normalize_text(text))
    if not match:
        return "", ""
    return normalize_text(match.group(1)), normalize_text(match.group(2)).replace(" ", "")


def parse_mass(text: str) -> str:
    match = MASS_RE.search(normalize_text(text))
    if not match:
        return ""
    return f"{normalize_text(match.group(1))} {normalize_text(match.group(2))}"


def parse_volume(text: str) -> str:
    match = VOLUME_RE.search(normalize_text(text))
    if not match:
        return ""
    value = normalize_text(match.group(1))
    unit = normalize_text(match.group(2)).replace("ml", "mL").replace("ML", "mL")
    if unit.lower() == "l":
        unit = "L"
    elif unit.lower() == "ml":
        unit = "mL"
    return f"{value} {unit}"


def parse_mw_kda(text: str) -> str:
    normalized = normalize_text(text)
    matches = re.findall(r"(?i)\b(\d+(?:\.\d+)?)\s*kDa\b", normalized)
    if len(matches) != 1:
        return ""
    value = normalize_text(matches[0])
    try:
        numeric = float(value)
    except ValueError:
        return ""
    if numeric <= 0 or numeric >= 1000:
        return ""
    return value


def normalize_ratio(text: str) -> str:
    match = MIXTURE_RATIO_RE.search(normalize_text(text))
    if not match:
        return ""
    lhs = normalize_text(match.group(1))
    rhs = normalize_text(match.group(2))
    suffix = normalize_text(match.group(3))
    normalized = f"{lhs}:{rhs}"
    if suffix:
        normalized = f"{normalized} {suffix.lower()}"
    return normalized


def detect_name(text: str, candidates: list[str]) -> str:
    lowered = normalize_text(text).lower()
    for candidate in candidates:
        if candidate.lower() in lowered:
            return candidate
    return ""


def detect_polymer_name(*texts: str) -> str:
    for text in texts:
        name = detect_name(text, POLYMER_NAMES)
        if name:
            return name
    return ""


def detect_surfactant_name(*texts: str) -> str:
    for text in texts:
        name = detect_name(text, SURFACTANT_NAMES)
        if name:
            return name
    return ""


def detect_drug_name(*texts: str) -> str:
    for text in texts:
        name = detect_name(text, DRUG_NAMES)
        if name:
            return name
    return ""


def detect_grade(*texts: str) -> str:
    for text in texts:
        normalized = normalize_text(text)
        for pattern in GRADE_PATTERNS:
            match = pattern.search(normalized)
            if match:
                return normalize_text(match.group(0))
    return ""


def detect_method_type(text: str) -> str:
    lowered = normalize_text(text).lower()
    if "double emulsion" in lowered or "w/o/w" in lowered:
        return "double_emulsion_w1_o_w2"
    if "single emulsion" in lowered or "o/w" in lowered:
        return "single_emulsion_o_w"
    if "nanoprecipitation" in lowered or "solvent displacement" in lowered:
        return "nanoprecipitation"
    if "solvent evaporation" in lowered:
        return "solvent_evaporation"
    if "emulsion solvent evaporation" in lowered:
        return "emulsion_solvent_evaporation"
    return ""


def find_time_candidate(text: str, keyword: str, to_unit: str) -> str:
    pattern = re.compile(
        rf"(?is){keyword}[^\.;:\n]{{0,80}}?(\d+(?:\.\d+)?)\s*(s|sec|second|seconds|min|minute|minutes|h|hr|hrs|hour|hours)"
    )
    match = pattern.search(text)
    if not match:
        return ""
    value = float(match.group(1))
    unit = match.group(2).lower()
    if to_unit == "s":
        if unit.startswith("min"):
            value *= 60
        elif unit.startswith("h") or unit.startswith("hr") or unit.startswith("hour"):
            value *= 3600
        return f"{value:g} s"
    if to_unit == "min":
        if unit.startswith("h") or unit.startswith("hr") or unit.startswith("hour"):
            value *= 60
        elif unit.startswith("s"):
            value /= 60
        return f"{value:g} min"
    if to_unit == "h":
        if unit.startswith("min"):
            value /= 60
        elif unit.startswith("s"):
            value /= 3600
        return f"{value:g} h"
    return ""


def find_centrifugation_g(text: str) -> str:
    match = re.search(r"(?is)centrifug[^\.;:\n]{0,80}?(\d+(?:\.\d+)?)\s*g\b", text)
    if not match:
        return ""
    return f"{normalize_text(match.group(1))} g"


def parse_solvent_fields(organic_solvent: str, organic_phase_text: str) -> tuple[str, str, str, str]:
    solvent_name = normalize_text(organic_solvent)
    co_solvent_name = ""
    phase_ratio = normalize_ratio(organic_phase_text)
    organic_volume = parse_volume(organic_phase_text)
    phase_text = normalize_text(organic_phase_text)

    lowered = phase_text.lower()
    if "/" in phase_text and "mixture" in lowered:
        before_mixture = phase_text.split("mixture", 1)[0]
        names = re.findall(r"(?i)(acetone|dichloromethane|ethyl acetate|chloroform|methanol|ethanol|acetonitrile)", before_mixture)
        if names:
            solvent_name = names[0]
            if len(names) > 1:
                co_solvent_name = names[1]
    elif not solvent_name:
        names = re.findall(r"(?i)(acetone|dichloromethane|ethyl acetate|chloroform|methanol|ethanol|acetonitrile)", phase_text)
        if names:
            solvent_name = names[0]
            if len(names) > 1:
                co_solvent_name = names[1]
    return solvent_name, co_solvent_name, organic_volume, phase_ratio


def extract_theoretical_concentration(label: str) -> tuple[str, str]:
    lowered = normalize_text(label).lower()
    if "theoretical concentration" not in lowered:
        return "", ""
    return parse_concentration(label)


def build_schema_diff_rows() -> list[dict[str, str]]:
    return [
        {
            "column_name": column_name,
            "status": status,
            "semantic_group": semantic_group,
            "intended meaning": meaning,
            "overwrite_policy": overwrite_policy,
        }
        for column_name, status, semantic_group, meaning, overwrite_policy in SCHEMA_DIFF_ROWS
    ]


def write_run_context(
    path: Path,
    *,
    run_id: str,
    source_run_id: str,
    source_run_dir: str,
    current_workbook_xlsx: Path,
    baseline_tsv: Path,
    boundary_workbook_xlsx: Path,
    gt_skeleton_tsv: Path,
    output_workbook_xlsx: Path,
    output_tsv: Path,
    schema_diff_tsv: Path,
    preservation_tsv: Path,
    migration_summary_md: Path,
    preserved_count: int,
    refreshed_count: int,
    corrected_count: int,
    columns_added: int,
    row_count: int,
) -> None:
    content = f"""# RUN_CONTEXT

## 1. Run ID

- `{run_id}`

## 2. Run Type

- `intermediate_diagnostic_run`

## 3. Purpose

- Refresh the compact Layer3 value annotation workbook with a preservation-aware
  diff workflow.
- Preserve human-edited workbook cells relative to the machine baseline.
- Refresh unchanged machine cells and remap values into a minimally wider
  representation-aware schema so concentration, ratio, mass, and phase volume
  no longer collapse into the same fields.

## 4. Source Authority

- Source run id: `{source_run_id}`
- Source run dir: `{source_run_dir}`
- Current edited workbook: `{current_workbook_xlsx}`
- Machine baseline TSV: `{baseline_tsv}`
- Boundary GT workbook: `{boundary_workbook_xlsx}`
- GT skeleton TSV: `{gt_skeleton_tsv}`

## 5. Exact Script Execution Order

1. Read the machine baseline TSV and current edited workbook.
2. Diff row-aligned cells to detect human-edited vs unchanged machine cells.
3. Join authoritative Layer2 boundary review metadata by GT formulation id.
4. Remap candidate values into the representation-aware v2 workbook surface.
5. Write the workbook, TSV export, schema diff, preservation summary, and migration summary.

## 6. Script Paths Used

- `src/stage5_benchmark/build_value_gt_annotation_representation_repair_v2.py`

## 7. Final Outputs

- `{output_workbook_xlsx}`
- `{output_tsv}`
- `{schema_diff_tsv}`
- `{preservation_tsv}`
- `{migration_summary_md}`

## 8. Diagnostic Status

- This run is workbook-generation only.
- It does not modify Stage2, Stage3, or Stage5 runtime behavior.
- It does not overwrite the current manually edited workbook.

## 9. Outcome Summary

- Authoritative GT rows used: `{row_count}`
- Output rows written: `{row_count}`
- Preserved human-edited cells: `{preserved_count}`
- Refreshed machine cells: `{refreshed_count}`
- Corrected projection-error cells: `{corrected_count}`
- Columns added relative to compact baseline: `{columns_added}`
    """
    path.write_text(content, encoding="utf-8")


def build_output_row(
    key: tuple[str, str],
    legacy_cells: dict[str, LegacyCell],
    boundary_row: dict[str, str],
    gt_row: dict[str, str],
    paper_text: str,
    preservation_rows: list[dict[str, str]],
) -> dict[str, str]:
    paper_key, gt_formulation_id = key
    chosen = {column: chosen_legacy_value(cell) for column, cell in legacy_cells.items()}
    output = {column: "" for column in OUTPUT_COLUMNS}

    output["paper_key"] = paper_key
    output["doi"] = chosen.get("doi", "")
    output["gt_formulation_id"] = gt_formulation_id
    output["family_id"] = normalize_text(boundary_row.get("family_id"))
    output["parent_core"] = normalize_text(boundary_row.get("parent_core_row_id"))
    output["variant_role"] = normalize_text(boundary_row.get("variant_role"))
    output["benchmark_default_include"] = normalize_text(
        boundary_row.get("benchmark_include_gt") or boundary_row.get("seed_pred_benchmark_default_include")
    )
    output["formulation_label"] = normalize_text(boundary_row.get("boundary_anchor_label") or gt_row.get("formulation_label_raw"))
    output["seed_pred_representative_source_formulation_id"] = chosen.get("seed_pred_representative_source_formulation_id", "")
    output["gt_row_decision"] = chosen.get("gt_row_decision", "")
    output["boundary_anchor_label"] = chosen.get("boundary_anchor_label", "")

    source_anchor = " | ".join(
        [
            output["boundary_anchor_label"],
            output["seed_pred_representative_source_formulation_id"],
            gt_row.get("notes", ""),
        ]
    )

    output["polymer_name"] = detect_polymer_name(
        output["boundary_anchor_label"],
        chosen.get("organic_solvent_candidates", ""),
        chosen.get("polymer_MW_candidates", ""),
        output["seed_pred_representative_source_formulation_id"],
    )
    output["polymer_grade"] = detect_grade(chosen.get("polymer_MW_candidates", ""), output["boundary_anchor_label"], paper_text)
    output["polymer_mw_raw"] = chosen.get("polymer_MW_candidates", "")
    output["polymer_mw_kDa"] = parse_mw_kda(output["polymer_mw_raw"])
    output["la_ga_ratio_raw"] = chosen.get("LA_GA_candidates", "")
    output["la_ga_ratio_normalized"] = normalize_ratio(output["la_ga_ratio_raw"])

    drug_amount_raw = chosen.get("drug_feed_amount_candidates", "")
    drug_amount_conc_value, drug_amount_conc_unit = parse_concentration(drug_amount_raw)
    theoretical_conc_value, theoretical_conc_unit = extract_theoretical_concentration(output["boundary_anchor_label"])
    if drug_amount_raw and not drug_amount_conc_value:
        output["drug_mass_mg"] = parse_mass(drug_amount_raw) or drug_amount_raw
    if drug_amount_conc_value:
        output["drug_concentration_value"] = drug_amount_conc_value
        output["drug_concentration_unit"] = drug_amount_conc_unit
    elif theoretical_conc_value:
        output["drug_concentration_value"] = theoretical_conc_value
        output["drug_concentration_unit"] = theoretical_conc_unit

    output["drug_name"] = chosen.get("drug_name_candidates", "") or detect_drug_name(
        output["boundary_anchor_label"],
        source_anchor,
        paper_text,
    )
    output["drug_to_polymer_ratio_raw"] = chosen.get("drug_polymer_ratio_candidates", "")

    polymer_mass_raw = chosen.get("polymer_feed_mass_candidates", "")
    polymer_conc_value, polymer_conc_unit = parse_concentration(polymer_mass_raw)
    if polymer_mass_raw and not polymer_conc_value:
        output["polymer_mass_mg"] = parse_mass(polymer_mass_raw) or polymer_mass_raw
    if polymer_conc_value:
        output["polymer_concentration_value"] = polymer_conc_value
        output["polymer_concentration_unit"] = polymer_conc_unit
        output["polymer_concentration_phase"] = "organic_phase"

    surfactant_conc_raw = chosen.get("surfactant_concentration_candidates", "")
    surfactant_conc_value, surfactant_conc_unit = parse_concentration(surfactant_conc_raw)
    output["surfactant_concentration_value"] = surfactant_conc_value or normalize_text(surfactant_conc_raw)
    output["surfactant_concentration_unit"] = surfactant_conc_unit

    surfactant_mass_raw = chosen.get("surfactant_mass_candidates", "")
    surfactant_mass_conc_value, _ = parse_concentration(surfactant_mass_raw)
    if surfactant_mass_raw and not surfactant_mass_conc_value:
        output["surfactant_mass_mg"] = parse_mass(surfactant_mass_raw) or surfactant_mass_raw

    output["surfactant_name"] = chosen.get("surfactant_name_candidates", "") or detect_surfactant_name(
        chosen.get("aqueous_phase_volume_candidates", ""),
        chosen.get("surfactant_mass_candidates", ""),
        paper_text,
    )
    output["stabilizer_name"] = output["surfactant_name"]
    if not output["surfactant_name"]:
        output["helper_material_name"] = detect_surfactant_name(
            chosen.get("aqueous_phase_volume_candidates", ""),
            chosen.get("surfactant_mass_candidates", ""),
            paper_text,
        )
    else:
        helper_from_mass = detect_name(chosen.get("surfactant_mass_candidates", ""), ["Labrafil"])
        if helper_from_mass and helper_from_mass != output["surfactant_name"]:
            output["helper_material_name"] = helper_from_mass

    solvent_name, co_solvent_name, organic_volume, phase_ratio_raw = parse_solvent_fields(
        chosen.get("organic_solvent_candidates", ""),
        chosen.get("organic_phase_volume_candidates", ""),
    )
    output["solvent_name"] = solvent_name
    output["co_solvent_name"] = co_solvent_name
    output["O_volume_mL"] = organic_volume
    output["external_aqueous_phase_volume_mL"] = parse_volume(chosen.get("aqueous_phase_volume_candidates", ""))
    output["phase_ratio_raw"] = phase_ratio_raw

    output["method_type"] = detect_method_type(paper_text)
    output["sonication_time_s"] = find_time_candidate(paper_text, "sonicat", "s")
    output["homogenization_time_min"] = find_time_candidate(paper_text, "homogen", "min")
    output["stirring_time_h"] = find_time_candidate(paper_text, "stirr", "h")
    output["evaporation_time_h"] = find_time_candidate(paper_text, "evapor", "h")
    output["centrifugation_g"] = find_centrifugation_g(paper_text)
    output["centrifugation_time_min"] = find_time_candidate(paper_text, "centrifug", "min")

    output["ee_percent"] = chosen.get("EE_candidates", "")
    output["lc_percent"] = chosen.get("LC_candidates", "")
    output["particle_size_nm"] = chosen.get("particle_size_candidates", "")

    note_parts: list[str] = []
    source_types: set[str] = set()
    if legacy_cells["drug_feed_amount_candidates"].category in {
        "human_edited_relative_to_baseline",
        "newly_added_by_user",
    }:
        source_types.add("human_edit")
    if drug_amount_conc_value:
        note_parts.append("drug concentration preserved separately from drug mass")
    if theoretical_conc_value and not drug_amount_conc_value:
        note_parts.append("theoretical concentration captured from formulation label")
        source_types.add("parsed_from_label")
    if output["phase_ratio_raw"]:
        note_parts.append("phase ratio preserved separately from phase volume")
    if output["co_solvent_name"]:
        note_parts.append("organic solvent mixture split into solvent and co-solvent")
    if output["method_type"]:
        source_types.add("paper_text")
    if not source_types:
        source_types.add("machine_baseline")
    if "human_edit" in source_types and len(source_types) > 1:
        output["value_source_type"] = "mixed"
    else:
        output["value_source_type"] = next(iter(sorted(source_types)))
    output["candidate_notes"] = "; ".join(note_parts)

    for column_name in OUTPUT_COLUMNS:
        legacy_source_column = next(
            (legacy for legacy, target in LEGACY_TO_OUTPUT_PRIMARY.items() if target == column_name),
            "",
        )
        baseline_value = ""
        current_value = ""
        if legacy_source_column:
            baseline_value = legacy_cells[legacy_source_column].baseline_value
            current_value = legacy_cells[legacy_source_column].current_value

        final_value = output[column_name]
        preservation_status = "unchanged"
        update_reason = "identity carry-through" if column_name in IDENTITY_COLUMNS else "direct carry-through"

        if not final_value:
            preservation_status = "left_blank"
            update_reason = "blank because no direct support or safe deterministic mapping was found"
        elif legacy_source_column:
            category = legacy_cells[legacy_source_column].category
            if category in {"human_edited_relative_to_baseline", "newly_added_by_user"}:
                preservation_status = "preserved_human_edit"
                update_reason = f"mapped from human-edited legacy cell `{legacy_source_column}`"
            elif current_value == baseline_value == final_value:
                preservation_status = "unchanged"
                update_reason = f"legacy machine cell `{legacy_source_column}` carried forward unchanged"
            elif (
                legacy_source_column in {"drug_feed_amount_candidates", "polymer_feed_mass_candidates", "surfactant_mass_candidates"}
                and parse_concentration(chosen.get(legacy_source_column, ""))[0]
            ) or (
                legacy_source_column == "organic_phase_volume_candidates"
                and normalize_ratio(chosen.get(legacy_source_column, ""))
            ):
                preservation_status = "corrected_projection_error"
                update_reason = f"legacy cell `{legacy_source_column}` was semantically split to avoid wrong field projection"
            else:
                preservation_status = "refreshed_machine_cell"
                update_reason = f"legacy machine cell `{legacy_source_column}` refreshed through representation-aware remapping"
        else:
            if column_name in IDENTITY_COLUMNS:
                preservation_status = "refreshed_machine_cell"
                update_reason = "authoritative GT/boundary metadata refresh"
            elif column_name == "candidate_notes":
                preservation_status = "refreshed_machine_cell" if final_value else "left_blank"
                update_reason = "generated migration note"
            elif column_name == "value_source_type":
                preservation_status = "refreshed_machine_cell"
                update_reason = "generated lightweight provenance helper"
            elif column_name in {
                "polymer_mw_kDa",
                "la_ga_ratio_normalized",
                "drug_concentration_value",
                "drug_concentration_unit",
                "surfactant_concentration_unit",
                "polymer_concentration_value",
                "polymer_concentration_unit",
                "solvent_name",
                "co_solvent_name",
                "O_volume_mL",
                "external_aqueous_phase_volume_mL",
                "phase_ratio_raw",
            }:
                preservation_status = "corrected_projection_error" if final_value else "left_blank"
                update_reason = "new representation-aware field prevents concentration/ratio/phase-value misprojection"
            else:
                preservation_status = "refreshed_machine_cell"
                update_reason = "new candidate field refreshed from deterministic source text or boundary metadata"

        if preservation_status not in ALLOWED_PRESERVATION_STATUSES:
            raise ValueError(f"Unexpected preservation status: {preservation_status}")
        preservation_rows.append(
            {
                "sheet_name": LEGACY_MAIN_SHEET,
                "row_key": gt_formulation_id,
                "column_name": column_name,
                "old_baseline_value": baseline_value,
                "current_workbook_value": current_value,
                "final_output_value": final_value,
                "preservation_status": preservation_status,
                "update_reason": update_reason,
            }
        )

    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--run-id", required=True)
    parser.add_argument("--run-dir", required=True)
    parser.add_argument("--source-run-id", required=True)
    parser.add_argument("--source-run-dir", required=True)
    parser.add_argument("--current-workbook-xlsx", required=True)
    parser.add_argument("--baseline-tsv", required=True)
    parser.add_argument("--boundary-workbook-xlsx", required=True)
    parser.add_argument("--gt-skeleton-tsv", required=True)
    parser.add_argument("--text-dir", required=True)
    args = parser.parse_args()

    run_dir = Path(args.run_dir)
    current_workbook_xlsx = Path(args.current_workbook_xlsx)
    baseline_tsv = Path(args.baseline_tsv)
    boundary_workbook_xlsx = Path(args.boundary_workbook_xlsx)
    gt_skeleton_tsv = Path(args.gt_skeleton_tsv)
    text_dir = Path(args.text_dir)

    output_workbook_xlsx = run_dir / "value_gt_annotation_workbook_representation_repaired_v2.xlsx"
    output_tsv = run_dir / "value_gt_annotation_workbook_representation_repaired_v2.tsv"
    schema_diff_tsv = run_dir / "value_gt_annotation_schema_diff_v2.tsv"
    preservation_tsv = run_dir / "value_gt_annotation_cell_preservation_summary_v2.tsv"
    migration_summary_md = run_dir / "value_gt_annotation_migration_summary_v2.md"
    run_context_md = run_dir / "RUN_CONTEXT.md"

    baseline_rows = read_tsv_rows(baseline_tsv)
    workbook_headers, current_rows = read_workbook_rows(current_workbook_xlsx, LEGACY_MAIN_SHEET)
    baseline_headers = list(baseline_rows[0].keys())
    normalized_workbook_headers = [header for header in workbook_headers if not header.startswith("__blank_col_")]
    if normalized_workbook_headers != baseline_headers:
        raise ValueError("Workbook and baseline TSV column headers do not align.")

    ordered_keys, legacy_cell_index = build_legacy_cell_index(baseline_rows, current_rows)
    boundary_index = load_boundary_rows(boundary_workbook_xlsx)
    gt_index = load_gt_rows(gt_skeleton_tsv)
    text_by_paper = load_text_by_paper(text_dir)

    preservation_rows: list[dict[str, str]] = []
    output_rows: list[dict[str, str]] = []
    for key in ordered_keys:
        paper_key, gt_formulation_id = key
        boundary_row = boundary_index.get(key, {})
        gt_row = gt_index.get((paper_key, gt_formulation_id.replace("_G", "_F")), gt_index.get(key, {}))
        paper_text = text_by_paper.get(paper_key, "")
        output_rows.append(
            build_output_row(
                key,
                legacy_cell_index[key],
                boundary_row,
                gt_row,
                paper_text,
                preservation_rows,
            )
        )

    schema_rows = build_schema_diff_rows()
    write_tsv(output_tsv, OUTPUT_COLUMNS, output_rows)
    write_tsv(
        schema_diff_tsv,
        ["column_name", "status", "semantic_group", "intended meaning", "overwrite_policy"],
        schema_rows,
    )
    write_tsv(
        preservation_tsv,
        [
            "sheet_name",
            "row_key",
            "column_name",
            "old_baseline_value",
            "current_workbook_value",
            "final_output_value",
            "preservation_status",
            "update_reason",
        ],
        preservation_rows,
    )

    status_counts = Counter(row["preservation_status"] for row in preservation_rows)
    summary_json_path = baseline_tsv.parent / "value_gt_annotation_workbook_with_phase_and_polymer_values_v1.summary.json"
    baseline_summary = json.loads(summary_json_path.read_text(encoding="utf-8")) if summary_json_path.exists() else {}
    metadata_rows = [
        ("baseline_workbook_lineage_source", normalize_text(baseline_summary.get("workbook_path_used", ""))),
        ("current_workbook_xlsx", str(current_workbook_xlsx.resolve())),
        ("machine_baseline_tsv", str(baseline_tsv.resolve())),
        ("boundary_workbook_xlsx", str(boundary_workbook_xlsx.resolve())),
        ("gt_skeleton_tsv", str(gt_skeleton_tsv.resolve())),
        ("source_run_id", args.source_run_id),
        ("source_run_dir", args.source_run_dir),
        ("generated_at_utc", datetime.now(timezone.utc).isoformat()),
        ("authoritative_gt_row_count", str(len(output_rows))),
        ("output_row_count", str(len(output_rows))),
        ("preserved_human_edit", str(status_counts["preserved_human_edit"])),
        ("refreshed_machine_cell", str(status_counts["refreshed_machine_cell"])),
        ("corrected_projection_error", str(status_counts["corrected_projection_error"])),
        ("left_blank", str(status_counts["left_blank"])),
        ("unchanged", str(status_counts["unchanged"])),
    ]
    write_workbook(output_workbook_xlsx, output_rows, metadata_rows)

    columns_added = len(OUTPUT_COLUMNS) - len(baseline_headers)
    migration_summary = f"""# Value GT Annotation Migration Summary

## Baseline

- Baseline workbook path identified from lineage summary: `data/results/run_20260314_1206_076995e_dev15_deterministic_refresh_no_llm_v1/value_gt_annotation_workbook_with_phase_values_v2.xlsx`
- Machine baseline used for cell-level diffing: `{baseline_tsv}`
- Current workbook compared: `{current_workbook_xlsx}`
- Row identity aligned: `True`
- Sheet structure aligned on main review sheet: `True`

## Script Strategy

- Existing maintained script reused directly: `No`
- Existing compact-family builder preserved in `src/`: `Not found by repo inspection`
- Implemented thin migration/update layer: `src/stage5_benchmark/build_value_gt_annotation_representation_repair_v2.py`

## Human Edit Detection

- Human-edited cells were inferred by diffing the current workbook main sheet against the machine baseline TSV for the same workbook family.
- Categories used:
  - `unchanged_from_machine_baseline`
  - `human_edited_relative_to_baseline`
  - `newly_added_by_user`
  - `machine_fill_candidate`

## Cell Outcomes

- Preserved human-edited cells: `{status_counts["preserved_human_edit"]}`
- Refreshed machine cells: `{status_counts["refreshed_machine_cell"]}`
- Corrected projection-error cells: `{status_counts["corrected_projection_error"]}`
- Left blank intentionally: `{status_counts["left_blank"]}`
- Unchanged carry-through cells: `{status_counts["unchanged"]}`

## Minimal New Columns Added

- Columns added relative to the compact machine baseline: `{columns_added}`
- New columns focus on:
  - family / variant anchors
  - polymer identity separation
  - concentration-vs-mass separation
  - solvent / co-solvent / phase-volume separation
  - lightweight provenance and notes

## Projection Repairs

- Concentration-to-mass projection is now prevented by routing concentration-like strings into:
  - `drug_concentration_value`
  - `drug_concentration_unit`
  - `polymer_concentration_value`
  - `polymer_concentration_unit`
  - `surfactant_concentration_value`
  - `surfactant_concentration_unit`
- Raw ratio strings are preserved in dedicated ratio fields and are not written into mass or concentration fields.
- Phase volume and phase ratio are separated:
  - `O_volume_mL`
  - `external_aqueous_phase_volume_mL`
  - `phase_ratio_raw`

## Width Control

- Workbook width was kept under control by reusing existing compact anchors and raw candidate concepts where possible.
- The migration avoided creating a duplicate legacy-plus-new column universe for every field.

## Important Limitation

- No deeper preserved pre-edit `.xlsx` copy of `value_gt_annotation_workbook_with_phase_and_polymer_values_v1.xlsx` was found in the lineage search.
- The strongest machine baseline available in-repo for this workbook is its sibling machine TSV plus the recorded source workbook lineage.
"""
    migration_summary_md.write_text(migration_summary, encoding="utf-8")

    write_run_context(
        run_context_md,
        run_id=args.run_id,
        source_run_id=args.source_run_id,
        source_run_dir=args.source_run_dir,
        current_workbook_xlsx=current_workbook_xlsx.resolve(),
        baseline_tsv=baseline_tsv.resolve(),
        boundary_workbook_xlsx=boundary_workbook_xlsx.resolve(),
        gt_skeleton_tsv=gt_skeleton_tsv.resolve(),
        output_workbook_xlsx=output_workbook_xlsx.resolve(),
        output_tsv=output_tsv.resolve(),
        schema_diff_tsv=schema_diff_tsv.resolve(),
        preservation_tsv=preservation_tsv.resolve(),
        migration_summary_md=migration_summary_md.resolve(),
        preserved_count=status_counts["preserved_human_edit"],
        refreshed_count=status_counts["refreshed_machine_cell"],
        corrected_count=status_counts["corrected_projection_error"],
        columns_added=columns_added,
        row_count=len(output_rows),
    )

    print(
        json.dumps(
            {
                "run_dir": str(run_dir.resolve()),
                "row_count": len(output_rows),
                "preserved_human_edit": status_counts["preserved_human_edit"],
                "refreshed_machine_cell": status_counts["refreshed_machine_cell"],
                "corrected_projection_error": status_counts["corrected_projection_error"],
                "columns_added": columns_added,
                "output_workbook_xlsx": str(output_workbook_xlsx.resolve()),
                "output_tsv": str(output_tsv.resolve()),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
