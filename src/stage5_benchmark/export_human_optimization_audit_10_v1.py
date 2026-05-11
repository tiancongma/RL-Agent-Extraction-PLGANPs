#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation


TARGET_FIELDS = ["encapsulation_efficiency_percent", "size_nm"]
OUTPUT_COLUMNS = [
    "bucket",
    "zotero_key",
    "doi",
    "title",
    "year",
    "field_name",
    "extracted_value_raw",
    "extracted_value_canon",
    "qc_fail_type",
    "qc_detail",
    "evidence_source_type_inferred",
    "evidence_pointer_raw",
    "evidence_text",
    "evidence_context_before",
    "evidence_context_after",
    "evidence_span_start",
    "evidence_span_end",
    "evidence_section",
    "table_csv_path",
    "table_filename",
    "table_title_or_caption",
    "table_row_text",
    "table_cell_text",
    "value_source_field",
    "nearby_numeric_tokens",
    "table_key_consistency",
    "reviewer_true_source",
    "reviewer_match_type",
    "reviewer_root_issue",
    "suspected_layer",
    "recommended_fix_target",
    "human_notes",
]


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Export a 10-row DEV optimization audit workbook for numeric token mismatches.")
    p.add_argument("--run-id", required=True)
    p.add_argument("--out-xlsx", default="")
    p.add_argument("--seed", type=int, default=13)
    return p.parse_args()


def discover_run_inputs(run_id: str) -> dict[str, Path]:
    run_dir = Path("data/results") / run_id
    if not run_dir.exists():
        raise FileNotFoundError(f"Run folder not found: {run_dir}")

    weak_candidates = sorted(run_dir.rglob("step1_dev/weak_labels__gemini.tsv"))
    qc_candidates = sorted(run_dir.rglob("step1_dev/benchmark_goren_2025/derivation_v1/evidence_token_qc_checks.tsv"))
    if not weak_candidates:
        raise FileNotFoundError("Could not find step1_dev/weak_labels__gemini.tsv under run folder.")
    if not qc_candidates:
        raise FileNotFoundError("Could not find step1_dev/.../evidence_token_qc_checks.tsv under run folder.")

    weak_path = weak_candidates[0]
    qc_path = qc_candidates[0]
    step1_dev_dir = weak_path.parent
    audit_dir = step1_dev_dir / "audit_pack"
    audit_candidates = sorted(audit_dir.glob("audit_pack__human_evidence_v1*.xlsx"))
    manifest_path = Path("data/cleaned/goren_2025/index/manifest.tsv")
    if not manifest_path.exists():
        raise FileNotFoundError(f"Dataset manifest missing: {manifest_path}")

    return {
        "run_dir": run_dir,
        "step1_dev_dir": step1_dev_dir,
        "weak_path": weak_path,
        "qc_path": qc_path,
        "audit_path": audit_candidates[-1] if audit_candidates else Path(""),
        "manifest_path": manifest_path,
    }


def safe_text(v: Any) -> str:
    return str(v or "").strip()


def stable_hash_token(*parts: str) -> str:
    s = "||".join(parts)
    return hashlib.sha1(s.encode("utf-8", errors="ignore")).hexdigest()


def infer_source_type(value_source: str, evidence_source_type: str) -> str:
    vs = safe_text(value_source).lower()
    est = safe_text(evidence_source_type).lower()
    if vs == "table_csv_cell":
        return "table"
    if vs in {"fulltext_span", "derived_rule", "proxy_compose", "derived_doe_decode"}:
        return "text"
    if est == "table":
        return "table"
    if est in {"fulltext", "text", "unknown", "proxy_compose"}:
        return "text"
    return "none_or_unknown"


def build_weak_match(df_weak: pd.DataFrame, field_name: str) -> pd.DataFrame:
    out = df_weak.copy()
    out["field_name"] = field_name
    out["field_value_for_join"] = out[field_name].astype(str) if field_name in out.columns else ""
    out["weak_span_hash"] = out["evidence_span_text"].map(lambda x: stable_hash_token(safe_text(x)))
    out["weak_row_id"] = out.index.astype(str)
    return out


def choose_first_deterministic(df: pd.DataFrame, sort_cols: list[str]) -> pd.Series | None:
    if df.empty:
        return None
    use_cols = [c for c in sort_cols if c in df.columns]
    if use_cols:
        return df.sort_values(use_cols, kind="mergesort").iloc[0]
    return df.iloc[0]


def extract_nearby_numeric_tokens(text: str, before: str, after: str, limit: int = 8) -> str:
    merged = " ".join([safe_text(before), safe_text(text), safe_text(after)])
    nums = re.findall(r"[+-]?\d+(?:[.,]\d+)?", merged)
    out: list[str] = []
    seen: set[str] = set()
    for n in nums:
        n2 = n.replace(",", ".")
        if n2 not in seen:
            seen.add(n2)
            out.append(n2)
        if len(out) >= limit:
            break
    return "|".join(out)


def is_table_key_consistent(zotero_key: str, table_csv_path: str) -> bool:
    p = safe_text(table_csv_path).replace("\\", "/")
    k = safe_text(zotero_key)
    if not p:
        return True
    return f"/tables/{k}/" in p


def sample_with_table_mix(df: pd.DataFrame, n: int, seed: int, field_name: str) -> pd.DataFrame:
    if df.empty:
        return df
    work = df.copy()
    work["_sel_rank"] = work.apply(
        lambda r: int(
            stable_hash_token(
                str(seed),
                field_name,
                safe_text(r.get("zotero_key", "")),
                safe_text(r.get("field_name", "")),
                safe_text(r.get("evidence_span_start", "")),
                safe_text(r.get("evidence_span_end", "")),
                safe_text(r.get("evidence_text", ""))[:64],
            )[:12],
            16,
        ),
        axis=1,
    )
    work = work.sort_values("_sel_rank", kind="mergesort")
    table_rows = work[work["table_candidate"] == True]
    table_pick_n = min(2, len(table_rows), n)
    picked_idx: list[int] = []
    if table_pick_n > 0:
        picked_idx.extend(table_rows.head(table_pick_n).index.tolist())
    remain_n = n - len(picked_idx)
    if remain_n > 0:
        remaining = work[~work.index.isin(picked_idx)]
        picked_idx.extend(remaining.head(remain_n).index.tolist())
    return work.loc[picked_idx].drop(columns=["_sel_rank"])


def apply_excel_formatting(path_xlsx: Path) -> None:
    wb = load_workbook(path_xlsx)
    ws = wb["audit10"]
    ws.freeze_panes = "A2"

    text_wrap_cols = {
        "qc_detail",
        "evidence_pointer_raw",
        "evidence_text",
        "evidence_context_before",
        "evidence_context_after",
        "table_csv_path",
        "table_filename",
        "table_title_or_caption",
        "table_row_text",
        "table_cell_text",
        "nearby_numeric_tokens",
        "recommended_fix_target",
        "human_notes",
    }
    header = [c.value for c in ws[1]]
    col_idx = {name: idx + 1 for idx, name in enumerate(header)}

    widths = {
        "bucket": 14,
        "zotero_key": 12,
        "doi": 28,
        "title": 32,
        "year": 8,
        "field_name": 34,
        "extracted_value_raw": 16,
        "extracted_value_canon": 16,
        "qc_fail_type": 22,
        "qc_detail": 40,
        "evidence_source_type_inferred": 22,
        "evidence_pointer_raw": 42,
        "evidence_text": 52,
        "evidence_context_before": 42,
        "evidence_context_after": 42,
        "evidence_span_start": 14,
        "evidence_span_end": 14,
        "evidence_section": 16,
        "table_csv_path": 52,
        "table_filename": 24,
        "table_title_or_caption": 32,
        "table_row_text": 52,
        "table_cell_text": 32,
        "value_source_field": 20,
        "nearby_numeric_tokens": 28,
        "table_key_consistency": 18,
        "reviewer_true_source": 22,
        "reviewer_match_type": 22,
        "reviewer_root_issue": 26,
        "suspected_layer": 30,
        "recommended_fix_target": 30,
        "human_notes": 42,
    }
    for name, w in widths.items():
        if name in col_idx:
            ws.column_dimensions[ws.cell(row=1, column=col_idx[name]).column_letter].width = w

    max_row = ws.max_row
    for name in text_wrap_cols:
        if name not in col_idx:
            continue
        cidx = col_idx[name]
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=cidx).alignment = Alignment(wrap_text=True, vertical="top")

    dropdowns = {
        "reviewer_true_source": ["table", "text", "none_or_unknown"],
        "reviewer_match_type": ["exact", "soft_plusminus", "soft_rounding", "range_match", "no_match"],
        "reviewer_root_issue": [
            "wrong_anchor",
            "wrong_source_binding",
            "wrong_value",
            "multiple_numbers_ambiguous",
            "qc_too_strict",
            "qc_bug",
            "retrieval_insufficient",
        ],
        "suspected_layer": [
            "A_extraction",
            "B_resolver_binding",
            "B_evidence_pointer_normalization",
            "B_qc_matcher",
            "retrieval_insufficient",
            "unknown",
        ],
    }
    for name, options in dropdowns.items():
        if name not in col_idx:
            continue
        cidx = col_idx[name]
        dv = DataValidation(type="list", formula1=f"\"{','.join(options)}\"", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=cidx).column_letter}2:{ws.cell(row=max_row, column=cidx).column_letter}{max_row}")

    wb.save(path_xlsx)


def main() -> None:
    args = parse_args()
    discovered = discover_run_inputs(args.run_id)
    weak_path = discovered["weak_path"]
    qc_path = discovered["qc_path"]
    manifest_path = discovered["manifest_path"]
    audit_path = discovered["audit_path"]

    step1_dev_dir = discovered["step1_dev_dir"]
    out_xlsx = (
        Path(args.out_xlsx)
        if safe_text(args.out_xlsx)
        else (step1_dev_dir / "audit_pack" / "dev_human_optimization_audit_10__numeric_mismatch_v1.xlsx")
    )
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    df_weak = pd.read_csv(weak_path, sep="\t", dtype=str).fillna("")
    df_qc = pd.read_csv(qc_path, sep="\t", dtype=str).fillna("")
    manifest = pd.read_csv(manifest_path, sep="\t", dtype=str).fillna("")

    # Manifest mapping
    key_col = "key" if "key" in manifest.columns else ("zotero_key" if "zotero_key" in manifest.columns else "")
    meta = pd.DataFrame(columns=["zotero_key", "doi", "title", "year"])
    if key_col:
        meta = pd.DataFrame(
            {
                "zotero_key": manifest[key_col].astype(str),
                "doi": manifest["doi"].astype(str) if "doi" in manifest.columns else "",
                "title": manifest["title"].astype(str) if "title" in manifest.columns else "",
                "year": manifest["year"].astype(str) if "year" in manifest.columns else "",
            }
        ).drop_duplicates(subset=["zotero_key"])

    # Optional richer audit data
    audit_lookup: dict[tuple[str, str], dict[str, Any]] = {}
    if audit_path and audit_path.exists():
        try:
            audit_df = pd.read_excel(audit_path, sheet_name="audit_cases", dtype=str).fillna("")
            if {"zotero_key", "derived_field_name"}.issubset(set(audit_df.columns)):
                for _, row in audit_df.iterrows():
                    k = (safe_text(row.get("zotero_key")), safe_text(row.get("derived_field_name")))
                    if k not in audit_lookup:
                        audit_lookup[k] = row.to_dict()
        except Exception:
            audit_lookup = {}

    df_qc["fail_numeric_token"] = pd.to_numeric(df_qc.get("fail_numeric_token", 0), errors="coerce").fillna(0).astype(int)
    qc_target = df_qc[(df_qc["field_name"].isin(TARGET_FIELDS)) & (df_qc["fail_numeric_token"] == 1)].copy()
    qc_target["qc_fail_type"] = "numeric_token_mismatch"
    qc_target["qc_detail"] = qc_target.apply(
        lambda r: (
            f"main_numeric={safe_text(r.get('main_numeric_token'))}; "
            f"required_unit={safe_text(r.get('required_unit_tokens'))}; "
            f"source={safe_text(r.get('evidence_source_type'))}"
        ),
        axis=1,
    )
    qc_target["qc_span_hash"] = qc_target["evidence_span_text"].map(lambda x: stable_hash_token(safe_text(x)))

    weak_expanded = pd.concat([build_weak_match(df_weak, f) for f in TARGET_FIELDS], ignore_index=True)
    weak_expanded["key"] = weak_expanded["key"].astype(str)
    weak_expanded["formulation_id"] = weak_expanded["formulation_id"].astype(str)

    matched_rows: list[dict[str, Any]] = []
    unmatched = 0
    ambiguous = 0
    for _, q in qc_target.iterrows():
        key = safe_text(q.get("key"))
        fid = safe_text(q.get("formulation_id"))
        field = safe_text(q.get("field_name"))
        field_value = safe_text(q.get("field_value"))
        q_start = safe_text(q.get("field_span_start"))
        q_hash = safe_text(q.get("qc_span_hash"))

        cand = weak_expanded[
            (weak_expanded["key"].astype(str) == key)
            & (weak_expanded["formulation_id"].astype(str) == fid)
            & (weak_expanded["field_name"].astype(str) == field)
        ].copy()
        if field_value:
            cand2 = cand[cand["field_value_for_join"].astype(str) == field_value]
            if not cand2.empty:
                cand = cand2
        if q_start:
            cand2 = cand[cand["evidence_span_start"].astype(str) == q_start]
            if not cand2.empty:
                cand = cand2
        if q_hash:
            cand2 = cand[cand["weak_span_hash"].astype(str) == q_hash]
            if not cand2.empty:
                cand = cand2

        picked = choose_first_deterministic(cand, ["weak_row_id", "evidence_span_start", "evidence_span_end"])
        if picked is None:
            unmatched += 1
            base = {
                "zotero_key": key,
                "formulation_id": fid,
                "field_name": field,
                "extracted_value_raw": field_value,
                "extracted_value_canon": safe_text(q.get("main_numeric_token")),
                "evidence_text": safe_text(q.get("evidence_span_text")),
                "evidence_span_start": q_start,
                "evidence_span_end": "",
                "evidence_section": "",
                "evidence_pointer_raw": "",
                "table_csv_path": "",
                "table_filename": "",
                "table_title_or_caption": "",
                "table_row_text": "",
                "table_cell_text": "",
                "evidence_context_before": "",
                "evidence_context_after": "",
                "value_source_field": "",
                "evidence_source_type_raw": safe_text(q.get("evidence_source_type")),
                "qc_fail_type": "numeric_token_mismatch",
                "qc_detail": safe_text(q.get("qc_detail")),
            }
        else:
            if len(cand) > 1:
                ambiguous += 1
            value_source_col = "value_source_EE" if field == "encapsulation_efficiency_percent" else "value_source_size"
            base = {
                "zotero_key": key,
                "formulation_id": fid,
                "field_name": field,
                "extracted_value_raw": safe_text(picked.get(field, "")),
                "extracted_value_canon": safe_text(q.get("main_numeric_token")),
                "evidence_text": safe_text(picked.get("evidence_span_text", "")),
                "evidence_span_start": safe_text(picked.get("evidence_span_start", q_start)),
                "evidence_span_end": safe_text(picked.get("evidence_span_end", "")),
                "evidence_section": safe_text(picked.get("evidence_section", "")),
                "evidence_pointer_raw": "",
                "table_csv_path": "",
                "table_filename": "",
                "table_title_or_caption": "",
                "table_row_text": "",
                "table_cell_text": "",
                "evidence_context_before": "",
                "evidence_context_after": "",
                "value_source_field": safe_text(picked.get(value_source_col, "")),
                "evidence_source_type_raw": safe_text(q.get("evidence_source_type", "")),
                "qc_fail_type": "numeric_token_mismatch",
                "qc_detail": safe_text(q.get("qc_detail")),
            }

        audit_row = audit_lookup.get((key, field), {})
        if audit_row:
            base["evidence_pointer_raw"] = safe_text(audit_row.get("evidence_pointer_raw", base["evidence_pointer_raw"]))
            base["table_csv_path"] = safe_text(audit_row.get("table_csv_path", base["table_csv_path"]))
            base["table_filename"] = safe_text(audit_row.get("table_filename", base["table_filename"]))
            base["table_title_or_caption"] = safe_text(audit_row.get("table_title_or_caption", base["table_title_or_caption"]))
            base["table_row_text"] = safe_text(audit_row.get("table_row_text", base["table_row_text"]))
            base["table_cell_text"] = safe_text(audit_row.get("table_cell_text", base["table_cell_text"]))
            base["evidence_context_before"] = safe_text(audit_row.get("evidence_context_before", base["evidence_context_before"]))
            base["evidence_context_after"] = safe_text(audit_row.get("evidence_context_after", base["evidence_context_after"]))
            if not base["value_source_field"]:
                value_source_col = "value_source_EE" if field == "encapsulation_efficiency_percent" else "value_source_size"
                base["value_source_field"] = safe_text(audit_row.get(value_source_col, ""))
            if not base["evidence_source_type_raw"]:
                base["evidence_source_type_raw"] = safe_text(audit_row.get("evidence_source_type", ""))
            if not base["evidence_text"]:
                base["evidence_text"] = safe_text(audit_row.get("evidence_text", ""))
            if not base["evidence_span_start"]:
                base["evidence_span_start"] = safe_text(audit_row.get("evidence_span_start", ""))
            if not base["evidence_span_end"]:
                base["evidence_span_end"] = safe_text(audit_row.get("evidence_span_end", ""))
            if not base["evidence_section"]:
                base["evidence_section"] = safe_text(audit_row.get("evidence_section", ""))

        base["table_candidate"] = bool(base["table_csv_path"] or base["table_cell_text"] or base["table_row_text"])
        base["evidence_source_type_inferred"] = infer_source_type(base["value_source_field"], base["evidence_source_type_raw"])
        base["nearby_numeric_tokens"] = extract_nearby_numeric_tokens(
            base["evidence_text"],
            base["evidence_context_before"],
            base["evidence_context_after"],
        )
        base["table_key_consistency"] = is_table_key_consistent(base["zotero_key"], base["table_csv_path"])
        matched_rows.append(base)

    merged = pd.DataFrame(matched_rows)
    merged = merged.merge(meta, how="left", on="zotero_key")
    merged["title"] = merged["title"].fillna("")
    merged["doi"] = merged["doi"].fillna("")
    merged["year"] = merged["year"].fillna("")

    selected_parts: list[pd.DataFrame] = []
    for field in TARGET_FIELDS:
        sub = merged[merged["field_name"] == field].copy()
        picked = sample_with_table_mix(sub, n=5, seed=args.seed, field_name=field)
        picked["bucket"] = "ee_numeric_mismatch" if field == "encapsulation_efficiency_percent" else "size_numeric_mismatch"
        selected_parts.append(picked)
    selected = pd.concat(selected_parts, ignore_index=True) if selected_parts else pd.DataFrame(columns=OUTPUT_COLUMNS)
    if len(selected) != 10:
        raise RuntimeError(f"Expected exactly 10 rows, got {len(selected)}")

    selected["recommended_fix_target"] = ""
    selected["human_notes"] = ""
    selected["reviewer_true_source"] = ""
    selected["reviewer_match_type"] = ""
    selected["reviewer_root_issue"] = ""
    selected["suspected_layer"] = ""

    # Ensure all output columns exist
    for c in OUTPUT_COLUMNS:
        if c not in selected.columns:
            selected[c] = ""
    out_df = selected[OUTPUT_COLUMNS].copy()
    out_df.to_excel(out_xlsx, sheet_name="audit10", index=False)
    apply_excel_formatting(out_xlsx)

    print(f"run_id={args.run_id}")
    print(f"weak_labels_tsv={weak_path}")
    print(f"qc_checks_tsv={qc_path}")
    print(f"manifest_tsv={manifest_path}")
    print(f"audit_source={'none' if not (audit_path and audit_path.exists()) else audit_path}")
    print(f"join_unmatched_count={unmatched}")
    print(f"join_ambiguous_count={ambiguous}")
    print(f"output_xlsx={out_xlsx}")
    print("selected_rows:")
    for _, r in out_df.iterrows():
        print(
            f"- {safe_text(r.get('zotero_key'))}\t{safe_text(r.get('field_name'))}\t"
            f"{safe_text(r.get('qc_fail_type'))}\ttable_candidate={bool(r.get('table_csv_path') or r.get('table_cell_text') or r.get('table_row_text'))}"
        )


if __name__ == "__main__":
    main()
