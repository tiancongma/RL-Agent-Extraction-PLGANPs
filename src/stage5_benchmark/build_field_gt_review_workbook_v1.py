#!/usr/bin/env python3
from __future__ import annotations

"""
Build a human-reviewable Layer 3 field-GT workbook from frozen Stage 5 rows.

Purpose:
- seed field-level GT review from an existing frozen `final_formulation_table_v1.tsv`
- keep workbook columns compact for human review
- preserve row-level identity while expanding to one row per `(formulation_id, field_name)`

Inputs:
- frozen Stage 5 `final_formulation_table_v1.tsv`
- optional scope manifest TSV for DOI/title enrichment

Outputs:
- `field_gt_review_workbook_v1.xlsx`
- companion seed and summary TSVs

Stage role:
- supporting Stage 5 / Layer 3 GT review surface
- not part of the benchmark-governing final-output path
"""

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from src.stage5_benchmark.audit_evidence_resolver_v1 import AuditEvidenceResolverV1
    from src.utils.active_data_source import (
        build_artifact_metadata,
        resolve_artifact_path,
        resolve_run_context,
        write_artifact_metadata_json,
    )
    from src.utils.run_id import validate_artifact_subdir
except ModuleNotFoundError:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from src.stage5_benchmark.audit_evidence_resolver_v1 import AuditEvidenceResolverV1
    from src.utils.active_data_source import (
        build_artifact_metadata,
        resolve_artifact_path,
        resolve_run_context,
        write_artifact_metadata_json,
    )
    from src.utils.run_id import validate_artifact_subdir


GT_STATUS_OPTIONS = ["correct", "incorrect", "not_reported", "unclear"]
GT_UNIT_OPTIONS = [
    "",
    "none",
    "ratio",
    "w/w",
    "%",
    "% w/v",
    "% w/w",
    "kDa",
    "mg",
    "mg/mL",
    "ug/mg",
    "nm",
]

WORKBOOK_COLUMNS = [
    "paper_key",
    "formulation_id",
    "formulation_label_stage5",
    "formulation_label_params",
    "article_formulation_id",
    "article_formulation_label",
    "paper_risk_level",
    "layer3_inclusion_flag",
    "field_name",
    "extracted_value",
    "extracted_unit",
    "evidence_text",
    "evidence_anchor_text",
    "evidence_source_type",
    "evidence_status_detail",
    "review_warning",
    "normalization_status",
    "gt_status",
    "gt_value",
    "gt_unit",
    "notes",
]

SEED_COLUMNS = [
    "paper_key",
    "doi",
    "paper_title",
    "formulation_id",
    "formulation_label_stage5",
    "formulation_label_params",
    "article_formulation_id",
    "article_formulation_label",
    "family_id",
    "parent_core_row_id",
    "variant_role",
    "payload_state",
    "paper_risk_level",
    "risk_source",
    "layer3_inclusion_flag",
    "field_name",
    "source_value_column",
    "source_value_text_column",
    "source_evidence_region_column",
    "extracted_value",
    "extracted_unit",
    "evidence_text",
    "evidence_anchor_text",
    "evidence_source_type",
    "evidence_status_detail",
    "evidence_section",
    "evidence_span_start",
    "evidence_span_end",
    "relation_resolution_rule",
    "relation_resolution_confidence",
    "relation_resolution_source_ids",
    "derivation_status",
    "derivation_rule",
    "derivation_inputs",
    "review_warning",
    "normalization_status",
    "evidence_support_status",
    "gt_status",
    "gt_value",
    "gt_unit",
    "notes",
]

LEGACY_COLUMN_ALIASES = {
    "polymer_mw_kDa_value": ["plga_mw_kDa_value"],
    "polymer_mw_kDa_value_text": ["plga_mw_kDa_value_text"],
    "polymer_mw_kDa_evidence_region_type": ["plga_mw_kDa_evidence_region_type"],
}


@dataclass(frozen=True)
class FieldSpec:
    field_name: str
    value_column: str = ""
    value_text_column: str = ""
    evidence_region_column: str = ""
    default_unit: str = ""
    include_when_blank: bool = True
    allow_derivation: bool = False


FIELD_SPECS = [
    FieldSpec(
        field_name="polymer_MW",
        value_column="polymer_mw_kDa_value",
        value_text_column="polymer_mw_kDa_value_text",
        evidence_region_column="polymer_mw_kDa_evidence_region_type",
        default_unit="kDa",
    ),
    FieldSpec(
        field_name="LA/GA",
        value_column="la_ga_ratio_value",
        value_text_column="la_ga_ratio_value_text",
        evidence_region_column="la_ga_ratio_evidence_region_type",
        default_unit="ratio",
    ),
    FieldSpec(
        field_name="drug_polymer_ratio",
        allow_derivation=True,
        default_unit="w/w",
    ),
    FieldSpec(
        field_name="surfactant_name",
        value_column="surfactant_name_value",
        value_text_column="surfactant_name_value_text",
        evidence_region_column="surfactant_name_evidence_region_type",
    ),
    FieldSpec(
        field_name="surfactant_concentration",
        value_column="surfactant_concentration_text_value",
        value_text_column="surfactant_concentration_text_value_text",
        evidence_region_column="surfactant_concentration_text_evidence_region_type",
    ),
    FieldSpec(
        field_name="solvent",
        value_column="organic_solvent_value",
        value_text_column="organic_solvent_value_text",
        evidence_region_column="organic_solvent_evidence_region_type",
    ),
    FieldSpec(
        field_name="particle_size",
        value_column="size_nm_value",
        value_text_column="size_nm_value_text",
        evidence_region_column="size_nm_evidence_region_type",
        default_unit="nm",
    ),
    FieldSpec(
        field_name="EE",
        value_column="encapsulation_efficiency_percent_value",
        value_text_column="encapsulation_efficiency_percent_value_text",
        evidence_region_column="encapsulation_efficiency_percent_evidence_region_type",
        default_unit="%",
    ),
    FieldSpec(
        field_name="LC",
        value_column="loading_content_percent_value",
        value_text_column="loading_content_percent_value_text",
        evidence_region_column="loading_content_percent_evidence_region_type",
        default_unit="%",
    ),
    FieldSpec(field_name="preparation_method", value_column="preparation_method"),
    FieldSpec(field_name="emulsion_structure", value_column="emulsion_structure"),
]


def normalize_text(value: Any) -> str:
    text = str(value or "")
    text = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def get_row_value(row: dict[str, str], column_name: str) -> str:
    for candidate in [column_name, *LEGACY_COLUMN_ALIASES.get(column_name, [])]:
        value = normalize_text(row.get(candidate))
        if value:
            return value
    return ""


def sanitize_out_subdir(value: str) -> str:
    return validate_artifact_subdir(value, param_name="--out-subdir")


def workbook_name_for_version(version: int) -> str:
    return f"field_gt_review_workbook_v{int(version)}.xlsx"


def seed_tsv_name_for_version(version: int) -> str:
    return f"field_gt_review_seed_rows_v{int(version)}.tsv"


def summary_tsv_name_for_version(version: int) -> str:
    return f"field_gt_review_source_summary_v{int(version)}.tsv"


def read_tsv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def load_manifest_map(path: Path | None) -> dict[str, dict[str, str]]:
    if path is None or not path.exists():
        return {}
    out: dict[str, dict[str, str]] = {}
    for row in read_tsv_rows(path):
        key = normalize_text(row.get("key") or row.get("paper_key") or row.get("zotero_key"))
        if key:
            out[key] = row
    return out


def load_paper_risk_map(path: Path | None) -> dict[str, dict[str, str]]:
    if path is None or not path.exists():
        return {}
    out: dict[str, dict[str, str]] = {}
    for row in read_tsv_rows(path):
        key = normalize_text(row.get("paper_key"))
        if key:
            out[key] = row
    return out


def discover_resolved_relation_fields(run_dir: Path) -> Path | None:
    candidate = run_dir / "formulation_relation_v1" / "resolved_relation_fields_v1.tsv"
    return candidate if candidate.exists() else None


def load_resolved_relation_map(path: Path | None) -> dict[tuple[str, str, str], dict[str, str]]:
    if path is None or not path.exists():
        return {}
    out: dict[tuple[str, str, str], dict[str, str]] = {}
    for row in read_tsv_rows(path):
        paper_key = normalize_text(row.get("paper_key") or row.get("key"))
        formulation_candidate_id = normalize_text(row.get("formulation_candidate_id") or row.get("formulation_id"))
        field_name = normalize_text(row.get("field_name"))
        if not paper_key or not formulation_candidate_id or not field_name:
            continue
        out[(paper_key, formulation_candidate_id, field_name)] = row
    return out


def parse_json_object_list(value: Any) -> list[dict[str, str]]:
    text = normalize_text(value)
    if not text:
        return []
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        return []
    if not isinstance(parsed, list):
        return []
    out: list[dict[str, str]] = []
    for item in parsed:
        if isinstance(item, dict):
            out.append({str(k): normalize_text(v) for k, v in item.items()})
    return out


def parse_mass_to_mg(raw: str) -> float | None:
    text = normalize_text(raw).lower().replace("渭", "u")
    if not text:
        return None
    match = re.search(r"(-?\d+(?:\.\d+)?)\s*(mg|g|ug|mcg|ng)\b", text)
    if not match:
        numeric = re.search(r"-?\d+(?:\.\d+)?", text)
        return float(numeric.group(0)) if numeric else None
    value = float(match.group(1))
    unit = match.group(2)
    if unit == "g":
        return value * 1000.0
    if unit in {"ug", "mcg"}:
        return value / 1000.0
    if unit == "ng":
        return value / 1_000_000.0
    return value


def format_number(value: float | None, decimals: int = 6) -> str:
    if value is None:
        return ""
    return f"{value:.{decimals}f}".rstrip("0").rstrip(".")


def derive_drug_polymer_ratio(row: dict[str, str]) -> tuple[str, str, str]:
    drug_raw = normalize_text(row.get("drug_feed_amount_text_value") or row.get("drug_feed_amount_text_value_text"))
    polymer_raw = normalize_text(row.get("plga_mass_mg_value") or row.get("plga_mass_mg_value_text"))
    drug_mg = parse_mass_to_mg(drug_raw)
    polymer_mg = parse_mass_to_mg(polymer_raw)
    if drug_mg is None or polymer_mg is None or polymer_mg == 0:
        return "", "", ""
    ratio = drug_mg / polymer_mg
    derivation_inputs = json.dumps(
        {
            "drug_feed_amount_text_value": drug_raw,
            "plga_mass_mg_value": polymer_raw,
        },
        ensure_ascii=True,
    )
    return format_number(ratio), "derived_from_final_row", derivation_inputs


def infer_unit_from_text(value: str) -> str:
    text = normalize_text(value).lower()
    if not text:
        return ""
    if "kda" in text:
        return "kDa"
    if "% w/v" in text or "%w/v" in text:
        return "% w/v"
    if "% w/w" in text or "%w/w" in text:
        return "% w/w"
    if "%" in text:
        return "%"
    if "mg/ml" in text:
        return "mg/mL"
    if "ug/mg" in text:
        return "ug/mg"
    if re.search(r"\bnm\b", text):
        return "nm"
    if re.search(r"\bmg\b", text):
        return "mg"
    return ""


def looks_like_polymer_grade_text(value: str) -> bool:
    text = normalize_text(value).lower()
    if not text:
        return False
    patterns = [
        r"\bresomer\b",
        r"\bpurasorb\b",
        r"\bpolymer grade\b",
        r"\bplga grade\b",
        r"\brg\s*50[0-9]{1,2}\b",
    ]
    return any(re.search(pattern, text) for pattern in patterns)


def build_article_formulation_identity(row: dict[str, str]) -> tuple[str, str]:
    article_formulation_id = normalize_text(row.get("representative_source_formulation_id"))
    article_formulation_label = normalize_text(
        row.get("representative_source_raw_formulation_label")
        or row.get("raw_formulation_label")
        or article_formulation_id
    )
    return article_formulation_id, article_formulation_label


def extract_numeric_tokens(text: str) -> list[str]:
    return re.findall(r"[-+]?\d+(?:[.,]\d+)?", normalize_text(text))


def normalize_ratio_token(value: str) -> str:
    text = normalize_text(value).lower()
    text = text.replace(":", "/").replace(" ", "")
    return text


def field_has_local_anchor_relationship(field_name: str, candidate_text: str) -> bool:
    text = normalize_text(candidate_text).lower()
    if not text:
        return False
    keyword_map = {
        "polymer_MW": ["mw", "molecular weight", "kda", "plga", "polymer", "viscosity"],
        "LA/GA": ["la/ga", "lactic", "glycolic", "lactide", "glycolide", "plga"],
        "surfactant_concentration": ["surfactant", "stabilizer", "pva", "polysorbate", "labrafil", "poloxamer", "concentration"],
        "surfactant_name": ["surfactant", "stabilizer", "pva", "polysorbate", "labrafil", "poloxamer", "pluronic"],
        "solvent": ["solvent", "acetone", "dichloromethane", "ethyl acetate", "organic phase"],
        "particle_size": ["size", "nm", "diameter", "particle"],
        "EE": ["encapsulation efficiency", "ee", "entrapment efficiency"],
        "LC": ["loading content", "drug loading", "lc", "dl"],
        "drug_polymer_ratio": ["ratio", "drug/polymer", "drug to polymer", "feed"],
    }
    return any(token in text for token in keyword_map.get(field_name, []))


def field_value_matches_text(field_name: str, extracted_value: str, candidate_text: str) -> bool:
    value = normalize_text(extracted_value)
    text = normalize_text(candidate_text)
    if not value or not text:
        return False
    lowered = text.lower()
    if field_name == "LA/GA":
        ratio = normalize_ratio_token(value)
        if not ratio or "/" not in ratio:
            return False
        compact_text = lowered.replace(" ", "").replace(":", "/")
        return ratio in compact_text
    if field_name == "polymer_MW":
        value_nums = extract_numeric_tokens(value)
        text_nums = set(extract_numeric_tokens(text))
        if value_nums and not all(token in text_nums for token in value_nums):
            return False
        return any(token in lowered for token in ["kda", "da", "mw", "molecular weight", "viscosity"])
    value_nums = extract_numeric_tokens(value)
    if value_nums:
        text_nums = set(extract_numeric_tokens(text))
        return all(token in text_nums for token in value_nums)
    pattern = r"(?<![A-Za-z0-9])" + re.escape(value.lower()) + r"(?![A-Za-z0-9])"
    return re.search(pattern, text.lower()) is not None


def make_safe_token(value: str) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = text.replace("% w/v", "%w/v").replace("% w/w", "%w/w").replace("%", "pct")
    text = re.sub(r"[^A-Za-z0-9/+.-]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def compact_family_token(family_index: int | None, family_id: str, formulation_id: str) -> str:
    if family_index is not None and family_index > 0:
        return f"F{family_index:02d}"
    family = normalize_text(family_id)
    formulation = normalize_text(formulation_id)
    if family:
        suffix = family.split("::", 1)[-1]
        if suffix:
            return make_safe_token(suffix)
    if formulation:
        return make_safe_token(formulation.split("__fo__", 1)[0])
    return ""


def compact_role_token(value: str) -> str:
    token = normalize_text(value).lower()
    mapping = {
        "family_core": "core",
        "family_variant": "variant",
        "payload_variant": "payload_variant",
        "comparator": "comparator",
    }
    if token in mapping:
        return mapping[token]
    return make_safe_token(token)


def compact_payload_token(value: str) -> str:
    token = normalize_text(value).lower()
    mapping = {
        "drug_loaded": "drug",
        "blank_control": "blank",
        "unknown": "unknown",
    }
    if token in mapping:
        return mapping[token]
    return make_safe_token(token)


def build_stage5_label(row: dict[str, str], family_index: int | None) -> str:
    parts = [
        compact_family_token(family_index, row.get("family_id", ""), row.get("final_formulation_id", "")),
        compact_role_token(row.get("variant_role", "")),
        compact_payload_token(row.get("payload_state", "")),
    ]
    return "__".join(part for part in parts if part)


def abbreviate_polymer_identity(polymer: str, la_ga_ratio: str) -> str:
    polymer_token = make_safe_token(polymer).upper()
    ratio = normalize_text(la_ga_ratio)
    if polymer_token == "PLGA" and ratio:
        return f"PLGA{make_safe_token(ratio)}"
    if polymer_token:
        return polymer_token
    return ""


def build_parameter_label(row: dict[str, str]) -> str:
    parts: list[str] = []
    polymer_token = abbreviate_polymer_identity(
        row.get("polymer_identity_final", ""),
        row.get("la_ga_ratio_value", ""),
    )
    if polymer_token:
        parts.append(polymer_token)
    polymer_mass = normalize_text(row.get("plga_mass_mg_value"))
    if polymer_mass:
        parts.append(f"Poly{make_safe_token(polymer_mass)}")
    surfactant_name = normalize_text(row.get("surfactant_name_value"))
    surfactant_conc = normalize_text(row.get("surfactant_concentration_text_value"))
    if surfactant_name or surfactant_conc:
        surf_token = make_safe_token(surfactant_name or "surfactant")
        conc_token = make_safe_token(surfactant_conc)
        parts.append(f"{surf_token}{conc_token}" if conc_token else surf_token)
    drug_amount = normalize_text(row.get("drug_feed_amount_text_value"))
    if drug_amount:
        parts.append(f"Drug{make_safe_token(drug_amount)}")
    solvent = normalize_text(row.get("organic_solvent_value"))
    if solvent and solvent.lower() not in {"unknown", "not reported", "na", "n/a"}:
        parts.append(make_safe_token(solvent))
    prep = normalize_text(row.get("preparation_method"))
    if prep and prep.lower() not in {"unknown", "none"} and len(parts) < 4:
        parts.append(make_safe_token(prep))
    return "_".join(part for part in parts if part)[:72]


def split_sentences(text: str) -> list[str]:
    compact = normalize_text(text)
    if not compact:
        return []
    parts = re.split(r"(?<=[.!?])\s+", compact)
    return [part.strip() for part in parts if part.strip()]


def pick_relevant_sentence(field_name: str, text: str, extracted_value: str) -> str:
    sentences = split_sentences(text)
    if not sentences:
        return ""
    for sentence in sentences:
        if field_value_matches_text(field_name, extracted_value, sentence):
            return sentence
    return sentences[0]


def has_title_like_noise(text: str) -> bool:
    compact = normalize_text(text).lower()
    noisy_markers = [
        "to link to this article",
        "download full-size image",
        "previous article in issue",
        "published:",
        "received:",
    ]
    return any(marker in compact for marker in noisy_markers)


def map_evidence_source_type(value: str) -> str:
    text = normalize_text(value).lower()
    if text.startswith("table"):
        return "table"
    if text.startswith("figure"):
        return "figure"
    if text:
        return "text"
    return "unknown"


def is_table_derived_field(row: dict[str, str], spec: FieldSpec) -> bool:
    if spec.evidence_region_column:
        region = get_row_value(row, spec.evidence_region_column)
        if region.lower().startswith("table"):
            return True
    return False


def field_target_values(spec: FieldSpec, extracted_value: str) -> dict[str, str]:
    if not extracted_value:
        return {}
    mapping = {
        "particle_size": {"size": extracted_value},
        "EE": {"ee": extracted_value},
        "LC": {"ee": extracted_value},
        "surfactant_concentration": {"surfactant": extracted_value},
        "polymer_MW": {"polymer": extracted_value},
        "drug_polymer_ratio": {"drug": extracted_value},
    }
    return mapping.get(spec.field_name, {})


def find_matching_table_ref(row: dict[str, str], field_name: str, extracted_value: str) -> dict[str, str] | None:
    refs = parse_json_object_list(row.get("supporting_evidence_refs"))
    preferred_region_order = ["table_cell", "table_row"] if extract_numeric_tokens(extracted_value) else ["table_cell"]
    for region_type in preferred_region_order:
        for ref in refs:
            if normalize_text(ref.get("region_type")).lower() != region_type:
                continue
            span_text = normalize_text(ref.get("span_text"))
            if field_value_matches_text(field_name, extracted_value, span_text):
                return ref
    return None


def evidence_supports_value(field_name: str, extracted_value: str, evidence_text: str) -> bool:
    value = normalize_text(extracted_value)
    evidence = normalize_text(evidence_text)
    if not value:
        return not evidence
    if not evidence:
        return False
    return field_value_matches_text(field_name, value, evidence)


def resolve_text_field_evidence(
    resolver: AuditEvidenceResolverV1,
    row: dict[str, str],
    spec: FieldSpec,
    extracted_value: str,
    max_span_chars: int,
) -> tuple[str, str]:
    if not extracted_value:
        return "", "blank_value"
    numeric_tokens = extract_numeric_tokens(extracted_value)
    if numeric_tokens:
        text_evidence = resolver.resolve_text_evidence_numeric_locate(
            zotero_key=normalize_text(row.get("key")),
            evidence_pointer_raw="",
            numeric_values=numeric_tokens,
            max_span_chars=max_span_chars,
        )
        candidate = normalize_text(text_evidence.evidence_text)
        if candidate and not has_title_like_noise(candidate):
            sentence = pick_relevant_sentence(spec.field_name, candidate, extracted_value)
            if sentence and evidence_supports_value(spec.field_name, extracted_value, sentence):
                return sentence, "text"
    text_evidence = resolver.resolve_text_evidence(
        zotero_key=normalize_text(row.get("key")),
        evidence_span_start=row.get("evidence_span_start", ""),
        evidence_span_end=row.get("evidence_span_end", ""),
        evidence_section=row.get("evidence_section", ""),
        evidence_pointer_raw="",
        max_span_chars=max_span_chars,
        fallback_hint_text=normalize_text(row.get("evidence_span_text")),
    )
    evidence_text = normalize_text(text_evidence.evidence_text) or normalize_text(row.get("evidence_span_text"))
    evidence_text = pick_relevant_sentence(spec.field_name, evidence_text, extracted_value) if evidence_text else ""
    if has_title_like_noise(evidence_text):
        evidence_text = ""
    if evidence_supports_value(spec.field_name, extracted_value, evidence_text):
        return evidence_text, "text"
    return "", "unsupported_text"


def build_table_evidence_text(table_cell_text: str, table_row_text: str, max_chars: int) -> str:
    cell = normalize_text(table_cell_text)
    row = normalize_text(table_row_text)
    if cell and row and cell not in row:
        return normalize_text(f"cell={cell} | row={row}")[:max_chars]
    return (cell or row)[:max_chars]


def resolve_field_evidence(
    resolver: AuditEvidenceResolverV1,
    row: dict[str, str],
    manifest_row: dict[str, str],
    spec: FieldSpec,
    extracted_value: str,
    max_span_chars: int,
) -> tuple[str, str]:
    if spec.allow_derivation:
        return "", "unknown"

    if not extracted_value:
        return "", "blank_value"

    if is_table_derived_field(row, spec):
        direct_ref = find_matching_table_ref(row, spec.field_name, extracted_value)
        if direct_ref is not None:
            evidence_text = build_table_evidence_text(
                table_cell_text=direct_ref.get("span_text", "") if normalize_text(direct_ref.get("region_type")).lower() == "table_cell" else "",
                table_row_text=direct_ref.get("span_text", ""),
                max_chars=max_span_chars,
            )
            if evidence_text and evidence_supports_value(spec.field_name, extracted_value, evidence_text):
                return evidence_text, "table"
        return "", "unresolved_table"

    return resolve_text_field_evidence(
        resolver=resolver,
        row=row,
        spec=spec,
        extracted_value=extracted_value,
        max_span_chars=max_span_chars,
    )


def extract_evidence_anchor_text(
    row: dict[str, str],
    spec: FieldSpec,
    extracted_value: str,
    max_span_chars: int,
) -> str:
    def with_section(section: str, text: str) -> str:
        clean_section = normalize_text(section)
        clean_text = normalize_text(text)
        if clean_section and clean_text:
            return f"[{clean_section}] {clean_text}"[:max_span_chars]
        return clean_text[:max_span_chars]

    if normalize_text(get_row_value(row, spec.evidence_region_column)).lower() == "relation_resolved":
        return ""

    refs = parse_json_object_list(row.get("supporting_evidence_refs"))
    for ref in refs:
        span_text = normalize_text(ref.get("span_text"))
        if span_text and field_value_matches_text(spec.field_name, extracted_value, span_text):
            sentence = pick_relevant_sentence(spec.field_name, span_text[:max_span_chars], extracted_value)
            preview = sentence if len(sentence) >= 20 else span_text[:max_span_chars]
            return with_section(ref.get("section", ""), preview)
    fallback = normalize_text(row.get("evidence_span_text"))
    if fallback and not has_title_like_noise(fallback) and field_has_local_anchor_relationship(spec.field_name, fallback):
        sentence = pick_relevant_sentence(spec.field_name, fallback[:max_span_chars], extracted_value)
        if field_value_matches_text(spec.field_name, extracted_value, sentence):
            preview = sentence if len(sentence) >= 20 else fallback[:max_span_chars]
            return with_section(row.get("evidence_section", ""), preview)
    return ""


def classify_evidence_status_detail(
    extracted_value: str,
    evidence_text: str,
    evidence_anchor_text: str,
    evidence_source_type: str,
    derivation_status: str,
) -> str:
    if not extracted_value:
        return "blank_value"
    if derivation_status and derivation_status != "none":
        return "derived_without_direct_text"
    if evidence_text:
        return "supported"
    if evidence_source_type == "unresolved_table":
        return "unresolved_table"
    if evidence_source_type == "unsupported_text":
        return "unsupported_text" if evidence_anchor_text else "missing_evidence_anchor"
    return "missing_evidence_anchor" if not evidence_anchor_text else normalize_text(evidence_source_type) or "unknown"


def choose_extracted_value(
    row: dict[str, str], spec: FieldSpec
) -> tuple[str, str, str, str, str, str, str, str]:
    if spec.allow_derivation:
        value, derivation_status, derivation_inputs = derive_drug_polymer_ratio(row)
        return value, spec.default_unit, derivation_status, derivation_inputs, "", "", "", ""

    primary_value = get_row_value(row, spec.value_column) if spec.value_column else ""
    fallback_text = get_row_value(row, spec.value_text_column) if spec.value_text_column else ""
    review_warning = ""
    if spec.field_name == "polymer_MW" and not primary_value and looks_like_polymer_grade_text(fallback_text):
        review_warning = "polymer_grade_recorded_not_mw"
        return "", "", "none", "", primary_value, fallback_text, review_warning, ""
    extracted_value = primary_value or fallback_text
    extracted_unit = spec.default_unit or infer_unit_from_text(primary_value or fallback_text)
    normalization_status = ""
    if spec.field_name == "surfactant_concentration" and extracted_value:
        if extracted_unit in {"mg", "g"}:
            normalization_status = "normalization_pending"
        elif extracted_unit in {"%", "% w/v", "% w/w", "mg/mL", "ug/mg"}:
            normalization_status = "reported_as_concentration"
    return extracted_value, extracted_unit, "none", "", primary_value, fallback_text, review_warning, normalization_status


def workbook_field_to_relation_field(field_name: str) -> str:
    mapping = {
        "polymer_MW": "polymer_mw_kDa",
        "LA/GA": "la_ga_ratio",
        "surfactant_name": "surfactant_name",
        "solvent": "organic_solvent",
    }
    return mapping.get(field_name, "")


def lookup_relation_resolution(
    resolved_relation_map: dict[tuple[str, str, str], dict[str, str]],
    paper_key: str,
    representative_source_formulation_id: str,
    field_name: str,
) -> dict[str, str]:
    relation_field_name = workbook_field_to_relation_field(field_name)
    if not relation_field_name:
        return {}
    return resolved_relation_map.get((paper_key, representative_source_formulation_id, relation_field_name), {})


def build_seed_rows(
    final_rows: list[dict[str, str]],
    manifest_map: dict[str, dict[str, str]],
    paper_risk_map: dict[str, dict[str, str]],
    resolved_relation_map: dict[tuple[str, str, str], dict[str, str]],
    resolver: AuditEvidenceResolverV1,
    max_span_chars: int,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    workbook_rows: list[dict[str, str]] = []
    seed_rows: list[dict[str, str]] = []
    family_index_map: dict[tuple[str, str], int] = {}
    next_family_index_by_paper: dict[str, int] = {}
    for row in final_rows:
        paper_key = normalize_text(row.get("key"))
        family_id = normalize_text(row.get("family_id"))
        if not paper_key or not family_id:
            continue
        family_key = (paper_key, family_id)
        if family_key in family_index_map:
            continue
        next_index = next_family_index_by_paper.get(paper_key, 0) + 1
        next_family_index_by_paper[paper_key] = next_index
        family_index_map[family_key] = next_index
    for row in final_rows:
        paper_key = normalize_text(row.get("key"))
        manifest_row = manifest_map.get(paper_key, {})
        risk_row = paper_risk_map.get(paper_key, {})
        doi = normalize_text(manifest_row.get("doi") or row.get("doi"))
        paper_title = normalize_text(manifest_row.get("title") or manifest_row.get("paper_title"))
        formulation_id = normalize_text(row.get("final_formulation_id"))
        representative_source_formulation_id = normalize_text(row.get("representative_source_formulation_id"))
        article_formulation_id, article_formulation_label = build_article_formulation_identity(row)
        paper_risk_level = normalize_text(risk_row.get("paper_risk_level"))
        risk_source = normalize_text(risk_row.get("risk_source"))
        layer3_inclusion_flag = normalize_text(risk_row.get("layer3_inclusion_flag"))
        family_index = family_index_map.get((paper_key, normalize_text(row.get("family_id"))))
        stage5_label = build_stage5_label(row, family_index)
        parameter_label = build_parameter_label(row)
        for spec in FIELD_SPECS:
            (
                extracted_value,
                extracted_unit,
                derivation_status,
                derivation_inputs,
                primary_value,
                fallback_text,
                review_warning,
                normalization_status,
            ) = choose_extracted_value(row, spec)
            relation_resolution = lookup_relation_resolution(
                resolved_relation_map=resolved_relation_map,
                paper_key=paper_key,
                representative_source_formulation_id=representative_source_formulation_id,
                field_name=spec.field_name,
            )
            evidence_text, evidence_source_type = resolve_field_evidence(
                resolver=resolver,
                row=row,
                manifest_row=manifest_row,
                spec=spec,
                extracted_value=extracted_value,
                max_span_chars=max_span_chars,
            )
            derivation_rule = ""
            field_evidence_text = evidence_text
            evidence_anchor_text = ""
            evidence_support_status = "supported" if evidence_supports_value(spec.field_name, extracted_value, field_evidence_text) else ""
            if spec.allow_derivation:
                derivation_rule = "drug_feed_amount_text_value / plga_mass_mg_value"
                if extracted_value:
                    field_evidence_text = (
                        "Derived from final-row values: "
                        + f"drug_feed_amount_text_value={normalize_text(row.get('drug_feed_amount_text_value') or row.get('drug_feed_amount_text_value_text'))}; "
                        + f"plga_mass_mg_value={normalize_text(row.get('plga_mass_mg_value') or row.get('plga_mass_mg_value_text'))}"
                    )
                    evidence_source_type = "derived"
                    evidence_support_status = "derived_from_final_row"
                else:
                    evidence_source_type = "blank_value"
                    evidence_support_status = "blank_extracted_value"
            elif not extracted_value:
                field_evidence_text = ""
                evidence_support_status = "blank_extracted_value"
            elif evidence_source_type in {"unresolved_table", "unsupported_text"}:
                evidence_anchor_text = extract_evidence_anchor_text(
                    row=row,
                    spec=spec,
                    extracted_value=primary_value or fallback_text or extracted_value,
                    max_span_chars=max_span_chars,
                )
                field_evidence_text = ""
                evidence_support_status = evidence_source_type
            elif evidence_supports_value(spec.field_name, extracted_value, field_evidence_text):
                evidence_support_status = "supported"
            else:
                evidence_anchor_text = extract_evidence_anchor_text(
                    row=row,
                    spec=spec,
                    extracted_value=primary_value or fallback_text or extracted_value,
                    max_span_chars=max_span_chars,
                )
                field_evidence_text = ""
                evidence_source_type = "unsupported_text"
                evidence_support_status = "unsupported_text"
            evidence_status_detail = classify_evidence_status_detail(
                extracted_value=extracted_value,
                evidence_text=field_evidence_text,
                evidence_anchor_text=evidence_anchor_text,
                evidence_source_type=evidence_source_type,
                derivation_status=derivation_status,
            )

            workbook_rows.append(
                {
                    "paper_key": paper_key,
                    "formulation_id": formulation_id,
                    "formulation_label_stage5": stage5_label,
                    "formulation_label_params": parameter_label,
                    "article_formulation_id": article_formulation_id,
                    "article_formulation_label": article_formulation_label,
                    "paper_risk_level": paper_risk_level,
                    "layer3_inclusion_flag": layer3_inclusion_flag,
                    "field_name": spec.field_name,
                    "extracted_value": extracted_value,
                    "extracted_unit": extracted_unit,
                    "evidence_text": field_evidence_text,
                    "evidence_anchor_text": evidence_anchor_text,
                    "evidence_source_type": evidence_source_type,
                    "evidence_status_detail": evidence_status_detail,
                    "review_warning": review_warning,
                    "normalization_status": normalization_status,
                    "gt_status": "",
                    "gt_value": "",
                    "gt_unit": "",
                    "notes": "",
                }
            )
            seed_rows.append(
                {
                    "paper_key": paper_key,
                    "doi": doi,
                    "paper_title": paper_title,
                    "formulation_id": formulation_id,
                    "formulation_label_stage5": stage5_label,
                    "formulation_label_params": parameter_label,
                    "article_formulation_id": article_formulation_id,
                    "article_formulation_label": article_formulation_label,
                    "family_id": normalize_text(row.get("family_id")),
                    "parent_core_row_id": normalize_text(row.get("parent_core_row_id")),
                    "variant_role": normalize_text(row.get("variant_role")),
                    "payload_state": normalize_text(row.get("payload_state")),
                    "paper_risk_level": paper_risk_level,
                    "risk_source": risk_source,
                    "layer3_inclusion_flag": layer3_inclusion_flag,
                    "field_name": spec.field_name,
                    "source_value_column": spec.value_column,
                    "source_value_text_column": spec.value_text_column,
                    "source_evidence_region_column": spec.evidence_region_column,
                    "extracted_value": extracted_value,
                    "extracted_unit": extracted_unit,
                    "evidence_text": field_evidence_text,
                    "evidence_anchor_text": evidence_anchor_text,
                    "evidence_source_type": evidence_source_type,
                    "evidence_status_detail": evidence_status_detail,
                    "evidence_section": normalize_text(row.get("evidence_section")),
                    "evidence_span_start": normalize_text(row.get("evidence_span_start")),
                    "evidence_span_end": normalize_text(row.get("evidence_span_end")),
                    "relation_resolution_rule": normalize_text(relation_resolution.get("resolution_rule")),
                    "relation_resolution_confidence": normalize_text(relation_resolution.get("deterministic_confidence")),
                    "relation_resolution_source_ids": normalize_text(relation_resolution.get("source_relation_row_ids")),
                    "derivation_status": derivation_status,
                    "derivation_rule": derivation_rule,
                    "derivation_inputs": derivation_inputs,
                    "review_warning": review_warning,
                    "normalization_status": normalization_status,
                    "evidence_support_status": evidence_support_status,
                    "gt_status": "",
                    "gt_value": "",
                    "gt_unit": "",
                    "notes": "",
                }
            )
    return workbook_rows, seed_rows


def build_source_summary(seed_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in seed_rows:
        paper_key = row["paper_key"]
        bucket = grouped.setdefault(
            paper_key,
            {
                "paper_key": paper_key,
                "doi": row["doi"],
                "paper_title": row["paper_title"],
                "paper_risk_level": row.get("paper_risk_level", ""),
                "risk_source": row.get("risk_source", ""),
                "layer3_inclusion_flag": row.get("layer3_inclusion_flag", ""),
                "formulation_count": set(),
                "field_rows": 0,
            },
        )
        bucket["formulation_count"].add(row["formulation_id"])
        bucket["field_rows"] += 1
    out: list[dict[str, str]] = []
    for paper_key in sorted(grouped):
        bucket = grouped[paper_key]
        out.append(
            {
                "paper_key": paper_key,
                "doi": bucket["doi"],
                "paper_title": bucket["paper_title"],
                "paper_risk_level": bucket["paper_risk_level"],
                "risk_source": bucket["risk_source"],
                "layer3_inclusion_flag": bucket["layer3_inclusion_flag"],
                "formulation_count": str(len(bucket["formulation_count"])),
                "field_rows": str(bucket["field_rows"]),
            }
        )
    return out


def write_sheet_rows(ws, columns: list[str], rows: list[dict[str, str]]) -> dict[str, int]:
    for column_index, column_name in enumerate(columns, start=1):
        ws.cell(row=1, column=column_index, value=column_name)
    for row_index, row in enumerate(rows, start=2):
        for column_index, column_name in enumerate(columns, start=1):
            ws.cell(row=row_index, column=column_index, value=row.get(column_name, ""))
    return {column_name: index for index, column_name in enumerate(columns, start=1)}


def create_dropdown_sheet(ws) -> None:
    ws["A1"] = "gt_status_options"
    for index, value in enumerate(GT_STATUS_OPTIONS, start=2):
        ws[f"A{index}"] = value
    ws["B1"] = "gt_unit_options"
    for index, value in enumerate(GT_UNIT_OPTIONS, start=2):
        ws[f"B{index}"] = value
    ws.sheet_state = "hidden"


def add_data_validation(ws, header_map: dict[str, int], row_start: int, row_end: int) -> None:
    status_letter = get_column_letter(header_map["gt_status"])
    unit_letter = get_column_letter(header_map["gt_unit"])
    dv_status = DataValidation(type="list", formula1="=dropdown_options!$A$2:$A$5", allow_blank=True)
    dv_unit = DataValidation(
        type="list",
        formula1=f"=dropdown_options!$B$2:$B${len(GT_UNIT_OPTIONS) + 1}",
        allow_blank=True,
        showErrorMessage=False,
    )
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_unit)
    dv_status.add(f"{status_letter}{row_start}:{status_letter}{row_end}")
    dv_unit.add(f"{unit_letter}{row_start}:{unit_letter}{row_end}")


def style_review_sheet(ws) -> None:
    group_fills = {
        1: PatternFill("solid", fgColor="1F2937"),
        2: PatternFill("solid", fgColor="0F766E"),
        3: PatternFill("solid", fgColor="7C2D12"),
        4: PatternFill("solid", fgColor="7E22CE"),
    }
    locked_fill = PatternFill("solid", fgColor="F8FAFC")
    editable_fill = PatternFill("solid", fgColor="FFF7ED")
    groups = {
        "paper_key": 1,
        "formulation_id": 1,
        "formulation_label_stage5": 1,
        "formulation_label_params": 1,
        "article_formulation_id": 1,
        "article_formulation_label": 1,
        "paper_risk_level": 1,
        "layer3_inclusion_flag": 1,
        "field_name": 1,
        "extracted_value": 2,
        "extracted_unit": 2,
        "evidence_text": 3,
        "evidence_anchor_text": 3,
        "evidence_source_type": 3,
        "evidence_status_detail": 3,
        "review_warning": 3,
        "normalization_status": 3,
        "gt_status": 4,
        "gt_value": 4,
        "gt_unit": 4,
        "notes": 4,
    }
    editable_columns = {"gt_status", "gt_value", "gt_unit", "notes"}
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "J2"
    for cell in ws[1]:
        header = normalize_text(cell.value)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = group_fills[groups[header]]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            header = normalize_text(ws.cell(row=1, column=cell.column).value)
            editable = header in editable_columns
            cell.protection = Protection(locked=not editable)
            cell.fill = editable_fill if editable else locked_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.protection.sheet = True
    ws.protection.autoFilter = True
    ws.protection.sort = True
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True
    widths = {
        "A": 14,
        "B": 18,
        "C": 28,
        "D": 32,
        "E": 20,
        "F": 24,
        "G": 12,
        "H": 14,
        "I": 16,
        "J": 18,
        "K": 12,
        "L": 56,
        "M": 56,
        "N": 18,
        "O": 22,
        "P": 22,
        "Q": 22,
        "R": 14,
        "S": 18,
        "T": 12,
        "U": 30,
    }
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width


def style_simple_sheet(ws) -> None:
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in ws.columns:
        max_len = max(len(normalize_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 12), 42)


def build_workbook(
    workbook_path: Path,
    workbook_rows: list[dict[str, str]],
    seed_rows: list[dict[str, str]],
    source_summary_rows: list[dict[str, str]],
) -> None:
    wb = Workbook()
    ws_review = wb.active
    ws_review.title = "field_gt_review"
    ws_seed = wb.create_sheet("seed_reference")
    ws_summary = wb.create_sheet("source_summary")
    ws_instr = wb.create_sheet("instructions")
    ws_dropdown = wb.create_sheet("dropdown_options")

    review_header = write_sheet_rows(ws_review, WORKBOOK_COLUMNS, workbook_rows)
    write_sheet_rows(ws_seed, SEED_COLUMNS, seed_rows)
    summary_columns = list(source_summary_rows[0].keys()) if source_summary_rows else ["paper_key", "doi", "paper_title"]
    write_sheet_rows(ws_summary, summary_columns, source_summary_rows)

    create_dropdown_sheet(ws_dropdown)
    if ws_review.max_row >= 2:
        add_data_validation(ws_review, review_header, 2, ws_review.max_row)
    style_review_sheet(ws_review)
    style_simple_sheet(ws_seed)
    style_simple_sheet(ws_summary)

    instructions = [
        "Layer 3 field GT starts from frozen Stage 5 formulation rows and does not change row identity.",
        "Each row is one `(formulation_id, field_name)` review item.",
        "Columns A-I are frozen so paper, system identity, article-native labels, Layer 2 risk flags, and field_name stay visible while reviewing.",
        "Use gt_status dropdown values: correct, incorrect, not_reported, unclear.",
        "gt_unit provides common unit suggestions but allows free-text fallback when needed.",
        "formulation_id and formulation_label_stage5 remain the canonical system identity columns.",
        "article_formulation_id and article_formulation_label are reviewer aids only and do not replace system identity.",
        "evidence_text holds direct supporting evidence only when the extracted value is directly supported.",
        "evidence_anchor_text may preserve the closest available source anchor for manual lookup even when direct support is missing.",
        "Rows with unresolved_table, unsupported_text, derived_without_direct_text, or missing_evidence_anchor need manual paper lookup before GT completion.",
        "drug_polymer_ratio is a deterministic derived seed only when final-row drug and polymer masses are both present.",
        "polymer_grade_recorded_not_mw flags product-grade text that was carried in polymer MW text fields upstream and is intentionally not surfaced as a molecular-weight value here.",
        "normalization_pending marks concentration-like fields that are still reported as raw mass and need manual interpretation rather than silent normalization.",
        "This workbook does not yet seed phase-ratio review rows because the frozen final table does not carry a safe explicit phase-ratio field.",
        "paper_risk_level and layer3_inclusion_flag come from the post-comparison Layer 2 risk stratification layer and help prioritize field review only.",
        "Risk metadata does not change Stage 5 inclusion or benchmark-valid final rows.",
    ]
    ws_instr["A1"] = "Layer 3 Field GT Review Instructions"
    ws_instr["A1"].font = Font(bold=True)
    for index, text in enumerate(instructions, start=3):
        ws_instr[f"A{index}"] = text
    ws_instr.column_dimensions["A"].width = 140
    ws_instr.freeze_panes = "A3"

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--run-id",
        default="",
        help="Optional exact run_id compatibility alias. When omitted, the repository active source pointer is used.",
    )
    parser.add_argument(
        "--out-subdir",
        required=True,
        help="Relative output folder under data/results/<run_id>/ for this review surface.",
    )
    parser.add_argument(
        "--run-dir",
        type=Path,
        default=None,
        help="Optional explicit run directory when this run lives under a lineage child path.",
    )
    parser.add_argument(
        "--final-table-tsv",
        type=Path,
        default=None,
        help="Explicit frozen final_formulation_table_v1.tsv to seed from.",
    )
    parser.add_argument(
        "--scope-manifest-tsv",
        type=Path,
        default=None,
        help="Optional scope manifest TSV for DOI/title enrichment.",
    )
    parser.add_argument(
        "--paper-risk-tsv",
        type=Path,
        default=None,
        help="Optional Layer 2 paper risk assessment TSV for Layer 3 review prioritization metadata.",
    )
    parser.add_argument(
        "--resolved-relation-fields-tsv",
        type=Path,
        default=None,
        help="Optional Stage 3 resolved_relation_fields_v1.tsv for relation-resolved field provenance carry-through.",
    )
    parser.add_argument(
        "--max-span-chars",
        type=int,
        default=320,
        help="Maximum evidence text characters carried into the workbook.",
    )
    parser.add_argument(
        "--artifact-version",
        type=int,
        default=1,
        help="Version number for workbook and companion Layer 3 review artifacts.",
    )
    return parser


def main() -> None:
    args = build_arg_parser().parse_args()
    run_context = resolve_run_context(
        explicit_run_dir=args.run_dir,
        explicit_run_id=normalize_text(args.run_id),
    )
    run_id = str(run_context["run_id"])
    run_dir = Path(run_context["run_dir"])
    out_subdir = sanitize_out_subdir(args.out_subdir)
    out_dir = run_dir / out_subdir
    artifact_version = max(1, int(args.artifact_version))
    final_table_tsv = resolve_artifact_path(
        explicit_path=args.final_table_tsv,
        run_context=run_context,
        pointer_key="stage5_final_table_tsv",
        canonical_relative="final_formulation_table_v1.tsv",
    )
    scope_manifest_tsv = resolve_artifact_path(
        explicit_path=args.scope_manifest_tsv,
        run_context=run_context,
        pointer_key="scope_manifest_tsv",
        preferred_run_local_names=["dev15_scope.tsv", "scope.tsv", "scope_manifest.tsv"],
        required=False,
    )
    paper_risk_tsv = args.paper_risk_tsv
    resolved_relation_fields_tsv = resolve_artifact_path(
        explicit_path=args.resolved_relation_fields_tsv,
        run_context=run_context,
        pointer_key="resolved_relation_fields_tsv",
        canonical_relative="formulation_relation_v1/resolved_relation_fields_v1.tsv",
        required=False,
    )

    print(
        json.dumps(
            {
                "resolved_source_run_dir": str(run_dir),
                "resolved_source_run_id": run_id,
                "source_resolution": str(run_context["resolution_source"]),
                "active_run_pointer_path": str(run_context.get("pointer_path") or ""),
                "resolved_input_files": {
                    "final_table_tsv": str(final_table_tsv),
                    "scope_manifest_tsv": str(scope_manifest_tsv) if scope_manifest_tsv else "",
                    "paper_risk_tsv": str(paper_risk_tsv) if paper_risk_tsv else "",
                    "resolved_relation_fields_tsv": str(resolved_relation_fields_tsv) if resolved_relation_fields_tsv else "",
                },
            },
            indent=2,
        )
    )

    final_rows = read_tsv_rows(final_table_tsv)
    manifest_map = load_manifest_map(scope_manifest_tsv)
    paper_risk_map = load_paper_risk_map(paper_risk_tsv)
    resolved_relation_map = load_resolved_relation_map(resolved_relation_fields_tsv)
    resolver = AuditEvidenceResolverV1(project_root=Path(__file__).resolve().parents[2])
    workbook_rows, seed_rows = build_seed_rows(
        final_rows=final_rows,
        manifest_map=manifest_map,
        paper_risk_map=paper_risk_map,
        resolved_relation_map=resolved_relation_map,
        resolver=resolver,
        max_span_chars=max(120, int(args.max_span_chars)),
    )
    source_summary_rows = build_source_summary(seed_rows)

    seed_tsv_name = seed_tsv_name_for_version(artifact_version)
    summary_tsv_name = summary_tsv_name_for_version(artifact_version)
    workbook_name = workbook_name_for_version(artifact_version)

    write_tsv(out_dir / seed_tsv_name, SEED_COLUMNS, seed_rows)
    if source_summary_rows:
        write_tsv(out_dir / summary_tsv_name, list(source_summary_rows[0].keys()), source_summary_rows)

    workbook_path = out_dir / workbook_name
    build_workbook(
        workbook_path=workbook_path,
        workbook_rows=workbook_rows,
        seed_rows=seed_rows,
        source_summary_rows=source_summary_rows,
    )
    metadata_path = write_artifact_metadata_json(
        workbook_path,
        build_artifact_metadata(
            source_run_context=run_context,
            source_files={
                "final_table_tsv": str(final_table_tsv),
                "scope_manifest_tsv": str(scope_manifest_tsv) if scope_manifest_tsv else "",
                "paper_risk_tsv": str(paper_risk_tsv) if paper_risk_tsv else "",
                "resolved_relation_fields_tsv": str(resolved_relation_fields_tsv) if resolved_relation_fields_tsv else "",
            },
            generated_by="src/stage5_benchmark/build_field_gt_review_workbook_v1.py",
            note="Layer3 field GT workbook authority metadata.",
            extra={
                "seed_rows_tsv": str(out_dir / seed_tsv_name),
                "source_summary_tsv": str(out_dir / summary_tsv_name) if source_summary_rows else "",
            },
        ),
    )
    print(
        json.dumps(
            {
                "run_id": run_id,
                "run_dir": str(run_dir),
                "out_dir": str(out_dir),
                "final_table_tsv": str(final_table_tsv),
                "scope_manifest_tsv": str(scope_manifest_tsv) if scope_manifest_tsv else "",
                "paper_risk_tsv": str(paper_risk_tsv) if paper_risk_tsv else "",
                "resolved_relation_fields_tsv": str(resolved_relation_fields_tsv) if resolved_relation_fields_tsv else "",
                "workbook_path": str(workbook_path),
                "workbook_metadata_json": str(metadata_path),
                "field_rows": len(workbook_rows),
                "formulation_rows": len(final_rows),
                "field_catalog": [spec.field_name for spec in FIELD_SPECS],
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
