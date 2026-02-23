#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation


DEFAULT_RUN_ID = "run_20260219_1623_780eb83_goren18_weaklabels_v1"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export benchmark audit TSV(s) to a reviewer-friendly XLSX workbook."
    )
    parser.add_argument("--run-id", required=True, help="Run identifier under data/results/<run_id>/...")
    parser.add_argument(
        "--audit-tsv",
        default="",
        help="Default: data/results/<run_id>/benchmark_goren_2025/audit_parsing_derivation_samples.tsv",
    )
    parser.add_argument(
        "--alignment-tsv",
        default="",
        help="Default: data/results/<run_id>/benchmark_goren_2025/alignment_rows.tsv (optional)",
    )
    parser.add_argument(
        "--out-xlsx",
        default="",
        help="Default: data/results/<run_id>/benchmark_goren_2025/audit_parsing_derivation_samples.xlsx",
    )
    return parser.parse_args()


def read_tsv_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter="\t")
        rows = list(reader)
    if not rows:
        return [], []
    header = [str(x) for x in rows[0]]
    body = [[str(v) for v in row] for row in rows[1:]]
    return header, body


def set_column_widths(ws, header: list[str]) -> None:
    width_rules = {
        "evidence_excerpt": 70,
        "source_value_texts": 70,
        "trace_pointer": 32,
        "derived_from": 28,
        "value_source": 28,
        "key": 14,
        "formulation_id": 14,
        "derived_value": 16,
        "derived_unit": 14,
        "rule_id": 16,
        "field_name": 16,
        "sample_category": 16,
        "human_decision": 16,
        "human_notes": 40,
    }
    for idx, name in enumerate(header, start=1):
        col_letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[col_letter].width = width_rules.get(name, 22)


def apply_audit_samples_formatting(ws, header: list[str], n_rows: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wrap_cols = {"evidence_excerpt", "source_value_texts", "notes_for_reviewer", "human_notes"}
    wrap_idx = {i + 1 for i, name in enumerate(header) if name in wrap_cols}

    if n_rows > 0:
        for r in range(2, n_rows + 2):
            ws.row_dimensions[r].height = 60
            for c in wrap_idx:
                ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    decision_col = None
    for i, name in enumerate(header, start=1):
        if name == "human_decision":
            decision_col = i
            break
    if decision_col is not None:
        col_letter = ws.cell(row=1, column=decision_col).column_letter
        start_row = 2
        end_row = max(2, n_rows + 1)
        formula = '"pass,fail,unclear"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Please choose pass, fail, or unclear."
        dv.errorTitle = "Invalid decision"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def write_sheet(ws, header: list[str], rows: list[list[str]]) -> None:
    ws.append(header)
    for row in rows:
        if len(row) < len(header):
            row = row + [""] * (len(header) - len(row))
        elif len(row) > len(header):
            row = row[: len(header)]
        ws.append(row)


def main() -> None:
    args = parse_args()
    run_id = args.run_id.strip() or DEFAULT_RUN_ID

    base_dir = Path(f"data/results/{run_id}/benchmark_goren_2025")
    audit_tsv = Path(args.audit_tsv) if args.audit_tsv else base_dir / "audit_parsing_derivation_samples.tsv"
    alignment_tsv = Path(args.alignment_tsv) if args.alignment_tsv else base_dir / "alignment_rows.tsv"
    out_xlsx = Path(args.out_xlsx) if args.out_xlsx else base_dir / "audit_parsing_derivation_samples.xlsx"

    if not audit_tsv.exists():
        raise FileNotFoundError(
            f"audit TSV not found: {audit_tsv}. Please confirm --run-id or provide --audit-tsv."
        )

    audit_header, audit_rows = read_tsv_rows(audit_tsv)
    reviewer_cols = ["human_decision", "human_notes"]
    audit_header_out = list(audit_header) + reviewer_cols
    audit_rows_out = [list(r) + ["", ""] for r in audit_rows]

    wb = Workbook()
    ws_audit = wb.active
    ws_audit.title = "audit_samples"
    write_sheet(ws_audit, audit_header_out, audit_rows_out)
    set_column_widths(ws_audit, audit_header_out)
    apply_audit_samples_formatting(ws_audit, audit_header_out, len(audit_rows_out))

    alignment_included = False
    if alignment_tsv.exists():
        alignment_header, alignment_rows = read_tsv_rows(alignment_tsv)
        ws_alignment = wb.create_sheet("alignment_rows")
        write_sheet(ws_alignment, alignment_header, alignment_rows)
        ws_alignment.freeze_panes = "A2"
        ws_alignment.auto_filter.ref = ws_alignment.dimensions
        alignment_included = True

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)

    print(f"input_row_count={len(audit_rows)}")
    print(f"output_path={out_xlsx}")
    print(f"alignment_sheet_included={str(alignment_included).lower()}")


if __name__ == "__main__":
    main()
