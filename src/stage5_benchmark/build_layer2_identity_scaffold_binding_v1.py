#!/usr/bin/env python3
from __future__ import annotations

"""
Build a diagnostic-only Layer2 identity scaffold binding surface.

Purpose:
- freeze a reviewed-boundary-style identity anchor from the GT workbook
- bind current-system final rows to that anchor using a strict ladder
- avoid coarse value-level fallback for benchmark-style identity comparison

Scope:
- supporting Stage5 / Layer3 audit utility only
- does not modify final tables or benchmark-valid outputs
"""

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


GT_DEFAULT_SHEET = "value_gt_annotation"
SCAFFOLD_ROWS_NAME = "layer2_identity_scaffold_rows_v1.tsv"
SUMMARY_NAME = "layer2_identity_scaffold_summary_v1.tsv"
REPORT_NAME = "layer2_identity_scaffold_validation_v1.md"


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_label_token(value: Any) -> str:
    text = normalize_text(value).lower()
    if not text:
        return ""
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def normalize_namespaced_label(paper_key: str, value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    lowered = text.lower()
    prefix = normalize_text(paper_key).lower()
    if prefix and lowered.startswith(prefix):
        remainder = text[len(paper_key):]
        remainder = re.sub(r"^[\s_\-:]+", "", remainder)
        if remainder:
            return normalize_text(remainder)
    return text


def scaffold_key_for_row(row: dict[str, str]) -> str:
    paper_key = normalize_text(row.get("paper_key"))
    native_label = normalize_text(row.get("scaffold_native_label"))
    if native_label:
        return f"{paper_key}::native::{normalize_label_token(native_label)}"
    gt_formulation_id = normalize_text(row.get("gt_formulation_id"))
    return f"{paper_key}::gt::{normalize_label_token(gt_formulation_id)}"


def detect_gt_sheet(workbook_path: Path, requested: str) -> str:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if requested and requested in workbook.sheetnames:
            return requested
        if GT_DEFAULT_SHEET in workbook.sheetnames:
            return GT_DEFAULT_SHEET
        return workbook.sheetnames[0]
    finally:
        workbook.close()


def load_workbook_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        iterator = sheet.iter_rows(values_only=True)
        header = [normalize_text(value) for value in next(iterator)]
        rows: list[dict[str, str]] = []
        for values in iterator:
            row = {}
            for index, column in enumerate(header):
                row[column] = normalize_text(values[index] if index < len(values) else "")
            rows.append(row)
        return rows
    finally:
        workbook.close()


def read_tsv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: str(row.get(key, "")) for key in fieldnames})


def build_final_indexes(rows: list[dict[str, str]]) -> dict[str, dict[tuple[str, str], list[dict[str, str]]]]:
    exact_index: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    namespaced_index: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    strict_index: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        paper_key = normalize_text(row.get("key") or row.get("paper_key"))
        representative_id = normalize_text(row.get("representative_source_formulation_id"))
        raw_label = normalize_text(row.get("representative_source_raw_formulation_label"))
        if paper_key and representative_id:
            exact_index[(paper_key, representative_id.lower())].append(row)
            normalized_namespaced = normalize_namespaced_label(paper_key, representative_id)
            if normalized_namespaced:
                namespaced_index[(paper_key, normalized_namespaced.lower())].append(row)
        if paper_key and raw_label:
            strict_index[(paper_key, normalize_label_token(raw_label))].append(row)
    return {
        "exact": exact_index,
        "namespaced": namespaced_index,
        "strict": strict_index,
    }


def gt_native_label(row: dict[str, str]) -> str:
    return (
        normalize_text(row.get("seed_pred_representative_source_formulation_id"))
        or normalize_text(row.get("formulation_label"))
        or normalize_text(row.get("gt_formulation_id"))
    )


def collect_matches(
    *,
    paper_key: str,
    gt_row: dict[str, str],
    final_indexes: dict[str, dict[tuple[str, str], list[dict[str, str]]]],
) -> dict[str, list[dict[str, str]]]:
    native_label = gt_native_label(gt_row)
    formulation_label = normalize_text(gt_row.get("formulation_label"))
    return {
        "exact": final_indexes["exact"].get((paper_key, native_label.lower()), []) if native_label else [],
        "namespaced": final_indexes["namespaced"].get((paper_key, native_label.lower()), []) if native_label else [],
        "strict": final_indexes["strict"].get((paper_key, normalize_label_token(formulation_label)))
        if formulation_label
        else [],
    }


def row_id_list(rows: list[dict[str, str]]) -> str:
    return " | ".join(
        normalize_text(row.get("representative_source_formulation_id")) for row in rows if normalize_text(row.get("representative_source_formulation_id"))
    )


def final_id_list(rows: list[dict[str, str]]) -> str:
    return " | ".join(normalize_text(row.get("final_formulation_id")) for row in rows if normalize_text(row.get("final_formulation_id")))


def select_binding_rule(matches: dict[str, list[dict[str, str]]]) -> tuple[str, list[dict[str, str]], bool]:
    for rule in ("exact", "namespaced", "strict"):
        selected = matches[rule]
        if not selected:
            continue
        return rule, selected, len(selected) > 1
    return "unresolved", [], False


def build_scaffold_rows(
    *,
    gt_rows: list[dict[str, str]],
    baseline_rows: list[dict[str, str]],
    new_rows: list[dict[str, str]],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    baseline_indexes = build_final_indexes(baseline_rows)
    new_indexes = build_final_indexes(new_rows)
    detail_rows: list[dict[str, str]] = []
    summary_rows: list[dict[str, str]] = []

    gt_by_paper: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in gt_rows:
        gt_by_paper[normalize_text(row.get("paper_key"))].append(row)

    baseline_by_paper: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in baseline_rows:
        baseline_by_paper[normalize_text(row.get("key") or row.get("paper_key"))].append(row)

    new_by_paper: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in new_rows:
        new_by_paper[normalize_text(row.get("key") or row.get("paper_key"))].append(row)

    for paper_key in sorted(gt_by_paper):
        baseline_exact_count = 0
        new_exact_count = 0
        new_namespaced_count = 0
        new_strict_count = 0
        resolved_count = 0
        ambiguous_count = 0
        unresolved_count = 0
        examples: list[str] = []

        for gt_row in gt_by_paper[paper_key]:
            native_label = gt_native_label(gt_row)
            baseline_matches = collect_matches(paper_key=paper_key, gt_row=gt_row, final_indexes=baseline_indexes)
            new_matches = collect_matches(paper_key=paper_key, gt_row=gt_row, final_indexes=new_indexes)
            selected_rule, selected_matches, ambiguous = select_binding_rule(new_matches)
            if baseline_matches["exact"]:
                baseline_exact_count += 1
            if new_matches["exact"]:
                new_exact_count += 1
            elif new_matches["namespaced"]:
                new_namespaced_count += 1
                if len(examples) < 3:
                    normalized = normalize_namespaced_label(paper_key, new_matches["namespaced"][0].get("representative_source_formulation_id"))
                    examples.append(
                        f"{normalize_text(new_matches['namespaced'][0].get('representative_source_formulation_id'))} -> {normalized}"
                    )
            elif new_matches["strict"]:
                new_strict_count += 1
            if selected_matches:
                resolved_count += 1
            else:
                unresolved_count += 1
            if ambiguous:
                ambiguous_count += 1

            selected_ids = row_id_list(selected_matches)
            selected_final_ids = final_id_list(selected_matches)
            detail = {
                "paper_key": paper_key,
                "gt_formulation_id": normalize_text(gt_row.get("gt_formulation_id")),
                "gt_formulation_label": normalize_text(gt_row.get("formulation_label")),
                "scaffold_native_label": native_label,
                "scaffold_key": "",
                "baseline_exact_match_count": str(len(baseline_matches["exact"])),
                "baseline_exact_match_ids": row_id_list(baseline_matches["exact"]),
                "new_exact_match_count": str(len(new_matches["exact"])),
                "new_exact_match_ids": row_id_list(new_matches["exact"]),
                "new_namespaced_match_count": str(len(new_matches["namespaced"])),
                "new_namespaced_match_ids": row_id_list(new_matches["namespaced"]),
                "new_strict_match_count": str(len(new_matches["strict"])),
                "new_strict_match_ids": row_id_list(new_matches["strict"]),
                "selected_new_binding_rule": selected_rule,
                "selected_new_final_formulation_ids": selected_final_ids,
                "selected_new_representative_source_formulation_ids": selected_ids,
                "fallback_required": "no",
                "ambiguous_after_normalized_binding": "yes" if ambiguous else "no",
                "binding_status": "resolved" if selected_matches and not ambiguous else ("ambiguous" if ambiguous else "unresolved"),
            }
            detail["scaffold_key"] = scaffold_key_for_row(detail)
            detail_rows.append(detail)

        summary_rows.append(
            {
                "paper_key": paper_key,
                "gt_rows": str(len(gt_by_paper[paper_key])),
                "baseline_final_rows": str(len(baseline_by_paper.get(paper_key, []))),
                "new_final_rows": str(len(new_by_paper.get(paper_key, []))),
                "baseline_exact_bind_rows": str(baseline_exact_count),
                "new_exact_bind_rows": str(new_exact_count),
                "new_namespaced_bind_rows": str(new_namespaced_count),
                "new_strict_bind_rows": str(new_strict_count),
                "new_resolved_rows_after_scaffold": str(resolved_count),
                "new_unresolved_rows": str(unresolved_count),
                "new_ambiguous_rows": str(ambiguous_count),
                "coarse_fallback_needed_rows": "0",
                "normalization_examples": " ; ".join(examples),
            }
        )

    return detail_rows, summary_rows


def markdown_table(rows: list[dict[str, str]], columns: list[str]) -> str:
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join(["---"] * len(columns)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(normalize_text(row.get(column)).replace("|", "\\|") for column in columns) + " |")
    return "\n".join(lines)


def build_report(
    *,
    gt_workbook_path: Path,
    gt_sheet_name: str,
    baseline_path: Path,
    new_path: Path,
    paper_keys: list[str],
    detail_rows: list[dict[str, str]],
    summary_rows: list[dict[str, str]],
) -> str:
    lines = [
        "# Layer2 Identity Scaffold Binding Validation v1",
        "",
        "Diagnostic-only, benchmark-safe binding audit.",
        "",
        "## Frozen Identity Source",
        f"- GT workbook: `{gt_workbook_path}`",
        f"- GT sheet: `{gt_sheet_name}`",
        "- Frozen identity anchor field: `seed_pred_representative_source_formulation_id` when present, otherwise `formulation_label`, otherwise `gt_formulation_id`.",
        "",
        "## Input Prediction Surfaces",
        f"- Baseline final table: `{baseline_path}`",
        f"- New experiment final table: `{new_path}`",
        "",
        "## Binding Ladder",
        "1. exact article-native formulation label match",
        "2. normalized namespaced-label match",
        "3. strict identity-equivalent label match",
        "4. coarse fallback is disallowed for this diagnostic surface",
        "",
        "## Per-Paper Summary",
        markdown_table(
            summary_rows,
            [
                "paper_key",
                "gt_rows",
                "baseline_final_rows",
                "new_final_rows",
                "baseline_exact_bind_rows",
                "new_exact_bind_rows",
                "new_namespaced_bind_rows",
                "new_resolved_rows_after_scaffold",
                "new_unresolved_rows",
                "new_ambiguous_rows",
                "normalization_examples",
            ],
        ),
        "",
    ]

    for paper_key in paper_keys:
        paper_rows = [row for row in detail_rows if normalize_text(row.get("paper_key")) == paper_key]
        lines.extend(
            [
                f"## {paper_key}",
                "",
                markdown_table(
                    paper_rows[:20],
                    [
                        "gt_formulation_id",
                        "scaffold_native_label",
                        "baseline_exact_match_ids",
                        "new_exact_match_ids",
                        "new_namespaced_match_ids",
                        "selected_new_binding_rule",
                        "selected_new_representative_source_formulation_ids",
                        "binding_status",
                    ],
                ),
                "",
            ]
        )
    return "\n".join(lines).strip() + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build diagnostic Layer2 identity scaffold binding surfaces.")
    parser.add_argument("--gt-workbook-xlsx", type=Path, required=True)
    parser.add_argument("--baseline-final-tsv", type=Path, required=True)
    parser.add_argument("--new-final-tsv", type=Path, required=True)
    parser.add_argument("--paper-key", action="append", required=True, help="Repeatable paper key filter.")
    parser.add_argument("--out-dir", type=Path, required=True)
    parser.add_argument("--sheet-name", default=GT_DEFAULT_SHEET)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    gt_workbook_path = args.gt_workbook_xlsx.resolve()
    baseline_path = args.baseline_final_tsv.resolve()
    new_path = args.new_final_tsv.resolve()
    out_dir = args.out_dir.resolve()
    paper_keys = [normalize_text(value) for value in args.paper_key if normalize_text(value)]
    gt_sheet_name = detect_gt_sheet(gt_workbook_path, args.sheet_name)

    gt_rows = [
        row
        for row in load_workbook_rows(gt_workbook_path, gt_sheet_name)
        if normalize_text(row.get("paper_key")) in set(paper_keys) and normalize_text(row.get("gt_formulation_id"))
    ]
    baseline_rows = [
        row for row in read_tsv_rows(baseline_path) if normalize_text(row.get("key") or row.get("paper_key")) in set(paper_keys)
    ]
    new_rows = [
        row for row in read_tsv_rows(new_path) if normalize_text(row.get("key") or row.get("paper_key")) in set(paper_keys)
    ]

    detail_rows, summary_rows = build_scaffold_rows(
        gt_rows=gt_rows,
        baseline_rows=baseline_rows,
        new_rows=new_rows,
    )

    out_dir.mkdir(parents=True, exist_ok=True)
    write_tsv(
        out_dir / SCAFFOLD_ROWS_NAME,
        [
            "paper_key",
            "gt_formulation_id",
            "gt_formulation_label",
            "scaffold_native_label",
            "scaffold_key",
            "baseline_exact_match_count",
            "baseline_exact_match_ids",
            "new_exact_match_count",
            "new_exact_match_ids",
            "new_namespaced_match_count",
            "new_namespaced_match_ids",
            "new_strict_match_count",
            "new_strict_match_ids",
            "selected_new_binding_rule",
            "selected_new_final_formulation_ids",
            "selected_new_representative_source_formulation_ids",
            "fallback_required",
            "ambiguous_after_normalized_binding",
            "binding_status",
        ],
        detail_rows,
    )
    write_tsv(
        out_dir / SUMMARY_NAME,
        [
            "paper_key",
            "gt_rows",
            "baseline_final_rows",
            "new_final_rows",
            "baseline_exact_bind_rows",
            "new_exact_bind_rows",
            "new_namespaced_bind_rows",
            "new_strict_bind_rows",
            "new_resolved_rows_after_scaffold",
            "new_unresolved_rows",
            "new_ambiguous_rows",
            "coarse_fallback_needed_rows",
            "normalization_examples",
        ],
        summary_rows,
    )
    (out_dir / REPORT_NAME).write_text(
        build_report(
            gt_workbook_path=gt_workbook_path,
            gt_sheet_name=gt_sheet_name,
            baseline_path=baseline_path,
            new_path=new_path,
            paper_keys=paper_keys,
            detail_rows=detail_rows,
            summary_rows=summary_rows,
        ),
        encoding="utf-8",
    )

    diagnostics = {
        "gt_workbook_path": str(gt_workbook_path),
        "gt_sheet_name": gt_sheet_name,
        "baseline_final_tsv": str(baseline_path),
        "new_final_tsv": str(new_path),
        "paper_keys": paper_keys,
        "output_dir": str(out_dir),
    }
    print(json.dumps(diagnostics, ensure_ascii=True))
    for row in summary_rows:
        print(
            "paper={paper_key} gt_rows={gt_rows} baseline_rows={baseline_final_rows} "
            "new_rows={new_final_rows} exact={new_exact_bind_rows} namespaced={new_namespaced_bind_rows} "
            "strict={new_strict_bind_rows} resolved={new_resolved_rows_after_scaffold} "
            "fallback={coarse_fallback_needed_rows} ambiguous={new_ambiguous_rows}".format(**row)
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
