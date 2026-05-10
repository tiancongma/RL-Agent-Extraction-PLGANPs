#!/usr/bin/env python3
from __future__ import annotations

"""
Layer 3 post-annotation cross-audit framework.

This script is audit-only. It does not modify workbook contents or benchmark-
valid pipeline outputs. It produces reviewer-facing risk reports that flag
potentially unsupported, derived, contaminated, or ambiguous workbook cells.
"""

import argparse
import csv
import json
import math
import os
import re
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import requests
from dotenv import load_dotenv

try:
    import google.generativeai as genai  # type: ignore
    HAS_GENAI = True
except Exception:
    genai = None
    HAS_GENAI = False

try:
    from src.stage5_benchmark.audit_evidence_resolver_v1 import AuditEvidenceResolverV1
    from src.utils.active_data_source import (
        build_artifact_metadata,
        resolve_artifact_path,
        resolve_run_context,
        write_artifact_metadata_json,
    )
    from src.utils.paths import PROJECT_ROOT
except ModuleNotFoundError:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from src.stage5_benchmark.audit_evidence_resolver_v1 import AuditEvidenceResolverV1
    from src.utils.active_data_source import (
        build_artifact_metadata,
        resolve_artifact_path,
        resolve_run_context,
        write_artifact_metadata_json,
    )
    from src.utils.paths import PROJECT_ROOT


REPORT_COLUMNS = [
    "paper_id",
    "doi",
    "formulation_id",
    "field_name",
    "current_value",
    "risk_level",
    "risk_type",
    "source_of_flag",
    "reason",
    "evidence_status",
    "evidence_snippet",
    "source_paths",
]

TEXT_FIELDS = {
    "polymer_name",
    "polymer_grade",
    "drug_name",
    "surfactant_name",
    "stabilizer_name",
    "helper_material_name",
    "method_type",
    "solvent_name",
    "co_solvent_name",
}

NUMERIC_FIELDS = {
    "polymer_mw_raw",
    "polymer_mw_kDa",
    "polymer_mass_mg",
    "polymer_concentration_value",
    "drug_mass_mg",
    "drug_concentration_value",
    "surfactant_mass_mg",
    "surfactant_concentration_value",
    "W1_volume_mL",
    "O_volume_mL",
    "W2_volume_mL",
    "external_aqueous_phase_volume_mL",
    "internal_aqueous_phase_volume_mL",
    "sonication_time_s",
    "homogenization_time_min",
    "stirring_time_h",
    "evaporation_time_h",
    "centrifugation_g",
    "centrifugation_time_min",
    "ee_percent",
    "lc_percent",
    "dl_percent",
    "particle_size_nm",
    "pdi",
    "zeta_mV",
}

RATIO_FIELDS = {
    "la_ga_ratio_raw",
    "la_ga_ratio_normalized",
    "polymer_to_solvent_ratio_raw",
    "polymer_to_drug_ratio_raw",
    "drug_to_polymer_ratio_raw",
    "phase_ratio_raw",
}

UNIT_FIELDS = {
    "polymer_concentration_unit",
    "drug_concentration_unit",
    "surfactant_concentration_unit",
}

META_FIELDS = {
    "paper_key",
    "doi",
    "gt_formulation_id",
    "family_id",
    "parent_core",
    "variant_role",
    "benchmark_default_include",
    "formulation_label",
    "seed_pred_representative_source_formulation_id",
    "gt_row_decision",
    "value_source_type",
    "candidate_notes",
}

AUDIT_FIELDS = [field for field in (list(TEXT_FIELDS) + list(NUMERIC_FIELDS) + list(RATIO_FIELDS) + list(UNIT_FIELDS)) if field not in META_FIELDS]

BLANK_ROW_PATTERN = re.compile(
    r"\b(blank|empty|unloaded|placebo|drug[- ]free|no drug|without [a-z0-9-]+)\b",
    re.IGNORECASE,
)

REFERENCE_LIKE_PATTERN = re.compile(
    r"\b(references|bibliography|et al\.|doi:|https?://doi\.org|vol\.|pp\.|\d{4}\))\b",
    re.IGNORECASE,
)

RISK_PRIORITY = {
    "blank_should_be_null": 80,
    "cross_paper_contamination": 75,
    "inheritance_contamination": 70,
    "unsupported_value": 60,
    "direction_mismatch": 55,
    "derived_value": 50,
    "unit_or_normalization_only": 40,
    "ambiguity": 30,
}

RISK_LEVEL_PRIORITY = {"low": 1, "medium": 2, "high": 3}
FIELD_PRIORITY = [
    "drug_name",
    "drug_mass_mg",
    "drug_concentration_value",
    "polymer_mw_kDa",
    "polymer_concentration_value",
    "la_ga_ratio_normalized",
    "drug_to_polymer_ratio_raw",
    "particle_size_nm",
    "ee_percent",
    "lc_percent",
]

MODEL_SYSTEM_PROMPT = (
    "You are a Layer 3 workbook auditor. You are not an editor and must not fill, "
    "normalize, infer, or compute missing values. Flag only cells whose current "
    "workbook value appears unsupported, derived, directionally mismatched, "
    "normalization-only, contaminated, or ambiguous. Treat only explicit source "
    "text or table evidence from the current paper as support."
)

MODEL_OUTPUT_SCHEMA = {
    "paper_id": "DOI or zotero_key",
    "doi": "DOI if available, else empty string",
    "formulation_id": "Workbook formulation identifier",
    "field_name": "Exact workbook field name",
    "current_value": "Current workbook cell value",
    "risk_level": "high | medium | low",
    "risk_type": (
        "unsupported_value | derived_value | direction_mismatch | "
        "unit_or_normalization_only | blank_should_be_null | "
        "cross_paper_contamination | inheritance_contamination | ambiguity"
    ),
    "source_of_flag": "gemini or nvidia",
    "reason": "Short explanation",
    "evidence_status": "supported | derived | unsupported | ambiguous",
    "evidence_snippet": "Short supporting snippet if any",
    "source_paths": "Semicolon-separated file paths used",
}

MAIN_RISK_TYPES = {
    "blank_should_be_null",
    "unsupported_value",
    "direction_mismatch",
    "cross_paper_contamination",
    "inheritance_contamination",
}

HIGH_PRECISION_FIELDS = {
    "drug_name",
    "drug_to_polymer_ratio_raw",
    "polymer_to_drug_ratio_raw",
    "la_ga_ratio_raw",
    "la_ga_ratio_normalized",
    "phase_ratio_raw",
}

HIGH_PRIORITY_SOURCE_OF_FLAG = {
    "rule+gemini+nvidia",
    "rule+gemini",
    "rule+nvidia",
    "gemini+nvidia",
}

NVIDIA_HOSTED_CHAT_COMPLETIONS_URL = "https://integrate.api.nvidia.com/v1/chat/completions"
DEFAULT_BATCH_SIZE = 5
DEFAULT_REQUEST_TIMEOUT_SECONDS = 60
DEFAULT_MAX_RETRIES = 1


@dataclass
class CellFlag:
    paper_id: str
    doi: str
    formulation_id: str
    field_name: str
    current_value: str
    risk_level: str
    risk_type: str
    source_of_flag: str
    reason: str
    evidence_status: str
    evidence_snippet: str
    source_paths: str
    extras: list[str]

    def to_row(self) -> dict[str, str]:
        return {
            "paper_id": self.paper_id,
            "doi": self.doi,
            "formulation_id": self.formulation_id,
            "field_name": self.field_name,
            "current_value": self.current_value,
            "risk_level": self.risk_level,
            "risk_type": self.risk_type,
            "source_of_flag": self.source_of_flag,
            "reason": self.reason,
            "evidence_status": self.evidence_status,
            "evidence_snippet": self.evidence_snippet,
            "source_paths": self.source_paths,
        }


@dataclass
class SearchHit:
    source_type: str
    source_path: str
    snippet: str
    variant: str
    exact: bool
    reference_like: bool


@dataclass
class PaperContext:
    paper_key: str
    doi: str
    text_path: Path | None
    text_raw: str
    text_norm: str
    text_norm_to_raw: list[int]
    table_paths: list[Path]
    table_raw_parts: list[tuple[str, str]]
    table_norm_parts: list[tuple[str, str, list[int]]]


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_lower(value: Any) -> str:
    return normalize_text(value).lower()


def clean_numeric_token(value: str) -> str:
    token = normalize_text(value)
    token = token.replace(",", "")
    return token


def try_float(value: Any) -> float | None:
    token = clean_numeric_token(str(value or ""))
    if not token:
        return None
    token = token.rstrip("%")
    try:
        return float(token)
    except ValueError:
        return None


def write_tsv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def read_tsv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def workbook_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]
    headers = [normalize_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, str]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            row[header] = normalize_text(values[index] if index < len(values) else "")
        rows.append(row)
    return rows


def choose_sheet(workbook_path: Path, explicit_sheet: str = "") -> str:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    if explicit_sheet:
        if explicit_sheet not in wb.sheetnames:
            raise ValueError(f"Workbook sheet {explicit_sheet!r} not found in {workbook_path}")
        return explicit_sheet
    if "value_gt_annotation" in wb.sheetnames:
        return "value_gt_annotation"
    return wb.sheetnames[0]


def build_paper_context(
    resolver: AuditEvidenceResolverV1,
    paper_key: str,
    doi: str,
) -> PaperContext:
    text_path = resolver.key2txt_map.get(paper_key)
    text_raw = ""
    if text_path is not None and text_path.exists():
        text_raw = text_path.read_text(encoding="utf-8", errors="replace")
    text_norm, text_map = normalize_with_mapping(text_raw)

    table_paths = resolver.table_paths_by_key.get(paper_key, [])
    if not table_paths:
        key_tables_dir = PROJECT_ROOT / "data" / "cleaned" / "goren_2025" / "tables" / paper_key
        if key_tables_dir.exists():
            table_paths = sorted(p for p in key_tables_dir.glob("*.csv"))
    table_raw_parts: list[tuple[str, str]] = []
    table_norm_parts: list[tuple[str, str, list[int]]] = []
    for path in table_paths:
        raw = load_table_text(path)
        norm, mapping = normalize_with_mapping(raw)
        table_raw_parts.append((str(path.resolve()), raw))
        table_norm_parts.append((str(path.resolve()), norm, mapping))
    return PaperContext(
        paper_key=paper_key,
        doi=doi,
        text_path=text_path,
        text_raw=text_raw,
        text_norm=text_norm,
        text_norm_to_raw=text_map,
        table_paths=table_paths,
        table_raw_parts=table_raw_parts,
        table_norm_parts=table_norm_parts,
    )


def normalize_with_mapping(text: str) -> tuple[str, list[int]]:
    chars: list[str] = []
    mapping: list[int] = []
    prev_space = True
    for index, char in enumerate(text):
        lowered = char.lower()
        if lowered.isspace():
            if prev_space:
                continue
            chars.append(" ")
            mapping.append(index)
            prev_space = True
            continue
        chars.append(lowered)
        mapping.append(index)
        prev_space = False
    normalized = "".join(chars).strip()
    if normalized and mapping and mapping[0] != 0:
        offset = len("".join(chars)) - len(normalized)
        if offset > 0:
            mapping = mapping[offset:]
    return normalized, mapping


def load_table_text(path: Path) -> str:
    try:
        df = pd.read_csv(path, dtype=str).fillna("")
    except Exception:
        return ""
    lines: list[str] = []
    if list(df.columns):
        lines.append(" | ".join([normalize_text(col) for col in df.columns if normalize_text(col)]))
    for _, row in df.iterrows():
        cell_values = [normalize_text(value) for value in row.tolist()]
        cell_values = [value for value in cell_values if value]
        if cell_values:
            lines.append(" | ".join(cell_values))
    return "\n".join(lines)


def build_value_variants(field_name: str, value: str, unit_hint: str = "") -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    variants: set[str] = {raw}
    variants.add(normalize_lower(raw))
    variants.add(raw.replace("–", "-"))
    variants.add(raw.replace("-", " "))
    variants.add(raw.replace("/", " / "))
    variants.add(raw.replace(":", " : "))

    unit_hint = normalize_text(unit_hint)
    if field_name in UNIT_FIELDS:
        lower = normalize_lower(raw)
        if lower == "w/v":
            variants.update({"w / v", "% w/v", "w v"})
        elif lower == "v/v":
            variants.update({"v / v", "% v/v", "v v"})
        elif lower == "mg/ml":
            variants.update({"mg mL-1", "mg / ml", "mg/ml"})
        return sorted({normalize_text(v).lower() for v in variants if normalize_text(v)})

    if field_name in NUMERIC_FIELDS:
        num = try_float(raw)
        if num is not None:
            variants.add(format_number(num))
            variants.add(format_number(num).rstrip("0").rstrip("."))
        if field_name.endswith("_mg"):
            variants.update(with_unit_variants(raw, ["mg", "milligram", "milligrams"]))
        elif field_name.endswith("_mL"):
            variants.update(with_unit_variants(raw, ["ml", "mL", "mL.", "milliliter", "milliliters"]))
        elif field_name.endswith("_nm"):
            variants.update(with_unit_variants(raw, ["nm", "nanometer", "nanometers"]))
        elif field_name.endswith("_mV"):
            variants.update(with_unit_variants(raw, ["mv", "mV", "millivolt", "millivolts"]))
        elif field_name.endswith("_percent"):
            variants.update(with_percent_variants(raw))
        elif field_name == "polymer_mw_kDa":
            variants.update(with_unit_variants(raw, ["kda", "kDa", "kd", "k dalton"]))
            if num is not None:
                variants.add(f"{format_number(num * 1000)} da")
        elif field_name.endswith("_concentration_value"):
            if unit_hint:
                variants.update(with_unit_variants(raw, [unit_hint]))
            variants.update(with_percent_variants(raw))

    if field_name in RATIO_FIELDS or field_name == "la_ga_ratio_normalized":
        ratio_variants, inverse_variants = ratio_variant_sets(raw)
        variants.update(ratio_variants)
        variants.update(inverse_variants)

    return sorted({normalize_text(v).lower() for v in variants if normalize_text(v)})


def with_unit_variants(value: str, unit_tokens: list[str]) -> set[str]:
    variants: set[str] = set()
    raw = normalize_text(value)
    if not raw:
        return variants
    for token in unit_tokens:
        token_clean = normalize_text(token)
        if not token_clean:
            continue
        variants.add(f"{raw} {token_clean}")
        variants.add(f"{raw}{token_clean}")
    return variants


def with_percent_variants(value: str) -> set[str]:
    raw = normalize_text(value)
    variants = {raw}
    if raw.endswith("%"):
        base = raw.rstrip("%").strip()
    else:
        base = raw
        variants.add(f"{raw}%")
    if base:
        variants.add(f"{base} %")
        variants.add(f"{base} percent")
        variants.add(f"{base} wt%")
    return variants


def format_number(number: float) -> str:
    if math.isclose(number, round(number)):
        return str(int(round(number)))
    return f"{number:.6g}"


def ratio_variant_sets(raw: str) -> tuple[set[str], set[str]]:
    normalized = normalize_text(raw)
    forward: set[str] = {normalized.lower()}
    inverse: set[str] = set()
    parts = [part.strip() for part in re.split(r"[:/]", normalized) if part.strip()]
    if len(parts) == 2:
        left, right = parts[0], parts[1]
        forward.update(
            {
                f"{left}:{right}".lower(),
                f"{left}/{right}".lower(),
                f"{left} : {right}".lower(),
                f"{left} / {right}".lower(),
                f"{left} to {right}".lower(),
            }
        )
        inverse.update(
            {
                f"{right}:{left}".lower(),
                f"{right}/{left}".lower(),
                f"{right} : {left}".lower(),
                f"{right} / {left}".lower(),
                f"{right} to {left}".lower(),
            }
        )
        left_num = try_float(left)
        right_num = try_float(right)
        if left_num not in {None, 0} and right_num is not None:
            forward.add(format_number(right_num / left_num).lower())
        if right_num not in {None, 0} and left_num is not None:
            inverse.add(format_number(left_num / right_num).lower())
    else:
        numeric = try_float(normalized)
        if numeric is not None and numeric not in {0, 1}:
            inverse.add(format_number(1 / numeric).lower())
    return forward, inverse


def extract_snippet(raw_text: str, start: int, end: int, pad: int = 120) -> str:
    left = max(0, start - pad)
    right = min(len(raw_text), end + pad)
    return normalize_text(raw_text[left:right])[:280]


def search_hits_for_variants(
    norm_text: str,
    raw_text: str,
    mapping: list[int],
    source_path: str,
    source_type: str,
    variants: list[str],
    primary_variant: str,
) -> list[SearchHit]:
    hits: list[SearchHit] = []
    for variant in variants:
        needle = normalize_lower(variant)
        if not needle:
            continue
        idx = norm_text.find(needle)
        if idx < 0 or idx >= len(mapping):
            continue
        raw_start = mapping[idx]
        raw_end = mapping[min(len(mapping) - 1, idx + len(needle) - 1)] + 1
        snippet = extract_snippet(raw_text, raw_start, raw_end)
        hits.append(
            SearchHit(
                source_type=source_type,
                source_path=source_path,
                snippet=snippet,
                variant=needle,
                exact=needle == primary_variant,
                reference_like=bool(REFERENCE_LIKE_PATTERN.search(snippet)),
            )
        )
    return hits


def dedupe_hits(hits: list[SearchHit]) -> list[SearchHit]:
    seen: set[tuple[str, str, str]] = set()
    out: list[SearchHit] = []
    for hit in hits:
        key = (hit.source_type, hit.source_path, hit.snippet)
        if key in seen:
            continue
        seen.add(key)
        out.append(hit)
    return out


def find_hits(context: PaperContext, field_name: str, value: str, unit_hint: str = "") -> list[SearchHit]:
    variants = build_value_variants(field_name, value, unit_hint=unit_hint)
    if not variants:
        return []
    primary_variant = normalize_lower(normalize_text(value))
    hits: list[SearchHit] = []
    if context.text_raw and context.text_path is not None:
        hits.extend(
            search_hits_for_variants(
                context.text_norm,
                context.text_raw,
                context.text_norm_to_raw,
                str(context.text_path.resolve()),
                "text",
                variants,
                primary_variant,
            )
        )
    for source_path, table_norm, mapping in context.table_norm_parts:
        raw_text = ""
        for raw_path, raw_content in context.table_raw_parts:
            if raw_path == source_path:
                raw_text = raw_content
                break
        hits.extend(search_hits_for_variants(table_norm, raw_text, mapping, source_path, "table", variants, primary_variant))
    return dedupe_hits(hits)


def field_search_unit_hint(row: dict[str, str], field_name: str) -> str:
    if field_name == "polymer_concentration_value":
        return row.get("polymer_concentration_unit", "")
    if field_name == "drug_concentration_value":
        return row.get("drug_concentration_unit", "")
    if field_name == "surfactant_concentration_value":
        return row.get("surfactant_concentration_unit", "")
    return ""


def is_blank_like_row(row: dict[str, str]) -> bool:
    blob = " ".join(
        [
            row.get("gt_formulation_id", ""),
            row.get("formulation_label", ""),
            row.get("candidate_notes", ""),
            row.get("variant_role", ""),
        ]
    )
    return bool(BLANK_ROW_PATTERN.search(blob))


def sibling_rows(rows_by_paper: dict[str, list[dict[str, str]]], row: dict[str, str]) -> list[dict[str, str]]:
    paper_key = row.get("paper_key", "")
    family_id = row.get("family_id", "")
    group = rows_by_paper.get(paper_key, [])
    if family_id:
        family_group = [candidate for candidate in group if candidate.get("family_id", "") == family_id]
        if len(family_group) > 1:
            return family_group
    return [candidate for candidate in group if candidate.get("gt_formulation_id", "") != row.get("gt_formulation_id", "")]


def first_snippet(hits: list[SearchHit]) -> str:
    return hits[0].snippet if hits else ""


def stringify_paths(paths: list[str]) -> str:
    ordered: list[str] = []
    seen: set[str] = set()
    for path in paths:
        if not path or path in seen:
            continue
        seen.add(path)
        ordered.append(path)
    return "; ".join(ordered)


def first_paths(context: PaperContext, hits: list[SearchHit]) -> str:
    if hits:
        return stringify_paths([hit.source_path for hit in hits])
    base_paths = [str(path.resolve()) for path in context.table_paths]
    if context.text_path is not None:
        base_paths.append(str(context.text_path.resolve()))
    return stringify_paths(base_paths)


def has_loaded_sibling_with_same_drug(
    row: dict[str, str],
    rows_by_paper: dict[str, list[dict[str, str]]],
) -> bool:
    target_value = normalize_lower(row.get("drug_name", ""))
    if not target_value:
        return False
    for sibling in sibling_rows(rows_by_paper, row):
        if normalize_lower(sibling.get("drug_name", "")) != target_value:
            continue
        if not is_blank_like_row(sibling):
            return True
    return False


def detect_reference_only_hits(hits: list[SearchHit]) -> list[SearchHit]:
    return [hit for hit in hits if hit.reference_like]


def build_debug_row(
    row: dict[str, str],
    field_name: str,
    value: str,
    reason: str,
    context: PaperContext,
) -> dict[str, str]:
    return {
        "paper_id": row.get("doi", "") or row.get("paper_key", ""),
        "doi": row.get("doi", ""),
        "formulation_id": row.get("gt_formulation_id", ""),
        "field_name": field_name,
        "current_value": value,
        "suppressed_reason": reason,
        "source_paths": first_paths(context, []),
    }


def valid_ratio_support(hit: SearchHit, current_value: str) -> bool:
    snippet = normalize_lower(hit.snippet)
    current = normalize_lower(current_value)
    if current in snippet:
        return True
    compact = current.replace(" ", "")
    return compact and compact in snippet.replace(" ", "")


def normalize_with_ascii_quotes(value: str) -> str:
    return normalize_text(
        value.replace("\u2018", "'").replace("\u2019", "'").replace("\u201c", '"').replace("\u201d", '"')
    )


def extract_json_array(text: str) -> list[dict[str, Any]]:
    raw = normalize_with_ascii_quotes(str(text or ""))
    if not raw:
        return []
    code_match = re.search(r"```(?:json)?\s*(\[[\s\S]*?\])\s*```", raw, re.IGNORECASE)
    if code_match:
        raw = code_match.group(1)
    else:
        match = re.search(r"(\[[\s\S]*\])", raw)
        if match:
            raw = match.group(1)
    try:
        payload = json.loads(raw)
    except Exception:
        return []
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    return []


def normalize_model_risk_level(value: str, fallback: str) -> str:
    token = normalize_lower(value)
    if token in {"high", "medium", "low"}:
        return token
    if token in {"critical", "severe"}:
        return "high"
    if token in {"moderate"}:
        return "medium"
    return fallback


def normalize_model_risk_type(value: str, fallback: str) -> str:
    token = normalize_lower(value).replace(" ", "_").replace("-", "_")
    mapping = {
        "blank_should_be_null": "blank_should_be_null",
        "unsupported_value": "unsupported_value",
        "direction_mismatch": "direction_mismatch",
        "cross_paper_contamination": "cross_paper_contamination",
        "inheritance_contamination": "inheritance_contamination",
        "ambiguity": "ambiguity",
        "derived_value": "derived_value",
        "unit_or_normalization_only": "unit_or_normalization_only",
        "data_inconsistency": fallback,
        "inconsistency": fallback,
    }
    return mapping.get(token, fallback)


def normalize_model_evidence_status(value: str, fallback: str) -> str:
    token = normalize_lower(value)
    if token in {"supported", "derived", "unsupported", "ambiguous"}:
        return token
    if token in {"confirmed", "present"}:
        return fallback
    return fallback


def load_env_file() -> None:
    env_path = PROJECT_ROOT / ".env"
    if env_path.exists():
        load_dotenv(dotenv_path=env_path, override=False)


def ensure_gemini_backend(model: str) -> None:
    load_env_file()
    if not str(model or "").strip():
        raise RuntimeError("Gemini model name is empty; pass --gemini-model explicitly or skip Gemini execution.")
    if not HAS_GENAI:
        raise RuntimeError("google-generativeai is not installed in this environment.")
    key = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
    if not key:
        raise RuntimeError("GEMINI_API_KEY / GOOGLE_API_KEY is missing in environment.")
    genai.configure(api_key=key)
    _ = genai.GenerativeModel(model)


def ensure_nvidia_backend(model: str) -> None:
    load_env_file()
    api_key = os.getenv("NVIDIA_API_KEY")
    if not api_key:
        raise RuntimeError("NVIDIA_API_KEY is missing in environment.")
    if not str(model or "").strip():
        raise RuntimeError("NVIDIA model name is empty.")


def call_gemini_backend(
    model: str,
    prompt: str,
    retries: int = DEFAULT_MAX_RETRIES,
    sleep_sec: float = 1.0,
    timeout_seconds: int = DEFAULT_REQUEST_TIMEOUT_SECONDS,
) -> str:
    ensure_gemini_backend(model)
    last_err: Exception | None = None
    for attempt in range(retries + 1):
        try:
            response = genai.GenerativeModel(model).generate_content(
                prompt,
                request_options={"timeout": timeout_seconds},
            )
            if hasattr(response, "text") and response.text:
                return str(response.text)
            try:
                candidate = response.candidates[0].content.parts[0].text
                if candidate:
                    return str(candidate)
            except Exception:
                pass
            raise RuntimeError("Gemini returned empty content.")
        except Exception as exc:
            last_err = exc
            if attempt < retries:
                time.sleep(sleep_sec)
    raise last_err or RuntimeError("Gemini call failed.")


def call_nvidia_backend(
    model: str,
    prompt: str,
    retries: int = DEFAULT_MAX_RETRIES,
    sleep_sec: float = 1.0,
    timeout_seconds: int = DEFAULT_REQUEST_TIMEOUT_SECONDS,
) -> str:
    ensure_nvidia_backend(model)
    api_key = os.getenv("NVIDIA_API_KEY") or ""
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": "Return only a valid JSON array."},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0,
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    last_err: Exception | None = None
    for attempt in range(retries + 1):
        try:
            response = requests.post(
                NVIDIA_HOSTED_CHAT_COMPLETIONS_URL,
                headers=headers,
                json=payload,
                timeout=timeout_seconds,
            )
            response.raise_for_status()
            body = response.json()
            choices = body.get("choices") or []
            if not choices:
                raise RuntimeError(f"NVIDIA response had no choices: {body}")
            message = choices[0].get("message") or {}
            content = message.get("content")
            if isinstance(content, str) and content.strip():
                return content
            raise RuntimeError(f"NVIDIA response content was empty: {body}")
        except Exception as exc:
            last_err = exc
            if attempt < retries:
                time.sleep(sleep_sec * (attempt + 1))
    raise last_err or RuntimeError("NVIDIA call failed.")


def classify_rule_flags(
    row: dict[str, str],
    field_name: str,
    value: str,
    context: PaperContext,
    rows_by_paper: dict[str, list[dict[str, str]]],
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    flags: list[dict[str, Any]] = []
    suppressed: list[dict[str, str]] = []
    if field_name not in HIGH_PRECISION_FIELDS:
        return flags, suppressed

    unit_hint = field_search_unit_hint(row, field_name)
    hits = find_hits(context, field_name, value, unit_hint=unit_hint)
    exact_hit = next((hit for hit in hits if hit.exact and not hit.reference_like), None)
    non_reference_hit = next((hit for hit in hits if not hit.reference_like), None)
    reference_hits = detect_reference_only_hits(hits)
    reference_only = bool(reference_hits) and non_reference_hit is None
    blank_like = is_blank_like_row(row)
    loaded_sibling_same_drug = has_loaded_sibling_with_same_drug(row, rows_by_paper) if field_name == "drug_name" else False

    if field_name == "drug_name" and blank_like and value:
        flags.append(
            {
                "risk_type": "blank_should_be_null",
                "risk_level": "high",
                "evidence_status": "unsupported" if non_reference_hit is None else "ambiguous",
                "reason": (
                    "Blank/empty/unloaded formulation carries a non-null drug value. "
                    "Layer 3 blank rows must preserve a null drug field."
                ),
                "snippet": first_snippet(hits),
                "paths": first_paths(context, hits),
                "extras": ["inheritance_contamination"] if loaded_sibling_same_drug else [],
            }
        )
        return flags, suppressed

    if field_name == "drug_name" and reference_only and value:
        flags.append(
            {
                "risk_type": "cross_paper_contamination",
                "risk_level": "high",
                "evidence_status": "ambiguous",
                "reason": "Value appears only in reference-like text and lacks direct paper-local formulation support.",
                "snippet": first_snippet(reference_hits),
                "paths": first_paths(context, hits),
                "extras": [],
            }
        )
        return flags, suppressed

    if field_name in RATIO_FIELDS and value:
        _, inverse_variants = ratio_variant_sets(value)
        inverse_hits: list[SearchHit] = []
        if context.text_raw and context.text_path is not None:
            inverse_hits.extend(
                search_hits_for_variants(
                    context.text_norm,
                    context.text_raw,
                    context.text_norm_to_raw,
                    str(context.text_path.resolve()),
                    "text",
                    sorted(inverse_variants),
                    normalize_lower(value),
                )
            )
        for source_path, table_norm, mapping in context.table_norm_parts:
            raw = ""
            for raw_path, raw_content in context.table_raw_parts:
                if raw_path == source_path:
                    raw = raw_content
                    break
            inverse_hits.extend(
                search_hits_for_variants(
                    table_norm,
                    raw,
                    mapping,
                    source_path,
                    "table",
                    sorted(inverse_variants),
                    normalize_lower(value),
                )
            )
        inverse_hits = [hit for hit in dedupe_hits(inverse_hits) if not hit.reference_like and not valid_ratio_support(hit, value)]
        direct_ratio_support = exact_hit is not None or any(valid_ratio_support(hit, value) for hit in hits if not hit.reference_like)
        if inverse_hits and not direct_ratio_support:
            flags.append(
                {
                    "risk_type": "direction_mismatch",
                    "risk_level": "medium",
                    "evidence_status": "derived",
                    "reason": "Inverse ratio representation appears in source, but the workbook direction is not explicitly supported.",
                    "snippet": first_snippet(inverse_hits),
                    "paths": first_paths(context, inverse_hits),
                    "extras": [],
                }
            )
        elif hits and not direct_ratio_support:
            suppressed.append(build_debug_row(row, field_name, value, "suppressed_non_exact_ratio_support", context))
        return flags, suppressed

    if field_name == "drug_name" and blank_like and loaded_sibling_same_drug and non_reference_hit is None:
        flags.append(
            {
                "risk_type": "inheritance_contamination",
                "risk_level": "high",
                "evidence_status": "unsupported",
                "reason": "Blank control appears to have inherited drug identity from a loaded sibling without row-local support.",
                "snippet": "",
                "paths": stringify_paths([str(path.resolve()) for path in context.table_paths] + ([str(context.text_path.resolve())] if context.text_path else [])),
                "extras": [],
            }
        )
        return flags, suppressed

    if field_name == "drug_name" and not hits and value:
        flags.append(
            {
                "risk_type": "unsupported_value",
                "risk_level": "medium",
                "evidence_status": "unsupported",
                "reason": "Drug value was not found explicitly in the current paper's cleaned text or cleaned tables.",
                "snippet": "",
                "paths": stringify_paths([str(path.resolve()) for path in context.table_paths] + ([str(context.text_path.resolve())] if context.text_path else [])),
                "extras": [],
            }
        )
    elif field_name == "drug_name" and hits and exact_hit is None:
        suppressed.append(build_debug_row(row, field_name, value, "suppressed_partial_drug_match", context))
    return flags, suppressed


def choose_best_flag(
    row: dict[str, str],
    field_name: str,
    value: str,
    candidate_flags: list[dict[str, Any]],
) -> CellFlag | None:
    if not candidate_flags:
        return None
    ranked = sorted(
        candidate_flags,
        key=lambda item: (
            RISK_LEVEL_PRIORITY.get(item["risk_level"], 0),
            RISK_PRIORITY.get(item["risk_type"], 0),
        ),
        reverse=True,
    )
    best = ranked[0]
    extras: list[str] = []
    for candidate in ranked:
        for extra in candidate.get("extras", []):
            if extra not in extras and extra != best["risk_type"]:
                extras.append(extra)
    secondary = [
        candidate["risk_type"]
        for candidate in ranked[1:]
        if candidate["risk_type"] != best["risk_type"] and candidate["risk_type"] not in extras
    ]
    if secondary:
        extras.extend(secondary[:2])
    reason = best["reason"]
    if extras:
        reason += " Secondary signals: " + ", ".join(extras) + "."
    return CellFlag(
        paper_id=row.get("doi", "") or row.get("paper_key", ""),
        doi=row.get("doi", ""),
        formulation_id=row.get("gt_formulation_id", ""),
        field_name=field_name,
        current_value=value,
        risk_level=best["risk_level"],
        risk_type=best["risk_type"],
        source_of_flag="rule",
        reason=reason,
        evidence_status=best["evidence_status"],
        evidence_snippet=best["snippet"],
        source_paths=best["paths"],
        extras=extras,
    )


def rule_audit(
    workbook_path: Path,
    sheet_name: str,
    out_tsv: Path,
    out_json: Path | None = None,
    debug_tsv: Path | None = None,
) -> tuple[list[CellFlag], dict[str, Any], list[dict[str, str]]]:
    rows = workbook_rows(workbook_path, sheet_name)
    rows_by_paper: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        rows_by_paper[row.get("paper_key", "")].append(row)

    resolver = AuditEvidenceResolverV1(PROJECT_ROOT)
    paper_contexts: dict[str, PaperContext] = {}
    flags: list[CellFlag] = []
    debug_rows: list[dict[str, str]] = []
    candidate_cells = 0

    for row in rows:
        paper_key = row.get("paper_key", "")
        if paper_key not in paper_contexts:
            paper_contexts[paper_key] = build_paper_context(resolver, paper_key, row.get("doi", ""))
        context = paper_contexts[paper_key]
        for field_name in sorted(HIGH_PRECISION_FIELDS):
            value = normalize_text(row.get(field_name, ""))
            if not value:
                continue
            candidate_cells += 1
            candidate_flags, suppressed_rows = classify_rule_flags(row, field_name, value, context, rows_by_paper)
            debug_rows.extend(suppressed_rows)
            best = choose_best_flag(row, field_name, value, candidate_flags)
            if best is not None and best.risk_type in MAIN_RISK_TYPES and best.risk_level in {"high", "medium"}:
                flags.append(best)

    flags = sorted(
        flags,
        key=lambda item: (
            item.paper_id,
            item.formulation_id,
            99 - FIELD_PRIORITY.index(item.field_name) if item.field_name in FIELD_PRIORITY else 0,
            item.field_name,
        ),
    )
    rows_out = [flag.to_row() for flag in flags]
    write_tsv(out_tsv, rows_out, REPORT_COLUMNS)
    if debug_tsv is not None and debug_rows:
        write_tsv(debug_tsv, debug_rows, ["paper_id", "doi", "formulation_id", "field_name", "current_value", "suppressed_reason", "source_paths"])
    if out_json is not None:
        out_json.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "workbook_path": str(workbook_path.resolve()),
            "sheet_name": sheet_name,
            "candidate_nonempty_cells": candidate_cells,
            "flagged_cells": len(flags),
            "debug_rows": len(debug_rows),
            "rows": rows_out,
        }
        out_json.write_text(json.dumps(payload, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    summary = {
        "candidate_nonempty_cells": candidate_cells,
        "flagged_cells": len(flags),
        "debug_rows": len(debug_rows),
        "by_risk_type": dict(Counter(flag.risk_type for flag in flags)),
        "by_source": {"rule": len(flags)},
        "workbook_path": str(workbook_path.resolve()),
        "sheet_name": sheet_name,
    }
    return flags, summary, debug_rows


def export_model_tasks(
    rule_tsv: Path,
    task_rows_tsv: Path,
    gemini_jsonl: Path,
    nvidia_jsonl: Path,
    max_candidates: int | None = None,
) -> dict[str, Any]:
    rows = slice_candidates(read_tsv(rule_tsv), max_candidates)
    task_rows: list[dict[str, str]] = []
    gemini_records: list[dict[str, Any]] = []
    nvidia_records: list[dict[str, Any]] = []
    for row in rows:
        record = {
            "paper_id": row.get("paper_id", ""),
            "doi": row.get("doi", ""),
            "formulation_id": row.get("formulation_id", ""),
            "field_name": row.get("field_name", ""),
            "current_value": row.get("current_value", ""),
            "risk_level": row.get("risk_level", ""),
            "risk_type": row.get("risk_type", ""),
            "reason": row.get("reason", ""),
            "evidence_status": row.get("evidence_status", ""),
            "evidence_snippet": row.get("evidence_snippet", ""),
            "source_paths": row.get("source_paths", ""),
        }
        task_rows.append(record)
        prompt = build_model_prompt(record)
        gemini_records.append(
            {
                "model_family": "gemini",
                "system_prompt": MODEL_SYSTEM_PROMPT,
                "user_prompt": prompt,
                "expected_schema": MODEL_OUTPUT_SCHEMA,
                "record": record,
            }
        )
        nvidia_records.append(
            {
                "model_family": "nvidia",
                "system_prompt": MODEL_SYSTEM_PROMPT,
                "user_prompt": prompt,
                "expected_schema": MODEL_OUTPUT_SCHEMA,
                "record": record,
            }
        )

    write_tsv(
        task_rows_tsv,
        task_rows,
        list(task_rows[0].keys()) if task_rows else [
            "paper_id",
            "doi",
            "formulation_id",
            "field_name",
            "current_value",
            "risk_level",
            "risk_type",
            "reason",
            "evidence_status",
            "evidence_snippet",
            "source_paths",
        ],
    )
    write_jsonl(gemini_jsonl, gemini_records)
    write_jsonl(nvidia_jsonl, nvidia_records)
    return {
        "task_rows": len(task_rows),
        "task_rows_tsv": str(task_rows_tsv.resolve()),
        "gemini_jsonl": str(gemini_jsonl.resolve()),
        "nvidia_jsonl": str(nvidia_jsonl.resolve()),
    }


def build_model_prompt(record: dict[str, str]) -> str:
    return (
        "Audit this workbook cell. Return only a JSON array. Use the exact enumerations from the schema. "
        "If the cell is not worth human review, return []. Do not edit, fill, compute, normalize, invert, "
        "or infer support.\n\n"
        f"paper_id: {record['paper_id']}\n"
        f"doi: {record.get('doi', '')}\n"
        f"formulation_id: {record['formulation_id']}\n"
        f"field_name: {record['field_name']}\n"
        f"current_value: {record['current_value']}\n"
        f"rule_risk_level: {record.get('risk_level', '')}\n"
        f"rule_risk_type: {record.get('risk_type', '')}\n"
        f"rule_reason: {record.get('reason', '')}\n"
        f"rule_evidence_status: {record.get('evidence_status', '')}\n"
        f"evidence_snippet: {record.get('evidence_snippet', '')}\n"
        f"source_paths: {record['source_paths']}\n\n"
        "Only flag cells that are genuinely worth manual re-checking. Allowed risk types are: "
        "blank_should_be_null, unsupported_value, direction_mismatch, cross_paper_contamination, inheritance_contamination. "
        "Allowed risk levels are high, medium, low. Allowed evidence_status values are supported, derived, unsupported, ambiguous. "
        "source_of_flag must be the backend name only. Return either [] or a JSON array matching this exact schema: "
        + json.dumps(MODEL_OUTPUT_SCHEMA, ensure_ascii=True)
    )


def write_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=True) + "\n")


def slice_candidates(rows: list[dict[str, str]], max_candidates: int | None) -> list[dict[str, str]]:
    if max_candidates is None or max_candidates <= 0:
        return list(rows)
    return list(rows[:max_candidates])


def batched_rows(rows: list[dict[str, str]], batch_size: int) -> list[list[dict[str, str]]]:
    size = max(1, batch_size)
    return [rows[index : index + size] for index in range(0, len(rows), size)]


def print_backend_progress(
    backend: str,
    *,
    message: str,
    batch_index: int | None = None,
    total_batches: int | None = None,
    request_count: int | None = None,
    success_count: int | None = None,
    failure_count: int | None = None,
    parse_empty_count: int | None = None,
    out_tsv: Path | None = None,
) -> None:
    parts = [f"backend_progress\t{backend}", message]
    if batch_index is not None and total_batches is not None:
        parts.append(f"batch={batch_index}/{total_batches}")
    if request_count is not None:
        parts.append(f"requests={request_count}")
    if success_count is not None:
        parts.append(f"success={success_count}")
    if failure_count is not None:
        parts.append(f"failure={failure_count}")
    if parse_empty_count is not None:
        parts.append(f"parse_empty={parse_empty_count}")
    if out_tsv is not None:
        parts.append(f"out_tsv={out_tsv}")
    print("\t".join(parts))


def execute_model_audit(
    candidates_tsv: Path,
    backend: str,
    out_tsv: Path,
    *,
    model_name: str,
    max_candidates: int | None = None,
    max_calls: int | None = None,
    batch_size: int = DEFAULT_BATCH_SIZE,
    request_timeout_seconds: int = DEFAULT_REQUEST_TIMEOUT_SECONDS,
    max_retries: int = DEFAULT_MAX_RETRIES,
    write_partial_every_batch: bool = False,
) -> dict[str, Any]:
    candidates = slice_candidates(read_tsv(candidates_tsv), max_candidates)
    if max_calls is not None and max_calls > 0:
        candidates = candidates[:max_calls]
    results: list[dict[str, str]] = []
    blocked_reason = ""
    request_count = 0
    success_count = 0
    failure_count = 0
    parse_empty_count = 0
    write_tsv(out_tsv, [], REPORT_COLUMNS)

    try:
        if backend == "gemini":
            ensure_gemini_backend(model_name)
        elif backend == "nvidia":
            ensure_nvidia_backend(model_name)
        else:
            raise ValueError(f"Unsupported backend: {backend}")
    except Exception as exc:
        blocked_reason = str(exc)
        print_backend_progress(backend, message=f"blocked_before_execution: {blocked_reason}", out_tsv=out_tsv)
        return {
            "backend": backend,
            "candidate_cells": len(candidates),
            "request_count": 0,
            "success_count": 0,
            "failure_count": 0,
            "parse_empty_count": 0,
            "actually_sent": 0,
            "returned_rows": 0,
            "blocked_reason": blocked_reason,
            "out_tsv": str(out_tsv.resolve()),
        }

    batches = batched_rows(candidates, batch_size)
    print_backend_progress(
        backend,
        message="starting",
        batch_index=0 if batches else None,
        total_batches=len(batches) if batches else None,
        request_count=request_count,
        success_count=success_count,
        failure_count=failure_count,
        parse_empty_count=parse_empty_count,
        out_tsv=out_tsv,
    )
    for batch_index, batch in enumerate(batches, start=1):
        print_backend_progress(
            backend,
            message="batch_start",
            batch_index=batch_index,
            total_batches=len(batches),
            request_count=request_count,
            success_count=success_count,
            failure_count=failure_count,
            parse_empty_count=parse_empty_count,
        )
        for candidate in batch:
            prompt = build_model_prompt(candidate)
            request_count += 1
            try:
                if backend == "gemini":
                    raw = call_gemini_backend(
                        model_name,
                        prompt,
                        retries=max_retries,
                        timeout_seconds=request_timeout_seconds,
                    )
                else:
                    raw = call_nvidia_backend(
                        model_name,
                        prompt,
                        retries=max_retries,
                        timeout_seconds=request_timeout_seconds,
                    )
                success_count += 1
                parsed = extract_json_array(raw)
                if not parsed:
                    parse_empty_count += 1
                    continue
                for item in parsed:
                    normalized = normalize_model_row(item, candidate, backend)
                    if normalized is not None:
                        results.append(normalized)
            except Exception as exc:
                failure_count += 1
                blocked_reason = str(exc)
                print_backend_progress(
                    backend,
                    message=f"request_failed: {blocked_reason}",
                    batch_index=batch_index,
                    total_batches=len(batches),
                    request_count=request_count,
                    success_count=success_count,
                    failure_count=failure_count,
                    parse_empty_count=parse_empty_count,
                )
                break
        if write_partial_every_batch or blocked_reason:
            write_tsv(out_tsv, results, REPORT_COLUMNS)
            print_backend_progress(
                backend,
                message="partial_write",
                batch_index=batch_index,
                total_batches=len(batches),
                request_count=request_count,
                success_count=success_count,
                failure_count=failure_count,
                parse_empty_count=parse_empty_count,
                out_tsv=out_tsv,
            )
        if blocked_reason:
            break

    write_tsv(out_tsv, results, REPORT_COLUMNS)
    print_backend_progress(
        backend,
        message="completed" if not blocked_reason else "stopped_after_failure",
        batch_index=len(batches) if batches else 0,
        total_batches=len(batches) if batches else 0,
        request_count=request_count,
        success_count=success_count,
        failure_count=failure_count,
        parse_empty_count=parse_empty_count,
        out_tsv=out_tsv,
    )
    return {
        "backend": backend,
        "candidate_cells": len(candidates),
        "request_count": request_count,
        "success_count": success_count,
        "failure_count": failure_count,
        "parse_empty_count": parse_empty_count,
        "actually_sent": request_count,
        "returned_rows": len(results),
        "blocked_reason": blocked_reason,
        "out_tsv": str(out_tsv.resolve()),
    }


def normalize_model_row(item: dict[str, Any], candidate: dict[str, str], backend: str) -> dict[str, str] | None:
    risk_type = normalize_model_risk_type(str(item.get("risk_type", "")), candidate.get("risk_type", "unsupported_value"))
    if risk_type not in MAIN_RISK_TYPES:
        return None
    risk_level = normalize_model_risk_level(str(item.get("risk_level", "")), candidate.get("risk_level", "medium"))
    evidence_status = normalize_model_evidence_status(str(item.get("evidence_status", "")), candidate.get("evidence_status", "ambiguous"))
    return {
        "paper_id": candidate.get("paper_id", ""),
        "doi": candidate.get("doi", ""),
        "formulation_id": candidate.get("formulation_id", ""),
        "field_name": candidate.get("field_name", ""),
        "current_value": candidate.get("current_value", ""),
        "risk_level": risk_level,
        "risk_type": risk_type,
        "source_of_flag": backend,
        "reason": normalize_text(item.get("reason", "")) or candidate.get("reason", ""),
        "evidence_status": evidence_status,
        "evidence_snippet": normalize_text(item.get("evidence_snippet", "")) or candidate.get("evidence_snippet", ""),
        "source_paths": normalize_text(item.get("source_paths", "")) if isinstance(item.get("source_paths", ""), str) else stringify_paths([str(x) for x in item.get("source_paths", [])]),
    }


def merge_reports(
    rule_tsv: Path,
    merged_tsv: Path,
    markdown_path: Path,
    *,
    gemini_tsv: Path | None = None,
    nvidia_tsv: Path | None = None,
    high_risk_tsv: Path | None = None,
    execution_summary: dict[str, Any] | None = None,
) -> dict[str, Any]:
    source_rows: dict[str, list[dict[str, str]]] = {
        "rule": read_tsv(rule_tsv),
        "gemini": read_tsv(gemini_tsv) if gemini_tsv else [],
        "nvidia": read_tsv(nvidia_tsv) if nvidia_tsv else [],
    }
    grouped: dict[tuple[str, str, str], list[tuple[str, dict[str, str]]]] = defaultdict(list)
    for source_name, rows in source_rows.items():
        for row in rows:
            key = (row.get("paper_id", ""), row.get("formulation_id", ""), row.get("field_name", ""))
            grouped[key].append((source_name, row))

    merged_rows: list[dict[str, str]] = []
    for key, source_group in grouped.items():
        paper_id, formulation_id, field_name = key
        sources = sorted({source_name for source_name, _ in source_group})
        source_of_flag = "+".join(sources)
        best_row = max(
            [row for _, row in source_group],
            key=lambda item: (
                RISK_LEVEL_PRIORITY.get(item.get("risk_level", "low"), 0),
                RISK_PRIORITY.get(item.get("risk_type", "ambiguity"), 0),
            ),
        )
        risk_level = best_row.get("risk_level", "low")
        if source_of_flag in HIGH_PRIORITY_SOURCE_OF_FLAG:
            risk_level = elevate_risk_level(risk_level)
        reasons = []
        snippets = []
        paths = []
        risk_types = []
        for source_name, row in source_group:
            risk_type = row.get("risk_type", "")
            if risk_type and risk_type not in risk_types:
                risk_types.append(risk_type)
            reason = normalize_text(row.get("reason", ""))
            if reason and reason not in reasons:
                reasons.append(f"{source_name}: {reason}")
            snippet = normalize_text(row.get("evidence_snippet", ""))
            if snippet and snippet not in snippets:
                snippets.append(snippet)
            source_path = normalize_text(row.get("source_paths", ""))
            if source_path and source_path not in paths:
                paths.append(source_path)
        primary_risk_type = sorted(risk_types, key=lambda item: RISK_PRIORITY.get(item, 0), reverse=True)[0]
        evidence_status = merge_evidence_status([row.get("evidence_status", "") for _, row in source_group])
        merged_rows.append(
            {
                "paper_id": paper_id,
                "doi": best_row.get("doi", ""),
                "formulation_id": formulation_id,
                "field_name": field_name,
                "current_value": best_row.get("current_value", ""),
                "risk_level": risk_level,
                "risk_type": primary_risk_type,
                "source_of_flag": source_of_flag,
                "reason": " | ".join(reasons),
                "evidence_status": evidence_status,
                "evidence_snippet": " || ".join(snippets[:3]),
                "source_paths": " ; ".join(paths),
            }
        )

    merged_rows = sorted(
        merged_rows,
        key=lambda row: (
            -review_priority_score(row),
            -RISK_LEVEL_PRIORITY.get(row.get("risk_level", "low"), 0),
            row.get("paper_id", ""),
            row.get("doi", ""),
            row.get("formulation_id", ""),
            row.get("field_name", ""),
        ),
    )
    write_tsv(merged_tsv, merged_rows, REPORT_COLUMNS)
    if high_risk_tsv is not None:
        high_rows = [row for row in merged_rows if review_priority_score(row) >= 2]
        write_tsv(high_risk_tsv, high_rows, REPORT_COLUMNS)

    summary = summarize_rows(merged_rows)
    markdown_path.parent.mkdir(parents=True, exist_ok=True)
    markdown_path.write_text(
        build_markdown_report(merged_rows, summary, source_rows, merged_tsv, high_risk_tsv, execution_summary=execution_summary),
        encoding="utf-8",
    )
    return summary


def merge_evidence_status(statuses: list[str]) -> str:
    ordered = ["unsupported", "ambiguous", "derived", "supported"]
    for name in ordered:
        if name in statuses:
            return name
    return statuses[0] if statuses else ""


def elevate_risk_level(level: str) -> str:
    if level == "low":
        return "medium"
    if level == "medium":
        return "high"
    return level


def summarize_rows(rows: list[dict[str, str]]) -> dict[str, Any]:
    return {
        "total_flagged": len(rows),
        "by_risk_type": dict(Counter(row.get("risk_type", "") for row in rows)),
        "by_source": dict(Counter(row.get("source_of_flag", "") for row in rows)),
        "by_risk_level": dict(Counter(row.get("risk_level", "") for row in rows)),
        "high_risk": sum(1 for row in rows if row.get("risk_level", "") == "high"),
    }


def review_priority_score(row: dict[str, str]) -> int:
    source = row.get("source_of_flag", "")
    risk_level = row.get("risk_level", "")
    if source in HIGH_PRIORITY_SOURCE_OF_FLAG:
        return 3
    if source == "rule" and risk_level == "high":
        return 2
    return 1


def build_markdown_report(
    rows: list[dict[str, str]],
    summary: dict[str, Any],
    source_rows: dict[str, list[dict[str, str]]],
    merged_tsv: Path,
    high_risk_tsv: Path | None,
    execution_summary: dict[str, Any] | None = None,
) -> str:
    lines: list[str] = []
    lines.append("# Layer 3 GT Cross-Audit Report v4")
    lines.append("")
    lines.append("## Scope")
    lines.append("")
    lines.append("- Surface: value GT annotation workbook v4")
    lines.append("- Principle: retain only values explicitly supported by source text or tables")
    lines.append("- Audit mode: report-only; no workbook edits and no value corrections")
    lines.append("")
    lines.append("## Outputs")
    lines.append("")
    lines.append(f"- Merged TSV: `{merged_tsv}`")
    if high_risk_tsv is not None:
        lines.append(f"- High-risk subset: `{high_risk_tsv}`")
    lines.append("")
    lines.append("## Counts")
    lines.append("")
    lines.append(f"- Total flagged cells: `{summary['total_flagged']}`")
    lines.append(f"- High-risk cells: `{summary['high_risk']}`")
    lines.append("")
    if execution_summary:
        lines.append("## Execution")
        lines.append("")
        lines.append(f"- Candidate cells after refined rule pass: `{execution_summary.get('rule_candidates', 0)}`")
        lines.append(f"- Candidate cells selected for model execution: `{execution_summary.get('effective_model_candidates', 0)}`")
        lines.append(f"- Gemini requests attempted: `{execution_summary.get('gemini_requests', 0)}`")
        lines.append(f"- NVIDIA requests attempted: `{execution_summary.get('nvidia_requests', 0)}`")
        lines.append(f"- Gemini successes / failures: `{execution_summary.get('gemini_success', 0)}` / `{execution_summary.get('gemini_failure', 0)}`")
        lines.append(f"- NVIDIA successes / failures: `{execution_summary.get('nvidia_success', 0)}` / `{execution_summary.get('nvidia_failure', 0)}`")
        lines.append(f"- Cells actually sent to Gemini: `{execution_summary.get('gemini_sent', 0)}`")
        lines.append(f"- Cells actually sent to NVIDIA: `{execution_summary.get('nvidia_sent', 0)}`")
        lines.append(f"- Cells returned by Gemini: `{execution_summary.get('gemini_rows', 0)}`")
        lines.append(f"- Cells returned by NVIDIA: `{execution_summary.get('nvidia_rows', 0)}`")
        lines.append(f"- Actually executed: `{execution_summary.get('executed_note', '')}`")
        blocked = execution_summary.get("blocked_note", "")
        if blocked:
            lines.append(f"- Blockers: `{blocked}`")
        lines.append("")
    lines.append("### By Risk Type")
    lines.append("")
    for key, count in sorted(summary["by_risk_type"].items(), key=lambda item: (-item[1], item[0])):
        lines.append(f"- `{key}`: `{count}`")
    lines.append("")
    lines.append("### By Source")
    lines.append("")
    for key, count in sorted(summary["by_source"].items(), key=lambda item: (-item[1], item[0])):
        lines.append(f"- `{key}`: `{count}`")
    lines.append("")
    lines.append("## Source Availability")
    lines.append("")
    for source_name, source_list in source_rows.items():
        lines.append(f"- `{source_name}` rows merged: `{len(source_list)}`")
    lines.append("")
    lines.append("## High-Risk Sample")
    lines.append("")
    high_rows = [row for row in rows if row.get("risk_level", "") == "high"][:20]
    if not high_rows:
        lines.append("No high-risk rows were produced.")
    else:
        lines.append("| paper_id | doi | formulation_id | field_name | current_value | risk_type | source_of_flag | reason |")
        lines.append("|---|---|---|---|---|---|---|---|")
        for row in high_rows:
            lines.append(
                f"| {escape_md(row.get('paper_id', ''))} | {escape_md(row.get('doi', ''))} | {escape_md(row.get('formulation_id', ''))} | "
                f"{escape_md(row.get('field_name', ''))} | {escape_md(row.get('current_value', ''))} | "
                f"{escape_md(row.get('risk_type', ''))} | {escape_md(row.get('source_of_flag', ''))} | "
                f"{escape_md(shorten(row.get('reason', ''), 140))} |"
            )
    lines.append("")
    lines.append("## Notes")
    lines.append("")
    lines.append("- `supported` means explicit source support was available and the cell was still flagged for another risk such as contamination or ambiguity.")
    lines.append("- `derived` means the workbook value appears computable or formatting-dependent rather than explicitly stated.")
    lines.append("- `unsupported` means the value was not found explicitly in cleaned text/tables for the current paper.")
    lines.append("- Model outputs are audit signals only and must not be treated as edits or truth labels.")
    lines.append("- Use the TSV to locate the cell in the existing v4 workbook and manually verify the cited paper-local evidence.")
    lines.append("")
    return "\n".join(lines) + "\n"


def escape_md(value: str) -> str:
    return normalize_text(value).replace("|", "\\|")


def shorten(value: str, limit: int) -> str:
    text = normalize_text(value)
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."


def resolve_workbook_and_run_context(args: argparse.Namespace) -> tuple[dict[str, Any], Path]:
    run_context = resolve_run_context(
        explicit_run_dir=args.run_dir.resolve() if getattr(args, "run_dir", None) else None,
        explicit_run_id=getattr(args, "run_id", ""),
    )
    workbook_path = resolve_artifact_path(
        explicit_path=args.workbook.resolve() if getattr(args, "workbook", None) else None,
        run_context=run_context,
        pointer_key="value_gt_annotation_workbook_v4",
        preferred_run_local_names=["value_gt_annotation_workbook_representation_repaired_v4.xlsx"],
        required=True,
    )
    assert workbook_path is not None
    return run_context, workbook_path


def default_analysis_dir(run_context: dict[str, Any]) -> Path:
    out_dir = Path(run_context["run_dir"]) / "analysis"
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def add_common_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--run-dir", type=Path, help="Explicit run directory. If omitted, ACTIVE_RUN.json is used.")
    parser.add_argument("--run-id", default="", help="Compatibility override for explicit run id.")
    parser.add_argument("--workbook", type=Path, help="Explicit workbook path. Overrides run-context resolution.")
    parser.add_argument("--sheet-name", default="", help="Workbook sheet to audit. Defaults to value_gt_annotation.")


def add_model_execution_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--skip-gemini", action="store_true", help="Skip Gemini execution.")
    parser.add_argument("--skip-nvidia", action="store_true", help="Skip NVIDIA execution.")
    parser.add_argument("--max-candidates", type=int, default=0, help="Limit candidate cells passed from rules into model execution. 0 means no limit.")
    parser.add_argument("--max-gemini-calls", type=int, default=0, help="Cap Gemini requests. 0 means no limit.")
    parser.add_argument("--max-nvidia-calls", type=int, default=0, help="Cap NVIDIA requests. 0 means no limit.")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE, help="Progress and partial-write batch size for model execution.")
    parser.add_argument("--request-timeout-seconds", type=int, default=DEFAULT_REQUEST_TIMEOUT_SECONDS, help="Per-request timeout for live model calls.")
    parser.add_argument("--max-retries", type=int, default=DEFAULT_MAX_RETRIES, help="Maximum retries per request after the first attempt.")
    parser.add_argument("--write-partial-every-batch", action="store_true", help="Write backend TSV outputs after each batch.")


def positive_or_none(value: int) -> int | None:
    if value is None or value <= 0:
        return None
    return value


def backend_skipped_summary(backend: str, out_tsv: Path, reason: str) -> dict[str, Any]:
    write_tsv(out_tsv, [], REPORT_COLUMNS)
    print_backend_progress(backend, message=reason, out_tsv=out_tsv)
    return {
        "backend": backend,
        "candidate_cells": 0,
        "request_count": 0,
        "success_count": 0,
        "failure_count": 0,
        "parse_empty_count": 0,
        "actually_sent": 0,
        "returned_rows": 0,
        "blocked_reason": reason,
        "out_tsv": str(out_tsv.resolve()),
    }


def command_rule_audit(args: argparse.Namespace) -> int:
    run_context, workbook_path = resolve_workbook_and_run_context(args)
    sheet_name = choose_sheet(workbook_path, args.sheet_name)
    analysis_dir = default_analysis_dir(run_context)
    out_tsv = args.out_tsv.resolve() if args.out_tsv else analysis_dir / "layer3_cross_audit_rule_flags_v4.tsv"
    out_json = args.out_json.resolve() if args.out_json else analysis_dir / "layer3_cross_audit_rule_flags_v4.json"
    debug_tsv = args.debug_tsv.resolve() if args.debug_tsv else analysis_dir / "layer3_cross_audit_rule_suppressed_v4.tsv"
    _, summary, _ = rule_audit(workbook_path, sheet_name, out_tsv, out_json, debug_tsv)
    metadata = build_artifact_metadata(
        source_run_context=run_context,
        source_files={"workbook": str(workbook_path.resolve())},
        generated_by="src/stage5_benchmark/run_layer3_cross_audit_v1.py rule-audit",
        note="Layer 3 rule-based cell risk audit. Audit-only output.",
        extra=summary,
    )
    write_artifact_metadata_json(out_tsv, metadata)
    print(f"resolved_run_dir\t{run_context['run_dir']}")
    print(f"workbook\t{workbook_path}")
    print(f"sheet_name\t{sheet_name}")
    print(f"rule_flags_tsv\t{out_tsv}")
    print(f"debug_tsv\t{debug_tsv}")
    print(f"candidate_nonempty_cells\t{summary['candidate_nonempty_cells']}")
    print(f"flagged_cells\t{summary['flagged_cells']}")
    print(f"debug_rows\t{summary['debug_rows']}")
    for risk_type, count in sorted(summary["by_risk_type"].items(), key=lambda item: (-item[1], item[0])):
        print(f"risk_type_count\t{risk_type}\t{count}")
    return 0


def command_export_model_tasks(args: argparse.Namespace) -> int:
    run_context, _ = resolve_workbook_and_run_context(args)
    analysis_dir = default_analysis_dir(run_context)
    task_rows_tsv = args.task_rows_tsv.resolve() if args.task_rows_tsv else analysis_dir / "layer3_cross_audit_model_tasks_v4.tsv"
    gemini_jsonl = args.gemini_jsonl.resolve() if args.gemini_jsonl else analysis_dir / "layer3_cross_audit_gemini_tasks_v4.jsonl"
    nvidia_jsonl = args.nvidia_jsonl.resolve() if args.nvidia_jsonl else analysis_dir / "layer3_cross_audit_nvidia_tasks_v4.jsonl"
    rule_tsv = args.rule_tsv.resolve() if args.rule_tsv else analysis_dir / "layer3_cross_audit_rule_flags_v4.tsv"
    summary = export_model_tasks(rule_tsv, task_rows_tsv, gemini_jsonl, nvidia_jsonl)
    metadata = build_artifact_metadata(
        source_run_context=run_context,
        source_files={
            "rule_flags_tsv": str(rule_tsv.resolve()) if rule_tsv.exists() else "",
        },
        generated_by="src/stage5_benchmark/run_layer3_cross_audit_v1.py export-model-tasks",
        note="Gemini/NVIDIA auditor task export for Layer 3 cross-audit. Task prompts only.",
        extra=summary,
    )
    write_artifact_metadata_json(task_rows_tsv, metadata)
    print(f"resolved_run_dir\t{run_context['run_dir']}")
    print(f"task_rows\t{summary['task_rows']}")
    print(f"task_rows_tsv\t{summary['task_rows_tsv']}")
    print(f"gemini_jsonl\t{summary['gemini_jsonl']}")
    print(f"nvidia_jsonl\t{summary['nvidia_jsonl']}")
    return 0


def command_execute_models(args: argparse.Namespace) -> int:
    run_context, _ = resolve_workbook_and_run_context(args)
    analysis_dir = default_analysis_dir(run_context)
    candidates_tsv = args.candidates_tsv.resolve() if args.candidates_tsv else analysis_dir / "layer3_cross_audit_rule_flags_v4.tsv"
    gemini_tsv = args.gemini_tsv.resolve() if args.gemini_tsv else analysis_dir / "layer3_cross_audit_gemini_results_v4.tsv"
    nvidia_tsv = args.nvidia_tsv.resolve() if args.nvidia_tsv else analysis_dir / "layer3_cross_audit_nvidia_results_v4.tsv"
    max_candidates = positive_or_none(args.max_candidates)
    candidate_count = len(slice_candidates(read_tsv(candidates_tsv), max_candidates))
    if args.skip_gemini:
        gemini_summary = backend_skipped_summary("gemini", gemini_tsv, "skipped_via_cli")
    else:
        gemini_summary = execute_model_audit(
            candidates_tsv,
            "gemini",
            gemini_tsv,
            model_name=args.gemini_model,
            max_candidates=max_candidates,
            max_calls=positive_or_none(args.max_gemini_calls),
            batch_size=max(1, args.batch_size),
            request_timeout_seconds=max(1, args.request_timeout_seconds),
            max_retries=max(0, args.max_retries),
            write_partial_every_batch=args.write_partial_every_batch,
        )
    if args.skip_nvidia:
        nvidia_summary = backend_skipped_summary("nvidia", nvidia_tsv, "skipped_via_cli")
    else:
        nvidia_summary = execute_model_audit(
            candidates_tsv,
            "nvidia",
            nvidia_tsv,
            model_name=args.nvidia_model,
            max_candidates=max_candidates,
            max_calls=positive_or_none(args.max_nvidia_calls),
            batch_size=max(1, args.batch_size),
            request_timeout_seconds=max(1, args.request_timeout_seconds),
            max_retries=max(0, args.max_retries),
            write_partial_every_batch=args.write_partial_every_batch,
        )
    print(f"resolved_run_dir\t{run_context['run_dir']}")
    print(f"candidate_cells\t{candidate_count}")
    print(f"gemini_requests\t{gemini_summary['request_count']}")
    print(f"gemini_sent\t{gemini_summary['actually_sent']}")
    print(f"gemini_success\t{gemini_summary['success_count']}")
    print(f"gemini_failure\t{gemini_summary['failure_count']}")
    print(f"gemini_parse_empty\t{gemini_summary['parse_empty_count']}")
    print(f"gemini_rows\t{gemini_summary['returned_rows']}")
    print(f"gemini_blocked\t{gemini_summary['blocked_reason']}")
    print(f"nvidia_requests\t{nvidia_summary['request_count']}")
    print(f"nvidia_sent\t{nvidia_summary['actually_sent']}")
    print(f"nvidia_success\t{nvidia_summary['success_count']}")
    print(f"nvidia_failure\t{nvidia_summary['failure_count']}")
    print(f"nvidia_parse_empty\t{nvidia_summary['parse_empty_count']}")
    print(f"nvidia_rows\t{nvidia_summary['returned_rows']}")
    print(f"nvidia_blocked\t{nvidia_summary['blocked_reason']}")
    return 0


def command_merge(args: argparse.Namespace) -> int:
    run_context = resolve_run_context(
        explicit_run_dir=args.run_dir.resolve() if args.run_dir else None,
        explicit_run_id=args.run_id,
    )
    analysis_dir = default_analysis_dir(run_context)
    rule_tsv = args.rule_tsv.resolve() if args.rule_tsv else analysis_dir / "layer3_cross_audit_rule_flags_v4.tsv"
    gemini_tsv = args.gemini_tsv.resolve() if args.gemini_tsv else analysis_dir / "layer3_cross_audit_gemini_results_v4.tsv"
    nvidia_tsv = args.nvidia_tsv.resolve() if args.nvidia_tsv else analysis_dir / "layer3_cross_audit_nvidia_results_v4.tsv"
    merged_tsv = args.merged_tsv.resolve() if args.merged_tsv else analysis_dir / "layer3_gt_cross_audit_report_v4.tsv"
    high_risk_tsv = args.high_risk_tsv.resolve() if args.high_risk_tsv else analysis_dir / "layer3_gt_cross_audit_report_v4_high_priority.tsv"
    markdown_path = args.markdown.resolve() if args.markdown else PROJECT_ROOT / "project" / "layer3_gt_cross_audit_report_v4.md"
    summary = merge_reports(
        rule_tsv,
        merged_tsv,
        markdown_path,
        gemini_tsv=gemini_tsv,
        nvidia_tsv=nvidia_tsv,
        high_risk_tsv=high_risk_tsv,
    )
    metadata = build_artifact_metadata(
        source_run_context=run_context,
        source_files={
            "rule_tsv": str(rule_tsv.resolve()),
            "gemini_tsv": str(gemini_tsv.resolve()) if gemini_tsv else "",
            "nvidia_tsv": str(nvidia_tsv.resolve()) if nvidia_tsv else "",
        },
        generated_by="src/stage5_benchmark/run_layer3_cross_audit_v1.py merge",
        note="Merged Layer 3 cross-audit report. Audit-only output.",
        extra=summary,
    )
    write_artifact_metadata_json(merged_tsv, metadata)
    print(f"resolved_run_dir\t{run_context['run_dir']}")
    print(f"merged_tsv\t{merged_tsv}")
    print(f"markdown\t{markdown_path}")
    print(f"total_flagged\t{summary['total_flagged']}")
    for risk_type, count in sorted(summary["by_risk_type"].items(), key=lambda item: (-item[1], item[0])):
        print(f"risk_type_count\t{risk_type}\t{count}")
    for source_name, count in sorted(summary["by_source"].items(), key=lambda item: (-item[1], item[0])):
        print(f"source_count\t{source_name}\t{count}")
    return 0


def command_run(args: argparse.Namespace) -> int:
    run_context, workbook_path = resolve_workbook_and_run_context(args)
    sheet_name = choose_sheet(workbook_path, args.sheet_name)
    analysis_dir = default_analysis_dir(run_context)
    rule_tsv = analysis_dir / "layer3_cross_audit_rule_flags_v4.tsv"
    rule_json = analysis_dir / "layer3_cross_audit_rule_flags_v4.json"
    debug_tsv = analysis_dir / "layer3_cross_audit_rule_suppressed_v4.tsv"
    task_rows_tsv = analysis_dir / "layer3_cross_audit_model_tasks_v4.tsv"
    gemini_jsonl = analysis_dir / "layer3_cross_audit_gemini_tasks_v4.jsonl"
    nvidia_jsonl = analysis_dir / "layer3_cross_audit_nvidia_tasks_v4.jsonl"
    gemini_results_tsv = analysis_dir / "layer3_cross_audit_gemini_results_v4.tsv"
    nvidia_results_tsv = analysis_dir / "layer3_cross_audit_nvidia_results_v4.tsv"
    merged_tsv = analysis_dir / "layer3_gt_cross_audit_report_v4.tsv"
    high_risk_tsv = analysis_dir / "layer3_gt_cross_audit_report_v4_high_priority.tsv"
    markdown_path = PROJECT_ROOT / "project" / "layer3_gt_cross_audit_report_v4.md"

    _, rule_summary, _ = rule_audit(workbook_path, sheet_name, rule_tsv, rule_json, debug_tsv)
    max_candidates = positive_or_none(args.max_candidates)
    effective_candidate_count = min(rule_summary["flagged_cells"], max_candidates) if max_candidates else rule_summary["flagged_cells"]
    if args.rules_only:
        export_summary = {
            "task_rows": 0,
            "task_rows_tsv": str(task_rows_tsv.resolve()),
            "gemini_jsonl": str(gemini_jsonl.resolve()),
            "nvidia_jsonl": str(nvidia_jsonl.resolve()),
            "note": "rules_only_mode_skipped_task_export",
        }
        gemini_summary = backend_skipped_summary("gemini", gemini_results_tsv, "skipped_via_rules_only")
        nvidia_summary = backend_skipped_summary("nvidia", nvidia_results_tsv, "skipped_via_rules_only")
    else:
        export_summary = export_model_tasks(
            rule_tsv,
            task_rows_tsv,
            gemini_jsonl,
            nvidia_jsonl,
            max_candidates=max_candidates,
        )
        if args.skip_gemini:
            gemini_summary = backend_skipped_summary("gemini", gemini_results_tsv, "skipped_via_cli")
        else:
            gemini_summary = execute_model_audit(
                rule_tsv,
                "gemini",
                gemini_results_tsv,
                model_name=args.gemini_model,
                max_candidates=max_candidates,
                max_calls=positive_or_none(args.max_gemini_calls),
                batch_size=max(1, args.batch_size),
                request_timeout_seconds=max(1, args.request_timeout_seconds),
                max_retries=max(0, args.max_retries),
                write_partial_every_batch=args.write_partial_every_batch,
            )
        if args.skip_nvidia:
            nvidia_summary = backend_skipped_summary("nvidia", nvidia_results_tsv, "skipped_via_cli")
        else:
            nvidia_summary = execute_model_audit(
                rule_tsv,
                "nvidia",
                nvidia_results_tsv,
                model_name=args.nvidia_model,
                max_candidates=max_candidates,
                max_calls=positive_or_none(args.max_nvidia_calls),
                batch_size=max(1, args.batch_size),
                request_timeout_seconds=max(1, args.request_timeout_seconds),
                max_retries=max(0, args.max_retries),
                write_partial_every_batch=args.write_partial_every_batch,
            )
    execution_summary = {
        "rule_candidates": rule_summary["flagged_cells"],
        "effective_model_candidates": effective_candidate_count,
        "gemini_requests": gemini_summary["request_count"],
        "nvidia_requests": nvidia_summary["request_count"],
        "gemini_success": gemini_summary["success_count"],
        "nvidia_success": nvidia_summary["success_count"],
        "gemini_failure": gemini_summary["failure_count"],
        "nvidia_failure": nvidia_summary["failure_count"],
        "gemini_sent": gemini_summary["actually_sent"],
        "nvidia_sent": nvidia_summary["actually_sent"],
        "gemini_rows": gemini_summary["returned_rows"],
        "nvidia_rows": nvidia_summary["returned_rows"],
        "executed_note": ",".join(
            [
                label
                for label, summary in [("gemini", gemini_summary), ("nvidia", nvidia_summary)]
                if not summary["blocked_reason"]
            ]
        ) or "none",
        "blocked_note": "; ".join(
            [
                f"{label}: {summary['blocked_reason']}"
                for label, summary in [("gemini", gemini_summary), ("nvidia", nvidia_summary)]
                if summary["blocked_reason"]
            ]
        ),
    }
    merge_summary = merge_reports(
        rule_tsv,
        merged_tsv,
        markdown_path,
        gemini_tsv=gemini_results_tsv,
        nvidia_tsv=nvidia_results_tsv,
        high_risk_tsv=high_risk_tsv,
        execution_summary=execution_summary,
    )

    metadata = build_artifact_metadata(
        source_run_context=run_context,
        source_files={"workbook": str(workbook_path.resolve())},
        generated_by="src/stage5_benchmark/run_layer3_cross_audit_v1.py run",
        note="Full Layer 3 cross-audit run. Rule report generated, model tasks exported, merged report written from available sources.",
        extra={
            "rule_summary": rule_summary,
            "task_export_summary": export_summary,
            "gemini_execution_summary": gemini_summary,
            "nvidia_execution_summary": nvidia_summary,
            "merge_summary": merge_summary,
        },
    )
    write_artifact_metadata_json(merged_tsv, metadata)
    print(f"resolved_run_dir\t{run_context['run_dir']}")
    print(f"workbook\t{workbook_path}")
    print(f"sheet_name\t{sheet_name}")
    print(f"rule_flags_tsv\t{rule_tsv}")
    print(f"debug_tsv\t{debug_tsv}")
    print(f"task_rows_tsv\t{task_rows_tsv}")
    print(f"gemini_jsonl\t{gemini_jsonl}")
    print(f"nvidia_jsonl\t{nvidia_jsonl}")
    print(f"gemini_results_tsv\t{gemini_results_tsv}")
    print(f"nvidia_results_tsv\t{nvidia_results_tsv}")
    print(f"merged_tsv\t{merged_tsv}")
    print(f"markdown\t{markdown_path}")
    print(f"candidate_nonempty_cells\t{rule_summary['candidate_nonempty_cells']}")
    print(f"rule_candidate_cells\t{rule_summary['flagged_cells']}")
    print(f"effective_model_candidates\t{effective_candidate_count}")
    print(f"gemini_requests\t{gemini_summary['request_count']}")
    print(f"gemini_sent\t{gemini_summary['actually_sent']}")
    print(f"gemini_success\t{gemini_summary['success_count']}")
    print(f"gemini_failure\t{gemini_summary['failure_count']}")
    print(f"gemini_parse_empty\t{gemini_summary['parse_empty_count']}")
    print(f"nvidia_requests\t{nvidia_summary['request_count']}")
    print(f"nvidia_sent\t{nvidia_summary['actually_sent']}")
    print(f"nvidia_success\t{nvidia_summary['success_count']}")
    print(f"nvidia_failure\t{nvidia_summary['failure_count']}")
    print(f"nvidia_parse_empty\t{nvidia_summary['parse_empty_count']}")
    print(f"gemini_rows\t{gemini_summary['returned_rows']}")
    print(f"nvidia_rows\t{nvidia_summary['returned_rows']}")
    print(f"gemini_blocked\t{gemini_summary['blocked_reason']}")
    print(f"nvidia_blocked\t{nvidia_summary['blocked_reason']}")
    print(f"total_flagged\t{merge_summary['total_flagged']}")
    for risk_type, count in sorted(merge_summary["by_risk_type"].items(), key=lambda item: (-item[1], item[0])):
        print(f"risk_type_count\t{risk_type}\t{count}")
    for source_name, count in sorted(merge_summary["by_source"].items(), key=lambda item: (-item[1], item[0])):
        print(f"source_count\t{source_name}\t{count}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Layer 3 post-annotation cross-audit framework.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    parser_rule = subparsers.add_parser("rule-audit", help="Run deterministic rule-based cell audit.")
    add_common_args(parser_rule)
    parser_rule.add_argument("--out-tsv", type=Path, help="Output TSV path for rule flags.")
    parser_rule.add_argument("--out-json", type=Path, help="Optional JSON payload path.")
    parser_rule.add_argument("--debug-tsv", type=Path, help="Optional debug TSV path for suppressed low-value rule signals.")
    parser_rule.set_defaults(func=command_rule_audit)

    parser_export = subparsers.add_parser("export-model-tasks", help="Export Gemini/NVIDIA auditor task packs.")
    add_common_args(parser_export)
    parser_export.add_argument("--rule-tsv", type=Path, help="Optional rule TSV for preflag context.")
    parser_export.add_argument("--task-rows-tsv", type=Path, help="Task row TSV output path.")
    parser_export.add_argument("--gemini-jsonl", type=Path, help="Gemini task JSONL output path.")
    parser_export.add_argument("--nvidia-jsonl", type=Path, help="NVIDIA task JSONL output path.")
    parser_export.set_defaults(func=command_export_model_tasks)

    parser_exec = subparsers.add_parser("execute-models", help="Execute Gemini and NVIDIA auditors on filtered candidates.")
    add_common_args(parser_exec)
    parser_exec.add_argument("--candidates-tsv", type=Path, help="Filtered candidate TSV path.")
    parser_exec.add_argument("--gemini-tsv", type=Path, help="Gemini result TSV path.")
    parser_exec.add_argument("--nvidia-tsv", type=Path, help="NVIDIA result TSV path.")
    parser_exec.add_argument("--gemini-model", default="", help="Gemini model name. Required unless --skip-gemini is set.")
    parser_exec.add_argument("--nvidia-model", default="", help="NVIDIA model name. Required unless --skip-nvidia is set.")
    add_model_execution_args(parser_exec)
    parser_exec.set_defaults(func=command_execute_models)

    parser_merge = subparsers.add_parser("merge", help="Merge rule and model audit outputs into final report.")
    parser_merge.add_argument("--run-dir", type=Path, help="Explicit run directory. If omitted, ACTIVE_RUN.json is used.")
    parser_merge.add_argument("--run-id", default="", help="Compatibility override for explicit run id.")
    parser_merge.add_argument("--rule-tsv", type=Path, help="Rule TSV path.")
    parser_merge.add_argument("--gemini-tsv", type=Path, help="Optional Gemini audit TSV.")
    parser_merge.add_argument("--nvidia-tsv", type=Path, help="Optional NVIDIA audit TSV.")
    parser_merge.add_argument("--merged-tsv", type=Path, help="Merged TSV output path.")
    parser_merge.add_argument("--high-risk-tsv", type=Path, help="High-risk subset TSV output path.")
    parser_merge.add_argument("--markdown", type=Path, help="Markdown report path.")
    parser_merge.set_defaults(func=command_merge)

    parser_run = subparsers.add_parser("run", help="Run rule audit, export model tasks, and write merged report.")
    add_common_args(parser_run)
    parser_run.add_argument("--rules-only", action="store_true", help="Run only the refined rule audit and merge a rule-only report.")
    parser_run.add_argument("--gemini-model", default="", help="Gemini model name. Required unless --rules-only or --skip-gemini is set.")
    parser_run.add_argument("--nvidia-model", default="", help="NVIDIA model name. Required unless --rules-only or --skip-nvidia is set.")
    add_model_execution_args(parser_run)
    parser_run.set_defaults(func=command_run)
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
