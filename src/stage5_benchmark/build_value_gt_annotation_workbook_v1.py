#!/usr/bin/env python3
from __future__ import annotations

"""
Build a formulation-level value-focused GT annotation workbook from frozen
Layer 3 review artifacts.

Purpose:
- pivot field-level review rows into one row per formulation identity
- optimize manual numeric-value annotation for a small set of target fields
- preserve frozen benchmark semantics unchanged

This helper is annotation-surface only. It does not modify Stage 2, Stage 3,
or Stage 5 behavior, and it does not alter benchmark-valid outputs.
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from src.utils.active_data_source import (
        build_artifact_metadata,
        resolve_artifact_path,
        resolve_run_context,
        write_artifact_metadata_json,
    )
    from src.utils.run_id import validate_artifact_subdir
except ModuleNotFoundError:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from src.utils.active_data_source import (
        build_artifact_metadata,
        resolve_artifact_path,
        resolve_run_context,
        write_artifact_metadata_json,
    )
    from src.utils.run_id import validate_artifact_subdir


TARGET_FIELD_ORDER = [
    "drug_name",
    "drug_mass",
    "polymer_MW",
    "LA/GA",
    "surfactant_name",
    "surfactant_concentration",
    "drug_polymer_ratio",
    "particle_size",
    "EE",
    "LC",
]

FLAG_OPTIONS = ["correct", "incorrect", "missing", "unclear"]

BB3JUVW7_IDENTITY_OVERRIDE = {
    "BB3JUVW7_F04": "F2.2",
    "BB3JUVW7_F05": "F2.4",
    "BB3JUVW7_F06": "F2.5",
    "BB3JUVW7_F08": "F2.6",
    "BB3JUVW7_F10": "F2.7",
}

MAIN_COLUMNS = [
    "paper_key",
    "formulation_id",
    "formulation_label_stage5",
    "article_formulation_id",
    "article_formulation_label",
    "l2_gt_formulation_id",
    "l2_gt_formulation_label",
    "seed_pred_representative_source_formulation_id",
    "matched_system_formulation_id",
    "l2_gt_alignment_status",
    "paper_risk_level",
    "layer3_inclusion_flag",
    "sys_drug_name",
    "sys_drug_mass",
    "sys_polymer_MW",
    "sys_LA_GA",
    "sys_surfactant_name",
    "sys_surfactant_concentration",
    "sys_drug_polymer_ratio",
    "sys_particle_size",
    "sys_EE",
    "sys_LC",
    "gt_drug_name",
    "gt_drug_mass",
    "gt_polymer_MW",
    "gt_LA_GA",
    "gt_surfactant_name",
    "gt_surfactant_concentration",
    "gt_drug_polymer_ratio",
    "gt_particle_size",
    "gt_EE",
    "gt_LC",
    "gt_flag_drug_name",
    "gt_flag_drug_mass",
    "gt_flag_polymer_MW",
    "gt_flag_LA_GA",
    "gt_flag_surfactant_name",
    "gt_flag_surfactant_concentration",
    "gt_flag_drug_polymer_ratio",
    "gt_flag_particle_size",
    "gt_flag_EE",
    "gt_flag_LC",
    "source_locator_drug_name",
    "source_locator_drug_mass",
    "source_locator_polymer_MW",
    "source_locator_LA_GA",
    "source_locator_surfactant_name",
    "source_locator_surfactant_concentration",
    "source_locator_drug_polymer_ratio",
    "source_locator_particle_size",
    "source_locator_EE",
    "source_locator_LC",
    "annotation_source_workbook",
    "annotation_match_status",
    "annotation_match_key",
    "annotation_conflict_flag",
]

REFERENCE_COLUMNS = [
    "paper_key",
    "formulation_id",
    "formulation_label_stage5",
    "article_formulation_id",
    "article_formulation_label",
    "l2_gt_formulation_id",
    "l2_gt_formulation_label",
    "seed_pred_representative_source_formulation_id",
    "matched_system_formulation_id",
    "l2_gt_alignment_status",
    "field_name",
    "extracted_value",
    "extracted_unit",
    "evidence_text",
    "evidence_anchor_text",
    "evidence_source_type",
    "evidence_status_detail",
    "relation_resolution_rule",
    "relation_resolution_confidence",
    "review_warning",
    "normalization_status",
    "paper_risk_level",
    "layer3_inclusion_flag",
]

HUMAN_ANNOTATION_COLUMNS = [
    "gt_drug_name",
    "gt_drug_mass",
    "gt_polymer_MW",
    "gt_LA_GA",
    "gt_surfactant_name",
    "gt_surfactant_concentration",
    "gt_drug_polymer_ratio",
    "gt_particle_size",
    "gt_EE",
    "gt_LC",
    "gt_flag_drug_name",
    "gt_flag_drug_mass",
    "gt_flag_polymer_MW",
    "gt_flag_LA_GA",
    "gt_flag_surfactant_name",
    "gt_flag_surfactant_concentration",
    "gt_flag_drug_polymer_ratio",
    "gt_flag_particle_size",
    "gt_flag_EE",
    "gt_flag_LC",
]

PROVENANCE_COLUMNS = [
    "annotation_source_workbook",
    "annotation_match_status",
    "annotation_match_key",
    "annotation_conflict_flag",
]

OLD_COLUMN_ALIASES = {
    "gt_drug_name": ["drug_name_gt", "gt_value_drug_name"],
    "gt_drug_mass": ["drug_mass_gt", "gt_value_drug_mass"],
    "gt_polymer_MW": ["polymer_mw_gt", "gt_value_polymer_mw"],
    "gt_LA_GA": ["la_ga_gt", "gt_value_la_ga"],
    "gt_surfactant_name": ["surfactant_name_gt", "gt_value_surfactant_name"],
    "gt_surfactant_concentration": ["surfactant_concentration_gt", "gt_value_surfactant_concentration"],
    "gt_drug_polymer_ratio": ["drug_polymer_ratio_gt", "gt_value_drug_polymer_ratio"],
    "gt_particle_size": ["particle_size_gt", "gt_value_particle_size"],
    "gt_EE": ["ee_gt", "gt_value_ee"],
    "gt_LC": ["lc_gt", "gt_value_lc"],
    "gt_flag_drug_name": ["flag_drug_name"],
    "gt_flag_drug_mass": ["flag_drug_mass"],
    "gt_flag_polymer_MW": ["flag_polymer_mw"],
    "gt_flag_LA_GA": ["flag_la_ga"],
    "gt_flag_surfactant_name": ["flag_surfactant_name"],
    "gt_flag_surfactant_concentration": ["flag_surfactant_concentration"],
    "gt_flag_drug_polymer_ratio": ["flag_drug_polymer_ratio"],
    "gt_flag_particle_size": ["flag_particle_size"],
    "gt_flag_EE": ["flag_ee"],
    "gt_flag_LC": ["flag_lc"],
}


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def read_tsv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def sanitize_out_subdir(value: str) -> str:
    return validate_artifact_subdir(value, param_name="--out-subdir")


def workbook_name_for_version(version: int) -> str:
    return f"value_gt_annotation_workbook_v{int(version)}.xlsx"


def main_tsv_name_for_version(version: int) -> str:
    return f"value_gt_annotation_rows_v{int(version)}.tsv"


def reference_tsv_name_for_version(version: int) -> str:
    return f"value_gt_reference_rows_v{int(version)}.tsv"


def extra_tsv_name_for_version(version: int) -> str:
    return f"value_gt_extra_in_system_rows_v{int(version)}.tsv"


def merge_audit_csv_name_for_version(version: int) -> str:
    return f"annotation_merge_audit_v{int(version)}.csv"


def merge_summary_json_name_for_version(version: int) -> str:
    return f"annotation_merge_summary_v{int(version)}.json"


def merge_report_md_name_for_version(version: int) -> str:
    return f"annotation_merge_report_v{int(version)}.md"


def alignment_resolution_tsv_name_for_version(version: int) -> str:
    return f"value_gt_alignment_resolution_audit_v{int(version)}.tsv"


def resolve_workbook_name(version: int, explicit_name: str | None) -> str:
    if explicit_name:
        return explicit_name
    return workbook_name_for_version(version)


def normalize_key_token(value: str) -> str:
    return re.sub(r"\s+", " ", normalize_text(value)).strip().lower()


def normalize_label_token(value: str) -> str:
    text = normalize_key_token(value)
    text = re.sub(r"[\[\]\(\),;:]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text


def count_nonempty_annotations(row: dict[str, str]) -> int:
    return sum(1 for column in HUMAN_ANNOTATION_COLUMNS if normalize_text(row.get(column)))


def format_system_value(extracted_value: str, extracted_unit: str) -> str:
    value = normalize_text(extracted_value)
    unit = normalize_text(extracted_unit)
    if not value:
        return ""
    if not unit or unit.lower() in {"ratio", "none"}:
        return value
    if unit.lower() in value.lower():
        return value
    return f"{value} {unit}"


def detect_source_locator(audit_row: dict[str, str]) -> str:
    table_id = normalize_text(audit_row.get("table_id"))
    if table_id:
        return f"Table {table_id}"

    locator = normalize_text(audit_row.get("evidence_locator"))
    figure_match = re.search(r"(?i)\bfig(?:ure)?\.?\s*(\d+)", locator)
    if figure_match:
        return f"Fig {figure_match.group(1)}"

    source_type = normalize_text(audit_row.get("evidence_source_type")).lower()
    if source_type.startswith("table"):
        return "table"
    if "figure" in source_type:
        return "figure"
    if source_type:
        return "text"

    candidate_source = normalize_text(audit_row.get("candidate_source")).lower()
    if "figure" in candidate_source:
        return "figure"
    if "table" in candidate_source:
        return "table"
    return "text"


def field_to_main_column(field_name: str) -> str:
    mapping = {
        "drug_name": "sys_drug_name",
        "drug_mass": "sys_drug_mass",
        "polymer_MW": "sys_polymer_MW",
        "LA/GA": "sys_LA_GA",
        "surfactant_name": "sys_surfactant_name",
        "surfactant_concentration": "sys_surfactant_concentration",
        "drug_polymer_ratio": "sys_drug_polymer_ratio",
        "particle_size": "sys_particle_size",
        "EE": "sys_EE",
        "LC": "sys_LC",
    }
    return mapping[field_name]


def field_to_locator_column(field_name: str) -> str:
    mapping = {
        "drug_name": "source_locator_drug_name",
        "drug_mass": "source_locator_drug_mass",
        "polymer_MW": "source_locator_polymer_MW",
        "LA/GA": "source_locator_LA_GA",
        "surfactant_name": "source_locator_surfactant_name",
        "surfactant_concentration": "source_locator_surfactant_concentration",
        "drug_polymer_ratio": "source_locator_drug_polymer_ratio",
        "particle_size": "source_locator_particle_size",
        "EE": "source_locator_EE",
        "LC": "source_locator_LC",
    }
    return mapping[field_name]


def build_seed_index(seed_rows: list[dict[str, str]]) -> dict[tuple[str, str], dict[str, dict[str, str]]]:
    index: dict[tuple[str, str], dict[str, dict[str, str]]] = {}
    for row in seed_rows:
        key = (normalize_text(row.get("paper_key")), normalize_text(row.get("formulation_id")))
        field_name = normalize_text(row.get("field_name"))
        if field_name not in TARGET_FIELD_ORDER:
            continue
        index.setdefault(key, {})[field_name] = row
    return index


def build_final_row_index(final_rows: list[dict[str, str]]) -> dict[tuple[str, str], dict[str, str]]:
    index: dict[tuple[str, str], dict[str, str]] = {}
    for row in final_rows:
        key = (
            normalize_text(row.get("key") or row.get("paper_key")),
            normalize_text(row.get("final_formulation_id")),
        )
        if key[0] and key[1]:
            index[key] = row
    return index


SWEEP_SIGNAL_PATTERN = re.compile(
    r"(?i)\b("
    r"sweep|varying|varied|effect of|optimization|optimized variables|"
    r"doe|design matrix|box-behnken|response surface|factorial|"
    r"experimental runs|run order|parameter scan|concentration study|"
    r"polymer content study|drug amount study|stabilizer concentration study"
    r")\b"
)

CORE_ARTICLE_ID_PATTERN = re.compile(r"(?i)^f0*\d+(?:\.\d+)?$")


def has_sweep_signal(*values: str) -> bool:
    for value in values:
        text = normalize_text(value)
        if text and SWEEP_SIGNAL_PATTERN.search(text):
            return True
    return False


def classify_gt_semantic_role(
    *,
    paper_key: str,
    gt_formulation_id: str,
    gt_row: dict[str, str],
) -> str:
    explicit_candidates = infer_gt_article_id_candidates(
        paper_key=paper_key,
        gt_formulation_id=gt_formulation_id,
        gt_row=gt_row,
    )
    if has_sweep_signal(
        gt_row.get("formulation_label_raw", ""),
        gt_row.get("gt_evidence_note", ""),
    ):
        return "sweep"
    if explicit_candidates:
        return "core"
    return "unknown"


def classify_audit_candidate_semantic_role(audit_row: dict[str, str]) -> str:
    candidate_source = normalize_text(audit_row.get("candidate_source")).lower()
    article_id = normalize_text(audit_row.get("article_formulation_id"))
    representative_id = normalize_text(audit_row.get("representative_source_formulation_id"))
    article_label = normalize_text(audit_row.get("article_formulation_label"))
    needs_review_reason = normalize_text(audit_row.get("needs_review_reason"))
    evidence_source_type = normalize_text(audit_row.get("evidence_source_type"))
    table_id = normalize_text(audit_row.get("table_id"))

    if candidate_source in {"figure_variable_sweep", "synthetic_figure_sweep"}:
        return "sweep"
    if has_sweep_signal(
        candidate_source,
        article_label,
        needs_review_reason,
        evidence_source_type,
    ):
        return "sweep"
    if "[" in article_label and "]" in article_label:
        return "sweep"
    if CORE_ARTICLE_ID_PATTERN.fullmatch(article_id) or CORE_ARTICLE_ID_PATTERN.fullmatch(representative_id):
        return "core"
    if table_id and not has_sweep_signal(article_label, needs_review_reason):
        return "core"
    if article_label and any(
        token in article_label.lower()
        for token in ["(empty)", "(drug loaded)", "(blank)", "(control)"]
    ):
        return "core"
    return "unknown"


def build_audit_identity_indexes(
    audit_rows: list[dict[str, str]],
) -> dict[str, dict[tuple[str, str], list[dict[str, str]]]]:
    article_id_index: dict[tuple[str, str], list[dict[str, str]]] = {}
    representative_id_index: dict[tuple[str, str], list[dict[str, str]]] = {}
    formulation_id_index: dict[tuple[str, str], dict[str, str]] = {}
    for row in audit_rows:
        paper_key = normalize_text(row.get("paper_id") or row.get("paper_key") or row.get("key"))
        article_id = normalize_text(row.get("article_formulation_id"))
        representative_id = normalize_text(row.get("representative_source_formulation_id"))
        formulation_id = normalize_text(row.get("formulation_id"))
        if paper_key and article_id:
            article_id_index.setdefault((paper_key, article_id.lower()), []).append(row)
        if paper_key and representative_id:
            representative_id_index.setdefault((paper_key, representative_id.lower()), []).append(row)
        if paper_key and formulation_id:
            formulation_id_index[(paper_key, formulation_id)] = row
    return {
        "article_id": article_id_index,
        "representative_id": representative_id_index,
        "formulation_id": formulation_id_index,
    }


def infer_gt_article_id_candidates(
    *,
    paper_key: str,
    gt_formulation_id: str,
    gt_row: dict[str, str],
) -> list[str]:
    candidates: list[str] = []

    explicit_label = normalize_text(gt_row.get("formulation_label_raw"))
    if explicit_label and re.fullmatch(r"(?i)f0*\d+(?:\.\d+)?", explicit_label):
        candidates.append(explicit_label.upper())

    match = re.fullmatch(rf"{re.escape(paper_key)}_F0*(\d+)", gt_formulation_id)
    if match:
        number = int(match.group(1))
        candidates.append(f"F{number}")
        candidates.append(f"F{number:02d}")

    deduped: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        normalized = normalize_text(candidate)
        if not normalized:
            continue
        key = normalized.lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(normalized)
    return deduped


def resolve_direct_canonical_alignment(
    *,
    paper_key: str,
    gt_formulation_id: str,
    gt_row: dict[str, str],
    audit_identity_indexes: dict[str, dict[tuple[str, str], list[dict[str, str]]]],
) -> tuple[dict[str, str], str]:
    candidate_ids = infer_gt_article_id_candidates(
        paper_key=paper_key,
        gt_formulation_id=gt_formulation_id,
        gt_row=gt_row,
    )
    for candidate in candidate_ids:
        for index_name in ("article_id", "representative_id"):
            matches = audit_identity_indexes[index_name].get((paper_key, candidate.lower()), [])
            if len(matches) == 1:
                return matches[0], f"canonical_{index_name}_direct_match:{candidate}"
    return {}, ""


def add_alignment_candidate(
    candidate_rows: list[dict[str, str]],
    *,
    seen_formulation_ids: set[str],
    audit_row: dict[str, str],
) -> None:
    formulation_id = normalize_text(audit_row.get("formulation_id"))
    if not formulation_id or formulation_id in seen_formulation_ids:
        return
    seen_formulation_ids.add(formulation_id)
    candidate_rows.append(audit_row)


def collect_alignment_candidates(
    *,
    paper_key: str,
    gt_formulation_id: str,
    gt_row: dict[str, str],
    scaffold_pred_row_id: str,
    trusted_alignment_row: dict[str, str],
    direct_canonical_audit_row: dict[str, str],
    audit_identity_indexes: dict[str, dict[tuple[str, str], list[dict[str, str]]]],
) -> list[dict[str, str]]:
    candidate_rows: list[dict[str, str]] = []
    seen_formulation_ids: set[str] = set()

    add_alignment_candidate(
        candidate_rows,
        seen_formulation_ids=seen_formulation_ids,
        audit_row=direct_canonical_audit_row,
    )

    for candidate_id in infer_gt_article_id_candidates(
        paper_key=paper_key,
        gt_formulation_id=gt_formulation_id,
        gt_row=gt_row,
    ):
        for index_name in ("article_id", "representative_id"):
            for audit_row in audit_identity_indexes[index_name].get((paper_key, candidate_id.lower()), []):
                add_alignment_candidate(
                    candidate_rows,
                    seen_formulation_ids=seen_formulation_ids,
                    audit_row=audit_row,
                )

    if scaffold_pred_row_id:
        scaffold_row = audit_identity_indexes["formulation_id"].get((paper_key, scaffold_pred_row_id), {})
        add_alignment_candidate(
            candidate_rows,
            seen_formulation_ids=seen_formulation_ids,
            audit_row=scaffold_row,
        )

    trusted_pred_row_id = normalize_text(trusted_alignment_row.get("matched_system_formulation_id"))
    if trusted_pred_row_id:
        trusted_pred_row = audit_identity_indexes["formulation_id"].get((paper_key, trusted_pred_row_id), {})
        add_alignment_candidate(
            candidate_rows,
            seen_formulation_ids=seen_formulation_ids,
            audit_row=trusted_pred_row,
        )

    return candidate_rows


def resolve_semantic_alignment_candidate(
    *,
    paper_key: str,
    gt_formulation_id: str,
    gt_row: dict[str, str],
    scaffold_pred_row_id: str,
    trusted_alignment_row: dict[str, str],
    direct_canonical_audit_row: dict[str, str],
    direct_resolution_rule: str,
    audit_identity_indexes: dict[str, dict[tuple[str, str], list[dict[str, str]]]],
) -> tuple[dict[str, str], str, dict[str, Any]]:
    gt_semantic_role = classify_gt_semantic_role(
        paper_key=paper_key,
        gt_formulation_id=gt_formulation_id,
        gt_row=gt_row,
    )
    candidate_rows = collect_alignment_candidates(
        paper_key=paper_key,
        gt_formulation_id=gt_formulation_id,
        gt_row=gt_row,
        scaffold_pred_row_id=scaffold_pred_row_id,
        trusted_alignment_row=trusted_alignment_row,
        direct_canonical_audit_row=direct_canonical_audit_row,
        audit_identity_indexes=audit_identity_indexes,
    )
    candidate_details: list[dict[str, str]] = []
    for audit_row in candidate_rows:
        candidate_details.append(
            {
                "formulation_id": normalize_text(audit_row.get("formulation_id")),
                "article_formulation_id": normalize_text(audit_row.get("article_formulation_id")),
                "article_formulation_label": normalize_text(audit_row.get("article_formulation_label")),
                "semantic_role": classify_audit_candidate_semantic_role(audit_row),
            }
        )

    core_candidates = [row for row in candidate_details if row["semantic_role"] == "core"]
    sweep_candidates = [row for row in candidate_details if row["semantic_role"] == "sweep"]
    filtered_candidates = candidate_details
    semantic_filter_applied = "no"
    if gt_semantic_role == "core" and core_candidates:
        filtered_candidates = [row for row in candidate_details if row["semantic_role"] != "sweep"]
        semantic_filter_applied = "yes" if sweep_candidates else "no"

    selected = filtered_candidates[0] if filtered_candidates else {}
    selected_formulation_id = normalize_text(selected.get("formulation_id"))
    selected_audit_row = (
        audit_identity_indexes["formulation_id"].get((paper_key, selected_formulation_id), {})
        if selected_formulation_id
        else {}
    )

    if selected_formulation_id and selected_formulation_id != scaffold_pred_row_id and semantic_filter_applied == "yes":
        resolution_rule = f"s1_core_candidate_preferred:{selected_formulation_id}"
    elif selected_formulation_id and selected_formulation_id != scaffold_pred_row_id and direct_resolution_rule:
        resolution_rule = direct_resolution_rule
    else:
        resolution_rule = direct_resolution_rule

    return selected_audit_row, resolution_rule, {
        "gt_semantic_role": gt_semantic_role,
        "selected_candidate_semantic_role": normalize_text(selected.get("semantic_role")),
        "semantic_filter_applied": semantic_filter_applied,
        "core_candidate_ids": json.dumps([row["formulation_id"] for row in core_candidates], ensure_ascii=True),
        "sweep_candidate_ids": json.dumps([row["formulation_id"] for row in sweep_candidates], ensure_ascii=True),
        "candidate_pool_ids": json.dumps([row["formulation_id"] for row in candidate_details], ensure_ascii=True),
    }


def format_final_table_value(row: dict[str, str], value_key: str, default_unit: str = "") -> str:
    value = normalize_text(row.get(value_key))
    if not value:
        return ""
    if not default_unit:
        return value
    if default_unit.lower() in value.lower():
        return value
    return f"{value} {default_unit}"


def build_final_table_field_values(row: dict[str, str]) -> dict[str, str]:
    return {
        "drug_name": normalize_text(row.get("drug_name_value")),
        "drug_mass": normalize_text(row.get("drug_feed_amount_text_value")),
        "surfactant_name": normalize_text(row.get("surfactant_name_value")),
        "particle_size": format_final_table_value(row, "size_nm_value", "nm"),
    }


def append_reference_row(
    reference_rows: list[dict[str, Any]],
    *,
    paper_key: str,
    formulation_id: str,
    formulation_label_stage5: str,
    article_formulation_id: str,
    article_formulation_label: str,
    l2_gt_formulation_id: str,
    l2_gt_formulation_label: str,
    seed_pred_representative_source_formulation_id: str,
    matched_system_formulation_id: str,
    field_name: str,
    extracted_value: str,
    extracted_unit: str,
    paper_risk_level: str,
    layer3_inclusion_flag: str,
    l2_gt_alignment_status: str = "",
) -> None:
    reference_rows.append(
        {
            "paper_key": paper_key,
            "formulation_id": formulation_id,
            "formulation_label_stage5": formulation_label_stage5,
            "article_formulation_id": article_formulation_id,
            "article_formulation_label": article_formulation_label,
            "l2_gt_formulation_id": l2_gt_formulation_id,
            "l2_gt_formulation_label": l2_gt_formulation_label,
            "seed_pred_representative_source_formulation_id": seed_pred_representative_source_formulation_id,
            "matched_system_formulation_id": matched_system_formulation_id,
            "l2_gt_alignment_status": normalize_text(l2_gt_alignment_status),
            "field_name": field_name,
            "extracted_value": normalize_text(extracted_value),
            "extracted_unit": normalize_text(extracted_unit),
            "evidence_text": "",
            "evidence_anchor_text": "",
            "evidence_source_type": "",
            "evidence_status_detail": "",
            "relation_resolution_rule": "",
            "relation_resolution_confidence": "",
            "review_warning": "",
            "normalization_status": "",
            "paper_risk_level": paper_risk_level,
            "layer3_inclusion_flag": layer3_inclusion_flag,
        }
    )


def load_workbook_rows(path: Path, sheet_name: str) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=False)
    worksheet = workbook[sheet_name]
    headers = [normalize_text(cell.value) for cell in worksheet[1]]
    rows: list[dict[str, str]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        rows.append({header: normalize_text(value) for header, value in zip(headers, row)})
    return rows


def canonicalize_prior_row(row: dict[str, str]) -> dict[str, str]:
    out = {column: normalize_text(row.get(column)) for column in MAIN_COLUMNS}
    for target, aliases in OLD_COLUMN_ALIASES.items():
        if out.get(target):
            continue
        for alias in aliases:
            alias_value = normalize_text(row.get(alias))
            if alias_value:
                out[target] = alias_value
                break
    if not out.get("l2_gt_formulation_id"):
        out["l2_gt_formulation_id"] = normalize_text(row.get("formulation_id"))
    return out


def load_prior_sheet_rows(path: Path, sheet_name: str) -> list[dict[str, str]]:
    try:
        rows = load_workbook_rows(path, sheet_name)
    except KeyError:
        return []
    return [canonicalize_prior_row(row) for row in rows]


def build_prior_row_indexes(rows: list[dict[str, str]]) -> dict[str, dict[tuple[str, str], list[dict[str, str]]]]:
    indexes: dict[str, dict[tuple[str, str], list[dict[str, str]]]] = {
        "exact": {},
        "article_id": {},
        "article_label": {},
    }
    for row in rows:
        if count_nonempty_annotations(row) == 0:
            continue
        paper_key = normalize_text(row.get("paper_key"))
        exact_formulation_id = normalize_text(row.get("formulation_id") or row.get("l2_gt_formulation_id"))
        if paper_key and exact_formulation_id:
            indexes["exact"].setdefault((paper_key, exact_formulation_id), []).append(row)
        article_formulation_id = normalize_text(row.get("article_formulation_id"))
        if paper_key and article_formulation_id:
            indexes["article_id"].setdefault((paper_key, article_formulation_id), []).append(row)
        article_label = normalize_label_token(row.get("article_formulation_label", ""))
        if paper_key and article_label:
            indexes["article_label"].setdefault((paper_key, article_label), []).append(row)
    return indexes


def build_trusted_alignment_index(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    index: dict[str, dict[str, str]] = {}
    for row in rows:
        gt_formulation_id = normalize_text(row.get("l2_gt_formulation_id") or row.get("formulation_id"))
        if not gt_formulation_id:
            continue
        index[gt_formulation_id] = row
    return index


def build_alignment_index(alignment_rows: list[dict[str, str]]) -> tuple[dict[str, dict[str, str]], list[dict[str, str]]]:
    gt_index: dict[str, dict[str, str]] = {}
    extras: list[dict[str, str]] = []
    for row in alignment_rows:
        gt_formulation_id = normalize_text(row.get("gt_formulation_id"))
        pred_row_id = normalize_text(row.get("pred_row_id"))
        if gt_formulation_id:
            gt_index[gt_formulation_id] = row
        elif pred_row_id:
            extras.append(row)
    return gt_index, extras


def classify_gt_alignment_status(
    alignment_row: dict[str, str],
    *,
    pred_row_id: str,
    representative_source_formulation_id: str,
) -> str:
    alignment_decision = normalize_text(alignment_row.get("alignment_decision")).lower()
    alignment_confidence = normalize_text(alignment_row.get("alignment_confidence")).lower()

    trustworthy_aligned = (
        bool(pred_row_id)
        and bool(representative_source_formulation_id)
        and alignment_decision == "matched"
        and alignment_confidence in {"high", "medium"}
    )
    if trustworthy_aligned:
        return "aligned"
    if pred_row_id and alignment_decision == "matched":
        return "ambiguous_match"
    if pred_row_id:
        return "missing_in_system"
    return "missing_in_system"


def is_generated_gt_row_id(value: str, paper_key: str) -> bool:
    normalized = normalize_text(value)
    if not normalized:
        return False
    return bool(re.fullmatch(rf"{re.escape(paper_key)}_F\d+", normalized))


def has_valid_bridge_identity(
    *,
    paper_key: str,
    gt_formulation_id: str,
    gt_formulation_label: str,
    trusted_alignment_row: dict[str, str],
    article_formulation_id: str,
    article_formulation_label: str,
    seed_pred_representative_source_formulation_id: str,
) -> bool:
    seed_anchor = normalize_text(seed_pred_representative_source_formulation_id)
    article_id = normalize_text(article_formulation_id)
    article_label = normalize_text(article_formulation_label)
    gt_label = normalize_text(gt_formulation_label)
    trusted_l2_label = normalize_text(trusted_alignment_row.get("l2_gt_formulation_label"))

    if article_id and not is_generated_gt_row_id(article_id, paper_key):
        return True
    if seed_anchor and not is_generated_gt_row_id(seed_anchor, paper_key):
        return True
    if article_label and article_label not in {gt_label, gt_formulation_id}:
        return True
    if trusted_l2_label and trusted_l2_label not in {gt_label, gt_formulation_id}:
        return True
    return False


def resolve_canonical_presence(
    *,
    paper_key: str,
    gt_formulation_id: str,
    gt_formulation_label: str,
    pred_row_id: str,
    audit_row: dict[str, str],
    final_row: dict[str, str],
) -> tuple[bool, str, str, str]:
    if not pred_row_id or (not audit_row and not final_row):
        return False, "", "", ""

    canonical_article_id = normalize_text(audit_row.get("article_formulation_id"))
    canonical_article_label = normalize_text(audit_row.get("article_formulation_label"))
    canonical_seed_anchor = normalize_text(audit_row.get("representative_source_formulation_id")) or normalize_text(
        final_row.get("representative_source_formulation_id")
    )
    canonical_identity_valid = has_valid_bridge_identity(
        paper_key=paper_key,
        gt_formulation_id=gt_formulation_id,
        gt_formulation_label=gt_formulation_label,
        trusted_alignment_row={},
        article_formulation_id=canonical_article_id,
        article_formulation_label=canonical_article_label,
        seed_pred_representative_source_formulation_id=canonical_seed_anchor,
    )
    return canonical_identity_valid, canonical_article_id, canonical_article_label, canonical_seed_anchor


def bb3_identity_override(gt_formulation_id: str) -> tuple[str, str]:
    article_id = BB3JUVW7_IDENTITY_OVERRIDE.get(gt_formulation_id, "")
    if not article_id:
        return "", ""
    return article_id, article_id


def initialize_annotation_provenance(row: dict[str, Any]) -> None:
    row["annotation_source_workbook"] = ""
    row["annotation_match_status"] = "unmatched"
    row["annotation_match_key"] = ""
    row["annotation_conflict_flag"] = "false"


def preserve_columns() -> list[str]:
    return list(HUMAN_ANNOTATION_COLUMNS)


def merge_preserved_cells(
    new_row: dict[str, Any],
    prior_row: dict[str, str] | None,
    preserve_columns: list[str],
) -> int:
    if prior_row is None:
        return 0
    preserved = 0
    for column in preserve_columns:
        if column not in new_row:
            continue
        old_value = normalize_text(prior_row.get(column))
        if old_value:
            new_row[column] = old_value
            preserved += 1
    return preserved


def exact_match_key(row: dict[str, str]) -> tuple[str, str]:
    return (
        normalize_text(row.get("paper_key")),
        normalize_text(row.get("formulation_id") or row.get("l2_gt_formulation_id")),
    )


def article_id_match_key(row: dict[str, str]) -> tuple[str, str]:
    return (
        normalize_text(row.get("paper_key")),
        normalize_text(row.get("article_formulation_id")),
    )


def article_label_match_key(row: dict[str, str]) -> tuple[str, str]:
    return (
        normalize_text(row.get("paper_key")),
        normalize_label_token(row.get("article_formulation_label", "")),
    )


def match_prior_row(
    *,
    new_row: dict[str, Any],
    prior_indexes: dict[str, dict[tuple[str, str], list[dict[str, str]]]],
) -> tuple[dict[str, str] | None, str, str, bool, list[dict[str, str]]]:
    key_builders = [
        ("exact", exact_match_key, "paper_key+formulation_id"),
        ("fallback", article_id_match_key, "paper_key+article_formulation_id"),
        ("fallback", article_label_match_key, "paper_key+article_formulation_label_normalized"),
    ]
    for status, builder, key_name in key_builders:
        key = builder(new_row)
        if not all(key):
            continue
        index_name = "exact" if status == "exact" else ("article_id" if "article_formulation_id" in key_name else "article_label")
        matches = prior_indexes[index_name].get(key, [])
        if not matches:
            continue
        if len(matches) > 1:
            return None, "conflict", key_name, True, matches
        return matches[0], status, key_name, False, matches
    return None, "unmatched", "", False, []


def build_main_rows(
    audit_rows: list[dict[str, str]],
    seed_index: dict[tuple[str, str], dict[str, dict[str, str]]],
    final_row_index: dict[tuple[str, str], dict[str, str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    main_rows: list[dict[str, Any]] = []
    reference_rows: list[dict[str, Any]] = []

    seen_keys: set[tuple[str, str]] = set()
    for audit_row in audit_rows:
        paper_key = normalize_text(audit_row.get("paper_id") or audit_row.get("paper_key"))
        formulation_id = normalize_text(audit_row.get("formulation_id"))
        key = (paper_key, formulation_id)
        if not paper_key or not formulation_id or key in seen_keys:
            continue
        seen_keys.add(key)

        field_rows = seed_index.get(key, {})
        identity_seed_row = next(iter(field_rows.values()), {})
        final_row = final_row_index.get(key, {})
        final_field_values = build_final_table_field_values(final_row)
        default_locator = detect_source_locator(audit_row)
        main_row: dict[str, Any] = {
            "paper_key": paper_key,
            "formulation_id": formulation_id,
            "formulation_label_stage5": normalize_text(audit_row.get("formulation_label_stage5"))
            or normalize_text(identity_seed_row.get("formulation_label_stage5")),
            "article_formulation_id": normalize_text(audit_row.get("article_formulation_id"))
            or normalize_text(identity_seed_row.get("article_formulation_id")),
            "article_formulation_label": normalize_text(audit_row.get("article_formulation_label"))
            or normalize_text(identity_seed_row.get("article_formulation_label")),
            "l2_gt_formulation_id": formulation_id,
            "l2_gt_formulation_label": normalize_text(audit_row.get("article_formulation_label"))
            or normalize_text(identity_seed_row.get("article_formulation_label")),
            "seed_pred_representative_source_formulation_id": normalize_text(
                audit_row.get("representative_source_formulation_id")
            ),
            "matched_system_formulation_id": formulation_id,
            "l2_gt_alignment_status": "aligned",
            "paper_risk_level": normalize_text(identity_seed_row.get("paper_risk_level")),
            "layer3_inclusion_flag": normalize_text(identity_seed_row.get("layer3_inclusion_flag")),
            "sys_drug_name": final_field_values.get("drug_name", ""),
            "sys_drug_mass": final_field_values.get("drug_mass", ""),
            "sys_polymer_MW": "",
            "sys_LA_GA": "",
            "sys_surfactant_name": final_field_values.get("surfactant_name", ""),
            "sys_surfactant_concentration": "",
            "sys_drug_polymer_ratio": "",
            "sys_particle_size": final_field_values.get("particle_size", ""),
            "sys_EE": "",
            "sys_LC": "",
            "gt_drug_name": "",
            "gt_drug_mass": "",
            "gt_polymer_MW": "",
            "gt_LA_GA": "",
            "gt_surfactant_name": "",
            "gt_surfactant_concentration": "",
            "gt_drug_polymer_ratio": "",
            "gt_particle_size": "",
            "gt_EE": "",
            "gt_LC": "",
            "gt_flag_drug_name": "",
            "gt_flag_drug_mass": "",
            "gt_flag_polymer_MW": "",
            "gt_flag_LA_GA": "",
            "gt_flag_surfactant_name": "",
            "gt_flag_surfactant_concentration": "",
            "gt_flag_drug_polymer_ratio": "",
            "gt_flag_particle_size": "",
            "gt_flag_EE": "",
            "gt_flag_LC": "",
            "source_locator_drug_name": default_locator if final_field_values.get("drug_name") else "",
            "source_locator_drug_mass": default_locator if final_field_values.get("drug_mass") else "",
            "source_locator_polymer_MW": "",
            "source_locator_LA_GA": "",
            "source_locator_surfactant_name": default_locator if final_field_values.get("surfactant_name") else "",
            "source_locator_surfactant_concentration": "",
            "source_locator_drug_polymer_ratio": "",
            "source_locator_particle_size": default_locator if final_field_values.get("particle_size") else "",
            "source_locator_EE": "",
            "source_locator_LC": "",
        }
        initialize_annotation_provenance(main_row)

        for field_name in ["drug_name", "drug_mass", "surfactant_name", "particle_size"]:
            existing_field_row = field_rows.get(field_name)
            if existing_field_row is not None and normalize_text(existing_field_row.get("extracted_value")):
                continue
            field_value = main_row[field_to_main_column(field_name)]
            if field_value:
                append_reference_row(
                    reference_rows,
                paper_key=paper_key,
                formulation_id=formulation_id,
                formulation_label_stage5=main_row["formulation_label_stage5"],
                article_formulation_id=main_row["article_formulation_id"],
                article_formulation_label=main_row["article_formulation_label"],
                l2_gt_formulation_id=main_row["formulation_id"],
                l2_gt_formulation_label=main_row["article_formulation_label"],
                seed_pred_representative_source_formulation_id="",
                matched_system_formulation_id=formulation_id,
                field_name=field_name,
                extracted_value=field_value,
                extracted_unit="",
                    paper_risk_level=main_row["paper_risk_level"],
                    layer3_inclusion_flag=main_row["layer3_inclusion_flag"],
                    l2_gt_alignment_status=main_row["l2_gt_alignment_status"],
                )

        for field_name in TARGET_FIELD_ORDER:
            field_row = field_rows.get(field_name)
            if field_row is None:
                continue
            main_row[field_to_main_column(field_name)] = format_system_value(
                extracted_value=field_row.get("extracted_value", ""),
                extracted_unit=field_row.get("extracted_unit", ""),
            )
            main_row[field_to_locator_column(field_name)] = default_locator
            append_reference_row(
                reference_rows,
                paper_key=paper_key,
                formulation_id=formulation_id,
                formulation_label_stage5=main_row["formulation_label_stage5"],
                article_formulation_id=main_row["article_formulation_id"],
                article_formulation_label=main_row["article_formulation_label"],
                l2_gt_formulation_id=main_row["formulation_id"],
                l2_gt_formulation_label=main_row["article_formulation_label"],
                seed_pred_representative_source_formulation_id="",
                matched_system_formulation_id=formulation_id,
                field_name=field_name,
                extracted_value=normalize_text(field_row.get("extracted_value")),
                extracted_unit=normalize_text(field_row.get("extracted_unit")),
                paper_risk_level=normalize_text(field_row.get("paper_risk_level")),
                layer3_inclusion_flag=normalize_text(field_row.get("layer3_inclusion_flag")),
                l2_gt_alignment_status=main_row["l2_gt_alignment_status"],
            )
            reference_rows[-1]["evidence_text"] = normalize_text(field_row.get("evidence_text"))
            reference_rows[-1]["evidence_anchor_text"] = normalize_text(field_row.get("evidence_anchor_text"))
            reference_rows[-1]["evidence_source_type"] = normalize_text(field_row.get("evidence_source_type"))
            reference_rows[-1]["evidence_status_detail"] = normalize_text(field_row.get("evidence_status_detail"))
            reference_rows[-1]["relation_resolution_rule"] = normalize_text(field_row.get("relation_resolution_rule"))
            reference_rows[-1]["relation_resolution_confidence"] = normalize_text(field_row.get("relation_resolution_confidence"))
            reference_rows[-1]["review_warning"] = normalize_text(field_row.get("review_warning"))
            reference_rows[-1]["normalization_status"] = normalize_text(field_row.get("normalization_status"))

        main_rows.append(main_row)

    main_rows.sort(key=lambda row: (row["paper_key"], row["formulation_id"]))
    reference_rows.sort(key=lambda row: (row["paper_key"], row["formulation_id"], row["field_name"]))
    return main_rows, reference_rows


def blank_main_row() -> dict[str, Any]:
    row = {column: "" for column in MAIN_COLUMNS}
    initialize_annotation_provenance(row)
    return row


def build_gt_skeleton_rows(
    gt_rows: list[dict[str, str]],
    alignment_index: dict[str, dict[str, str]],
    extra_alignment_rows: list[dict[str, str]],
    audit_rows: list[dict[str, str]],
    seed_index: dict[tuple[str, str], dict[str, dict[str, str]]],
    final_row_index: dict[tuple[str, str], dict[str, str]],
    prior_rows: list[dict[str, str]],
    trusted_alignment_rows: list[dict[str, str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], int, list[dict[str, Any]]]:
    audit_by_key = {
        (normalize_text(row.get("paper_id") or row.get("paper_key")), normalize_text(row.get("formulation_id"))): row
        for row in audit_rows
    }
    audit_identity_indexes = build_audit_identity_indexes(audit_rows)
    trusted_alignment_index = build_trusted_alignment_index(trusted_alignment_rows)

    main_rows: list[dict[str, Any]] = []
    reference_rows: list[dict[str, Any]] = []
    extra_rows: list[dict[str, Any]] = []
    alignment_resolution_rows: list[dict[str, Any]] = []

    for gt_row in gt_rows:
        paper_key = normalize_text(gt_row.get("paper_key"))
        gt_formulation_id = normalize_text(gt_row.get("formulation_id"))
        gt_formulation_label = normalize_text(gt_row.get("formulation_label_raw"))
        trusted_alignment_row = trusted_alignment_index.get(gt_formulation_id, {})
        alignment_row = alignment_index.get(gt_formulation_id, {})
        scaffold_pred_row_id = normalize_text(alignment_row.get("pred_row_id"))
        trusted_status = normalize_text(trusted_alignment_row.get("l2_gt_alignment_status")).lower()
        direct_canonical_audit_row, direct_resolution_rule = resolve_direct_canonical_alignment(
            paper_key=paper_key,
            gt_formulation_id=gt_formulation_id,
            gt_row=gt_row,
            audit_identity_indexes=audit_identity_indexes,
        )
        semantic_audit_row, semantic_resolution_rule, semantic_audit_details = resolve_semantic_alignment_candidate(
            paper_key=paper_key,
            gt_formulation_id=gt_formulation_id,
            gt_row=gt_row,
            scaffold_pred_row_id=scaffold_pred_row_id,
            trusted_alignment_row=trusted_alignment_row,
            direct_canonical_audit_row=direct_canonical_audit_row,
            direct_resolution_rule=direct_resolution_rule,
            audit_identity_indexes=audit_identity_indexes,
        )
        pred_row_id = (
            normalize_text(semantic_audit_row.get("formulation_id"))
            or normalize_text(direct_canonical_audit_row.get("formulation_id"))
            or scaffold_pred_row_id
        )
        audit_row = semantic_audit_row or direct_canonical_audit_row or (
            audit_by_key.get((paper_key, pred_row_id), {}) if pred_row_id else {}
        )
        canonical_final_row = final_row_index.get((paper_key, pred_row_id), {}) if pred_row_id else {}
        candidate_seed_source_id = normalize_text(audit_row.get("representative_source_formulation_id"))
        article_label = (
            normalize_text(semantic_audit_row.get("article_formulation_label"))
            or normalize_text(direct_canonical_audit_row.get("article_formulation_label"))
            or normalize_text(trusted_alignment_row.get("article_formulation_label"))
            or normalize_text(audit_row.get("article_formulation_label"))
            or gt_formulation_label
        )
        article_formulation_id = (
            normalize_text(semantic_audit_row.get("article_formulation_id"))
            or normalize_text(direct_canonical_audit_row.get("article_formulation_id"))
            or normalize_text(trusted_alignment_row.get("article_formulation_id"))
            or normalize_text(audit_row.get("article_formulation_id"))
            or normalize_text(trusted_alignment_row.get("seed_pred_representative_source_formulation_id"))
            or candidate_seed_source_id
        )
        seed_pred_representative_source_formulation_id = (
            normalize_text(semantic_audit_row.get("representative_source_formulation_id"))
            or normalize_text(direct_canonical_audit_row.get("representative_source_formulation_id"))
            or normalize_text(trusted_alignment_row.get("seed_pred_representative_source_formulation_id"))
            or candidate_seed_source_id
        )
        bridge_identity_valid = has_valid_bridge_identity(
            paper_key=paper_key,
            gt_formulation_id=gt_formulation_id,
            gt_formulation_label=gt_formulation_label,
            trusted_alignment_row=trusted_alignment_row,
            article_formulation_id=article_formulation_id,
            article_formulation_label=article_label,
            seed_pred_representative_source_formulation_id=seed_pred_representative_source_formulation_id,
        )
        canonical_system_present, canonical_article_id, canonical_article_label, canonical_seed_anchor = (
            resolve_canonical_presence(
                paper_key=paper_key,
                gt_formulation_id=gt_formulation_id,
                gt_formulation_label=gt_formulation_label,
                pred_row_id=pred_row_id,
                audit_row=audit_row,
                final_row=canonical_final_row,
            )
        )
        if canonical_system_present:
            # Contract: current canonical Stage5 artifacts own system-presence truth
            # for Layer 3 workbook population. Historical alignment scaffolds,
            # trusted prior workbook rows, and paper-specific bridge patches may
            # help resolve the GT -> system mapping, but they must never
            # downgrade a row back to missing_in_system after the latest
            # final-table/audit-ready artifacts confirm a valid current row.
            article_formulation_id = canonical_article_id or article_formulation_id
            article_label = canonical_article_label or article_label
            seed_pred_representative_source_formulation_id = (
                canonical_seed_anchor or seed_pred_representative_source_formulation_id
            )
        inherited_trusted_alignment = (
            not direct_canonical_audit_row
            and trusted_status == "aligned"
            and bool(pred_row_id)
            and bridge_identity_valid
        )
        if canonical_system_present:
            l2_status = "aligned"
        elif inherited_trusted_alignment:
            l2_status = "aligned"
        else:
            l2_status = classify_gt_alignment_status(
                alignment_row,
                pred_row_id=pred_row_id,
                representative_source_formulation_id=candidate_seed_source_id,
            )
        if paper_key == "BB3JUVW7" and not canonical_system_present and not bridge_identity_valid:
            override_article_id, override_article_label = bb3_identity_override(gt_formulation_id)
            if override_article_id:
                article_formulation_id = override_article_id
                article_label = override_article_label
                seed_pred_representative_source_formulation_id = override_article_id
        resolution_rule = semantic_resolution_rule or (
            "trusted_alignment_bridge"
            if inherited_trusted_alignment
            else "alignment_scaffold_fallback"
        )
        alignment_resolution_rows.append(
            {
                "paper_key": paper_key,
                "gt_formulation_id": gt_formulation_id,
                "gt_formulation_label": gt_formulation_label,
                "scaffold_pred_row_id": scaffold_pred_row_id,
                "resolved_pred_row_id": pred_row_id,
                "resolved_article_formulation_id": article_formulation_id,
                "resolved_article_formulation_label": article_label,
                "scaffold_alignment_decision": normalize_text(alignment_row.get("alignment_decision")),
                "scaffold_alignment_confidence": normalize_text(alignment_row.get("alignment_confidence")),
                "resolution_rule": resolution_rule,
                "override_applied": (
                    "yes"
                    if (
                        (semantic_audit_row or direct_canonical_audit_row)
                        and scaffold_pred_row_id != pred_row_id
                    )
                    else "no"
                ),
                **semantic_audit_details,
            }
        )
        aligned = l2_status == "aligned"
        field_rows = seed_index.get((paper_key, pred_row_id), {}) if aligned else {}
        final_row = canonical_final_row if aligned else {}
        final_field_values = build_final_table_field_values(final_row)
        default_locator = (
            detect_source_locator(audit_row)
            if audit_row
            else normalize_text(gt_row.get("source_locator"))
        )
        identity_seed_row = next(iter(field_rows.values()), {})

        row = blank_main_row()
        row.update(
            {
                "paper_key": paper_key,
                "formulation_id": gt_formulation_id,
                "formulation_label_stage5": normalize_text(trusted_alignment_row.get("formulation_label_stage5"))
                or normalize_text(identity_seed_row.get("formulation_label_stage5")),
                "article_formulation_id": article_formulation_id,
                "article_formulation_label": article_label,
                "l2_gt_formulation_id": gt_formulation_id,
                "l2_gt_formulation_label": gt_formulation_label,
                "seed_pred_representative_source_formulation_id": seed_pred_representative_source_formulation_id,
                "matched_system_formulation_id": (
                    pred_row_id
                    if (semantic_audit_row or direct_canonical_audit_row)
                    else (
                        normalize_text(trusted_alignment_row.get("matched_system_formulation_id"))
                        or (pred_row_id if aligned else "")
                    )
                ),
                "l2_gt_alignment_status": l2_status,
                "paper_risk_level": normalize_text(trusted_alignment_row.get("paper_risk_level"))
                or normalize_text(identity_seed_row.get("paper_risk_level")),
                "layer3_inclusion_flag": normalize_text(trusted_alignment_row.get("layer3_inclusion_flag"))
                or normalize_text(identity_seed_row.get("layer3_inclusion_flag")),
                "sys_drug_name": final_field_values.get("drug_name", ""),
                "sys_drug_mass": final_field_values.get("drug_mass", ""),
                "sys_surfactant_name": final_field_values.get("surfactant_name", ""),
                "sys_particle_size": final_field_values.get("particle_size", ""),
            }
        )

        for field_name in TARGET_FIELD_ORDER:
            field_row = field_rows.get(field_name)
            if field_row is None:
                continue
            row[field_to_main_column(field_name)] = format_system_value(
                extracted_value=field_row.get("extracted_value", ""),
                extracted_unit=field_row.get("extracted_unit", ""),
            )
            row[field_to_locator_column(field_name)] = default_locator
            append_reference_row(
                reference_rows,
                paper_key=paper_key,
                formulation_id=gt_formulation_id,
                formulation_label_stage5=row["formulation_label_stage5"],
                article_formulation_id=row["article_formulation_id"],
                article_formulation_label=row["article_formulation_label"],
                l2_gt_formulation_id=row["l2_gt_formulation_id"],
                l2_gt_formulation_label=row["l2_gt_formulation_label"],
                seed_pred_representative_source_formulation_id=row["seed_pred_representative_source_formulation_id"],
                matched_system_formulation_id=row["matched_system_formulation_id"],
                field_name=field_name,
                extracted_value=normalize_text(field_row.get("extracted_value")),
                extracted_unit=normalize_text(field_row.get("extracted_unit")),
                paper_risk_level=row["paper_risk_level"],
                layer3_inclusion_flag=row["layer3_inclusion_flag"],
                l2_gt_alignment_status=l2_status,
            )
            reference_rows[-1]["evidence_text"] = normalize_text(field_row.get("evidence_text"))
            reference_rows[-1]["evidence_anchor_text"] = normalize_text(field_row.get("evidence_anchor_text"))
            reference_rows[-1]["evidence_source_type"] = normalize_text(field_row.get("evidence_source_type"))
            reference_rows[-1]["evidence_status_detail"] = normalize_text(field_row.get("evidence_status_detail"))
            reference_rows[-1]["relation_resolution_rule"] = normalize_text(field_row.get("relation_resolution_rule"))
            reference_rows[-1]["relation_resolution_confidence"] = normalize_text(field_row.get("relation_resolution_confidence"))
            reference_rows[-1]["review_warning"] = normalize_text(field_row.get("review_warning"))
            reference_rows[-1]["normalization_status"] = normalize_text(field_row.get("normalization_status"))

        if not aligned and default_locator:
            for field_name in TARGET_FIELD_ORDER:
                row[field_to_locator_column(field_name)] = default_locator

        main_rows.append(row)

    for alignment_row in extra_alignment_rows:
        paper_key = normalize_text(alignment_row.get("doi"))
        pred_row_id = normalize_text(alignment_row.get("pred_row_id"))
        if not pred_row_id:
            continue
        final_row = next((row for (pk, fid), row in final_row_index.items() if pk and row and fid == pred_row_id), {})
        resolved_paper_key = normalize_text(final_row.get("key") or paper_key)
        audit_row = audit_by_key.get((resolved_paper_key, pred_row_id), {})
        field_rows = seed_index.get((resolved_paper_key, pred_row_id), {})
        final_field_values = build_final_table_field_values(final_row)
        identity_seed_row = next(iter(field_rows.values()), {})
        default_locator = detect_source_locator(audit_row) if audit_row else normalize_text(alignment_row.get("pred_source_text"))
        row = blank_main_row()
        row.update(
            {
                "paper_key": resolved_paper_key,
                "formulation_id": pred_row_id,
                "formulation_label_stage5": normalize_text(identity_seed_row.get("formulation_label_stage5")),
                "article_formulation_id": normalize_text(audit_row.get("article_formulation_id")),
                "article_formulation_label": normalize_text(audit_row.get("article_formulation_label")),
                "l2_gt_formulation_id": "",
                "l2_gt_formulation_label": "",
                "seed_pred_representative_source_formulation_id": normalize_text(
                    audit_row.get("representative_source_formulation_id")
                ),
                "matched_system_formulation_id": pred_row_id,
                "l2_gt_alignment_status": "extra_in_system",
                "paper_risk_level": normalize_text(identity_seed_row.get("paper_risk_level")),
                "layer3_inclusion_flag": normalize_text(identity_seed_row.get("layer3_inclusion_flag")),
                "sys_drug_name": final_field_values.get("drug_name", ""),
                "sys_drug_mass": final_field_values.get("drug_mass", ""),
                "sys_surfactant_name": final_field_values.get("surfactant_name", ""),
                "sys_particle_size": final_field_values.get("particle_size", ""),
            }
        )
        for field_name in TARGET_FIELD_ORDER:
            field_row = field_rows.get(field_name)
            if field_row is None:
                continue
            row[field_to_main_column(field_name)] = format_system_value(
                extracted_value=field_row.get("extracted_value", ""),
                extracted_unit=field_row.get("extracted_unit", ""),
            )
            row[field_to_locator_column(field_name)] = default_locator

        extra_rows.append(row)

    main_rows.sort(key=lambda row: (row["paper_key"], row["formulation_id"]))
    reference_rows.sort(key=lambda row: (row["paper_key"], row["formulation_id"], row["field_name"]))
    extra_rows.sort(key=lambda row: (row["paper_key"], row["formulation_id"]))
    alignment_resolution_rows.sort(key=lambda row: (row["paper_key"], row["gt_formulation_id"]))
    return main_rows, reference_rows, extra_rows, 0, alignment_resolution_rows


def style_sheet(worksheet, freeze_panes: str = "E2") -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = freeze_panes
    worksheet.auto_filter.ref = worksheet.dimensions


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(normalize_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 42)


def write_sheet(worksheet, columns: list[str], rows: list[dict[str, Any]], freeze_panes: str = "E2") -> None:
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column, "") for column in columns])
    style_sheet(worksheet, freeze_panes=freeze_panes)
    autosize_columns(worksheet)


def validate_main_rows(
    main_rows: list[dict[str, Any]],
    alignment_resolution_rows: list[dict[str, Any]],
) -> None:
    for row in main_rows:
        if normalize_text(row.get("l2_gt_alignment_status")) != "aligned":
            continue
        if not has_valid_bridge_identity(
            paper_key=normalize_text(row.get("paper_key")),
            gt_formulation_id=normalize_text(row.get("l2_gt_formulation_id") or row.get("formulation_id")),
            gt_formulation_label=normalize_text(row.get("l2_gt_formulation_label")),
            trusted_alignment_row={},
            article_formulation_id=normalize_text(row.get("article_formulation_id")),
            article_formulation_label=normalize_text(row.get("article_formulation_label")),
            seed_pred_representative_source_formulation_id=normalize_text(
                row.get("seed_pred_representative_source_formulation_id")
            ),
        ):
            raise ValueError(
                f"Aligned row lacks valid explicit identity anchor: {normalize_text(row.get('paper_key'))} "
                f"{normalize_text(row.get('formulation_id'))}"
            )

    for row in alignment_resolution_rows:
        if normalize_text(row.get("gt_semantic_role")) != "core":
            continue
        if normalize_text(row.get("selected_candidate_semantic_role")) != "sweep":
            continue
        core_candidate_ids = normalize_text(row.get("core_candidate_ids"))
        if core_candidate_ids in {"", "[]"}:
            continue
        raise ValueError(
            "Semantic alignment contract violation: core GT row resolved to sweep candidate "
            f"despite available core candidates: {normalize_text(row.get('paper_key'))} "
            f"{normalize_text(row.get('gt_formulation_id'))} core_candidates={core_candidate_ids} "
            f"resolved_pred_row_id={normalize_text(row.get('resolved_pred_row_id'))}"
        )

    bb3_expected = {
        "F1.1",
        "F1.2",
        "F1.3",
        "F1.4",
        "F1.5",
        "F2.1",
        "F2.2",
        "F2.3",
        "F2.4",
        "F2.5",
        "F2.6",
        "F2.7",
    }
    bb3_found = {
        normalize_text(row.get("article_formulation_id"))
        for row in main_rows
        if normalize_text(row.get("paper_key")) == "BB3JUVW7" and normalize_text(row.get("article_formulation_id"))
    }
    missing_bb3 = sorted(bb3_expected - bb3_found)
    if missing_bb3:
        raise ValueError(f"BB3JUVW7 missing explicit article identities: {', '.join(missing_bb3)}")


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")


def write_markdown_report(path: Path, summary: dict[str, Any]) -> None:
    lines = [
        "# Annotation Merge Report",
        "",
        f"- Matching strategy: {summary['matching_strategy']}",
        f"- Human annotation columns preserved: {', '.join(summary['human_annotation_columns'])}",
        f"- Old workbook rows: {summary['old_workbook_rows']}",
        f"- New workbook rows: {summary['new_workbook_rows']}",
        f"- Total old annotated cells: {summary['total_old_annotated_cells']}",
        f"- Preserved annotation cells: {summary['preserved_annotation_cells']}",
        f"- Exact row matches: {summary['total_matched_rows_exact']}",
        f"- Fallback row matches: {summary['total_fallback_matches']}",
        f"- Conflicts: {summary['total_conflicts']}",
        f"- New rows without annotation: {summary['total_new_rows_without_annotation']}",
        f"- Unmatched old annotation rows: {summary['total_unmatched_old_annotation_rows']}",
        f"- Unmatched old annotated cells: {summary['total_unmatched_old_annotations']}",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def merge_annotations_into_rows(
    *,
    new_rows: list[dict[str, Any]],
    prior_rows: list[dict[str, str]],
    workbook_label: str,
) -> tuple[list[dict[str, Any]], list[dict[str, str]], dict[str, Any]]:
    prior_indexes = build_prior_row_indexes(prior_rows)
    used_prior_ids: set[str] = set()
    total_old_annotated_cells = sum(count_nonempty_annotations(row) for row in prior_rows)
    total_old_annotated_rows = sum(1 for row in prior_rows if count_nonempty_annotations(row) > 0)
    audit_rows: list[dict[str, str]] = []
    matched_examples: list[dict[str, str]] = []
    unmatched_examples: list[dict[str, str]] = []

    summary: dict[str, Any] = {
        "matching_strategy": "exact: paper_key+formulation_id; fallback: paper_key+article_formulation_id; fallback: paper_key+article_formulation_label_normalized",
        "human_annotation_columns": HUMAN_ANNOTATION_COLUMNS,
        "old_workbook_rows": len(prior_rows),
        "new_workbook_rows": len(new_rows),
        "total_old_annotated_cells": total_old_annotated_cells,
        "total_old_annotated_rows": total_old_annotated_rows,
        "preserved_annotation_cells": 0,
        "total_matched_rows_exact": 0,
        "total_fallback_matches": 0,
        "total_conflicts": 0,
        "total_new_rows_without_annotation": 0,
        "total_unmatched_old_annotation_rows": 0,
        "total_unmatched_old_annotations": 0,
    }

    for row in new_rows:
        initialize_annotation_provenance(row)
        prior_row, match_status, match_key_name, conflict_flag, matched_rows = match_prior_row(
            new_row=row,
            prior_indexes=prior_indexes,
        )
        copied_cells = 0
        source_row_id = ""
        if conflict_flag:
            row["annotation_match_status"] = "conflict"
            row["annotation_conflict_flag"] = "true"
            row["annotation_match_key"] = match_key_name
            row["annotation_source_workbook"] = workbook_label
            summary["total_conflicts"] += 1
        elif prior_row is not None:
            copied_cells = merge_preserved_cells(row, prior_row, preserve_columns())
            prior_row_id = exact_match_key(prior_row)
            source_row_id = "::".join(prior_row_id)
            used_prior_ids.add(source_row_id)
            row["annotation_source_workbook"] = workbook_label
            row["annotation_match_status"] = match_status
            row["annotation_match_key"] = match_key_name
            row["annotation_conflict_flag"] = "false"
            if match_status == "exact":
                summary["total_matched_rows_exact"] += 1
            elif match_status == "fallback":
                summary["total_fallback_matches"] += 1
            summary["preserved_annotation_cells"] += copied_cells
            if copied_cells > 0 and len(matched_examples) < 5:
                matched_examples.append(
                    {
                        "paper_key": normalize_text(row.get("paper_key")),
                        "formulation_id": normalize_text(row.get("formulation_id")),
                        "article_formulation_id": normalize_text(row.get("article_formulation_id")),
                        "match_status": match_status,
                        "match_key": match_key_name,
                        "copied_cells": str(copied_cells),
                    }
                )
        if copied_cells == 0:
            summary["total_new_rows_without_annotation"] += 1
            if len(unmatched_examples) < 5:
                unmatched_examples.append(
                    {
                        "paper_key": normalize_text(row.get("paper_key")),
                        "formulation_id": normalize_text(row.get("formulation_id")),
                        "article_formulation_id": normalize_text(row.get("article_formulation_id")),
                        "match_status": row.get("annotation_match_status", "unmatched"),
                        "match_key": row.get("annotation_match_key", ""),
                    }
                )
        audit_rows.append(
            {
                "paper_key": normalize_text(row.get("paper_key")),
                "formulation_id": normalize_text(row.get("formulation_id")),
                "article_formulation_id": normalize_text(row.get("article_formulation_id")),
                "article_formulation_label": normalize_text(row.get("article_formulation_label")),
                "annotation_source_workbook": normalize_text(row.get("annotation_source_workbook")),
                "annotation_match_status": normalize_text(row.get("annotation_match_status")),
                "annotation_match_key": normalize_text(row.get("annotation_match_key")),
                "annotation_conflict_flag": normalize_text(row.get("annotation_conflict_flag")),
                "matched_old_row_count": str(len(matched_rows)),
                "matched_old_row_id": source_row_id,
                "copied_annotation_cells": str(copied_cells),
            }
        )

    unmatched_old_rows = 0
    unmatched_old_cells = 0
    for prior_row in prior_rows:
        if count_nonempty_annotations(prior_row) == 0:
            continue
        prior_row_id = "::".join(exact_match_key(prior_row))
        if prior_row_id in used_prior_ids:
            continue
        unmatched_old_rows += 1
        unmatched_old_cells += count_nonempty_annotations(prior_row)
    summary["total_unmatched_old_annotation_rows"] = unmatched_old_rows
    summary["total_unmatched_old_annotations"] = unmatched_old_cells
    summary["matched_examples"] = matched_examples
    summary["unmatched_examples"] = unmatched_examples
    return new_rows, audit_rows, summary


def add_flag_validation(worksheet, header_name: str) -> None:
    headers = [cell.value for cell in worksheet[1]]
    if header_name not in headers:
        return
    column_index = headers.index(header_name) + 1
    letter = get_column_letter(column_index)
    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(FLAG_OPTIONS) + '"',
        allow_blank=True,
    )
    validation.prompt = "Select a structured GT flag."
    validation.error = "Use one of: correct, incorrect, missing, unclear."
    worksheet.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}{max(worksheet.max_row, 2)}")


def build_workbook(
    path: Path,
    main_rows: list[dict[str, Any]],
    reference_rows: list[dict[str, Any]],
    extra_rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    main_sheet = workbook.active
    main_sheet.title = "value_gt_annotation"
    write_sheet(main_sheet, MAIN_COLUMNS, main_rows, freeze_panes="H2")
    for header_name in [
        "gt_flag_drug_name",
        "gt_flag_drug_mass",
        "gt_flag_polymer_MW",
        "gt_flag_LA_GA",
        "gt_flag_surfactant_name",
        "gt_flag_surfactant_concentration",
        "gt_flag_drug_polymer_ratio",
        "gt_flag_particle_size",
        "gt_flag_EE",
        "gt_flag_LC",
    ]:
        add_flag_validation(main_sheet, header_name)

    reference_sheet = workbook.create_sheet("field_reference")
    write_sheet(reference_sheet, REFERENCE_COLUMNS, reference_rows, freeze_panes="F2")

    if extra_rows:
        extra_sheet = workbook.create_sheet("extra_in_system")
        write_sheet(extra_sheet, MAIN_COLUMNS, extra_rows, freeze_panes="H2")
        for header_name in [
            "gt_flag_drug_name",
            "gt_flag_drug_mass",
            "gt_flag_polymer_MW",
            "gt_flag_LA_GA",
            "gt_flag_surfactant_name",
            "gt_flag_surfactant_concentration",
            "gt_flag_drug_polymer_ratio",
            "gt_flag_particle_size",
            "gt_flag_EE",
            "gt_flag_LC",
        ]:
            add_flag_validation(extra_sheet, header_name)

    instructions_sheet = workbook.create_sheet("instructions")
    instructions_rows = [
        {"instruction": "Purpose", "details": "This workbook is for fast value-only GT construction using the Layer 2 GT skeleton as the authoritative main-sheet row set."},
        {"instruction": "Scope", "details": "Evidence adjudication remains a separate later phase; this sheet is optimized for rapid numeric/value entry."},
        {"instruction": "Identity", "details": "Main-sheet formulation_id comes from Layer 2 GT. Preserve l2_gt_formulation_* and seed_pred_representative_source_formulation_id as the traceable skeleton identity. Article-native labels are additive only."},
        {"instruction": "Drug mass", "details": "Drug mass is included to support later drug/polymer ratio checking or derivation."},
        {"instruction": "Alignment", "details": "The latest Stage5 final table plus audit-ready export are canonical for current-system presence. Historical alignment scaffolds and prior workbook bridges may help map GT rows, but they must not downgrade a canonically present row to missing_in_system."},
        {"instruction": "Flags", "details": "Use gt_flag_* dropdowns with: correct, incorrect, missing, unclear."},
    ]
    write_sheet(instructions_sheet, ["instruction", "details"], instructions_rows, freeze_panes="A2")

    dropdown_sheet = workbook.create_sheet("dropdown_options")
    dropdown_sheet.append(["flag_options"])
    for option in FLAG_OPTIONS:
        dropdown_sheet.append([option])
    style_sheet(dropdown_sheet, freeze_panes="A2")
    autosize_columns(dropdown_sheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build a formulation-level value GT annotation workbook.")
    parser.add_argument("--run-id", default="", help="Optional exact run_id compatibility alias for this annotation export.")
    parser.add_argument("--run-dir", type=Path, default=None, help="Explicit run directory. Defaults to data/results/<run_id>.")
    parser.add_argument(
        "--out-subdir",
        default="value_gt_v1",
        type=sanitize_out_subdir,
        help="Functional artifact subdirectory under the run root.",
    )
    parser.add_argument("--audit-ready-tsv", type=Path, default=None, help="Optional explicit final_formulation_table_audit_ready_v1.tsv")
    parser.add_argument("--seed-rows-tsv", type=Path, default=None, help="Optional explicit field_gt_review_seed_rows_v*.tsv")
    parser.add_argument("--final-table-tsv", type=Path, default=None, help="Optional explicit frozen final_formulation_table_v1.tsv")
    parser.add_argument("--gt-skeleton-tsv", type=Path, default=None, help="Optional Layer 2 GT skeleton TSV. When provided, main-sheet rows come from GT.")
    parser.add_argument("--alignment-scaffold-tsv", type=Path, default=None, help="Optional GT-to-system alignment scaffold TSV for GT-skeleton mode.")
    parser.add_argument("--prior-workbook-xlsx", type=Path, default=None, help="Optional prior annotation workbook to merge forward.")
    parser.add_argument("--trusted-alignment-tsv", type=Path, default=None, help="Optional prior GT-skeleton workbook/TSV rows used only as a trusted alignment bridge.")
    parser.add_argument("--workbook-name", default=None, help="Optional explicit workbook filename.")
    parser.add_argument("--artifact-version", type=int, default=1, help="Artifact version suffix.")
    args = parser.parse_args()

    run_context = resolve_run_context(
        explicit_run_dir=args.run_dir,
        explicit_run_id=normalize_text(args.run_id),
    )
    run_dir = Path(run_context["run_dir"])
    run_id = str(run_context["run_id"])
    out_dir = run_dir / args.out_subdir

    audit_ready_tsv = resolve_artifact_path(
        explicit_path=args.audit_ready_tsv,
        run_context=run_context,
        pointer_key="stage5_audit_ready_tsv",
        required=True,
    )
    seed_rows_tsv = resolve_artifact_path(
        explicit_path=args.seed_rows_tsv,
        run_context=run_context,
        pointer_key="field_gt_seed_rows_tsv",
        required=True,
    )
    final_table_tsv = resolve_artifact_path(
        explicit_path=args.final_table_tsv,
        run_context=run_context,
        pointer_key="stage5_final_table_tsv",
        canonical_relative="final_formulation_table_v1.tsv",
        required=True,
    )
    gt_skeleton_tsv = resolve_artifact_path(
        explicit_path=args.gt_skeleton_tsv,
        run_context=run_context,
        pointer_key="gt_skeleton_tsv",
        required=False,
    )
    alignment_scaffold_tsv = resolve_artifact_path(
        explicit_path=args.alignment_scaffold_tsv,
        run_context=run_context,
        pointer_key="alignment_scaffold_tsv",
        required=False,
    )
    trusted_alignment_tsv = resolve_artifact_path(
        explicit_path=args.trusted_alignment_tsv,
        run_context=run_context,
        pointer_key="trusted_alignment_tsv",
        required=False,
    )

    print(
        json.dumps(
            {
                "resolved_source_run_dir": str(run_dir),
                "resolved_source_run_id": run_id,
                "source_resolution": str(run_context["resolution_source"]),
                "active_run_pointer_path": str(run_context.get("pointer_path") or ""),
                "resolved_input_files": {
                    "audit_ready_tsv": str(audit_ready_tsv),
                    "seed_rows_tsv": str(seed_rows_tsv),
                    "final_table_tsv": str(final_table_tsv),
                    "gt_skeleton_tsv": str(gt_skeleton_tsv) if gt_skeleton_tsv else "",
                    "alignment_scaffold_tsv": str(alignment_scaffold_tsv) if alignment_scaffold_tsv else "",
                    "prior_workbook_xlsx": str(args.prior_workbook_xlsx.resolve()) if args.prior_workbook_xlsx else "",
                    "trusted_alignment_tsv": str(trusted_alignment_tsv) if trusted_alignment_tsv else "",
                },
            },
            indent=2,
        )
    )

    audit_rows = read_tsv_rows(audit_ready_tsv)
    seed_rows = read_tsv_rows(seed_rows_tsv)
    final_rows = read_tsv_rows(final_table_tsv)
    seed_index = build_seed_index(seed_rows)
    final_row_index = build_final_row_index(final_rows)
    extra_rows: list[dict[str, Any]] = []
    preserved_nonempty_cells = 0
    merge_audit_rows: list[dict[str, str]] = []
    merge_summary: dict[str, Any] | None = None
    alignment_resolution_rows: list[dict[str, Any]] = []

    if gt_skeleton_tsv is not None:
        if alignment_scaffold_tsv is None:
            raise ValueError("--alignment-scaffold-tsv is required when --gt-skeleton-tsv is provided.")
        gt_rows = read_tsv_rows(gt_skeleton_tsv)
        alignment_rows = read_tsv_rows(alignment_scaffold_tsv)
        alignment_index, extra_alignment_rows = build_alignment_index(alignment_rows)
        prior_rows = (
            load_workbook_rows(args.prior_workbook_xlsx.resolve(), "value_gt_annotation")
            if args.prior_workbook_xlsx is not None
            else []
        )
        trusted_alignment_rows = (
            read_tsv_rows(trusted_alignment_tsv)
            if trusted_alignment_tsv is not None
            else []
        )
        main_rows, reference_rows, extra_rows, preserved_nonempty_cells, alignment_resolution_rows = build_gt_skeleton_rows(
            gt_rows=gt_rows,
            alignment_index=alignment_index,
            extra_alignment_rows=extra_alignment_rows,
            audit_rows=audit_rows,
            seed_index=seed_index,
            final_row_index=final_row_index,
            prior_rows=prior_rows,
            trusted_alignment_rows=trusted_alignment_rows,
        )
    else:
        main_rows, reference_rows = build_main_rows(audit_rows, seed_index, final_row_index)

    if args.prior_workbook_xlsx is not None:
        prior_main_rows = load_prior_sheet_rows(args.prior_workbook_xlsx.resolve(), "value_gt_annotation")
        prior_extra_rows = load_prior_sheet_rows(args.prior_workbook_xlsx.resolve(), "extra_in_system")
        prior_rows = prior_main_rows + prior_extra_rows
        merged_main_rows, merge_audit_rows, merge_summary = merge_annotations_into_rows(
            new_rows=main_rows,
            prior_rows=prior_rows,
            workbook_label=str(args.prior_workbook_xlsx.resolve()),
        )
        main_rows = merged_main_rows
        if extra_rows:
            extra_rows, extra_audit_rows, extra_summary = merge_annotations_into_rows(
                new_rows=extra_rows,
                prior_rows=prior_rows,
                workbook_label=str(args.prior_workbook_xlsx.resolve()),
            )
            merge_audit_rows.extend(extra_audit_rows)
            merge_summary["new_workbook_rows"] += extra_summary["new_workbook_rows"]
            merge_summary["total_matched_rows_exact"] += extra_summary["total_matched_rows_exact"]
            merge_summary["total_fallback_matches"] += extra_summary["total_fallback_matches"]
            merge_summary["total_conflicts"] += extra_summary["total_conflicts"]
            merge_summary["total_new_rows_without_annotation"] += extra_summary["total_new_rows_without_annotation"]
            merge_summary["preserved_annotation_cells"] += extra_summary["preserved_annotation_cells"]
            merge_summary["matched_examples"] = (merge_summary["matched_examples"] + extra_summary["matched_examples"])[:5]
            merge_summary["unmatched_examples"] = (merge_summary["unmatched_examples"] + extra_summary["unmatched_examples"])[:5]
        preserved_nonempty_cells = int(merge_summary["preserved_annotation_cells"])

    validate_main_rows(main_rows, alignment_resolution_rows)

    main_tsv = out_dir / main_tsv_name_for_version(args.artifact_version)
    reference_tsv = out_dir / reference_tsv_name_for_version(args.artifact_version)
    extra_tsv = out_dir / extra_tsv_name_for_version(args.artifact_version)
    alignment_resolution_tsv = out_dir / alignment_resolution_tsv_name_for_version(args.artifact_version)
    workbook_path = out_dir / resolve_workbook_name(args.artifact_version, args.workbook_name)
    merge_audit_csv = out_dir / merge_audit_csv_name_for_version(args.artifact_version)
    merge_summary_json = out_dir / merge_summary_json_name_for_version(args.artifact_version)
    merge_report_md = out_dir / merge_report_md_name_for_version(args.artifact_version)

    write_tsv(main_tsv, MAIN_COLUMNS, main_rows)
    write_tsv(reference_tsv, REFERENCE_COLUMNS, reference_rows)
    if extra_rows:
        write_tsv(extra_tsv, MAIN_COLUMNS, extra_rows)
    if alignment_resolution_rows:
        write_tsv(
            alignment_resolution_tsv,
            list(alignment_resolution_rows[0].keys()),
            alignment_resolution_rows,
        )
    build_workbook(workbook_path, main_rows, reference_rows, extra_rows)
    metadata_path = write_artifact_metadata_json(
        workbook_path,
        build_artifact_metadata(
            source_run_context=run_context,
            source_files={
                "audit_ready_tsv": str(audit_ready_tsv),
                "seed_rows_tsv": str(seed_rows_tsv),
                "final_table_tsv": str(final_table_tsv),
                "gt_skeleton_tsv": str(gt_skeleton_tsv) if gt_skeleton_tsv else "",
                "alignment_scaffold_tsv": str(alignment_scaffold_tsv) if alignment_scaffold_tsv else "",
                "prior_workbook_xlsx": str(args.prior_workbook_xlsx.resolve()) if args.prior_workbook_xlsx else "",
                "trusted_alignment_tsv": str(trusted_alignment_tsv) if trusted_alignment_tsv else "",
            },
            generated_by="src/stage5_benchmark/build_value_gt_annotation_workbook_v1.py",
            note="Layer3 value GT annotation workbook authority metadata.",
            extra={
                "main_rows_tsv": str(main_tsv),
                "reference_rows_tsv": str(reference_tsv),
                "extra_rows_tsv": str(extra_tsv) if extra_rows else "",
                "alignment_resolution_tsv": str(alignment_resolution_tsv) if alignment_resolution_rows else "",
            },
        ),
    )
    if merge_summary is not None:
        audit_columns = list(merge_audit_rows[0].keys()) if merge_audit_rows else [
            "paper_key",
            "formulation_id",
            "article_formulation_id",
            "article_formulation_label",
            "annotation_source_workbook",
            "annotation_match_status",
            "annotation_match_key",
            "annotation_conflict_flag",
            "matched_old_row_count",
            "matched_old_row_id",
            "copied_annotation_cells",
        ]
        write_csv(merge_audit_csv, audit_columns, merge_audit_rows)
        write_json(merge_summary_json, merge_summary)
        write_markdown_report(merge_report_md, merge_summary)

        print(
            f"Annotation merge summary: old_rows={merge_summary['old_workbook_rows']} new_rows={merge_summary['new_workbook_rows']} matched_exact={merge_summary['total_matched_rows_exact']} fallback={merge_summary['total_fallback_matches']} conflicts={merge_summary['total_conflicts']} preserved_cells={merge_summary['preserved_annotation_cells']}"
        )
        print("Example matched rows:")
        for example in merge_summary.get("matched_examples", [])[:5]:
            print("  " + json.dumps(example, ensure_ascii=True))
        print("Example unmatched rows:")
        for example in merge_summary.get("unmatched_examples", [])[:5]:
            print("  " + json.dumps(example, ensure_ascii=True))

    print(
        {
            "run_id": run_id,
            "run_dir": str(run_dir),
            "out_dir": str(out_dir),
            "final_table_tsv": str(final_table_tsv),
            "main_tsv": str(main_tsv),
            "reference_tsv": str(reference_tsv),
            "extra_tsv": str(extra_tsv) if extra_rows else "",
            "alignment_resolution_tsv": str(alignment_resolution_tsv) if alignment_resolution_rows else "",
            "workbook_path": str(workbook_path),
            "workbook_metadata_json": str(metadata_path),
            "formulation_rows": len(main_rows),
            "reference_rows": len(reference_rows),
            "extra_in_system_rows": len(extra_rows),
            "preserved_nonempty_cells": preserved_nonempty_cells,
            "annotation_merge_audit_csv": str(merge_audit_csv) if merge_summary is not None else "",
            "annotation_merge_summary_json": str(merge_summary_json) if merge_summary is not None else "",
            "annotation_merge_report_md": str(merge_report_md) if merge_summary is not None else "",
            "trusted_alignment_tsv": str(trusted_alignment_tsv) if trusted_alignment_tsv else "",
            "target_fields": TARGET_FIELD_ORDER,
        }
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
