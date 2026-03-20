#!/usr/bin/env python3
from __future__ import annotations

"""
Build a reviewer-friendly Layer 2 boundary-GT workbook from Stage 5 final rows.

Purpose:
- seed a human-reviewable boundary GT workbook from the benchmark object
- keep prediction-reference fields visually separate from GT-authoritative fields
- preserve enough deterministic structure for later GT validation and alignment

Inputs:
- Stage 5 `final_formulation_table_v1.tsv`
- optional Stage 5 `final_output_decision_trace_v1.tsv`
- optional Stage 3 `formulation_relation_records_v1.tsv`
- optional scope manifest TSV for DOI/title/year enrichment

Outputs:
- `boundary_gt_review_workbook_v1.xlsx`
- companion TSV surfaces for seed rows, manual-addition templates, and references

Stage role:
- supporting Stage 5 / Layer 2 GT review surface
- not part of the canonical benchmark-valid production endpoint
"""

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from src.utils.paths import DATA_RESULTS_DIR
    from src.utils.run_id import is_valid_run_id
except ModuleNotFoundError:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from src.utils.paths import DATA_RESULTS_DIR
    from src.utils.run_id import is_valid_run_id


WORKBOOK_NAME = "boundary_gt_review_workbook_v1.xlsx"
REVIEW_SEED_TSV_NAME = "boundary_gt_review_seed_rows_v1.tsv"
MANUAL_TEMPLATE_TSV_NAME = "boundary_gt_manual_additions_template_v1.tsv"
PRED_REFERENCE_TSV_NAME = "boundary_gt_pred_reference_rows_v1.tsv"
FAMILY_REFERENCE_TSV_NAME = "boundary_gt_family_reference_v1.tsv"
SOURCE_SUMMARY_TSV_NAME = "boundary_gt_source_summary_v1.tsv"

GT_ROW_DECISION_OPTIONS = [
    "include_gt",
    "exclude_seed_spurious",
    "manual_added_gt",
    "unclear",
]
VARIANT_ROLE_OPTIONS = [
    "family_core",
    "true_family_variant",
    "optimized_variant",
    "checkpoint_validation_variant",
    "post_processing_variant",
    "measurement_variant",
    "unclear",
]
PAYLOAD_STATE_OPTIONS = [
    "drug_loaded",
    "blank_control",
    "empty",
    "fitc_assay_loaded",
    "unknown",
]
BENCHMARK_INCLUDE_OPTIONS = ["yes", "no", "unclear"]
BOUNDARY_SOURCE_TYPE_OPTIONS = [
    "table_row",
    "explicit_label",
    "text_described",
    "figure_sweep",
    "derived_family_context",
    "unclear",
]
REVIEW_STATUS_OPTIONS = ["pending", "reviewed", "needs_second_pass"]

REVIEW_COLUMNS = [
    "paper_key",
    "doi",
    "doi_url",
    "paper_title",
    "publication_year",
    "seed_row_origin",
    "seed_pred_final_formulation_id",
    "seed_pred_representative_source_formulation_id",
    "seed_pred_family_id",
    "seed_pred_parent_core_row_id",
    "seed_pred_variant_role",
    "seed_pred_payload_state",
    "seed_pred_benchmark_default_include",
    "seed_pred_final_output_rule",
    "seed_pred_review_needed",
    "seed_pred_field_source_type",
    "seed_pred_row_priority",
    "seed_pred_priority_reason",
    "seed_pred_suspicious_case_flag",
    "seed_pred_boundary_source_type",
    "seed_pred_boundary_source_locator",
    "seed_pred_boundary_anchor_label",
    "seed_pred_boundary_anchor_table_id",
    "seed_pred_boundary_anchor_row_index",
    "seed_pred_polymer_identity_anchor",
    "seed_pred_drug_identity_anchor",
    "seed_pred_surfactant_anchor",
    "seed_pred_solvent_phase_anchor",
    "seed_pred_drug_polymer_ratio_anchor",
    "seed_pred_polymer_mw_anchor",
    "seed_pred_la_ga_ratio_anchor",
    "seed_pred_variant_change_anchor",
    "seed_pred_boundary_signature",
    "gt_row_decision",
    "gt_formulation_id",
    "family_id",
    "parent_core_row_id",
    "variant_role",
    "payload_state",
    "benchmark_include_gt",
    "boundary_source_type",
    "boundary_source_locator",
    "boundary_anchor_label",
    "boundary_anchor_table_id",
    "boundary_anchor_row_index",
    "polymer_identity_anchor",
    "drug_identity_anchor",
    "surfactant_anchor",
    "solvent_phase_anchor",
    "drug_polymer_ratio_anchor",
    "polymer_mw_anchor",
    "la_ga_ratio_anchor",
    "variant_change_anchor",
    "gt_evidence_note",
    "review_status",
    "review_notes",
]

REFERENCE_COLUMNS = {
    "paper_key",
    "doi",
    "doi_url",
    "paper_title",
    "publication_year",
    "seed_row_origin",
    "seed_pred_final_formulation_id",
    "seed_pred_representative_source_formulation_id",
    "seed_pred_family_id",
    "seed_pred_parent_core_row_id",
    "seed_pred_variant_role",
    "seed_pred_payload_state",
    "seed_pred_benchmark_default_include",
    "seed_pred_final_output_rule",
    "seed_pred_review_needed",
    "seed_pred_boundary_source_type",
    "seed_pred_boundary_source_locator",
    "seed_pred_boundary_anchor_label",
    "seed_pred_boundary_anchor_table_id",
    "seed_pred_boundary_anchor_row_index",
    "seed_pred_polymer_identity_anchor",
    "seed_pred_drug_identity_anchor",
    "seed_pred_surfactant_anchor",
    "seed_pred_solvent_phase_anchor",
    "seed_pred_drug_polymer_ratio_anchor",
    "seed_pred_polymer_mw_anchor",
    "seed_pred_la_ga_ratio_anchor",
    "seed_pred_variant_change_anchor",
    "seed_pred_boundary_signature",
}

EDITABLE_COLUMNS = [column for column in REVIEW_COLUMNS if column not in REFERENCE_COLUMNS]


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_token(value: Any) -> str:
    text = normalize_text(value).lower()
    if not text:
        return ""
    text = re.sub(r"[^a-z0-9%:/.+-]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def normalize_doi(value: Any) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"^doi\s*:\s*", "", text)
    text = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", text)
    text = re.sub(r"^doi\.org/", "", text)
    return text


def read_tsv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def sanitize_out_subdir(value: str) -> str:
    text = normalize_text(value).replace("\\", "/")
    if not text:
        raise ValueError("--out-subdir is required.")
    if Path(text).is_absolute():
        raise ValueError("--out-subdir must be a relative path.")
    parts = [part for part in text.split("/") if part]
    if not parts or any(part == ".." for part in parts):
        raise ValueError("--out-subdir cannot contain path traversal.")
    return "/".join(parts)


def resolve_run_dir(run_id: str, explicit_run_dir: Path | None) -> Path:
    if explicit_run_dir is not None:
        return explicit_run_dir.resolve()
    return (DATA_RESULTS_DIR / run_id).resolve()


def default_input_path(run_dir: Path, filename: str) -> Path:
    return run_dir / filename


def discover_scope_manifest(run_dir: Path) -> Path | None:
    preferred = [run_dir / "dev15_scope.tsv", run_dir / "scope.tsv", run_dir / "scope_manifest.tsv"]
    for path in preferred:
        if path.exists():
            return path
    matches = sorted(run_dir.glob("*scope*.tsv"))
    return matches[0] if len(matches) == 1 else None


def load_manifest_map(path: Path | None) -> dict[str, dict[str, str]]:
    if path is None or not path.exists():
        return {}
    out: dict[str, dict[str, str]] = {}
    for row in read_tsv_rows(path):
        key = normalize_text(row.get("key") or row.get("paper_key") or row.get("zotero_key"))
        if key:
            out[key] = row
    return out


def load_decision_trace_map(path: Path | None) -> dict[tuple[str, str], dict[str, str]]:
    if path is None or not path.exists():
        return {}
    out: dict[tuple[str, str], dict[str, str]] = {}
    for row in read_tsv_rows(path):
        key = normalize_text(row.get("zotero_key"))
        source_id = normalize_text(row.get("source_formulation_id"))
        if key and source_id:
            out[(key, source_id)] = row
    return out


def load_relation_counts(path: Path | None) -> dict[tuple[str, str], dict[str, str]]:
    if path is None or not path.exists():
        return {}
    grouped: dict[tuple[str, str], dict[str, Any]] = defaultdict(
        lambda: {
            "relation_graph_ids": set(),
            "relation_method_group_ids": set(),
            "relation_row_count": 0,
        }
    )
    for row in read_tsv_rows(path):
        key = normalize_text(row.get("paper_key"))
        candidate_id = normalize_text(row.get("formulation_candidate_id"))
        if not key or not candidate_id:
            continue
        bucket = grouped[(key, candidate_id)]
        graph_id = normalize_text(row.get("relation_graph_id"))
        method_group_id = normalize_text(row.get("method_group_id"))
        if graph_id:
            bucket["relation_graph_ids"].add(graph_id)
        if method_group_id:
            bucket["relation_method_group_ids"].add(method_group_id)
        bucket["relation_row_count"] += 1
    out: dict[tuple[str, str], dict[str, str]] = {}
    for bucket_key, bucket in grouped.items():
        out[bucket_key] = {
            "relation_graph_ids": json.dumps(sorted(bucket["relation_graph_ids"]), ensure_ascii=True),
            "relation_method_group_ids": json.dumps(sorted(bucket["relation_method_group_ids"]), ensure_ascii=True),
            "relation_row_count": str(bucket["relation_row_count"]),
        }
    return out


def load_relation_fill_counts(
    path: Path | None,
) -> tuple[dict[str, int], dict[str, int]]:
    if path is None or not path.exists():
        return {}, {}
    per_final_id: dict[str, int] = defaultdict(int)
    per_paper: dict[str, int] = defaultdict(int)
    for row in read_tsv_rows(path):
        final_id = normalize_text(row.get("final_formulation_id"))
        paper_key = normalize_text(row.get("paper_key"))
        if final_id:
            per_final_id[final_id] += 1
        if paper_key:
            per_paper[paper_key] += 1
    return dict(per_final_id), dict(per_paper)


def load_suspicious_case_maps(
    path: Path | None,
) -> tuple[dict[str, list[str]], dict[str, list[str]]]:
    if path is None or not path.exists():
        return {}, {}
    per_final_id: dict[str, list[str]] = defaultdict(list)
    per_paper: dict[str, list[str]] = defaultdict(list)
    for row in read_tsv_rows(path):
        reason = normalize_text(row.get("reason"))
        details = normalize_text(row.get("details"))
        concern_rank = normalize_text(row.get("concern_rank"))
        message = " | ".join(part for part in [concern_rank, reason, details] if part)
        final_id = normalize_text(row.get("final_formulation_id"))
        paper_key = normalize_text(row.get("paper_key"))
        if final_id and message and message not in per_final_id[final_id]:
            per_final_id[final_id].append(message)
        if paper_key and message and message not in per_paper[paper_key]:
            per_paper[paper_key].append(message)
    return dict(per_final_id), dict(per_paper)


def classify_row_priority(
    row: dict[str, str],
    relation_fill_count: int,
    suspicious_messages: list[str],
) -> tuple[str, str]:
    reasons: list[str] = []
    field_source_type = normalize_text(row.get("field_source_type"))
    review_needed = normalize_text(row.get("review_needed")).lower()
    polymer_identity = normalize_text(row.get("polymer_identity_final"))
    if suspicious_messages:
        reasons.append("suspicious_relation_case")
    if field_source_type == "relation_resolved":
        reasons.append("relation_resolved")
    if relation_fill_count > 0:
        reasons.append(f"relation_resolved_fills={relation_fill_count}")
    if field_source_type == "unresolved_blank":
        reasons.append("unresolved_blank")
    if review_needed in {"yes", "true", "1"}:
        reasons.append("review_needed")
    if polymer_identity.lower() in {"", "unknown"}:
        reasons.append("weak_or_unknown_branch_identity")

    if suspicious_messages or field_source_type == "unresolved_blank":
        return "high", "; ".join(reasons)
    if field_source_type == "relation_resolved" or review_needed in {"yes", "true", "1"}:
        return "medium", "; ".join(reasons)
    return "normal", "; ".join(reasons)


def extract_table_id(section: str) -> str:
    text = normalize_text(section)
    if not text:
        return ""
    if "__table_" in text.lower():
        return text
    match = re.search(r"\btable\s+\d+\b", text, flags=re.IGNORECASE)
    return match.group(0) if match else ""


def extract_anchor_row_index(row: dict[str, str]) -> str:
    tokens = [
        normalize_text(row.get("representative_source_formulation_id")),
        normalize_text(row.get("representative_source_raw_formulation_label")),
        normalize_text(row.get("formulation_id")),
    ]
    patterns = [
        re.compile(r"doe_row_(\d+)$", re.IGNORECASE),
        re.compile(r"^f[\s_-]*(\d+)\b", re.IGNORECASE),
        re.compile(r"^(\d+)\.?\b"),
    ]
    for token in tokens:
        for pattern in patterns:
            match = pattern.search(token)
            if match:
                return str(int(match.group(1)))
    return ""


def guess_boundary_source_type(row: dict[str, str]) -> str:
    source = normalize_text(row.get("source_candidate_sources") or row.get("candidate_source")).lower()
    evidence_region = normalize_text(row.get("instance_evidence_region_type")).lower()
    if "figure_variable_sweep" in source or "figure_variable_sweep" in normalize_text(row.get("change_context_tags")).lower():
        return "figure_sweep"
    if evidence_region == "table_row":
        return "table_row"
    if extract_table_id(row.get("evidence_section", "")):
        return "table_row"
    if normalize_text(row.get("representative_source_raw_formulation_label")):
        return "explicit_label"
    return "text_described"


def build_boundary_signature(anchor_fields: dict[str, str]) -> str:
    return "|".join(
        value
        for value in [
            anchor_fields["polymer_identity_anchor"],
            anchor_fields["drug_identity_anchor"],
            anchor_fields["surfactant_anchor"],
            anchor_fields["solvent_phase_anchor"],
            anchor_fields["drug_polymer_ratio_anchor"],
            anchor_fields["polymer_mw_anchor"],
            anchor_fields["la_ga_ratio_anchor"],
            anchor_fields["boundary_anchor_label"],
            anchor_fields["boundary_anchor_table_id"],
            anchor_fields["boundary_anchor_row_index"],
        ]
        if value
    )


def build_anchor_fields(row: dict[str, str]) -> dict[str, str]:
    boundary_anchor_label = normalize_text(
        row.get("representative_source_raw_formulation_label") or row.get("representative_source_formulation_id")
    )
    boundary_anchor_table_id = extract_table_id(row.get("evidence_section", ""))
    boundary_anchor_row_index = extract_anchor_row_index(row)
    boundary_source_type = guess_boundary_source_type(row)
    boundary_source_locator = normalize_text(row.get("evidence_section") or boundary_anchor_table_id or boundary_anchor_label)
    variant_change_anchor = normalize_text(row.get("change_descriptions"))
    if not variant_change_anchor:
        variant_change_anchor = normalize_text(row.get("collapsed_variant_classes"))
    anchor_fields = {
        "boundary_source_type": boundary_source_type or "unclear",
        "boundary_source_locator": boundary_source_locator,
        "boundary_anchor_label": boundary_anchor_label,
        "boundary_anchor_table_id": boundary_anchor_table_id,
        "boundary_anchor_row_index": boundary_anchor_row_index,
        "polymer_identity_anchor": normalize_text(row.get("polymer_identity_final") or row.get("polymer_identity")),
        "drug_identity_anchor": normalize_text(row.get("drug_name_value")),
        "surfactant_anchor": normalize_text(row.get("surfactant_name_value")),
        "solvent_phase_anchor": normalize_text(row.get("organic_solvent_value")),
        "drug_polymer_ratio_anchor": normalize_text(row.get("drug_to_polymer_mass_ratio_value")),
        "polymer_mw_anchor": normalize_text(row.get("polymer_mw_kDa_value") or row.get("plga_mw_kDa_value")),
        "la_ga_ratio_anchor": normalize_text(row.get("la_ga_ratio_value")),
        "variant_change_anchor": variant_change_anchor,
    }
    anchor_fields["boundary_signature"] = build_boundary_signature(anchor_fields)
    return anchor_fields


def choice_if_valid(value: str, options: list[str], fallback: str = "") -> str:
    text = normalize_text(value)
    return text if text in options else fallback


def sort_final_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    def sort_key(row: dict[str, str]) -> tuple[str, str, str, str]:
        return (
            normalize_text(row.get("key")),
            normalize_text(row.get("family_id")),
            normalize_text(row.get("variant_role")),
            normalize_text(row.get("final_formulation_id")),
        )

    return sorted(rows, key=sort_key)


def build_seed_and_reference_rows(
    final_rows: list[dict[str, str]],
    manifest_map: dict[str, dict[str, str]],
    decision_map: dict[tuple[str, str], dict[str, str]],
    relation_map: dict[tuple[str, str], dict[str, str]],
    relation_fill_counts_by_final_id: dict[str, int],
    suspicious_messages_by_final_id: dict[str, list[str]],
    manual_template_rows_per_paper: int,
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    rows_by_key: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in sort_final_rows(final_rows):
        rows_by_key[normalize_text(row.get("key"))].append(row)

    review_rows: list[dict[str, str]] = []
    manual_rows: list[dict[str, str]] = []
    pred_reference_rows: list[dict[str, str]] = []
    family_reference_rows: list[dict[str, str]] = []

    for paper_key in sorted(rows_by_key):
        manifest_row = manifest_map.get(paper_key, {})
        paper_rows = rows_by_key[paper_key]
        paper_title = normalize_text(manifest_row.get("title") or manifest_row.get("paper_title") or paper_rows[0].get("paper_title"))
        publication_year = normalize_text(manifest_row.get("publication_year") or manifest_row.get("year"))
        doi = normalize_doi(manifest_row.get("doi") or paper_rows[0].get("doi"))
        doi_url = f"https://doi.org/{doi}" if doi else ""

        predicted_family_raw_values: list[str] = []
        for row in paper_rows:
            source_id = normalize_text(row.get("representative_source_formulation_id"))
            decision_row = decision_map.get((paper_key, source_id), {})
            predicted_family_raw_values.append(
                normalize_text(row.get("family_id") or decision_row.get("family_id") or row.get("final_formulation_id"))
            )

        family_order: list[str] = []
        for family_value in predicted_family_raw_values:
            if family_value not in family_order:
                family_order.append(family_value)
        family_label_map = {
            family_value: f"{paper_key}::FAM{index:02d}"
            for index, family_value in enumerate(family_order, start=1)
        }

        gt_id_map: dict[str, str] = {}
        group_rows: dict[str, list[dict[str, str]]] = defaultdict(list)
        for index, row in enumerate(paper_rows, start=1):
            gt_id = f"{paper_key}_G{index:03d}"
            source_id = normalize_text(row.get("representative_source_formulation_id"))
            gt_id_map[normalize_text(row.get("final_formulation_id"))] = gt_id
            gt_id_map[source_id] = gt_id
            gt_id_map[normalize_text(row.get("formulation_id"))] = gt_id
            decision_row = decision_map.get((paper_key, source_id), {})
            family_raw = normalize_text(row.get("family_id") or decision_row.get("family_id") or row.get("final_formulation_id"))
            group_rows[family_raw].append(row)

        family_core_gt_id: dict[str, str] = {}
        for family_raw, members in group_rows.items():
            preferred = members[0]
            for member in members:
                source_id = normalize_text(member.get("representative_source_formulation_id"))
                decision_row = decision_map.get((paper_key, source_id), {})
                predicted_variant_role = normalize_text(member.get("variant_role") or decision_row.get("variant_role"))
                if predicted_variant_role == "family_core":
                    preferred = member
                    break
            preferred_source_id = normalize_text(preferred.get("representative_source_formulation_id"))
            family_core_gt_id[family_raw] = gt_id_map.get(preferred_source_id) or gt_id_map.get(
                normalize_text(preferred.get("final_formulation_id"))
            )

        for index, row in enumerate(paper_rows, start=1):
            source_id = normalize_text(row.get("representative_source_formulation_id"))
            decision_row = decision_map.get((paper_key, source_id), {})
            final_formulation_id = normalize_text(row.get("final_formulation_id"))
            family_raw = normalize_text(row.get("family_id") or decision_row.get("family_id") or row.get("final_formulation_id"))
            predicted_family_label = family_label_map.get(family_raw, "")
            predicted_variant_role = normalize_text(row.get("variant_role") or decision_row.get("variant_role"))
            predicted_payload_state = normalize_text(row.get("payload_state") or decision_row.get("payload_state"))
            predicted_parent_raw = normalize_text(row.get("parent_core_row_id") or decision_row.get("parent_core_row_id"))
            suggested_gt_id = f"{paper_key}_G{index:03d}"
            suggested_parent_gt_id = family_core_gt_id.get(family_raw, suggested_gt_id)
            anchor_fields = build_anchor_fields(row)
            relation_fill_count = relation_fill_counts_by_final_id.get(final_formulation_id, 0)
            suspicious_messages = suspicious_messages_by_final_id.get(final_formulation_id, [])
            row_priority, priority_reason = classify_row_priority(
                row=row,
                relation_fill_count=relation_fill_count,
                suspicious_messages=suspicious_messages,
            )
            review_row = {column: "" for column in REVIEW_COLUMNS}
            review_row.update(
                {
                    "paper_key": paper_key,
                    "doi": doi,
                    "doi_url": doi_url,
                    "paper_title": paper_title,
                    "publication_year": publication_year,
                    "seed_row_origin": "pred_seed",
                    "seed_pred_final_formulation_id": final_formulation_id,
                    "seed_pred_representative_source_formulation_id": source_id,
                    "seed_pred_family_id": predicted_family_label or normalize_text(row.get("family_id") or decision_row.get("family_id")),
                    "seed_pred_parent_core_row_id": gt_id_map.get(predicted_parent_raw, predicted_parent_raw),
                    "seed_pred_variant_role": predicted_variant_role,
                    "seed_pred_payload_state": predicted_payload_state,
                    "seed_pred_benchmark_default_include": normalize_text(
                        row.get("benchmark_default_include") or decision_row.get("benchmark_default_include")
                    ),
                    "seed_pred_final_output_rule": normalize_text(row.get("final_output_rule")),
                    "seed_pred_review_needed": normalize_text(row.get("review_needed") or decision_row.get("review_needed")),
                    "seed_pred_field_source_type": normalize_text(row.get("field_source_type")),
                    "seed_pred_row_priority": row_priority,
                    "seed_pred_priority_reason": priority_reason,
                    "seed_pred_suspicious_case_flag": "yes" if suspicious_messages else "no",
                    "seed_pred_boundary_source_type": anchor_fields["boundary_source_type"],
                    "seed_pred_boundary_source_locator": anchor_fields["boundary_source_locator"],
                    "seed_pred_boundary_anchor_label": anchor_fields["boundary_anchor_label"],
                    "seed_pred_boundary_anchor_table_id": anchor_fields["boundary_anchor_table_id"],
                    "seed_pred_boundary_anchor_row_index": anchor_fields["boundary_anchor_row_index"],
                    "seed_pred_polymer_identity_anchor": anchor_fields["polymer_identity_anchor"],
                    "seed_pred_drug_identity_anchor": anchor_fields["drug_identity_anchor"],
                    "seed_pred_surfactant_anchor": anchor_fields["surfactant_anchor"],
                    "seed_pred_solvent_phase_anchor": anchor_fields["solvent_phase_anchor"],
                    "seed_pred_drug_polymer_ratio_anchor": anchor_fields["drug_polymer_ratio_anchor"],
                    "seed_pred_polymer_mw_anchor": anchor_fields["polymer_mw_anchor"],
                    "seed_pred_la_ga_ratio_anchor": anchor_fields["la_ga_ratio_anchor"],
                    "seed_pred_variant_change_anchor": anchor_fields["variant_change_anchor"],
                    "seed_pred_boundary_signature": anchor_fields["boundary_signature"],
                    "gt_formulation_id": suggested_gt_id,
                    "family_id": predicted_family_label,
                    "parent_core_row_id": suggested_parent_gt_id,
                    "variant_role": choice_if_valid(
                        predicted_variant_role if predicted_variant_role else ("family_core" if suggested_parent_gt_id == suggested_gt_id else ""),
                        VARIANT_ROLE_OPTIONS,
                        "unclear" if predicted_variant_role else "",
                    ),
                    "payload_state": choice_if_valid(predicted_payload_state, PAYLOAD_STATE_OPTIONS, "unknown" if predicted_payload_state else ""),
                    "benchmark_include_gt": choice_if_valid(
                        normalize_text(row.get("benchmark_default_include") or decision_row.get("benchmark_default_include")),
                        BENCHMARK_INCLUDE_OPTIONS,
                    ),
                    "boundary_source_type": choice_if_valid(anchor_fields["boundary_source_type"], BOUNDARY_SOURCE_TYPE_OPTIONS, "unclear"),
                    "boundary_source_locator": anchor_fields["boundary_source_locator"],
                    "boundary_anchor_label": anchor_fields["boundary_anchor_label"],
                    "boundary_anchor_table_id": anchor_fields["boundary_anchor_table_id"],
                    "boundary_anchor_row_index": anchor_fields["boundary_anchor_row_index"],
                    "polymer_identity_anchor": anchor_fields["polymer_identity_anchor"],
                    "drug_identity_anchor": anchor_fields["drug_identity_anchor"],
                    "surfactant_anchor": anchor_fields["surfactant_anchor"],
                    "solvent_phase_anchor": anchor_fields["solvent_phase_anchor"],
                    "drug_polymer_ratio_anchor": anchor_fields["drug_polymer_ratio_anchor"],
                    "polymer_mw_anchor": anchor_fields["polymer_mw_anchor"],
                    "la_ga_ratio_anchor": anchor_fields["la_ga_ratio_anchor"],
                    "variant_change_anchor": anchor_fields["variant_change_anchor"],
                    "review_status": "pending",
                }
            )
            review_rows.append(review_row)

            relation_key = (paper_key, source_id)
            relation_info = relation_map.get(relation_key, {})
            pred_reference_rows.append(
                {
                    "paper_key": paper_key,
                    "doi": doi,
                    "doi_url": doi_url,
                    "paper_title": paper_title,
                    "publication_year": publication_year,
                    "pred_final_formulation_id": normalize_text(row.get("final_formulation_id")),
                    "pred_representative_source_formulation_id": source_id,
                    "pred_raw_formulation_label": normalize_text(row.get("representative_source_raw_formulation_label")),
                    "pred_family_id": predicted_family_label or normalize_text(row.get("family_id") or decision_row.get("family_id")),
                    "pred_parent_core_row_id": gt_id_map.get(predicted_parent_raw, predicted_parent_raw),
                    "pred_variant_role": predicted_variant_role,
                    "pred_payload_state": predicted_payload_state,
                    "pred_benchmark_default_include": normalize_text(
                        row.get("benchmark_default_include") or decision_row.get("benchmark_default_include")
                    ),
                    "pred_final_output_rule": normalize_text(row.get("final_output_rule")),
                    "pred_review_needed": normalize_text(row.get("review_needed")),
                    "pred_field_source_type": normalize_text(row.get("field_source_type")),
                    "pred_relation_resolved_fill_count": str(relation_fill_count),
                    "pred_row_priority": row_priority,
                    "pred_priority_reason": priority_reason,
                    "pred_suspicious_case_flag": "yes" if suspicious_messages else "no",
                    "pred_suspicious_case_notes": " || ".join(suspicious_messages),
                    "pred_polymer_identity_final": normalize_text(row.get("polymer_identity_final")),
                    "pred_drug_name_value": normalize_text(row.get("drug_name_value")),
                    "pred_surfactant_name_value": normalize_text(row.get("surfactant_name_value")),
                    "pred_organic_solvent_value": normalize_text(row.get("organic_solvent_value")),
                    "pred_polymer_mw_kDa_value": normalize_text(row.get("polymer_mw_kDa_value") or row.get("plga_mw_kDa_value")),
                    "pred_la_ga_ratio_value": normalize_text(row.get("la_ga_ratio_value")),
                    "pred_preparation_method": normalize_text(row.get("preparation_method")),
                    "pred_source_candidate_count": normalize_text(row.get("source_candidate_count")),
                    "pred_source_candidate_ids": normalize_text(row.get("source_candidate_ids")),
                    "pred_source_candidate_labels": normalize_text(row.get("source_candidate_labels")),
                    "pred_source_candidate_sources": normalize_text(row.get("source_candidate_sources")),
                    "pred_collapsed_variant_count": normalize_text(row.get("collapsed_variant_count")),
                    "pred_collapsed_variant_source_ids": normalize_text(row.get("collapsed_variant_source_ids")),
                    "pred_collapsed_variant_classes": normalize_text(row.get("collapsed_variant_classes")),
                    "pred_boundary_source_type": anchor_fields["boundary_source_type"],
                    "pred_boundary_source_locator": anchor_fields["boundary_source_locator"],
                    "pred_boundary_anchor_label": anchor_fields["boundary_anchor_label"],
                    "pred_boundary_anchor_table_id": anchor_fields["boundary_anchor_table_id"],
                    "pred_boundary_anchor_row_index": anchor_fields["boundary_anchor_row_index"],
                    "pred_polymer_identity_anchor": anchor_fields["polymer_identity_anchor"],
                    "pred_drug_identity_anchor": anchor_fields["drug_identity_anchor"],
                    "pred_surfactant_anchor": anchor_fields["surfactant_anchor"],
                    "pred_solvent_phase_anchor": anchor_fields["solvent_phase_anchor"],
                    "pred_drug_polymer_ratio_anchor": anchor_fields["drug_polymer_ratio_anchor"],
                    "pred_polymer_mw_anchor": anchor_fields["polymer_mw_anchor"],
                    "pred_la_ga_ratio_anchor": anchor_fields["la_ga_ratio_anchor"],
                    "pred_variant_change_anchor": anchor_fields["variant_change_anchor"],
                    "pred_boundary_signature": anchor_fields["boundary_signature"],
                    "pred_relation_graph_ids": normalize_text(row.get("relation_graph_ids") or relation_info.get("relation_graph_ids")),
                    "pred_relation_method_group_ids": normalize_text(
                        row.get("relation_method_group_ids") or relation_info.get("relation_method_group_ids")
                    ),
                    "pred_relation_parent_candidate_ids": normalize_text(row.get("relation_parent_candidate_ids")),
                    "pred_relation_record_count": normalize_text(row.get("relation_record_count") or relation_info.get("relation_row_count")),
                }
            )

        for family_raw in family_order:
            members = group_rows.get(family_raw, [])
            member_final_ids = [normalize_text(member.get("final_formulation_id")) for member in members]
            member_source_ids = [normalize_text(member.get("representative_source_formulation_id")) for member in members]
            member_variant_roles = []
            member_payload_states = []
            for member in members:
                member_source_id = normalize_text(member.get("representative_source_formulation_id"))
                decision_row = decision_map.get((paper_key, member_source_id), {})
                member_variant_roles.append(normalize_text(member.get("variant_role") or decision_row.get("variant_role")))
                member_payload_states.append(normalize_text(member.get("payload_state") or decision_row.get("payload_state")))
            family_reference_rows.append(
                {
                    "paper_key": paper_key,
                    "doi": doi,
                    "paper_title": paper_title,
                    "pred_family_group_raw": family_raw,
                    "suggested_gt_family_id": family_label_map.get(family_raw, ""),
                    "suggested_parent_core_row_id": family_core_gt_id.get(family_raw, ""),
                    "member_count": str(len(members)),
                    "member_pred_final_formulation_ids": json.dumps(member_final_ids, ensure_ascii=True),
                    "member_pred_source_formulation_ids": json.dumps(member_source_ids, ensure_ascii=True),
                    "member_pred_variant_roles": json.dumps(member_variant_roles, ensure_ascii=True),
                    "member_pred_payload_states": json.dumps(member_payload_states, ensure_ascii=True),
                }
            )

        next_index = len(paper_rows) + 1
        for offset in range(manual_template_rows_per_paper):
            manual_row = {column: "" for column in REVIEW_COLUMNS}
            manual_row.update(
                {
                    "paper_key": paper_key,
                    "doi": doi,
                    "doi_url": doi_url,
                    "paper_title": paper_title,
                    "publication_year": publication_year,
                    "seed_row_origin": "manual_addition_template",
                    "gt_formulation_id": f"{paper_key}_G{next_index + offset:03d}",
                }
            )
            manual_rows.append(manual_row)

    return review_rows, manual_rows, pred_reference_rows, family_reference_rows


class CounterMap(dict[str, int]):
    def __init__(self, key_name: str, rows: list[dict[str, str]]) -> None:
        super().__init__()
        for row in rows:
            key = normalize_text(row.get(key_name))
            if key:
                self[key] = self.get(key, 0) + 1


def build_source_summary(
    review_rows: list[dict[str, str]],
    manual_rows: list[dict[str, str]],
    family_reference_rows: list[dict[str, str]],
    relation_fill_counts_by_paper: dict[str, int],
    suspicious_messages_by_paper: dict[str, list[str]],
) -> list[dict[str, str]]:
    per_key_seed = CounterMap("paper_key", review_rows)
    per_key_manual = CounterMap("paper_key", manual_rows)
    per_key_family = CounterMap("paper_key", family_reference_rows)
    keys = sorted(set(per_key_seed.keys()) | set(per_key_manual.keys()) | set(per_key_family.keys()))
    rows: list[dict[str, str]] = []
    for key in keys:
        sample = next((row for row in review_rows if row["paper_key"] == key), None)
        if sample is None:
            sample = next((row for row in manual_rows if row["paper_key"] == key), None)
        if sample is None:
            continue
        paper_review_rows = [row for row in review_rows if row["paper_key"] == key]
        direct_rows = sum(1 for row in paper_review_rows if normalize_text(row.get("seed_pred_field_source_type")) == "direct_extraction")
        relation_rows = sum(1 for row in paper_review_rows if normalize_text(row.get("seed_pred_field_source_type")) == "relation_resolved")
        unresolved_rows = sum(1 for row in paper_review_rows if normalize_text(row.get("seed_pred_field_source_type")) == "unresolved_blank")
        high_priority_rows = sum(1 for row in paper_review_rows if normalize_text(row.get("seed_pred_row_priority")) == "high")
        medium_priority_rows = sum(1 for row in paper_review_rows if normalize_text(row.get("seed_pred_row_priority")) == "medium")
        review_needed_rows = sum(
            1
            for row in paper_review_rows
            if normalize_text(row.get("seed_pred_review_needed")).lower() in {"yes", "true", "1"}
        )
        paper_suspicious_messages = suspicious_messages_by_paper.get(key, [])
        paper_priority_bucket = "high" if high_priority_rows > 0 else ("medium" if medium_priority_rows > 0 else "normal")
        rows.append(
            {
                "paper_key": key,
                "doi": sample["doi"],
                "doi_url": sample["doi_url"],
                "paper_title": sample["paper_title"],
                "publication_year": sample["publication_year"],
                "seed_predicted_rows": str(per_key_seed.get(key, 0)),
                "manual_addition_template_rows": str(per_key_manual.get(key, 0)),
                "predicted_family_groups": str(per_key_family.get(key, 0)),
                "relation_resolved_fill_events": str(relation_fill_counts_by_paper.get(key, 0)),
                "rows_direct_extraction": str(direct_rows),
                "rows_relation_resolved": str(relation_rows),
                "rows_unresolved_blank": str(unresolved_rows),
                "rows_review_needed": str(review_needed_rows),
                "rows_high_priority": str(high_priority_rows),
                "rows_medium_priority": str(medium_priority_rows),
                "paper_has_suspicious_relation_case": "yes" if paper_suspicious_messages else "no",
                "paper_priority_bucket": paper_priority_bucket,
                "paper_priority_notes": " || ".join(paper_suspicious_messages[:3]),
            }
        )
    return rows


def create_dropdown_sheet(ws) -> None:
    ws["A1"] = "gt_row_decision_options"
    for index, value in enumerate(GT_ROW_DECISION_OPTIONS, start=2):
        ws[f"A{index}"] = value
    ws["B1"] = "variant_role_options"
    for index, value in enumerate(VARIANT_ROLE_OPTIONS, start=2):
        ws[f"B{index}"] = value
    ws["C1"] = "payload_state_options"
    for index, value in enumerate(PAYLOAD_STATE_OPTIONS, start=2):
        ws[f"C{index}"] = value
    ws["D1"] = "benchmark_include_gt_options"
    for index, value in enumerate(BENCHMARK_INCLUDE_OPTIONS, start=2):
        ws[f"D{index}"] = value
    ws["E1"] = "boundary_source_type_options"
    for index, value in enumerate(BOUNDARY_SOURCE_TYPE_OPTIONS, start=2):
        ws[f"E{index}"] = value
    ws["F1"] = "review_status_options"
    for index, value in enumerate(REVIEW_STATUS_OPTIONS, start=2):
        ws[f"F{index}"] = value
    ws.sheet_state = "hidden"


def add_data_validation(ws, header_map: dict[str, int], row_start: int, row_end: int) -> None:
    validations = {
        "gt_row_decision": "=dropdown_options!$A$2:$A$5",
        "variant_role": "=dropdown_options!$B$2:$B$8",
        "payload_state": "=dropdown_options!$C$2:$C$6",
        "benchmark_include_gt": "=dropdown_options!$D$2:$D$4",
        "boundary_source_type": "=dropdown_options!$E$2:$E$7",
        "review_status": "=dropdown_options!$F$2:$F$4",
    }
    for column_name, formula in validations.items():
        column_index = header_map[column_name]
        column_letter = get_column_letter(column_index)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{column_letter}{row_start}:{column_letter}{row_end}")


def style_sheet(ws, editable_columns: set[str]) -> None:
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    header_reference_fill = PatternFill("solid", fgColor="1F2937")
    header_edit_fill = PatternFill("solid", fgColor="9A3412")
    reference_fill = PatternFill("solid", fgColor="F3F4F6")
    editable_fill = PatternFill("solid", fgColor="FFF7ED")
    for cell in ws[1]:
        header = normalize_text(cell.value)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_edit_fill if header in editable_columns else header_reference_fill
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            header = normalize_text(ws.cell(row=1, column=cell.column).value)
            editable = header in editable_columns
            cell.protection = Protection(locked=not editable)
            cell.fill = editable_fill if editable else reference_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.protection.sheet = True
    ws.protection.autoFilter = True
    ws.protection.sort = True
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True
    for column_cells in ws.columns:
        max_len = max(len(normalize_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 12), 38)


def write_sheet_rows(ws, columns: list[str], rows: list[dict[str, str]]) -> dict[str, int]:
    for column_index, column_name in enumerate(columns, start=1):
        ws.cell(row=1, column=column_index, value=column_name)
    for row_index, row in enumerate(rows, start=2):
        for column_index, column_name in enumerate(columns, start=1):
            ws.cell(row=row_index, column=column_index, value=row.get(column_name, ""))
    return {column_name: index for index, column_name in enumerate(columns, start=1)}


def build_workbook(
    workbook_path: Path,
    review_rows: list[dict[str, str]],
    manual_rows: list[dict[str, str]],
    pred_reference_rows: list[dict[str, str]],
    family_reference_rows: list[dict[str, str]],
    source_summary_rows: list[dict[str, str]],
) -> None:
    wb = Workbook()
    ws_review = wb.active
    ws_review.title = "review_gt_rows"
    ws_manual = wb.create_sheet("manual_gt_additions")
    ws_pred = wb.create_sheet("pred_reference_rows")
    ws_family = wb.create_sheet("family_reference")
    ws_summary = wb.create_sheet("source_summary")
    ws_instr = wb.create_sheet("instructions")
    ws_dropdown = wb.create_sheet("dropdown_options")

    review_header = write_sheet_rows(ws_review, REVIEW_COLUMNS, review_rows)
    manual_header = write_sheet_rows(ws_manual, REVIEW_COLUMNS, manual_rows)
    pred_columns = list(pred_reference_rows[0].keys()) if pred_reference_rows else [
        "paper_key",
        "doi",
        "paper_title",
        "pred_final_formulation_id",
    ]
    family_columns = list(family_reference_rows[0].keys()) if family_reference_rows else [
        "paper_key",
        "doi",
        "paper_title",
        "pred_family_group_raw",
    ]
    summary_columns = list(source_summary_rows[0].keys()) if source_summary_rows else [
        "paper_key",
        "doi",
        "paper_title",
    ]
    write_sheet_rows(ws_pred, pred_columns, pred_reference_rows)
    write_sheet_rows(ws_family, family_columns, family_reference_rows)
    write_sheet_rows(ws_summary, summary_columns, source_summary_rows)

    create_dropdown_sheet(ws_dropdown)
    if ws_review.max_row >= 2:
        add_data_validation(ws_review, review_header, 2, ws_review.max_row)
    if ws_manual.max_row >= 2:
        add_data_validation(ws_manual, manual_header, 2, ws_manual.max_row)

    style_sheet(ws_review, set(EDITABLE_COLUMNS))
    style_sheet(ws_manual, set(EDITABLE_COLUMNS))
    for ws in [ws_pred, ws_family, ws_summary]:
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in ws.columns:
            max_len = max(len(normalize_text(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 12), 42)

    instructions = [
        "Layer 2 boundary GT review is seeded from Stage 5 final formulation rows.",
        "Prediction-reference columns are locked and shaded gray; GT-authoritative columns are editable and shaded orange.",
        "On review_gt_rows, choose gt_row_decision for each seeded predicted row.",
        "Use include_gt when the seeded row should become a GT formulation row.",
        "Use exclude_seed_spurious when the seeded row should not become GT authority.",
        "Use unclear when the reviewer cannot yet decide and the row needs another pass.",
        "Use manual_gt_additions for rows created in the manual_gt_additions sheet.",
        "Only GT columns define authority. Prediction columns are reference only and may be overwritten by reviewer decisions.",
        "manual_gt_additions provides blank templates for missing GT rows that were not seeded by prediction.",
        "family_id should represent the reviewed GT family, not necessarily the seeded predicted family.",
        "parent_core_row_id should point to the GT family core row id. For family cores, self-reference is allowed.",
        "Core decision fields use dropdowns so later validation/export can remain machine-readable.",
        "Optional note fields are review_notes and gt_evidence_note.",
    ]
    ws_instr["A1"] = "Boundary GT Review Instructions"
    ws_instr["A1"].font = Font(bold=True)
    for index, text in enumerate(instructions, start=3):
        ws_instr[f"A{index}"] = text
    ws_instr.column_dimensions["A"].width = 140
    ws_instr.freeze_panes = "A3"

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--run-id", required=True, help="Existing run_id whose Stage 5 outputs will seed review.")
    parser.add_argument(
        "--out-subdir",
        required=True,
        help="Relative output folder under data/results/<run_id>/ for this review surface.",
    )
    parser.add_argument(
        "--run-dir",
        type=Path,
        default=None,
        help="Optional explicit run directory. Use this for nested child-lineage runs.",
    )
    parser.add_argument(
        "--final-table-tsv",
        type=Path,
        default=None,
        help="Optional explicit final_formulation_table_v1.tsv. Defaults to the run-local Stage 5 final table.",
    )
    parser.add_argument(
        "--decision-trace-tsv",
        type=Path,
        default=None,
        help="Optional explicit final_output_decision_trace_v1.tsv.",
    )
    parser.add_argument(
        "--relation-records-tsv",
        type=Path,
        default=None,
        help="Optional Stage 3 formulation_relation_records_v1.tsv for extra reference provenance.",
    )
    parser.add_argument(
        "--scope-manifest-tsv",
        type=Path,
        default=None,
        help="Optional scope manifest TSV for DOI/title/year enrichment. Auto-detected when omitted.",
    )
    parser.add_argument(
        "--manual-template-rows-per-paper",
        type=int,
        default=2,
        help="Blank manual-addition template rows to pre-seed per paper.",
    )
    parser.add_argument(
        "--workbook-name",
        default=WORKBOOK_NAME,
        help="Workbook filename to write inside the output subdirectory.",
    )
    parser.add_argument(
        "--relation-fills-tsv",
        type=Path,
        default=None,
        help="Optional relation_resolved_new_fills.tsv for review prioritization.",
    )
    parser.add_argument(
        "--suspicious-cases-tsv",
        type=Path,
        default=None,
        help="Optional suspicious_relation_resolved_cases.tsv for review prioritization.",
    )
    return parser


def main() -> None:
    args = build_arg_parser().parse_args()
    run_id = normalize_text(args.run_id)
    if not is_valid_run_id(run_id):
        raise ValueError(f"Invalid run_id: {run_id}")
    run_dir = resolve_run_dir(run_id, args.run_dir)
    if not run_dir.exists():
        raise FileNotFoundError(f"Run directory not found: {run_dir}")
    out_subdir = sanitize_out_subdir(args.out_subdir)
    out_dir = run_dir / out_subdir

    final_table_tsv = args.final_table_tsv or default_input_path(run_dir, "final_formulation_table_v1.tsv")
    decision_trace_tsv = args.decision_trace_tsv
    if decision_trace_tsv is None:
        candidate = default_input_path(run_dir, "final_output_decision_trace_v1.tsv")
        decision_trace_tsv = candidate if candidate.exists() else None
    scope_manifest_tsv = args.scope_manifest_tsv or discover_scope_manifest(run_dir)

    if not final_table_tsv.exists():
        raise FileNotFoundError(f"Final table TSV not found: {final_table_tsv}")
    if decision_trace_tsv is not None and not decision_trace_tsv.exists():
        raise FileNotFoundError(f"Decision trace TSV not found: {decision_trace_tsv}")
    if args.relation_records_tsv is not None and not args.relation_records_tsv.exists():
        raise FileNotFoundError(f"Relation records TSV not found: {args.relation_records_tsv}")
    if args.relation_fills_tsv is not None and not args.relation_fills_tsv.exists():
        raise FileNotFoundError(f"Relation fills TSV not found: {args.relation_fills_tsv}")
    if args.suspicious_cases_tsv is not None and not args.suspicious_cases_tsv.exists():
        raise FileNotFoundError(f"Suspicious cases TSV not found: {args.suspicious_cases_tsv}")

    final_rows = read_tsv_rows(final_table_tsv)
    manifest_map = load_manifest_map(scope_manifest_tsv)
    decision_map = load_decision_trace_map(decision_trace_tsv)
    relation_map = load_relation_counts(args.relation_records_tsv)
    relation_fill_counts_by_final_id, relation_fill_counts_by_paper = load_relation_fill_counts(args.relation_fills_tsv)
    suspicious_messages_by_final_id, suspicious_messages_by_paper = load_suspicious_case_maps(args.suspicious_cases_tsv)

    review_rows, manual_rows, pred_reference_rows, family_reference_rows = build_seed_and_reference_rows(
        final_rows=final_rows,
        manifest_map=manifest_map,
        decision_map=decision_map,
        relation_map=relation_map,
        relation_fill_counts_by_final_id=relation_fill_counts_by_final_id,
        suspicious_messages_by_final_id=suspicious_messages_by_final_id,
        manual_template_rows_per_paper=max(0, int(args.manual_template_rows_per_paper)),
    )
    source_summary_rows = build_source_summary(
        review_rows,
        manual_rows,
        family_reference_rows,
        relation_fill_counts_by_paper=relation_fill_counts_by_paper,
        suspicious_messages_by_paper=suspicious_messages_by_paper,
    )

    write_tsv(out_dir / REVIEW_SEED_TSV_NAME, REVIEW_COLUMNS, review_rows)
    write_tsv(out_dir / MANUAL_TEMPLATE_TSV_NAME, REVIEW_COLUMNS, manual_rows)
    if pred_reference_rows:
        write_tsv(out_dir / PRED_REFERENCE_TSV_NAME, list(pred_reference_rows[0].keys()), pred_reference_rows)
    if family_reference_rows:
        write_tsv(out_dir / FAMILY_REFERENCE_TSV_NAME, list(family_reference_rows[0].keys()), family_reference_rows)
    if source_summary_rows:
        write_tsv(out_dir / SOURCE_SUMMARY_TSV_NAME, list(source_summary_rows[0].keys()), source_summary_rows)

    workbook_path = out_dir / normalize_text(args.workbook_name or WORKBOOK_NAME)
    build_workbook(
        workbook_path=workbook_path,
        review_rows=review_rows,
        manual_rows=manual_rows,
        pred_reference_rows=pred_reference_rows,
        family_reference_rows=family_reference_rows,
        source_summary_rows=source_summary_rows,
    )

    print(
        json.dumps(
            {
                "run_id": run_id,
                "run_dir": str(run_dir),
                "out_dir": str(out_dir),
                "final_table_tsv": str(final_table_tsv),
                "decision_trace_tsv": str(decision_trace_tsv) if decision_trace_tsv else "",
                "relation_records_tsv": str(args.relation_records_tsv) if args.relation_records_tsv else "",
                "scope_manifest_tsv": str(scope_manifest_tsv) if scope_manifest_tsv else "",
                "seed_review_rows": len(review_rows),
                "manual_template_rows": len(manual_rows),
                "pred_reference_rows": len(pred_reference_rows),
                "family_reference_rows": len(family_reference_rows),
                "workbook_path": str(workbook_path),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
