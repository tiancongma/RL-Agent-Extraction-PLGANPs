#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from src.utils.active_data_source import (
        build_artifact_metadata,
        resolve_artifact_path,
        resolve_run_context,
        write_artifact_metadata_json,
    )
    from src.utils.paths import PROJECT_ROOT
except ModuleNotFoundError:
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from src.utils.active_data_source import (
        build_artifact_metadata,
        resolve_artifact_path,
        resolve_run_context,
        write_artifact_metadata_json,
    )
    from src.utils.paths import PROJECT_ROOT


COUNTS_NAME = "final_table_vs_gt_counts.tsv"
SUMMARY_NAME = "final_table_vs_gt_summary.md"
EE_SUBSET_NAME = "final_table_vs_gt_ee_subset.tsv"
AUDIT_COUNTS_NAME = "final_table_vs_gt_counts_by_doi.tsv"
AUDIT_SUMMARY_NAME = "final_table_vs_gt_count_audit.md"


def normalize_text(value: Any) -> str:
    return str(value or "").strip().lower()


def read_scope_manifest(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def read_final_table_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def read_gt_rows_from_workbook(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["review_formulations"]
    raw_rows = worksheet.iter_rows(values_only=True)
    header = [str(value) if value is not None else "" for value in next(raw_rows)]
    rows: list[dict[str, str]] = []
    for values in raw_rows:
        rows.append(
            {
                header[idx]: "" if idx >= len(values) or values[idx] is None else str(values[idx])
                for idx in range(len(header))
            }
        )
    return rows, header


def write_tsv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def build_summary_markdown(
    scope_name: str,
    manifest_path: Path,
    final_table_path: Path,
    gt_xlsx_path: Path,
    counts_rows: list[dict[str, str]],
    ee_supported: bool,
    summary_path: Path,
) -> None:
    totals = {
        "final_table_rows": sum(int(row["final_table_count"]) for row in counts_rows),
        "gt_rows": sum(int(row["gt_count"]) for row in counts_rows),
        "matched_papers": sum(1 for row in counts_rows if row["comparison_status"] == "match"),
        "mismatched_papers": sum(1 for row in counts_rows if row["comparison_status"] != "match"),
    }
    lines = [
        "# Final Table vs GT Summary",
        "",
        "## Declared scope",
        "",
        f"- scope_name: `{scope_name}`",
        f"- scope_manifest_tsv: `{manifest_path}`",
        f"- final_formulation_table_tsv: `{final_table_path}`",
        f"- gt_workbook: `{gt_xlsx_path}`",
        "",
        "## Benchmark-validity statement",
        "",
        "- This comparison is benchmark-valid for the declared scope because it evaluates only the complete-pipeline final formulation table produced by the full pipeline runner.",
        "- No intermediate Stage2 or other partial-layer artifacts are used as the official evaluation object.",
        "",
        "## Supported benchmark views",
        "",
        "- per-DOI final-formulation count comparison: supported",
        f"- EE subset comparison: {'supported' if ee_supported else 'not supported by the current authoritative GT artifact'}",
        "",
        "## Aggregate outcome",
        "",
        f"- total_final_table_rows: `{totals['final_table_rows']}`",
        f"- total_gt_rows: `{totals['gt_rows']}`",
        f"- matched_papers: `{totals['matched_papers']}`",
        f"- mismatched_papers: `{totals['mismatched_papers']}`",
        "",
        "## Per-paper counts",
        "",
    ]
    for row in counts_rows:
        lines.append(
            "- `{paper_key}`: final=`{final_table_count}` gt=`{gt_count}` diff=`{count_diff}` status=`{comparison_status}`".format(
                **row
            )
        )
    lines.extend(
        [
            "",
            "## Limitations",
            "",
            "- The current benchmark comparison is limited to final-formulation counts for this declared scope.",
            "- The authoritative fixed DEV15 skeleton workbook does not expose structured EE ground-truth fields, so no benchmark-valid EE subset comparison is emitted in this first full-pipeline run.",
            "- Any mismatch investigation must start from these final-table results and only then trace backward into Stage 5A decision-trace artifacts or Stage 2 candidate rows.",
        ]
    )
    summary_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_count_audit_markdown(
    counts_rows: list[dict[str, str]],
    gt_xlsx_path: Path,
    final_table_path: Path,
    audit_summary_path: Path,
) -> None:
    exact = sum(1 for row in counts_rows if row["count_status"] == "match")
    plus = sum(1 for row in counts_rows if row["count_status"] == "pred_gt_plus")
    minus = sum(1 for row in counts_rows if row["count_status"] == "pred_gt_minus")
    mismatches = sorted(
        [row for row in counts_rows if row["count_status"] != "match"],
        key=lambda row: abs(int(row["delta_count"])),
        reverse=True,
    )
    lines = [
        "# DEV15 GT vs Pred Count Audit",
        "",
        f"- gt_authority_file: `{gt_xlsx_path}`",
        f"- pred_source_file: `{final_table_path}`",
        f"- total_doi_count: `{len(counts_rows)}`",
        f"- exact_matches: `{exact}`",
        f"- positive_deltas: `{plus}`",
        f"- negative_deltas: `{minus}`",
        "",
        "## Top mismatches",
        "",
    ]
    if not mismatches:
        lines.append("- No DOI-level count mismatches.")
    else:
        for row in mismatches:
            lines.append(
                "- `{doi}` / `{paper_key}`: gt=`{gt_count}` pred=`{pred_count}` delta=`{delta_count}` status=`{count_status}`".format(
                    **row
                )
            )
    audit_summary_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_run_context(
    *,
    run_dir: Path,
    source_run_context: dict[str, Any],
    final_table_tsv: Path,
    gt_xlsx: Path,
    scope_manifest_tsv: Path,
    out_dir: Path,
    scope_name: str,
    result: dict[str, Any],
) -> str:
    generated_at = datetime.now().isoformat(timespec="seconds")
    return "\n".join(
        [
            "# RUN_CONTEXT",
            "",
            "## 1. Run Type",
            "",
            "- `diagnostic-only`",
            "",
            "## 2. Purpose",
            "",
            "- Compare the completed Stage 5 final table against the authoritative GT workbook for the declared scope.",
            "- Emit benchmark comparison evidence without modifying the final table or GT authority.",
            "",
            "## 3. Source Authority Resolution",
            "",
            f"- source_resolution: `{source_run_context['resolution_source']}`",
            f"- source_run_id: `{source_run_context['run_id']}`",
            f"- source_run_dir: `{source_run_context['run_dir']}`",
            f"- active_run_pointer_path: `{source_run_context.get('pointer_path') or ''}`",
            f"- scope_manifest_tsv: `{scope_manifest_tsv}`",
            f"- final_table_tsv: `{final_table_tsv}`",
            f"- gt_workbook_xlsx: `{gt_xlsx}`",
            "",
            "## 4. Exact Script Execution Order",
            "",
            "1. Run `src/stage5_benchmark/compare_final_table_to_gt_v1.py` on the completed Stage 5 final table.",
            "2. Refresh `RUN_CONTEXT.md` via `src/utils/update_run_context_with_feature_activation_v1.py` so feature activation lineage is recorded in the compare run.",
            "",
            "## 5. Outputs",
            "",
            f"- `{out_dir / COUNTS_NAME}`",
            f"- `{out_dir / SUMMARY_NAME}`",
            f"- `{out_dir / AUDIT_COUNTS_NAME}`",
            f"- `{out_dir / AUDIT_SUMMARY_NAME}`",
            f"- `{out_dir / 'RUN_CONTEXT.md'}`",
            "",
            "## 6. Benchmark Status",
            "",
            "- `diagnostic-only, not benchmark-valid final output`",
            "- Reason: this node compares the final table to GT; it does not alter Stage 2, Stage 3, or Stage 5 semantics.",
            "",
            "## 7. Reproduction Metadata",
            "",
            f"- generated_at: `{generated_at}`",
            f"- scope_name: `{scope_name}`",
            f"- final_table_rows: `{result['total_final_table_rows']}`",
            f"- gt_rows: `{result['total_gt_rows']}`",
            f"- matched_papers: `{result['papers_matching']}`",
            f"- mismatched_papers: `{result['papers_mismatching']}`",
        ]
    ) + "\n"


def compare_final_table_to_gt(
    final_table_tsv: Path,
    gt_xlsx: Path,
    scope_manifest_tsv: Path,
    out_dir: Path,
    scope_name: str,
    source_run_context: dict[str, Any],
) -> dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)

    manifest_rows = read_scope_manifest(scope_manifest_tsv)
    scope_by_key = {row["key"]: row for row in manifest_rows}
    scope_keys = set(scope_by_key)

    final_rows = [
        row for row in read_final_table_rows(final_table_tsv) if row.get("key", "") in scope_keys
    ]
    gt_rows, gt_header = read_gt_rows_from_workbook(gt_xlsx)
    gt_rows = [
        row
        for row in gt_rows
        if row.get("paper_key", "") in scope_keys
        and normalize_text(row.get("formulation_exists_gt")) == "yes"
    ]

    final_counts = Counter(row["key"] for row in final_rows)
    gt_counts = Counter(row["paper_key"] for row in gt_rows)

    ee_supported = any("encapsulation" in normalize_text(column) for column in gt_header)

    counts_rows: list[dict[str, str]] = []
    audit_rows: list[dict[str, str]] = []
    for key in sorted(scope_keys):
        manifest_row = scope_by_key[key]
        final_count = int(final_counts.get(key, 0))
        gt_count = int(gt_counts.get(key, 0))
        diff = final_count - gt_count
        if diff == 0:
            status = "match"
        elif diff > 0:
            status = "over"
        else:
            status = "under"
        count_status = "match" if diff == 0 else ("pred_gt_plus" if diff > 0 else "pred_gt_minus")
        counts_rows.append(
            {
                "paper_key": key,
                "doi": manifest_row.get("doi", ""),
                "paper_title": manifest_row.get("title", ""),
                "final_table_count": str(final_count),
                "gt_count": str(gt_count),
                "count_diff": str(diff),
                "comparison_status": status,
                "final_table_artifact": str(final_table_tsv),
                "gt_artifact": str(gt_xlsx),
                "notes": (
                    "count_match"
                    if status == "match"
                    else "final_table_vs_fixed_skeleton_count_mismatch"
                ),
            }
        )
        audit_rows.append(
            {
                "doi": manifest_row.get("doi", ""),
                "paper_key": key,
                "gt_count": str(gt_count),
                "pred_count": str(final_count),
                "delta_count": str(diff),
                "count_status": count_status,
                "gt_authority_file": str(gt_xlsx),
                "pred_source_file": str(final_table_tsv),
                "matched_rows": "",
                "missing_rows": "",
                "spurious_rows": "",
            }
        )

    counts_path = out_dir / COUNTS_NAME
    summary_path = out_dir / SUMMARY_NAME
    audit_counts_path = out_dir / AUDIT_COUNTS_NAME
    audit_summary_path = out_dir / AUDIT_SUMMARY_NAME
    write_tsv(
        counts_path,
        [
            "paper_key",
            "doi",
            "paper_title",
            "final_table_count",
            "gt_count",
            "count_diff",
            "comparison_status",
            "final_table_artifact",
            "gt_artifact",
            "notes",
        ],
        counts_rows,
    )
    write_tsv(
        audit_counts_path,
        [
            "doi",
            "paper_key",
            "gt_count",
            "pred_count",
            "delta_count",
            "count_status",
            "gt_authority_file",
            "pred_source_file",
            "matched_rows",
            "missing_rows",
            "spurious_rows",
        ],
        audit_rows,
    )
    build_summary_markdown(
        scope_name=scope_name,
        manifest_path=scope_manifest_tsv,
        final_table_path=final_table_tsv,
        gt_xlsx_path=gt_xlsx,
        counts_rows=counts_rows,
        ee_supported=ee_supported,
        summary_path=summary_path,
    )
    build_count_audit_markdown(
        counts_rows=audit_rows,
        gt_xlsx_path=gt_xlsx,
        final_table_path=final_table_tsv,
        audit_summary_path=audit_summary_path,
    )

    result = {
        "scope_name": scope_name,
        "final_table_path": str(final_table_tsv),
        "gt_xlsx_path": str(gt_xlsx),
        "counts_path": str(counts_path),
        "summary_path": str(summary_path),
        "audit_counts_path": str(audit_counts_path),
        "audit_summary_path": str(audit_summary_path),
        "ee_subset_path": "",
        "ee_subset_supported": ee_supported,
        "papers_in_scope": len(scope_keys),
        "papers_matching": sum(1 for row in counts_rows if row["comparison_status"] == "match"),
        "papers_mismatching": sum(1 for row in counts_rows if row["comparison_status"] != "match"),
        "total_final_table_rows": sum(int(row["final_table_count"]) for row in counts_rows),
        "total_gt_rows": sum(int(row["gt_count"]) for row in counts_rows),
    }
    (out_dir / "RUN_CONTEXT.md").write_text(
        build_run_context(
            run_dir=out_dir,
            source_run_context=source_run_context,
            final_table_tsv=final_table_tsv,
            gt_xlsx=gt_xlsx,
            scope_manifest_tsv=scope_manifest_tsv,
            out_dir=out_dir,
            scope_name=scope_name,
            result=result,
        ),
        encoding="utf-8",
    )
    subprocess.run(
        [
            sys.executable,
            str(PROJECT_ROOT / "src" / "utils" / "update_run_context_with_feature_activation_v1.py"),
            "--run-dir",
            str(out_dir),
        ],
        cwd=PROJECT_ROOT,
        text=True,
        check=True,
    )
    return result


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Compare a final formulation table against the authoritative fixed skeleton GT workbook."
    )
    parser.add_argument("--final-table-tsv", type=Path, default=None)
    parser.add_argument("--gt-xlsx", type=Path, default=None)
    parser.add_argument("--scope-manifest-tsv", type=Path, default=None)
    parser.add_argument("--out-dir", type=Path, default=None)
    parser.add_argument("--run-dir", type=Path, default=None)
    parser.add_argument("--run-id", default="")
    parser.add_argument("--scope-name", default="controlled_scope")
    return parser


def main() -> None:
    args = build_arg_parser().parse_args()
    run_context = resolve_run_context(
        explicit_run_dir=args.run_dir,
        explicit_run_id=str(args.run_id or "").strip(),
    )
    final_table_tsv = resolve_artifact_path(
        explicit_path=args.final_table_tsv,
        run_context=run_context,
        pointer_key="stage5_final_table_tsv",
        canonical_relative="final_formulation_table_v1.tsv",
    )
    gt_xlsx = resolve_artifact_path(
        explicit_path=args.gt_xlsx,
        run_context=run_context,
        pointer_key="gt_workbook_xlsx",
    )
    scope_manifest_tsv = resolve_artifact_path(
        explicit_path=args.scope_manifest_tsv,
        run_context=run_context,
        pointer_key="scope_manifest_tsv",
        preferred_run_local_names=["dev15_scope.tsv", "scope.tsv", "scope_manifest.tsv"],
    )
    out_dir = args.out_dir.resolve() if args.out_dir is not None else (Path(run_context["run_dir"]) / "gt_authority_v2_variantaware").resolve()
    print(
        json.dumps(
            {
                "resolved_source_run_dir": str(run_context["run_dir"]),
                "resolved_source_run_id": str(run_context["run_id"]),
                "source_resolution": str(run_context["resolution_source"]),
                "active_run_pointer_path": str(run_context.get("pointer_path") or ""),
                "resolved_input_files": {
                    "final_table_tsv": str(final_table_tsv),
                    "gt_xlsx": str(gt_xlsx),
                    "scope_manifest_tsv": str(scope_manifest_tsv),
                },
                "resolved_out_dir": str(out_dir),
            },
            indent=2,
        )
    )
    result = compare_final_table_to_gt(
        final_table_tsv=final_table_tsv,
        gt_xlsx=gt_xlsx,
        scope_manifest_tsv=scope_manifest_tsv,
        out_dir=out_dir,
        scope_name=args.scope_name,
        source_run_context=run_context,
    )
    metadata_path = write_artifact_metadata_json(
        Path(result["counts_path"]),
        build_artifact_metadata(
            source_run_context=run_context,
            source_files={
                "final_table_tsv": str(final_table_tsv),
                "gt_xlsx": str(gt_xlsx),
                "scope_manifest_tsv": str(scope_manifest_tsv),
            },
            generated_by="src/stage5_benchmark/compare_final_table_to_gt_v1.py",
            note="Stage5 final-table vs GT comparison authority metadata.",
            extra={
                "summary_path": str(result["summary_path"]),
                "audit_counts_path": str(result["audit_counts_path"]),
                "audit_summary_path": str(result["audit_summary_path"]),
                "scope_name": args.scope_name,
            },
        ),
    )
    result["counts_metadata_json"] = str(metadata_path)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
