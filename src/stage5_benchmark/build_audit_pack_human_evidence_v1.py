#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment


def _bootstrap_import_paths() -> None:
    here = Path(__file__).resolve()
    for p in [here] + list(here.parents):
        if (p / "src" / "utils" / "paths.py").exists():
            sys.path.insert(0, str(p))
            return


_bootstrap_import_paths()
from src.stage5_benchmark.audit_evidence_resolver_v1 import AuditEvidenceResolverV1  # noqa: E402
from src.utils import paths  # noqa: E402
from src.utils.run_id import is_valid_run_id  # noqa: E402

MISSING_FLAGS = ["missing_drug", "missing_polymer_identity", "missing_solvent", "missing_surfactant", "missing_feed_anchor"]
VALUE_SOURCES = {
    "table_csv_cell",
    "fulltext_span",
    "derived_doe_decode",
    "derived_rule",
    "proxy_compose",
    "unknown",
}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build human-readable audit pack with DoE/table evidence support.")
    p.add_argument("--run-id", default="")
    p.add_argument("--out-subdir", default="", help="Required subdirectory under data/results/<run_id>/ for this run variant.")
    p.add_argument("--input-tsv", required=True)
    p.add_argument("--out-xlsx", default="")
    p.add_argument("--n-per-bucket", type=int, default=15)
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--max-span-chars", type=int, default=500)
    p.add_argument("--max-table-row-chars", type=int, default=800)
    return p.parse_args()


def _sanitize_out_subdir(s: str) -> str:
    v = str(s or "").strip().replace("\\", "/")
    if not v:
        raise ValueError(
            "ERROR: --out-subdir is required when reusing a run_id. Use a stage/variant folder name, e.g. stage2_validation or stage5_signature_iter001."
        )
    if Path(v).is_absolute():
        raise ValueError("ERROR: --out-subdir must be a relative path.")
    parts = [p for p in v.split("/") if p]
    if not parts or any(p == ".." for p in parts):
        raise ValueError("ERROR: --out-subdir cannot contain path traversal ('..').")
    return "/".join(parts)


def parse_json_obj(v: Any) -> dict[str, Any]:
    try:
        s = str(v).strip()
        return json.loads(s) if s else {}
    except Exception:
        return {}


def short_text(v: Any, n: int) -> str:
    s = "" if v is None else str(v)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:n]


def console_safe(v: Any, n: int) -> str:
    s = short_text(v, n)
    return s.encode("ascii", errors="replace").decode("ascii")


def looks_nonempty(v: Any) -> bool:
    return str(v or "").strip() != ""


def detect_table_evidence_kind(table_csv_path: str, table_filename: str, table_row_text: str) -> str:
    p = str(table_csv_path or "").strip()
    fn = str(table_filename or "").strip().lower()
    row = str(table_row_text or "").strip()
    if p and Path(p).exists():
        return "table_csv_cell"
    if row and (fn.startswith("fulltext_table_proxy__") or not p):
        return "proxy_compose"
    return "none"


def infer_value_source_numeric(
    field_value: str,
    table_evidence_kind: str,
    evidence_text: str,
    derived_rule_id: str,
) -> str:
    v = str(field_value or "").strip()
    if not v:
        return "unknown"
    if table_evidence_kind == "table_csv_cell":
        return "table_csv_cell"
    if table_evidence_kind == "proxy_compose":
        return "proxy_compose"
    if str(derived_rule_id or "").strip():
        return "derived_rule"
    if str(evidence_text or "").strip():
        return "fulltext_span"
    return "unknown"


def is_cross_paper_table_path(zotero_key: str, table_csv_path: str) -> bool:
    p = str(table_csv_path or "").strip().replace("\\", "/")
    k = str(zotero_key or "").strip()
    if not p or not k:
        return False
    return f"/tables/{k}/" not in p


def is_mixed_table_fulltext_pointer(evidence_source_type: str, evidence_pointer_raw: str) -> bool:
    et = str(evidence_source_type or "").strip().lower()
    ptr = str(evidence_pointer_raw or "").strip().lower()
    if et != "table":
        return False
    return ptr.startswith("fulltext|") and "pattern_window" in ptr


def apply_provenance_hard_guards(row: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    out = dict(row)
    violations: list[str] = []

    evidence_source_type = str(out.get("evidence_source_type", "")).strip()
    zotero_key = str(out.get("zotero_key", "")).strip()
    table_csv_path = str(out.get("table_csv_path", "")).strip()
    pointer_raw = str(out.get("evidence_pointer_raw", "")).strip()

    if evidence_source_type == "table" and is_cross_paper_table_path(zotero_key, table_csv_path):
        violations.append("cross_paper_table_path")
        out["table_selection_status"] = "rejected_cross_paper"
        out["ownership_check_passed"] = False
        out["ownership_check_reason"] = "cross_paper_table_path_rejected"
        out["table_csv_path"] = ""
        out["table_filename"] = ""
        out["table_row_text"] = ""
        out["table_cell_text"] = ""
        out["table_evidence_kind"] = "none"
        out["evidence_source_type"] = "unknown"

    if is_mixed_table_fulltext_pointer(evidence_source_type=evidence_source_type, evidence_pointer_raw=pointer_raw):
        violations.append("mixed_table_fulltext_pointer")
        if str(out.get("table_selection_status", "")).strip() != "rejected_cross_paper":
            out["table_selection_status"] = "evidence_mixed_source"
        out["human_review_tag"] = "cross_paper_or_mixed_evidence"

    return out, violations


def norm_doi(v: Any) -> str:
    s = str(v or "").strip().lower()
    s = re.sub(r"^doi\s*:\s*", "", s)
    s = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", s)
    return re.sub(r"^doi\.org/", "", s)


def canonical_polymer_type(raw_polymer: str, notes: str, evidence: str) -> str:
    t = " ".join([str(raw_polymer), str(notes), str(evidence)]).lower()
    if re.search(r"plga[\s\-]*peg|peg[\s\-]*plga", t):
        return "PLGA-PEG"
    if re.search(r"\bpcl\b|poly\s*\(?caprolactone\)?", t):
        return "PCL"
    if re.search(r"\bpla\b|polylactic acid", t):
        return "PLA"
    if "plga" in t:
        return "PLGA"
    return ""


def canonical_mw_or_iv(mw_raw: str, notes: str) -> str:
    t = " ".join([str(mw_raw), str(notes)]).lower()
    m = re.search(r"(?:iv|intrinsic viscosity)[^0-9]*([-+]?\d+(?:\.\d+)?)\s*d[l1]/g", t)
    if m:
        return f"IV:{m.group(1)}"
    m = re.search(r"(?:molecular weight(?: of)?|mw)[^0-9]*([-+]?\d+(?:,\d{3})*(?:\.\d+)?)", t)
    if m:
        x = float(m.group(1).replace(",", ""))
        if "kda" in t:
            return str(int(x)) if x.is_integer() else f"{x:.6g}"
        x = x / 1000.0 if x > 10000 else x
        return str(int(x)) if float(x).is_integer() else f"{x:.6g}"
    m = re.search(r"[-+]?\d+(?:\.\d+)?", str(mw_raw))
    if m:
        x = float(m.group(0))
        x = x / 1000.0 if x > 10000 else x
        return str(int(x)) if float(x).is_integer() else f"{x:.6g}"
    return ""


def load_manifest_map() -> tuple[pd.DataFrame, str]:
    for p in [paths.DATA_CLEANED_INDEX_DIR / "manifest_goren_2025.tsv", paths.DATA_CLEANED_INDEX_DIR / "manifest_current.tsv"]:
        if not p.exists():
            continue
        df = pd.read_csv(p, sep="\t", dtype=str).fillna("")
        key_col = "key" if "key" in df.columns else ("zotero_key" if "zotero_key" in df.columns else "")
        if key_col and "doi" in df.columns:
            out = pd.DataFrame({
                "zotero_key": df[key_col].astype(str),
                "doi_mf": df["doi"].astype(str).map(norm_doi),
                "title_mf": df["title"].astype(str) if "title" in df.columns else "",
                "year_mf": df["year"].astype(str) if "year" in df.columns else "",
            })
            return out.drop_duplicates("zotero_key"), str(p)
    return pd.DataFrame(columns=["zotero_key", "doi_mf", "title_mf", "year_mf"]), ""


def add_missing_flags(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    def col(name: str) -> pd.Series:
        return out[name].astype(str) if name in out.columns else pd.Series([""] * len(out), index=out.index)

    if "critical_missing_json" in out.columns:
        crit = out["critical_missing_json"].map(parse_json_obj)
        for f in MISSING_FLAGS:
            out[f] = crit.map(lambda d: bool(d.get(f, False)))
    else:
        out["missing_drug"] = col("drug_name").str.strip().eq("")
        poly_ok = col("la_ga_ratio").str.strip().ne("") | col("plga_mw_kDa").str.strip().ne("") | col("polymer_name").str.strip().ne("")
        out["missing_polymer_identity"] = ~poly_ok
        out["missing_solvent"] = col("organic_solvent").str.strip().eq("")
        out["missing_surfactant"] = ~(col("surfactant_name").str.strip().ne("") | col("notes").str.contains("pva|pluronic|tween|poloxamer", case=False, regex=True))
        out["missing_feed_anchor"] = ~(col("drug_to_polymer_ratio").str.strip().ne("") | col("drug_to_polymer_mass_ratio").str.strip().ne("") | col("drug_feed_amount_text").str.strip().ne("") | col("plga_mass_mg").str.strip().ne(""))
    return out


def choose_buckets(df: pd.DataFrame, n: int, seed: int) -> dict[str, pd.DataFrame]:
    work = df.copy()
    work["quality_score"] = 0
    for c in ["drug_name", "la_ga_ratio", "plga_mw_kDa", "organic_solvent", "drug_feed_amount_text"]:
        if c in work.columns:
            work["quality_score"] += work[c].astype(str).str.strip().ne("").astype(int)

    a = work[(work.get("gate_used", "").astype(str).eq("C") | work.get("merge_reason", "").astype(str).str.contains("no_auto_merge", case=False, na=False) | work["missing_polymer_identity"])].copy()
    if len(a) > n:
        a = a.sample(n=n, random_state=seed)

    b = work[(work.get("gate_used", "").astype(str).eq("B") | work.get("merge_reason", "").astype(str).str.contains("strict_merge", case=False, na=False))].copy()
    if "formulation_core_id" in b.columns and b["formulation_core_id"].astype(str).str.strip().ne("").any():
        sizes = b.groupby("formulation_core_id").size().reset_index(name="sz").sort_values(["sz", "formulation_core_id"], ascending=[False, True])
        top_ids = sizes.head(10)["formulation_core_id"].tolist()
        b = b[b["formulation_core_id"].isin(top_ids)]
    b = b.sort_values(["quality_score", "missing_polymer_identity"], ascending=[False, True]).head(n)

    c = work[(work.get("gate_used", "").astype(str).eq("A") | work.get("gate_a_anchor", "").astype(str).str.strip().ne("") | work.get("formulation_id", "").astype(str).str.upper().str.match(r"^(F|NP|BATCH|RUN)[-_]?\d+$"))].copy()
    if len(c) > n:
        c = c.sample(n=n, random_state=seed + 1)

    used = set(a.index.tolist() + b.index.tolist() + c.index.tolist())
    rem = work.loc[~work.index.isin(list(used))]
    def fill(x: pd.DataFrame, seed_off: int) -> pd.DataFrame:
        if len(x) >= n:
            return x.head(n)
        need = n - len(x)
        extra = rem.sample(n=min(need, len(rem)), random_state=seed + seed_off) if len(rem) else rem
        return pd.concat([x, extra], axis=0).head(n)

    return {"A_unresolved": fill(a, 11), "B_strict_merge": fill(b, 12), "C_anchor_merge": fill(c, 13)}


def write_xlsx(audit_df: pd.DataFrame, summary_df: pd.DataFrame, out_xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "audit_cases"
    headers = list(audit_df.columns)
    ws.append(headers)
    for _, r in audit_df.iterrows():
        ws.append([r.get(c, "") for c in headers])
    ws.freeze_panes = "A2"
    wrap_cols = {"signature_string", "signature_quality", "evidence_pointer_raw", "evidence_text", "evidence_context_before", "evidence_context_after", "table_row_text", "table_cell_text", "trace_pointer", "human_notes"}
    widths = {"title": 40, "doi": 30, "signature_string": 55, "signature_quality": 45, "evidence_text": 70, "table_row_text": 70, "human_notes": 40}
    for i, h in enumerate(headers, start=1):
        col = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col].width = widths.get(h, 18)
        if h in wrap_cols:
            for rr in range(2, ws.max_row + 1):
                ws.cell(row=rr, column=i).alignment = Alignment(wrap_text=True, vertical="top")

    ws2 = wb.create_sheet("summary")
    ws2.append(list(summary_df.columns))
    for _, r in summary_df.iterrows():
        ws2.append([r.get(c, "") for c in summary_df.columns])
    ws2.freeze_panes = "A2"
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


def maybe_write_doe_derived(run_id: str, rows: list[dict[str, Any]]) -> tuple[str, int]:
    target = paths.DATA_RESULTS_DIR / run_id / "benchmark_goren_2025" / "derived_values.tsv"
    if not rows:
        return str(target), 0
    base = pd.read_csv(target, sep="\t", dtype=str).fillna("") if target.exists() else pd.DataFrame(columns=["run_id", "group_key", "key", "formulation_id", "field_name", "value", "rule_id", "derived_from", "value_source", "trace_pointer"])
    add = pd.DataFrame(rows)
    out = pd.concat([base, add], axis=0, ignore_index=True).drop_duplicates(subset=["group_key", "field_name", "value", "rule_id"], keep="last")
    target.parent.mkdir(parents=True, exist_ok=True)
    out.to_csv(target, sep="\t", index=False)
    return str(target), len(add)


def build_doe_signature_maps(derived_df: pd.DataFrame) -> tuple[dict[str, str], dict[str, str]]:
    if derived_df.empty or "group_key" not in derived_df.columns or "field_name" not in derived_df.columns:
        return {}, {}
    out_sig: dict[str, str] = {}
    out_decoded: dict[str, str] = {}
    w = derived_df.copy()
    w["field_name"] = w["field_name"].astype(str)
    w = w[w["field_name"].str.startswith("doe_factor::")].copy()
    if w.empty:
        return {}, {}
    w["value"] = w["value"].astype(str)
    # Prefer decoded values from existing stage5 DOE decode outputs.
    for gk, g in w.groupby("group_key", sort=False):
        factor_map_dec: dict[str, str] = {}
        factor_map_coded: dict[str, str] = {}
        for _, r in g.iterrows():
            fn = str(r.get("field_name", ""))
            m = re.match(r"doe_factor::(.+?)::(decoded|coded)$", fn)
            if not m:
                continue
            factor = m.group(1).strip()
            kind = m.group(2).strip()
            val = str(r.get("value", "")).strip()
            if not factor or val == "":
                continue
            if kind == "decoded":
                factor_map_dec[factor] = val
            else:
                factor_map_coded[factor] = val
        merged = factor_map_dec if factor_map_dec else factor_map_coded
        if merged:
            parts = [f"{k}={merged[k]}" for k in sorted(merged.keys())]
            out_sig[str(gk)] = "|".join(parts)
        if factor_map_dec:
            out_decoded[str(gk)] = "; ".join([f"{k}={factor_map_dec[k]}" for k in sorted(factor_map_dec.keys())])
    return out_sig, out_decoded


def main() -> None:
    args = parse_args()
    input_tsv = Path(args.input_tsv).resolve()
    if not input_tsv.exists():
        raise FileNotFoundError(f"input TSV not found: {input_tsv}")
    run_id = str(args.run_id or "").strip()
    if not run_id:
        raise ValueError(
            "ERROR: --run-id is required. Generate/reuse a run_id via: python -m src.utils.run_preflight ..."
        )
    if not is_valid_run_id(run_id):
        raise ValueError(f"Invalid --run-id (must match required regex): {run_id}")
    out_subdir = _sanitize_out_subdir(args.out_subdir)
    run_base = paths.DATA_RESULTS_DIR / run_id / out_subdir
    out_xlsx = Path(args.out_xlsx).resolve() if args.out_xlsx.strip() else (run_base / "audit_pack" / "audit_pack__human_evidence_v1.xlsx")
    if args.out_xlsx:
        try:
            out_xlsx.resolve().relative_to(run_base.resolve())
        except Exception:
            raise ValueError(
                f"ERROR: --out-xlsx must be under data/results/<run_id>/<out-subdir>/. Got: {out_xlsx}"
            )
    out_log = out_xlsx.parent / "build_log.json"
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    source_files = [str(input_tsv)]
    schema_mismatch: dict[str, Any] = {}

    base = run_base
    sig_dir = base / "formulation_core_signature_v1"
    p_assign = sig_dir / "instance_assignment_v1.tsv"
    p_core = sig_dir / "formulation_core_v1.tsv"
    p_trace = sig_dir / "signature_trace_v1.tsv"
    p_derived = base / "benchmark_goren_2025" / "derived_values.tsv"
    p_projection = base / "benchmark_goren_2025" / "projection_trace.tsv"

    raw = pd.read_csv(input_tsv, sep="\t", dtype=str).fillna("")
    if "key" not in raw.columns:
        raise RuntimeError("input TSV must include key")
    if "formulation_id" not in raw.columns:
        raw["formulation_id"] = ""
    raw["zotero_key"] = raw["key"].astype(str)
    raw["group_key"] = raw["key"].astype(str) + "::" + raw["formulation_id"].astype(str)
    merged = raw.copy()

    if p_assign.exists():
        a = pd.read_csv(p_assign, sep="\t", dtype=str).fillna("")
        source_files.append(str(p_assign))
        a["group_key"] = a["doc_key"].astype(str) + "::" + a["formulation_id"].astype(str)
        merged = merged.merge(a.drop_duplicates("group_key"), on="group_key", how="left", suffixes=("", "_assign"))
    else:
        schema_mismatch.setdefault("missing_optional_enrichment", []).append(str(p_assign))

    if p_core.exists():
        c = pd.read_csv(p_core, sep="\t", dtype=str).fillna("")
        source_files.append(str(p_core))
        keep = [x for x in ["formulation_core_id", "signature_quality", "provenance_map_json", "drug_name_canon", "polymer_type_canon", "la_ga_ratio_canon", "polymer_mw_kda_canon_or_iv", "organic_solvent_canon", "surfactant_name_canon", "feed_anchor_canon", "n_instances"] if x in c.columns]
        merged = merged.merge(c.drop_duplicates("formulation_core_id")[keep], on="formulation_core_id", how="left") if "formulation_core_id" in merged.columns else merged
    else:
        schema_mismatch.setdefault("missing_optional_enrichment", []).append(str(p_core))

    if p_trace.exists():
        t = pd.read_csv(p_trace, sep="\t", dtype=str).fillna("")
        source_files.append(str(p_trace))
        t["group_key"] = t["doc_key"].astype(str) + "::" + t["formulation_id"].astype(str)
        keep = [x for x in ["group_key", "evidence_ref", "trace_pointer"] if x in t.columns]
        merged = merged.merge(t.drop_duplicates("group_key")[keep], on="group_key", how="left", suffixes=("", "_trace"))
        if "evidence_ref_trace" in merged.columns:
            merged["evidence_ref"] = merged["evidence_ref"].where(merged["evidence_ref"].astype(str).str.strip().ne(""), merged["evidence_ref_trace"])
            merged = merged.drop(columns=["evidence_ref_trace"])
    else:
        schema_mismatch.setdefault("missing_optional_enrichment", []).append(str(p_trace))

    d_full = pd.DataFrame()
    if p_derived.exists():
        d = pd.read_csv(p_derived, sep="\t", dtype=str).fillna("")
        d_full = d.copy()
        source_files.append(str(p_derived))
        d = d.drop_duplicates("group_key")[["group_key", "field_name", "value", "rule_id", "derived_from", "trace_pointer"]].rename(columns={"field_name": "derived_field_name", "value": "derived_value"})
        merged = merged.merge(d, on="group_key", how="left")
    else:
        schema_mismatch.setdefault("missing_optional_enrichment", []).append(str(p_derived))
    doe_sig_map, doe_decoded_map = build_doe_signature_maps(d_full)

    if p_projection.exists():
        p = pd.read_csv(p_projection, sep="\t", dtype=str).fillna("")
        source_files.append(str(p_projection))
        p = p.drop_duplicates("group_key").set_index("group_key")
        if "derived_field_name" not in merged.columns:
            merged["derived_field_name"] = ""
            merged["derived_value"] = ""
            merged["rule_id"] = ""
            merged["derived_from"] = ""
            merged["trace_pointer"] = ""
        need = merged["derived_field_name"].astype(str).str.strip().eq("")
        for idx in merged[need].index:
            gk = merged.at[idx, "group_key"]
            if gk in p.index:
                rr = p.loc[gk]
                if isinstance(rr, pd.DataFrame):
                    rr = rr.iloc[0]
                merged.at[idx, "derived_field_name"] = rr.get("curated_column", "")
                merged.at[idx, "derived_value"] = rr.get("projected_value", "")
                merged.at[idx, "rule_id"] = rr.get("rule_id", "")
                merged.at[idx, "derived_from"] = rr.get("derived_from", "")
                merged.at[idx, "trace_pointer"] = rr.get("trace_pointer", "")

    manifest, manifest_used = load_manifest_map()
    if manifest_used:
        source_files.append(manifest_used)
    merged["doi"] = merged["doi"].astype(str).map(norm_doi) if "doi" in merged.columns else ""
    merged["title"] = merged["title"] if "title" in merged.columns else ""
    merged["year"] = merged["year"] if "year" in merged.columns else ""
    if not manifest.empty:
        merged = merged.merge(manifest, on="zotero_key", how="left")
        merged["doi"] = merged["doi"].where(merged["doi"].astype(str).str.strip().ne(""), merged["doi_mf"])
        merged["title"] = merged["title"].where(merged["title"].astype(str).str.strip().ne(""), merged["title_mf"])
        merged["year"] = merged["year"].where(merged["year"].astype(str).str.strip().ne(""), merged["year_mf"])
        merged = merged.drop(columns=[c for c in ["doi_mf", "title_mf", "year_mf"] if c in merged.columns])

    merged = add_missing_flags(merged)
    for c in ["gate_used", "merge_reason", "signature_hash", "signature_string", "signature_quality", "merge_risk_level", "evidence_span_start", "evidence_span_end", "evidence_section"]:
        if c not in merged.columns:
            merged[c] = ""

    if "evidence_ref" in merged.columns:
        for idx in merged.index:
            if str(merged.at[idx, "evidence_span_start"]).strip():
                continue
            ref = str(merged.at[idx, "evidence_ref"])
            parts = ref.split("|")
            if len(parts) >= 3 and parts[1].isdigit() and parts[2].isdigit():
                merged.at[idx, "evidence_section"] = parts[0]
                merged.at[idx, "evidence_span_start"] = parts[1]
                merged.at[idx, "evidence_span_end"] = parts[2]

    missing_expected = [c for c in ["key", "formulation_id", "drug_name", "la_ga_ratio", "plga_mw_kDa", "organic_solvent", "encapsulation_efficiency_percent", "size_nm", "pdi", "evidence_span_text"] if c not in merged.columns]
    if missing_expected:
        schema_mismatch["missing_expected_columns"] = missing_expected

    buckets = choose_buckets(merged, args.n_per_bucket, args.seed)
    resolver = AuditEvidenceResolverV1(project_root=paths.PROJECT_ROOT)

    audit_rows = []
    table_missing = []
    table_trace_rows = []
    doe_papers = set()
    doe_derived_rows = []

    for bucket_name, bdf in buckets.items():
        for _, r in bdf.iterrows():
            canon = parse_json_obj(r.get("canonical_components_json", ""))
            qual = parse_json_obj(r.get("signature_quality", ""))
            pointer_raw = " | ".join([str(r.get("evidence_ref", "")), str(r.get("trace_pointer", "")), str(r.get("evidence_method", ""))]).strip(" |")
            text_ev = resolver.resolve_text_evidence(
                zotero_key=str(r.get("zotero_key", "")),
                evidence_span_start=r.get("evidence_span_start", ""),
                evidence_span_end=r.get("evidence_span_end", ""),
                evidence_section=r.get("evidence_section", ""),
                evidence_pointer_raw=pointer_raw,
                max_span_chars=args.max_span_chars,
                fallback_hint_text=str(r.get("evidence_span_text", "")),
            )
            target_vals = {
                "ee": str(r.get("encapsulation_efficiency_percent", "")),
                "size": str(r.get("size_nm", "")),
                "pdi": str(r.get("pdi", "")),
                "drug": str(r.get("drug_feed_amount_text", "")),
                "polymer": str(r.get("plga_mass_mg", "")),
                "surfactant": str(r.get("pva_conc_percent", "")),
            }
            target_field = str(r.get("derived_field_name", "")).strip() or "EE_size_PDI"
            table_ev = resolver.resolve_table_evidence(
                zotero_key=str(r.get("zotero_key", "")),
                doi=str(r.get("doi", "")),
                title=str(r.get("title", "")),
                pointer_raw=pointer_raw,
                row_index=r.get("row_index", ""),
                col_name=str(r.get("derived_field_name", "")),
                target_values=target_vals,
                field_hint=target_field,
                target_field=target_field,
                notes_hint=str(r.get("notes", "")) + " " + str(r.get("evidence_span_text", "")),
                max_table_row_chars=args.max_table_row_chars,
            )
            doe_signature = doe_sig_map.get(str(r.get("group_key", "")), "").strip()
            is_doe = resolver.detect_doe_keywords(" ".join([str(r.get("notes", "")), str(r.get("evidence_span_text", "")), pointer_raw])) or bool(doe_signature)
            if is_doe:
                doe_papers.add(str(r.get("zotero_key", "")))
            if not table_ev.table_row_text:
                table_missing.append({"zotero_key": str(r.get("zotero_key", "")), "group_key": str(r.get("group_key", "")), "reason": table_ev.table_evidence_missing_reason or "no_table_match"})
            table_trace_rows.append(
                {
                    "zotero_key": str(r.get("zotero_key", "")),
                    "target_field": target_field,
                    "chosen_table_filename": table_ev.table_filename,
                    "chosen_score": table_ev.table_match_score,
                    "paper_local_candidate_count": int(table_ev.paper_local_candidate_count),
                    "ownership_check_passed": bool(table_ev.ownership_check_passed),
                    "ownership_check_reason": str(table_ev.ownership_check_reason),
                    "chosen_table_rejected": bool(table_ev.chosen_table_rejected),
                    "top5_candidates": "|".join(table_ev.top5_candidates),
                    "top5_scores": "|".join([str(x) for x in table_ev.top5_scores]),
                    "match_reason": table_ev.match_reason or table_ev.table_evidence_missing_reason,
                }
            )

            polymer_type = str(canon.get("polymer_type_canon", "")) or canonical_polymer_type(str(r.get("polymer_name", "")), str(r.get("notes", "")), str(r.get("evidence_span_text", "")))
            mw_canon = str(canon.get("polymer_mw_kda_canon_or_iv", "")) or canonical_mw_or_iv(str(r.get("plga_mw_kDa", "")), str(r.get("notes", "")))
            feed_anchor = str(canon.get("feed_anchor_canon", r.get("feed_anchor_canon", "")))
            if doe_signature:
                feed_anchor = f"doe_signature:{doe_signature}"

            table_evidence_kind = detect_table_evidence_kind(
                table_csv_path=table_ev.table_csv_path,
                table_filename=table_ev.table_filename,
                table_row_text=table_ev.table_row_text,
            )
            if not bool(table_ev.ownership_check_passed) and table_evidence_kind == "table_csv_cell":
                table_evidence_kind = "proxy_compose" if str(table_ev.table_row_text).strip() else "none"
            if table_evidence_kind == "table_csv_cell" and bool(table_ev.ownership_check_passed):
                table_selection_status = "accepted"
            elif bool(table_ev.chosen_table_rejected):
                table_selection_status = "rejected"
            elif table_evidence_kind == "proxy_compose":
                table_selection_status = "proxy"
            else:
                table_selection_status = "none"
            proxy_components: dict[str, str] = {}
            if table_evidence_kind == "proxy_compose":
                proxy_components = {
                    "EE": "extraction_field:encapsulation_efficiency_percent" if looks_nonempty(r.get("encapsulation_efficiency_percent", "")) else "not_available",
                    "size": "extraction_field:size_nm" if looks_nonempty(r.get("size_nm", "")) else "not_available",
                    "drug_mass": "extraction_field:drug_feed_amount_text" if looks_nonempty(r.get("drug_feed_amount_text", "")) else "not_available",
                    "polymer_mass": "extraction_field:plga_mass_mg" if looks_nonempty(r.get("plga_mass_mg", "")) else "not_available",
                    "doe_signature": (
                        "derived_doe_decode:derived_values.tsv"
                        if str(r.get("group_key", "")) in doe_sig_map
                        else "not_available"
                    ),
                }

            value_source_ee = infer_value_source_numeric(
                field_value=str(r.get("encapsulation_efficiency_percent", "")),
                table_evidence_kind=table_evidence_kind,
                evidence_text=text_ev.evidence_text,
                derived_rule_id="R_DIRECT_EE" if looks_nonempty(r.get("encapsulation_efficiency_percent", "")) else "",
            )
            value_source_size = infer_value_source_numeric(
                field_value=str(r.get("size_nm", "")),
                table_evidence_kind=table_evidence_kind,
                evidence_text=text_ev.evidence_text,
                derived_rule_id="R_DIRECT_PARTICLE_SIZE" if looks_nonempty(r.get("size_nm", "")) else "",
            )
            value_source_drug_mass = infer_value_source_numeric(
                field_value=str(r.get("drug_feed_amount_text", "")),
                table_evidence_kind=table_evidence_kind,
                evidence_text=text_ev.evidence_text,
                derived_rule_id="R_DRUG_MASS_PARSE" if looks_nonempty(r.get("drug_feed_amount_text", "")) else "",
            )
            value_source_polymer_mass = infer_value_source_numeric(
                field_value=str(r.get("plga_mass_mg", "")),
                table_evidence_kind=table_evidence_kind,
                evidence_text=text_ev.evidence_text,
                derived_rule_id="R_POLYMER_MASS_PARSE" if looks_nonempty(r.get("plga_mass_mg", "")) else "",
            )
            if doe_signature:
                value_source_doe_signature = "derived_doe_decode"
            elif is_doe:
                value_source_doe_signature = "unknown"
            else:
                value_source_doe_signature = "unknown"

            why_no_merge = ""
            if str(r.get("gate_used", "")) == "C" or "no_auto_merge" in str(r.get("merge_reason", "")):
                why_no_merge = "core_incomplete_no_anchor" if bool(r.get("missing_polymer_identity", False)) else "anchor_missing_or_low_quality"

            row_payload = {
                "bucket": bucket_name,
                "zotero_key": str(r.get("zotero_key", "")),
                "doi": str(r.get("doi", "")),
                "title": str(r.get("title", "")),
                "year": str(r.get("year", "")),
                "formulation_core_id": str(r.get("formulation_core_id", "")),
                "gate_used": str(r.get("gate_used", "")),
                "merge_reason": str(r.get("merge_reason", "")),
                "why_no_merge": why_no_merge,
                "signature_hash": str(r.get("signature_hash", "")),
                "signature_string": str(r.get("signature_string", "")),
                "signature_quality": json.dumps(qual, ensure_ascii=False, sort_keys=True) if qual else str(r.get("signature_quality", "")),
                "merge_risk_level": str(r.get("merge_risk_level", "")),
                "drug_name_raw": str(r.get("drug_name", "")),
                "polymer_name_raw": str(r.get("polymer_name", "")),
                "la_ga_ratio_raw": str(r.get("la_ga_ratio", "")),
                "mw_raw": str(r.get("plga_mw_kDa", "")),
                "polymer_code_raw/vendor_code_raw": str(r.get("vendor_product_code", "")),
                "solvent_raw": str(r.get("organic_solvent", "")),
                "surfactant_raw": str(r.get("surfactant_name", "")),
                "feed_raw": str(r.get("drug_feed_amount_text", "")),
                "EE_raw": str(r.get("encapsulation_efficiency_percent", "")),
                "size_raw": str(r.get("size_nm", "")),
                "pdi_raw": str(r.get("pdi", "")),
                "drug_name_canon": str(canon.get("drug_name_canon", r.get("drug_name_canon", ""))),
                "polymer_type_canon": polymer_type,
                "la_ga_ratio_canon": str(canon.get("la_ga_ratio_canon", r.get("la_ga_ratio_canon", ""))),
                "mw_kda_canon_or_iv_canon": mw_canon,
                "solvent_canon": str(canon.get("organic_solvent_canon", r.get("organic_solvent_canon", ""))),
                "surfactant_canon": str(canon.get("surfactant_name_canon", r.get("surfactant_name_canon", ""))),
                "feed_anchor_canon": feed_anchor,
                "missing_drug": bool(r.get("missing_drug", False)),
                "missing_polymer_identity": bool(r.get("missing_polymer_identity", False)),
                "missing_solvent": bool(r.get("missing_solvent", False)),
                "missing_surfactant": bool(r.get("missing_surfactant", False)),
                "missing_feed_anchor": bool(r.get("missing_feed_anchor", False)),
                "evidence_source_type": (
                    "table"
                    if table_evidence_kind == "table_csv_cell"
                    else ("proxy_compose" if table_evidence_kind == "proxy_compose" else text_ev.evidence_source_type)
                ),
                "evidence_pointer_raw": pointer_raw,
                "evidence_text": text_ev.evidence_text,
                "evidence_context_before": text_ev.evidence_context_before,
                "evidence_context_after": text_ev.evidence_context_after,
                "table_csv_path": table_ev.table_csv_path,
                "table_filename": table_ev.table_filename,
                "rejected_table_filename": table_ev.rejected_table_filename,
                "table_title_or_caption": table_ev.table_title_or_caption,
                "table_match_score": table_ev.table_match_score,
                "table_evidence_kind": table_evidence_kind,
                "ownership_check_passed": bool(table_ev.ownership_check_passed),
                "ownership_check_reason": str(table_ev.ownership_check_reason),
                "table_selection_status": table_selection_status,
                "table_row_text": table_ev.table_row_text,
                "table_cell_text": table_ev.table_cell_text,
                "proxy_components_json": json.dumps(proxy_components, ensure_ascii=False, sort_keys=True) if proxy_components else "",
                "evidence_block_id": text_ev.evidence_block_id,
                "evidence_span_id": text_ev.evidence_span_id,
                "evidence_span_start": text_ev.evidence_span_start,
                "evidence_span_end": text_ev.evidence_span_end,
                "evidence_section": text_ev.evidence_section,
                "derived_field_name": str(r.get("derived_field_name", "")),
                "derived_value": str(r.get("derived_value", "")),
                "rule_id": str(r.get("rule_id", "")),
                "derived_from": str(r.get("derived_from", "")),
                "trace_pointer": str(r.get("trace_pointer", "")),
                "doe_signature": doe_signature,
                "decoded_concentration_values": doe_decoded_map.get(str(r.get("group_key", "")), ""),
                "value_source_EE": value_source_ee if value_source_ee in VALUE_SOURCES else "unknown",
                "value_source_size": value_source_size if value_source_size in VALUE_SOURCES else "unknown",
                "value_source_drug_mass": value_source_drug_mass if value_source_drug_mass in VALUE_SOURCES else "unknown",
                "value_source_polymer_mass": value_source_polymer_mass if value_source_polymer_mass in VALUE_SOURCES else "unknown",
                "value_source_doe_signature": value_source_doe_signature if value_source_doe_signature in VALUE_SOURCES else "unknown",
                "human_review_tag": "",
                "human_notes": "",
            }
            row_payload, _ = apply_provenance_hard_guards(row_payload)
            audit_rows.append(row_payload)

    audit_df = pd.DataFrame(audit_rows)
    if audit_df.empty:
        audit_df = pd.DataFrame(columns=["bucket", "zotero_key", "doi"])

    derived_path, n_doe_written = maybe_write_doe_derived(run_id, doe_derived_rows)
    if n_doe_written:
        source_files.append(derived_path)

    bucket_counts = audit_df["bucket"].value_counts().rename_axis("bucket").reset_index(name="n_rows")
    table_rate = float(audit_df["table_row_text"].astype(str).str.strip().ne("").mean() * 100.0) if len(audit_df) else 0.0
    text_rate = float(audit_df["evidence_text"].astype(str).str.strip().ne("").mean() * 100.0) if len(audit_df) else 0.0
    doe_rate = float(audit_df["doe_signature"].astype(str).str.strip().ne("").mean() * 100.0) if len(audit_df) else 0.0
    kind_pct = (
        audit_df["table_evidence_kind"].astype(str).value_counts(normalize=True).mul(100.0).to_dict()
        if "table_evidence_kind" in audit_df.columns and len(audit_df)
        else {}
    )

    summary_rows = []
    for b, g in audit_df.groupby("bucket", dropna=False):
        summary_rows.append({"metric": "bucket_count", "bucket": b, "value": int(len(g))})
        for f in MISSING_FLAGS:
            summary_rows.append({"metric": f"{f}_rate_percent", "bucket": b, "value": round(float(g[f].astype(bool).mean() * 100.0), 2)})
    top_merge = audit_df["merge_reason"].astype(str).value_counts().head(20).rename_axis("merge_reason").reset_index(name="count")
    for _, rr in top_merge.iterrows():
        summary_rows.append({"metric": f"merge_reason::{rr['merge_reason']}", "bucket": "ALL", "value": int(rr['count'])})
    summary_rows += [
        {"metric": "table_availability_rate_percent", "bucket": "ALL", "value": round(table_rate, 2)},
        {"metric": "text_evidence_availability_rate_percent", "bucket": "ALL", "value": round(text_rate, 2)},
        {"metric": "doe_signature_rate_percent", "bucket": "ALL", "value": round(doe_rate, 2)},
        {"metric": "table_evidence_kind::table_csv_cell_percent", "bucket": "ALL", "value": round(float(kind_pct.get("table_csv_cell", 0.0)), 2)},
        {"metric": "table_evidence_kind::proxy_compose_percent", "bucket": "ALL", "value": round(float(kind_pct.get("proxy_compose", 0.0)), 2)},
        {"metric": "table_evidence_kind::none_percent", "bucket": "ALL", "value": round(float(kind_pct.get("none", 0.0)), 2)},
    ]
    summary_df = pd.DataFrame(summary_rows)

    diag = audit_df.copy()
    diag["diag_table_path_cross_key"] = diag.apply(
        lambda x: (
            str(x.get("evidence_source_type", "")).strip() == "table"
            and is_cross_paper_table_path(str(x.get("zotero_key", "")), str(x.get("table_csv_path", "")))
        ),
        axis=1,
    )
    diag["diag_table_filename_cross_key"] = diag.apply(
        lambda x: (
            str(x.get("evidence_source_type", "")).strip() == "table"
            and str(x.get("table_filename", "")).strip() != ""
            and not str(x.get("table_filename", "")).strip().startswith(str(x.get("zotero_key", "")).strip())
        ),
        axis=1,
    )
    diag["diag_mixed_source_pointer"] = diag.apply(
        lambda x: is_mixed_table_fulltext_pointer(
            evidence_source_type=str(x.get("evidence_source_type", "")),
            evidence_pointer_raw=str(x.get("evidence_pointer_raw", "")),
        ),
        axis=1,
    )
    diag["diag_value_source_conflict"] = diag.apply(
        lambda x: (
            str(x.get("evidence_pointer_raw", "")).strip().lower().startswith("fulltext|")
            and any(
                str(x.get(c, "")).strip() == "table_csv_cell"
                for c in [
                    "value_source_EE",
                    "value_source_size",
                    "value_source_drug_mass",
                    "value_source_polymer_mass",
                ]
            )
        ),
        axis=1,
    )
    diag_filtered = diag[
        diag["diag_table_path_cross_key"]
        | diag["diag_table_filename_cross_key"]
        | diag["diag_mixed_source_pointer"]
        | diag["diag_value_source_conflict"]
    ].copy()
    diag_path = run_base / "step1_dev_provenance_binding_diagnostics.tsv"
    diag_cols = [
        "zotero_key",
        "formulation_core_id",
        "evidence_source_type",
        "evidence_pointer_raw",
        "table_filename",
        "table_csv_path",
        "table_selection_status",
        "ownership_check_passed",
        "ownership_check_reason",
        "human_review_tag",
        "value_source_EE",
        "value_source_size",
        "value_source_drug_mass",
        "value_source_polymer_mass",
        "diag_table_path_cross_key",
        "diag_table_filename_cross_key",
        "diag_mixed_source_pointer",
        "diag_value_source_conflict",
    ]
    for c in diag_cols:
        if c not in diag_filtered.columns:
            diag_filtered[c] = ""
    diag_filtered[diag_cols].to_csv(diag_path, sep="\t", index=False)

    write_xlsx(audit_df, summary_df, out_xlsx)
    table_trace_path = out_xlsx.parent / "table_match_trace_v1.tsv"
    pd.DataFrame(
        table_trace_rows,
        columns=[
            "zotero_key",
            "target_field",
            "chosen_table_filename",
            "chosen_score",
            "paper_local_candidate_count",
            "ownership_check_passed",
            "ownership_check_reason",
            "chosen_table_rejected",
            "top5_candidates",
            "top5_scores",
            "match_reason",
        ],
    ).to_csv(table_trace_path, sep="\t", index=False)

    log = {
        "run_id": run_id,
        "input_tsv": str(input_tsv),
        "output_xlsx": str(out_xlsx),
        "table_match_trace_tsv": str(table_trace_path),
        "step1_dev_provenance_diagnostics_tsv": str(diag_path),
        "source_files": sorted(set(source_files)),
        "n_rows_audit": int(len(audit_df)),
        "n_rows_step1_dev_provenance_diagnostics": int(len(diag_filtered)),
        "bucket_counts": bucket_counts.to_dict(orient="records"),
        "table_row_text_non_empty_rate_percent": round(table_rate, 2),
        "evidence_text_non_empty_rate_percent": round(text_rate, 2),
        "doe_signature_non_empty_rate_percent": round(doe_rate, 2),
        "n_doe_detected_papers": int(len(doe_papers)),
        "n_doe_derived_written": int(n_doe_written),
        "table_evidence_kind_percent": {
            "table_csv_cell": round(float(kind_pct.get("table_csv_cell", 0.0)), 2),
            "proxy_compose": round(float(kind_pct.get("proxy_compose", 0.0)), 2),
            "none": round(float(kind_pct.get("none", 0.0)), 2),
        },
        "ownership_check_summary": {
            "table_csv_cell_and_ownership_true": int(
                ((audit_df["table_evidence_kind"].astype(str) == "table_csv_cell") & (audit_df["ownership_check_passed"].astype(bool))).sum()
            ) if len(audit_df) else 0,
            "ownership_false_total": int((~audit_df["ownership_check_passed"].astype(bool)).sum()) if len(audit_df) else 0,
            "ownership_false_downgraded_proxy": int(
                ((~audit_df["ownership_check_passed"].astype(bool)) & (audit_df["table_evidence_kind"].astype(str) == "proxy_compose")).sum()
            ) if len(audit_df) else 0,
            "ownership_false_downgraded_none": int(
                ((~audit_df["ownership_check_passed"].astype(bool)) & (audit_df["table_evidence_kind"].astype(str) == "none")).sum()
            ) if len(audit_df) else 0,
        },
        "table_evidence_missing": {"count": int(len(table_missing)), "examples": table_missing[:20]},
        "audit_pack_schema_mismatch": schema_mismatch,
        "params": {"n_per_bucket": int(args.n_per_bucket), "seed": int(args.seed), "max_span_chars": int(args.max_span_chars), "max_table_row_chars": int(args.max_table_row_chars)},
    }
    out_log.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")

    n_instances = int(len(merged))
    n_cores = int(merged["formulation_core_id"].astype(str).replace("", pd.NA).dropna().nunique()) if "formulation_core_id" in merged.columns else 0
    print(f"n_instances={n_instances}")
    print(f"n_cores={n_cores}")
    print(f"doe_detected_papers={len(doe_papers)}")
    print(f"table_row_text_non_empty_percent={round(table_rate, 2)}")
    print(f"doe_signature_non_empty_percent={round(doe_rate, 2)}")
    print(f"table_evidence_kind_table_csv_cell_percent={round(float(kind_pct.get('table_csv_cell', 0.0)), 2)}")
    print(f"table_evidence_kind_proxy_compose_percent={round(float(kind_pct.get('proxy_compose', 0.0)), 2)}")
    print(f"table_evidence_kind_none_percent={round(float(kind_pct.get('none', 0.0)), 2)}")
    print(f"step1_dev_provenance_diagnostics_rows={len(diag_filtered)}")
    print(f"step1_dev_provenance_diagnostics_tsv={diag_path}")
    print(f"excel_path={out_xlsx}")

    own_true_table = int(
        ((audit_df["table_evidence_kind"].astype(str) == "table_csv_cell") & (audit_df["ownership_check_passed"].astype(bool))).sum()
    )
    own_false = int((~audit_df["ownership_check_passed"].astype(bool)).sum())
    own_false_proxy = int(((~audit_df["ownership_check_passed"].astype(bool)) & (audit_df["table_evidence_kind"].astype(str) == "proxy_compose")).sum())
    own_false_none = int(((~audit_df["ownership_check_passed"].astype(bool)) & (audit_df["table_evidence_kind"].astype(str) == "none")).sum())
    print("\n[ownership_diagnostics]")
    print(f"table_csv_cell_and_ownership_true={own_true_table}")
    print(f"ownership_false_total={own_false}")
    print(f"ownership_false_downgraded_proxy={own_false_proxy}")
    print(f"ownership_false_downgraded_none={own_false_none}")

    yga = audit_df[audit_df["zotero_key"].astype(str) == "YGA8VQKU"].head(1)
    print("\n[YGA8VQKU_check]")
    if yga.empty:
        print("(empty)")
    else:
        yr = yga.iloc[0]
        prev_name = str(yr.get("rejected_table_filename", "") or yr.get("table_filename", ""))
        print(
            "zotero_key=YGA8VQKU"
            + f"\tprevious_chosen_table_filename={prev_name}"
            + f"\townership_check_passed={yr.get('ownership_check_passed', False)}"
            + f"\ttable_evidence_kind={yr.get('table_evidence_kind', '')}"
            + f"\townership_check_reason={yr.get('ownership_check_reason', '')}"
        )

    accepted_rows = audit_df[
        (audit_df["table_evidence_kind"].astype(str) == "table_csv_cell")
        & (audit_df["ownership_check_passed"].astype(bool))
    ].copy()
    print("\n[accepted_table_rows_sample_2]")
    if accepted_rows.empty:
        print("(empty)")
    else:
        s2 = accepted_rows.sample(n=min(2, len(accepted_rows)), random_state=42)
        print(s2[["zotero_key", "table_filename", "table_csv_path"]].to_string(index=False))


if __name__ == "__main__":
    main()
