#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment


def _bootstrap_import_paths() -> None:
    here = Path(__file__).resolve()
    for p in [here] + list(here.parents):
        if (p / "src" / "utils" / "paths.py").exists():
            sys.path.insert(0, str(p))
            return


_bootstrap_import_paths()
from src.utils import paths  # noqa: E402


RUN_ID = "run_20260219_1623_780eb83_goren18_weaklabels_v1"


def short_text(v: Any, n: int) -> str:
    s = "" if v is None else str(v)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:n]


def parse_evidence_ref(v: Any) -> dict[str, str]:
    s = str(v or "")
    parts = s.split("|")
    out = {
        "evidence_section": parts[0] if len(parts) > 0 else "",
        "evidence_span_start": parts[1] if len(parts) > 1 else "",
        "evidence_span_end": parts[2] if len(parts) > 2 else "",
        "evidence_span_text_short": "",
    }
    return out


def load_manifest() -> pd.DataFrame:
    for p in [
        paths.DATA_CLEANED_INDEX_DIR / "manifest_goren_2025.tsv",
        paths.DATA_CLEANED_INDEX_DIR / "manifest_current.tsv",
    ]:
        if not p.exists():
            continue
        df = pd.read_csv(p, sep="\t", dtype=str).fillna("")
        key_col = "key" if "key" in df.columns else ("zotero_key" if "zotero_key" in df.columns else "")
        if not key_col:
            continue
        out = pd.DataFrame(
            {
                "zotero_key": df[key_col].astype(str),
                "doi": df["doi"].astype(str) if "doi" in df.columns else "",
                "title": df["title"].astype(str) if "title" in df.columns else "",
            }
        )
        return out.drop_duplicates("zotero_key")
    return pd.DataFrame(columns=["zotero_key", "doi", "title"])


def write_sheet(ws, df: pd.DataFrame, wrap_cols: set[str], widths: dict[str, int]) -> None:
    ws.append(list(df.columns))
    for _, r in df.iterrows():
        ws.append([r.get(c, "") for c in df.columns])
    ws.freeze_panes = "A2"
    for i, col in enumerate(df.columns, start=1):
        letter = ws.cell(1, i).column_letter
        ws.column_dimensions[letter].width = widths.get(col, 18)
        if col in wrap_cols:
            for rr in range(2, ws.max_row + 1):
                ws.cell(rr, i).alignment = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    base = paths.DATA_RESULTS_DIR / RUN_ID / "formulation_core_signature_v1"
    p_core = base / "formulation_core_v1.tsv"
    p_assign = base / "instance_assignment_v1.tsv"
    p_trace = base / "signature_trace_v1.tsv"
    p_derived = paths.DATA_RESULTS_DIR / RUN_ID / "benchmark_goren_2025" / "derived_values.tsv"
    out_xlsx = base / "audit_pack__doe_signature_injection_v1.xlsx"

    core = pd.read_csv(p_core, sep="\t", dtype=str).fillna("")
    assign = pd.read_csv(p_assign, sep="\t", dtype=str).fillna("")
    trace = pd.read_csv(p_trace, sep="\t", dtype=str).fillna("")
    derived = pd.read_csv(p_derived, sep="\t", dtype=str).fillna("") if p_derived.exists() else pd.DataFrame()
    manifest = load_manifest()

    core_keep = [
        "formulation_core_id",
        "drug_name_canon",
        "polymer_type_canon",
        "la_ga_ratio_canon",
        "polymer_mw_kda_canon_or_iv",
        "organic_solvent_canon",
        "surfactant_name_canon",
        "feed_anchor_canon",
    ]
    merged = assign.merge(core[core_keep].drop_duplicates("formulation_core_id"), on="formulation_core_id", how="left")
    merged = merged.merge(
        trace[["instance_id", "evidence_ref"]].drop_duplicates("instance_id"),
        on="instance_id",
        how="left",
        suffixes=("", "_trace"),
    )
    if "evidence_ref_trace" in merged.columns:
        merged["evidence_ref"] = merged["evidence_ref"].where(merged["evidence_ref"].astype(str).str.strip().ne(""), merged["evidence_ref_trace"])
        merged = merged.drop(columns=["evidence_ref_trace"])
    merged["zotero_key"] = merged["doc_key"].astype(str)
    merged = merged.merge(manifest, on="zotero_key", how="left")

    ev = merged["evidence_ref"].map(parse_evidence_ref).apply(pd.Series)
    merged = pd.concat([merged, ev], axis=1)

    doe_rows = merged[merged["doe_signature_source"].astype(str).eq("derived_doe_decode")].copy()
    doe_rows["formulation_instance_id"] = doe_rows["formulation_id"].astype(str)
    empty_id = doe_rows["formulation_instance_id"].str.strip().eq("")
    doe_rows.loc[empty_id, "formulation_instance_id"] = doe_rows.loc[empty_id, "instance_id"]

    col_map = {
        "polymer_mw_kda_canon_or_iv": "mw_kda_canon_or_iv_canon",
        "organic_solvent_canon": "solvent_canon",
        "surfactant_name_canon": "surfactant_canon",
    }
    doe_rows = doe_rows.rename(columns=col_map)
    doe_rows["signature_string"] = doe_rows["signature_string"].map(lambda x: short_text(x, 300))
    doe_rows["human_review_tag"] = ""
    doe_rows["human_notes"] = ""

    optional_value_cols = [c for c in doe_rows.columns if c.startswith("value_source_")]
    ordered_cols = [
        "zotero_key",
        "doi",
        "title",
        "formulation_instance_id",
        "formulation_core_id",
        "doe_signature_canon",
        "doe_signature_source",
        "signature_string",
        "drug_name_canon",
        "polymer_type_canon",
        "la_ga_ratio_canon",
        "mw_kda_canon_or_iv_canon",
        "solvent_canon",
        "surfactant_canon",
        "feed_anchor_canon",
        "merge_reason",
        "gate_used",
        "evidence_section",
        "evidence_span_start",
        "evidence_span_end",
        "evidence_span_text_short",
    ] + optional_value_cols + ["human_review_tag", "human_notes"]
    ordered_cols = [c for c in ordered_cols if c in doe_rows.columns]
    doe_rows = doe_rows[ordered_cols]

    if not doe_rows.empty:
        core_summary = (
            doe_rows.groupby("doe_signature_canon", dropna=False)
            .agg(
                n_instances=("formulation_core_id", "size"),
                n_cores=("formulation_core_id", "nunique"),
                zotero_key_count=("zotero_key", "nunique"),
                example_formulation_ids=("formulation_instance_id", lambda s: "|".join(sorted(set([x for x in s.astype(str).tolist() if x]))[:5])),
            )
            .reset_index()
            .sort_values(["n_instances", "doe_signature_canon"], ascending=[False, True])
        )
    else:
        core_summary = pd.DataFrame(columns=["doe_signature_canon", "n_instances", "n_cores", "zotero_key_count", "example_formulation_ids"])

    core_sizes = assign.groupby("formulation_core_id").size().reset_index(name="core_size")
    core_has_doe = assign.groupby("formulation_core_id")["doe_signature_canon"].apply(lambda s: int(s.astype(str).str.strip().ne("").any())).reset_index(name="has_doe")
    core_size_df = core_sizes.merge(core_has_doe, on="formulation_core_id", how="left")
    core_size_df["group"] = core_size_df["has_doe"].map({1: "DoE", 0: "Non-DoE"})
    dist = (
        core_size_df.groupby(["group", "core_size"]).size().reset_index(name="n_cores").sort_values(["group", "core_size"])
    )
    non_doe_rows = [
        {"metric": "total_rows", "value": int(len(assign))},
        {"metric": "rows_with_doe_signature", "value": int(assign["doe_signature_canon"].astype(str).str.strip().ne("").sum())},
        {"metric": "total_cores", "value": int(assign["formulation_core_id"].astype(str).nunique())},
        {"metric": "cores_with_doe_signature", "value": int(core_has_doe["has_doe"].astype(int).sum())},
    ]
    non_doe_summary = pd.DataFrame(non_doe_rows)
    if not dist.empty:
        dist_rows = dist.rename(columns={"group": "metric", "core_size": "value", "n_cores": "n_cores_count"})
    else:
        dist_rows = pd.DataFrame(columns=["metric", "value", "n_cores_count"])

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "doe_rows"
    write_sheet(
        ws1,
        doe_rows,
        wrap_cols={"signature_string", "evidence_span_text_short", "human_notes"},
        widths={"title": 40, "signature_string": 60, "doe_signature_canon": 35, "human_notes": 40},
    )
    ws2 = wb.create_sheet("doe_core_summary")
    write_sheet(ws2, core_summary, wrap_cols={"example_formulation_ids"}, widths={"doe_signature_canon": 45, "example_formulation_ids": 40})
    ws3 = wb.create_sheet("non_doe_summary")
    ws3.append(["metric", "value", "n_cores_count"])
    for _, r in non_doe_summary.iterrows():
        ws3.append([r.get("metric", ""), r.get("value", ""), ""])
    ws3.append(["", "", ""])
    ws3.append(["distribution_group", "core_size", "n_cores"])
    for _, r in dist_rows.iterrows():
        ws3.append([r.get("metric", ""), r.get("value", ""), r.get("n_cores_count", "")])
    ws3.freeze_panes = "A2"
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 16

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)

    top10 = core_summary.head(10)
    print(f"output_xlsx={out_xlsx}")
    print("[top10_doe_signature_by_n_instances]")
    if top10.empty:
        print("(empty)")
    else:
        print(top10[["doe_signature_canon", "n_instances"]].to_string(index=False))
    print("[sample_doe_rows_5]")
    sample = doe_rows.head(5).copy()
    if sample.empty:
        print("(empty)")
    else:
        sample["doe_signature_canon"] = sample["doe_signature_canon"].map(lambda x: short_text(x, 80))
        sample["signature_string"] = sample["signature_string"].map(lambda x: short_text(x, 120))
        keep = ["zotero_key", "formulation_instance_id", "formulation_core_id", "doe_signature_canon", "signature_string"]
        keep = [c for c in keep if c in sample.columns]
        print(sample[keep].to_string(index=False))

    # Small reference print to show derived source availability.
    if not derived.empty:
        n_doe_trace = int(
            derived["field_name"].astype(str).str.contains(r"doe_factor", case=False, regex=True, na=False).sum()
        )
        print(f"derived_doe_factor_rows={n_doe_trace}")


if __name__ == "__main__":
    main()

