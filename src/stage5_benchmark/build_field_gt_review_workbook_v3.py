#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from src.stage5_benchmark import build_field_gt_review_workbook_v1 as v1
except ModuleNotFoundError:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from src.stage5_benchmark import build_field_gt_review_workbook_v1 as v1


WORKBOOK_COLUMNS = [
    "paper_key", "doi", "final_formulation_id", "formulation_label_stage5", "formulation_label_params",
    "family_id", "parent_core_row_id", "variant_role", "payload_state", "field_name",
    "predicted_value", "predicted_unit", "evidence_text", "evidence_source_type",
    "evidence_support_status", "field_source_type", "derivation_status",
    "gt_decision", "gt_value", "gt_unit", "gt_notes",
]

SEED_COLUMNS = [
    "paper_key", "doi", "paper_title", "final_formulation_id", "representative_source_formulation_id",
    "representative_source_raw_formulation_label", "formulation_label_stage5", "formulation_label_params",
    "family_id", "parent_core_row_id", "variant_role", "payload_state", "field_name",
    "source_value_column", "source_value_text_column", "source_evidence_region_column",
    "predicted_value", "predicted_unit", "evidence_text", "evidence_source_type",
    "evidence_support_status", "field_source_type", "relation_resolved_flag", "relation_fill_count",
    "relation_fill_resolution_rules", "relation_fill_scope_types", "relation_fill_source_relation_row_ids",
    "suspicious_relation_flag", "suspicious_relation_reasons", "suspicious_relation_details",
    "derivation_status", "derivation_rule", "derivation_inputs", "evidence_section",
    "evidence_span_start", "evidence_span_end", "final_output_rule", "review_needed",
    "benchmark_default_include", "relation_record_count", "relation_graph_ids",
    "relation_method_group_ids", "decision_trace_source_row_count", "decision_trace_decisions",
    "decision_trace_decision_rules", "decision_trace_review_needed", "decision_trace_notes",
    "gt_decision", "gt_value", "gt_unit", "gt_notes",
]

RELATION_FILL_FIELD_ALIASES = {
    "polymer_MW": {"polymer_mw_kDa", "plga_mw_kDa", "polymer_MW"},
    "LA/GA": {"la_ga_ratio", "LA/GA"},
    "drug_polymer_ratio": {"drug_polymer_ratio"},
    "surfactant_name": {"surfactant_name"},
    "surfactant_concentration": {"surfactant_concentration_text", "surfactant_concentration"},
    "solvent": {"organic_solvent", "solvent"},
    "particle_size": {"size_nm", "particle_size"},
    "EE": {"encapsulation_efficiency_percent", "EE"},
    "LC": {"loading_content_percent", "LC"},
    "preparation_method": {"preparation_method"},
    "emulsion_structure": {"emulsion_structure"},
}


def t(value: Any) -> str:
    return v1.normalize_text(value)


def versioned_name(prefix: str, version: int, suffix: str) -> str:
    return f"{prefix}_v{int(version)}{suffix}"


def load_decision_trace_map(path: Path | None) -> dict[str, dict[str, str]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in (v1.read_tsv_rows(path) if path and path.exists() else []):
        final_id = t(row.get("target_final_formulation_id"))
        if not final_id:
            continue
        bucket = grouped.setdefault(final_id, defaultdict(set))
        bucket["count"] = int(bucket.get("count", 0)) + 1
        for out_key, row_key in [
            ("decision_trace_decisions", "decision"),
            ("decision_trace_decision_rules", "decision_rule"),
            ("decision_trace_review_needed", "review_needed"),
            ("decision_trace_notes", "notes"),
        ]:
            value = t(row.get(row_key))
            if value:
                bucket[out_key].add(value)
    return {
        final_id: {
            "decision_trace_source_row_count": str(bucket.get("count", 0)),
            "decision_trace_decisions": "; ".join(sorted(bucket.get("decision_trace_decisions", set()))),
            "decision_trace_decision_rules": "; ".join(sorted(bucket.get("decision_trace_decision_rules", set()))),
            "decision_trace_review_needed": "; ".join(sorted(bucket.get("decision_trace_review_needed", set()))),
            "decision_trace_notes": " | ".join(sorted(bucket.get("decision_trace_notes", set()))),
        }
        for final_id, bucket in grouped.items()
    }


def load_relation_fill_map(path: Path | None) -> dict[tuple[str, str], dict[str, str]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in (v1.read_tsv_rows(path) if path and path.exists() else []):
        key = (t(row.get("final_formulation_id")), t(row.get("field_name")))
        if not key[0] or not key[1]:
            continue
        bucket = grouped.setdefault(key, defaultdict(set))
        bucket["count"] = int(bucket.get("count", 0)) + 1
        for out_key, row_key in [
            ("rules", "resolution_rule"),
            ("scopes", "scope_type"),
        ]:
            value = t(row.get(row_key))
            if value:
                bucket[out_key].add(value)
        raw_ids = t(row.get("source_relation_row_ids"))
        if raw_ids:
            try:
                parsed = json.loads(raw_ids)
            except json.JSONDecodeError:
                parsed = []
            if isinstance(parsed, list):
                for item in parsed:
                    value = t(item)
                    if value:
                        bucket["row_ids"].add(value)
    return {
        key: {
            "relation_fill_count": str(bucket.get("count", 0)),
            "relation_fill_resolution_rules": "; ".join(sorted(bucket.get("rules", set()))),
            "relation_fill_scope_types": "; ".join(sorted(bucket.get("scopes", set()))),
            "relation_fill_source_relation_row_ids": json.dumps(sorted(bucket.get("row_ids", set())), ensure_ascii=True),
        }
        for key, bucket in grouped.items()
    }


def load_suspicious_maps(path: Path | None) -> tuple[dict[tuple[str, str], dict[str, str]], dict[str, int]]:
    grouped: dict[tuple[str, str], dict[str, set[str]]] = {}
    per_paper: dict[str, int] = defaultdict(int)
    for row in (v1.read_tsv_rows(path) if path and path.exists() else []):
        paper_key = t(row.get("paper_key"))
        final_id = t(row.get("final_formulation_id"))
        field_name = t(row.get("field_name"))
        if paper_key:
            per_paper[paper_key] += 1
        if not final_id:
            continue
        bucket = grouped.setdefault((final_id, field_name), {"reasons": set(), "details": set()})
        for out_key, row_key in [("reasons", "reason"), ("details", "details")]:
            value = t(row.get(row_key))
            if value:
                bucket[out_key].add(value)
    return (
        {
            key: {
                "suspicious_relation_flag": "yes",
                "suspicious_relation_reasons": "; ".join(sorted(bucket["reasons"])),
                "suspicious_relation_details": " | ".join(sorted(bucket["details"])),
            }
            for key, bucket in grouped.items()
        },
        dict(per_paper),
    )


def get_relation_meta(fill_map: dict[tuple[str, str], dict[str, str]], final_id: str, field_name: str) -> dict[str, str]:
    rule_parts, scope_parts, row_ids, count = set(), set(), set(), 0
    for alias in RELATION_FILL_FIELD_ALIASES.get(field_name, {field_name}):
        payload = fill_map.get((final_id, alias))
        if not payload:
            continue
        count += int(payload["relation_fill_count"])
        rule_parts.update(filter(None, payload["relation_fill_resolution_rules"].split("; ")))
        scope_parts.update(filter(None, payload["relation_fill_scope_types"].split("; ")))
        try:
            row_ids.update(filter(None, map(t, json.loads(payload["relation_fill_source_relation_row_ids"]))))
        except json.JSONDecodeError:
            pass
    return {
        "relation_fill_count": str(count),
        "relation_fill_resolution_rules": "; ".join(sorted(rule_parts)),
        "relation_fill_scope_types": "; ".join(sorted(scope_parts)),
        "relation_fill_source_relation_row_ids": json.dumps(sorted(row_ids), ensure_ascii=True),
    }


def get_suspicious_meta(suspicious_map: dict[tuple[str, str], dict[str, str]], final_id: str, field_name: str) -> dict[str, str]:
    payload = suspicious_map.get((final_id, field_name)) or suspicious_map.get((final_id, ""))
    return payload or {"suspicious_relation_flag": "no", "suspicious_relation_reasons": "", "suspicious_relation_details": ""}


def infer_field_source_type(predicted_value: str, relation_fill_count: str) -> str:
    if int(relation_fill_count or "0") > 0:
        return "relation_resolved"
    return "unresolved_blank" if not t(predicted_value) else "direct_extraction"


def build_rows(
    final_rows: list[dict[str, str]],
    manifest_map: dict[str, dict[str, str]],
    resolver: v1.AuditEvidenceResolverV1,
    decision_trace_map: dict[str, dict[str, str]],
    relation_fill_map: dict[tuple[str, str], dict[str, str]],
    suspicious_map: dict[tuple[str, str], dict[str, str]],
    max_span_chars: int,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    workbook_rows, seed_rows = [], []
    family_index_map, next_family_index = {}, defaultdict(int)
    for row in final_rows:
        key = (t(row.get("key")), t(row.get("family_id")))
        if key[0] and key[1] and key not in family_index_map:
            next_family_index[key[0]] += 1
            family_index_map[key] = next_family_index[key[0]]
    for row in final_rows:
        paper_key = t(row.get("key"))
        manifest_row = manifest_map.get(paper_key, {})
        doi = t(row.get("doi") or manifest_row.get("doi"))
        final_id = t(row.get("final_formulation_id"))
        stage5_label = v1.build_stage5_label(row, family_index_map.get((paper_key, t(row.get("family_id")))))
        params_label = v1.build_parameter_label(row)
        trace_meta = decision_trace_map.get(final_id, {k: "" for k in SEED_COLUMNS if k.startswith("decision_trace_")})
        for spec in v1.FIELD_SPECS:
            predicted_value, predicted_unit, derivation_status, derivation_inputs = v1.choose_extracted_value(row, spec)
            evidence_text, evidence_source_type = v1.resolve_field_evidence(
                resolver=resolver, row=row, manifest_row=manifest_row, spec=spec,
                extracted_value=predicted_value, max_span_chars=max_span_chars,
            )
            derivation_rule = ""
            if spec.allow_derivation:
                derivation_rule = "drug_feed_amount_text_value / plga_mass_mg_value"
                if predicted_value:
                    evidence_text = (
                        "Derived from final-row values: "
                        + f"drug_feed_amount_text_value={t(row.get('drug_feed_amount_text_value') or row.get('drug_feed_amount_text_value_text'))}; "
                        + f"plga_mass_mg_value={t(row.get('plga_mass_mg_value') or row.get('plga_mass_mg_value_text'))}"
                    )
                    evidence_source_type, evidence_status = "derived", "derived_from_final_row"
                else:
                    evidence_source_type, evidence_status = "blank_value", "blank_extracted_value"
            elif not predicted_value:
                evidence_text, evidence_status = "", "blank_extracted_value"
            elif evidence_source_type in {"unresolved_table", "unsupported_text"}:
                evidence_text, evidence_status = "", evidence_source_type
            elif v1.evidence_supports_value(predicted_value, evidence_text):
                evidence_status = "supported"
            else:
                evidence_text, evidence_source_type, evidence_status = "", "unsupported_text", "unsupported_text"
            relation_meta = get_relation_meta(relation_fill_map, final_id, spec.field_name)
            suspicious_meta = get_suspicious_meta(suspicious_map, final_id, spec.field_name)
            field_source_type = infer_field_source_type(predicted_value, relation_meta["relation_fill_count"])
            common = {
                "paper_key": paper_key,
                "doi": doi,
                "paper_title": t(manifest_row.get("title") or manifest_row.get("paper_title")),
                "final_formulation_id": final_id,
                "representative_source_formulation_id": t(row.get("representative_source_formulation_id")),
                "representative_source_raw_formulation_label": t(row.get("representative_source_raw_formulation_label")),
                "formulation_label_stage5": stage5_label,
                "formulation_label_params": params_label,
                "family_id": t(row.get("family_id")),
                "parent_core_row_id": t(row.get("parent_core_row_id")),
                "variant_role": t(row.get("variant_role")),
                "payload_state": t(row.get("payload_state")),
                "field_name": spec.field_name,
                "source_value_column": spec.value_column,
                "source_value_text_column": spec.value_text_column,
                "source_evidence_region_column": spec.evidence_region_column,
                "predicted_value": predicted_value,
                "predicted_unit": predicted_unit,
                "evidence_text": evidence_text,
                "evidence_source_type": evidence_source_type,
                "evidence_support_status": evidence_status,
                "field_source_type": field_source_type,
                "relation_resolved_flag": "yes" if field_source_type == "relation_resolved" else "no",
                "derivation_status": derivation_status,
                "derivation_rule": derivation_rule,
                "derivation_inputs": derivation_inputs,
                "evidence_section": t(row.get("evidence_section")),
                "evidence_span_start": t(row.get("evidence_span_start")),
                "evidence_span_end": t(row.get("evidence_span_end")),
                "final_output_rule": t(row.get("final_output_rule")),
                "review_needed": t(row.get("review_needed")),
                "benchmark_default_include": t(row.get("benchmark_default_include")),
                "relation_record_count": t(row.get("relation_record_count")),
                "relation_graph_ids": t(row.get("relation_graph_ids")),
                "relation_method_group_ids": t(row.get("relation_method_group_ids")),
                "gt_decision": "", "gt_value": "", "gt_unit": "", "gt_notes": "",
                **relation_meta, **suspicious_meta, **trace_meta,
            }
            workbook_rows.append({key: common.get(key, "") for key in WORKBOOK_COLUMNS})
            seed_rows.append({key: common.get(key, "") for key in SEED_COLUMNS})
    return workbook_rows, seed_rows


def build_summary(seed_rows: list[dict[str, str]], suspicious_counts: dict[str, int]) -> list[dict[str, str]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in seed_rows:
        bucket = grouped.setdefault(row["paper_key"], {"paper_key": row["paper_key"], "doi": row["doi"], "paper_title": row["paper_title"], "formulation_ids": set(), "field_rows": 0, "direct": 0, "resolved": 0, "blank": 0, "unsupported": 0, "unresolved_table": 0, "suspicious": 0})
        bucket["formulation_ids"].add(row["final_formulation_id"])
        bucket["field_rows"] += 1
        bucket[{"direct_extraction": "direct", "relation_resolved": "resolved", "unresolved_blank": "blank"}[row["field_source_type"]]] += 1
        if row["evidence_support_status"] == "unsupported_text":
            bucket["unsupported"] += 1
        if row["evidence_support_status"] == "unresolved_table":
            bucket["unresolved_table"] += 1
        if row["suspicious_relation_flag"] == "yes":
            bucket["suspicious"] += 1
    return [
        {
            "paper_key": key,
            "doi": bucket["doi"],
            "paper_title": bucket["paper_title"],
            "formulation_count": str(len(bucket["formulation_ids"])),
            "field_rows": str(bucket["field_rows"]),
            "direct_extraction_field_rows": str(bucket["direct"]),
            "relation_resolved_field_rows": str(bucket["resolved"]),
            "unresolved_blank_field_rows": str(bucket["blank"]),
            "unsupported_text_field_rows": str(bucket["unsupported"]),
            "unresolved_table_field_rows": str(bucket["unresolved_table"]),
            "suspicious_field_rows": str(bucket["suspicious"]),
            "paper_level_suspicious_case_count": str(suspicious_counts.get(key, 0)),
        }
        for key, bucket in sorted(grouped.items())
    ]


def style_review_sheet(ws) -> None:
    header_fill = {1: "1F2937", 2: "0F766E", 3: "7C2D12", 4: "7E22CE"}
    groups = {name: (1 if idx <= 10 else 2 if idx <= 12 else 3 if idx <= 17 else 4) for idx, name in enumerate(WORKBOOK_COLUMNS, start=1)}
    editable = {"gt_decision", "gt_value", "gt_unit", "gt_notes"}
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "K2"
    for cell in ws[1]:
        name = t(cell.value)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=header_fill[groups[name]])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            name = t(ws.cell(row=1, column=cell.column).value)
            is_editable = name in editable
            cell.protection = Protection(locked=not is_editable)
            cell.fill = PatternFill("solid", fgColor="FFF7ED" if is_editable else "F8FAFC")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in {"A": 14, "B": 22, "C": 20, "D": 20, "E": 32, "F": 22, "G": 22, "H": 16, "I": 16, "J": 18, "K": 18, "L": 12, "M": 60, "N": 18, "O": 22, "P": 18, "Q": 16, "R": 14, "S": 18, "T": 12, "U": 30}.items():
        ws.column_dimensions[col].width = width
    ws.protection.sheet = True
    ws.protection.autoFilter = True
    ws.protection.sort = True


def build_workbook(path: Path, workbook_rows: list[dict[str, str]], seed_rows: list[dict[str, str]], summary_rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws_review, ws_seed = wb.active, wb.create_sheet("seed_reference")
    ws_review.title = "field_gt_review"
    ws_summary, ws_instr, ws_dropdown = wb.create_sheet("source_summary"), wb.create_sheet("instructions"), wb.create_sheet("dropdown_options")
    review_header = v1.write_sheet_rows(ws_review, WORKBOOK_COLUMNS, workbook_rows)
    v1.write_sheet_rows(ws_seed, SEED_COLUMNS, seed_rows)
    v1.write_sheet_rows(ws_summary, list(summary_rows[0].keys()) if summary_rows else ["paper_key", "doi", "paper_title"], summary_rows)
    ws_dropdown["A1"], ws_dropdown["B1"] = "gt_decision_options", "gt_unit_options"
    for i, value in enumerate(v1.GT_STATUS_OPTIONS, start=2):
        ws_dropdown[f"A{i}"] = value
    for i, value in enumerate(v1.GT_UNIT_OPTIONS, start=2):
        ws_dropdown[f"B{i}"] = value
    ws_dropdown.sheet_state = "hidden"
    if ws_review.max_row >= 2:
        decision_col = get_column_letter(review_header["gt_decision"])
        unit_col = get_column_letter(review_header["gt_unit"])
        dv_decision = DataValidation(type="list", formula1=f"=dropdown_options!$A$2:$A${len(v1.GT_STATUS_OPTIONS)+1}", allow_blank=True)
        dv_unit = DataValidation(type="list", formula1=f"=dropdown_options!$B$2:$B${len(v1.GT_UNIT_OPTIONS)+1}", allow_blank=True, showErrorMessage=False)
        ws_review.add_data_validation(dv_decision)
        ws_review.add_data_validation(dv_unit)
        dv_decision.add(f"{decision_col}2:{decision_col}{ws_review.max_row}")
        dv_unit.add(f"{unit_col}2:{unit_col}{ws_review.max_row}")
    style_review_sheet(ws_review)
    v1.style_simple_sheet(ws_seed)
    v1.style_simple_sheet(ws_summary)
    ws_instr["A1"] = "Layer 3 Field GT Review Instructions"
    ws_instr["A1"].font = Font(bold=True)
    for i, line in enumerate([
        "Layer 3 field GT starts from frozen relation-first Stage 5 rows and does not change row identity.",
        "Each row is one `(final_formulation_id, field_name)` review item.",
        "The main sheet stays compact; heavier provenance and relation metadata live on `seed_reference`.",
        "field_source_type is review-facing field provenance: direct_extraction, relation_resolved, or unresolved_blank.",
        "Table evidence remains row-local only; unresolved_table means no row-local table evidence matched the seeded value.",
        "unsupported_text means the seeded text evidence did not safely support the predicted value and needs manual lookup.",
        "drug_polymer_ratio remains the existing deterministic derived seed only when final-row drug and polymer masses are both present.",
    ], start=3):
        ws_instr[f"A{i}"] = line
    ws_instr.column_dimensions["A"].width = 140
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--run-id", required=True)
    parser.add_argument("--run-dir", type=Path, default=None)
    parser.add_argument("--out-subdir", required=True)
    parser.add_argument("--workbook-name", default=versioned_name("field_gt_review_workbook", 3, ".xlsx"))
    parser.add_argument("--final-table-tsv", type=Path, required=True)
    parser.add_argument("--decision-trace-tsv", type=Path, required=True)
    parser.add_argument("--scope-manifest-tsv", type=Path, required=True)
    parser.add_argument("--relation-records-tsv", type=Path, default=None)
    parser.add_argument("--relation-fills-tsv", type=Path, default=None)
    parser.add_argument("--suspicious-cases-tsv", type=Path, default=None)
    parser.add_argument("--max-span-chars", type=int, default=320)
    args = parser.parse_args()

    run_id = t(args.run_id)
    if not v1.is_valid_run_id(run_id):
        raise ValueError(f"Invalid run_id: {run_id}")
    run_dir = v1.resolve_run_dir(run_id, args.run_dir)
    out_dir = run_dir / v1.sanitize_out_subdir(args.out_subdir)
    resolver = v1.AuditEvidenceResolverV1(project_root=Path(__file__).resolve().parents[2])
    final_rows = v1.read_tsv_rows(args.final_table_tsv.resolve())
    manifest_map = v1.load_manifest_map(args.scope_manifest_tsv.resolve())
    decision_trace_map = load_decision_trace_map(args.decision_trace_tsv.resolve())
    relation_fill_map = load_relation_fill_map(args.relation_fills_tsv.resolve() if args.relation_fills_tsv else None)
    suspicious_map, suspicious_counts = load_suspicious_maps(args.suspicious_cases_tsv.resolve() if args.suspicious_cases_tsv else None)
    workbook_rows, seed_rows = build_rows(final_rows, manifest_map, resolver, decision_trace_map, relation_fill_map, suspicious_map, max(120, int(args.max_span_chars)))
    summary_rows = build_summary(seed_rows, suspicious_counts)
    v1.write_tsv(out_dir / versioned_name("field_gt_review_seed_rows", 3, ".tsv"), SEED_COLUMNS, seed_rows)
    if summary_rows:
        v1.write_tsv(out_dir / versioned_name("field_gt_review_source_summary", 3, ".tsv"), list(summary_rows[0].keys()), summary_rows)
    workbook_path = out_dir / (t(args.workbook_name) or versioned_name("field_gt_review_workbook", 3, ".xlsx"))
    build_workbook(workbook_path, workbook_rows, seed_rows, summary_rows)
    print(json.dumps({"run_id": run_id, "run_dir": str(run_dir), "out_dir": str(out_dir), "workbook_path": str(workbook_path), "field_rows": len(workbook_rows), "formulation_rows": len(final_rows), "field_catalog": [spec.field_name for spec in v1.FIELD_SPECS]}, indent=2))


if __name__ == "__main__":
    main()
