#!/usr/bin/env python3
"""
Shared helpers for DEV15 formulation skeleton annotation tools.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple


SOURCE_TYPE_OPTIONS = [
    "table_row",
    "explicit_label",
    "text_described",
    "inherited",
    "mixed_or_unclear",
]
FORMULATION_EXISTS_OPTIONS = ["yes", "no", "uncertain"]
BOUNDARY_CONFIDENCE_OPTIONS = ["high", "medium", "low"]
REVIEW_STATUS_OPTIONS = ["pending", "reviewed", "needs_second_pass"]

REVIEW_COLUMNS = [
    "paper_key",
    "doi",
    "paper_title",
    "formulation_id",
    "formulation_label_raw",
    "source_type",
    "source_locator",
    "formulation_exists_gt",
    "formulation_boundary_confidence",
    "review_status",
    "notes",
    "helper_incomplete",
    "helper_duplicate_id",
]

CANDIDATE_COLUMNS = [
    "paper_key",
    "doi",
    "candidate_formulation_label",
    "candidate_formulation_id",
    "source_type_candidate",
    "evidence_pointer_candidate",
    "notes_candidate",
]


def project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def norm_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm_doi(value: object) -> str:
    doi = norm_text(value).lower()
    doi = re.sub(r"^doi\s*:\s*", "", doi)
    doi = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", doi)
    doi = re.sub(r"^doi\.org/", "", doi)
    return doi.strip()


def slugify_doi(doi: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", norm_doi(doi)).strip("_")


def read_tsv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        if reader.fieldnames is None:
            raise ValueError(f"TSV has no header: {path}")
        rows: List[Dict[str, str]] = []
        for row in reader:
            rows.append({k: norm_text(v) for k, v in row.items()})
        return rows


def write_tsv(path: Path, fieldnames: Sequence[str], rows: Iterable[Dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=list(fieldnames),
            delimiter="\t",
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\n",
        )
        writer.writeheader()
        for row in rows:
            writer.writerow({c: norm_text(row.get(c, "")) for c in fieldnames})


def discover_dev15_manifest(explicit_path: Optional[Path] = None) -> Path:
    root = project_root()
    if explicit_path is not None:
        if not explicit_path.exists():
            raise FileNotFoundError(f"Manifest not found: {explicit_path}")
        return explicit_path

    preferred = root / "data/cleaned/goren_2025/index/splits/dev_manifest_v1.tsv"
    if preferred.exists():
        return preferred

    candidates = sorted((root / "data").glob("**/dev_manifest*.tsv"))
    for candidate in candidates:
        try:
            rows = read_tsv(candidate)
        except Exception:
            continue
        dois = {norm_doi(r.get("doi", "")) for r in rows if norm_doi(r.get("doi", ""))}
        if len(dois) == 15:
            return candidate
    raise FileNotFoundError("Could not find a DEV-15 manifest.")


def load_manifest_rows(manifest_path: Path) -> List[Dict[str, str]]:
    rows = read_tsv(manifest_path)
    out: List[Dict[str, str]] = []
    for row in rows:
        paper_key = (
            norm_text(row.get("paper_key"))
            or norm_text(row.get("zotero_key"))
            or norm_text(row.get("key"))
            or norm_text(row.get("doc_key"))
        )
        doi = norm_doi(row.get("doi", ""))
        title = norm_text(row.get("paper_title")) or norm_text(row.get("title"))
        if not paper_key:
            continue
        out.append({"paper_key": paper_key, "doi": doi, "paper_title": title})

    seen: set[str] = set()
    dedup: List[Dict[str, str]] = []
    for row in out:
        if row["paper_key"] in seen:
            continue
        seen.add(row["paper_key"])
        dedup.append(row)
    dedup.sort(key=lambda x: x["paper_key"])
    return dedup


def normalize_source_type(value: str) -> str:
    s = norm_text(value).lower()
    if not s:
        return ""
    if s in SOURCE_TYPE_OPTIONS:
        return s
    if "table" in s:
        return "table_row"
    if "inherit" in s:
        return "inherited"
    if "label" in s:
        return "explicit_label"
    if "text" in s or "narrative" in s:
        return "text_described"
    return "mixed_or_unclear"


def detect_candidate_source(dev_keys: Sequence[str], dev_dois: Sequence[str]) -> Optional[Path]:
    root = project_root()
    search_patterns = [
        "data/results/**/weak_labels*.tsv",
        "data/results/**/extracted_formulation_level*.tsv",
        "data/results/**/doi_level_ee_scaffold*.tsv",
        "data/benchmark/goren_2025/**/extracted_formulation_level*.tsv",
    ]
    candidates: List[Path] = []
    for pattern in search_patterns:
        candidates.extend(sorted(root.glob(pattern)))

    dev_key_set = {norm_text(k) for k in dev_keys if norm_text(k)}
    dev_doi_set = {norm_doi(d) for d in dev_dois if norm_doi(d)}
    best: Optional[Path] = None
    best_score = -1

    for path in candidates:
        try:
            rows = read_tsv(path)
        except Exception:
            continue
        if not rows:
            continue

        header = set(rows[0].keys())
        key_cols = [c for c in ["paper_key", "zotero_key", "key", "doc_key"] if c in header]
        doi_cols = [c for c in ["doi", "doi_norm", "reference_normalized_doi"] if c in header]
        has_formulation_id = 1 if "formulation_id" in header else 0

        overlap = 0
        for row in rows:
            row_key = ""
            for c in key_cols:
                row_key = norm_text(row.get(c, ""))
                if row_key:
                    break
            row_doi = ""
            for c in doi_cols:
                row_doi = norm_doi(row.get(c, ""))
                if row_doi:
                    break
            if row_key and row_key in dev_key_set:
                overlap += 1
            elif row_doi and row_doi in dev_doi_set:
                overlap += 1

        if overlap <= 0:
            continue

        score = overlap + (20 * has_formulation_id)
        if score > best_score:
            best_score = score
            best = path
    return best


def build_candidate_rows_from_source(
    source_path: Path,
    manifest_rows: Sequence[Dict[str, str]],
) -> List[Dict[str, str]]:
    manifest_by_key = {r["paper_key"]: r for r in manifest_rows}
    manifest_by_doi = {r["doi"]: r for r in manifest_rows if r["doi"]}
    rows = read_tsv(source_path)

    out: List[Dict[str, str]] = []
    seq_by_key: Dict[str, int] = {}
    for row in rows:
        row_key = ""
        for c in ["paper_key", "zotero_key", "key", "doc_key"]:
            row_key = norm_text(row.get(c, ""))
            if row_key:
                break

        row_doi = ""
        for c in ["doi", "doi_norm", "reference_normalized_doi"]:
            row_doi = norm_doi(row.get(c, ""))
            if row_doi:
                break

        manifest = None
        if row_key and row_key in manifest_by_key:
            manifest = manifest_by_key[row_key]
        elif row_doi and row_doi in manifest_by_doi:
            manifest = manifest_by_doi[row_doi]

        if manifest is None:
            continue

        paper_key = manifest["paper_key"]
        seq_by_key[paper_key] = seq_by_key.get(paper_key, 0) + 1
        fallback_id = f"{paper_key}_F{seq_by_key[paper_key]:02d}"
        candidate_id = norm_text(row.get("formulation_id")) or fallback_id

        label = (
            norm_text(row.get("formulation_label"))
            or norm_text(row.get("formulation_signature"))
            or norm_text(row.get("tracer_name"))
            or norm_text(row.get("drug_name"))
        )
        source_type = normalize_source_type(
            norm_text(row.get("source_type"))
            or norm_text(row.get("evidence_source"))
            or norm_text(row.get("evidence_section"))
        )
        evidence = (
            norm_text(row.get("evidence_pointer"))
            or norm_text(row.get("evidence_span_text_main"))
            or norm_text(row.get("evidence_section"))
        )
        notes = norm_text(row.get("notes"))

        out.append(
            {
                "paper_key": paper_key,
                "doi": manifest["doi"],
                "candidate_formulation_label": label,
                "candidate_formulation_id": candidate_id,
                "source_type_candidate": source_type,
                "evidence_pointer_candidate": evidence,
                "notes_candidate": notes,
            }
        )

    # Deduplicate by (paper_key, candidate_formulation_id)
    dedup: List[Dict[str, str]] = []
    seen: set[Tuple[str, str]] = set()
    for row in out:
        key = (row["paper_key"], row["candidate_formulation_id"])
        if key in seen:
            continue
        seen.add(key)
        dedup.append(row)
    return dedup


def ensure_candidate_rows(
    manifest_rows: Sequence[Dict[str, str]],
    candidate_rows: Sequence[Dict[str, str]],
    default_rows_per_paper: int,
) -> List[Dict[str, str]]:
    by_key: Dict[str, List[Dict[str, str]]] = {}
    for row in candidate_rows:
        by_key.setdefault(row["paper_key"], []).append(dict(row))

    out: List[Dict[str, str]] = []
    for paper in manifest_rows:
        paper_key = paper["paper_key"]
        doi = paper["doi"]
        existing = by_key.get(paper_key, [])
        target_n = max(default_rows_per_paper, len(existing))
        for idx in range(1, target_n + 1):
            source_row = existing[idx - 1] if idx <= len(existing) else {}
            out.append(
                {
                    "paper_key": paper_key,
                    "doi": doi,
                    "candidate_formulation_label": norm_text(source_row.get("candidate_formulation_label")),
                    # Always regenerate deterministic IDs for usability and stability.
                    "candidate_formulation_id": f"{paper_key}_F{idx:02d}",
                    "source_type_candidate": normalize_source_type(source_row.get("source_type_candidate", "")),
                    "evidence_pointer_candidate": norm_text(source_row.get("evidence_pointer_candidate")),
                    "notes_candidate": norm_text(source_row.get("notes_candidate")),
                }
            )

    out.sort(key=lambda r: (r["paper_key"], r["candidate_formulation_id"]))
    return out


def read_review_sheet_rows(xlsx_path: Path, sheet_name: str) -> List[Dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb[sheet_name]
    header = [norm_text(c.value) for c in ws[1]]
    rows: List[Dict[str, str]] = []
    for i in range(2, ws.max_row + 1):
        row = {}
        is_all_blank = True
        for j, col in enumerate(header, start=1):
            value = ws.cell(row=i, column=j).value
            text = norm_text(value)
            row[col] = text
            if text:
                is_all_blank = False
        if is_all_blank:
            continue
        row["_excel_row"] = str(i)
        rows.append(row)
    return rows
