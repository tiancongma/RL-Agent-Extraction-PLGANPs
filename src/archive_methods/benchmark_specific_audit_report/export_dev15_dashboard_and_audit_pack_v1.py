#!/usr/bin/env python3
from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl.styles import Alignment


REPO_ROOT = Path.cwd()
DEV15_MANIFEST = REPO_ROOT / "data/cleaned/goren_2025/index/splits/dev_manifest_v1.tsv"
CURATED_CSV = REPO_ROOT / "data/benchmark/goren_2025/NP_dataset_formulations.csv"
EXTRACTED_VIEW_XLSX = REPO_ROOT / "data/cleaned/labels/manual/dev15_extracted_formulation_view.xlsx"
OUT_DASHBOARD_TSV = REPO_ROOT / "data/cleaned/labels/manual/dev15_dashboard_counts.tsv"
OUT_AUDIT_XLSX = REPO_ROOT / "data/cleaned/labels/manual/dev15_highrisk_audit_pack.xlsx"


def norm_text(v: object) -> str:
    if v is None or pd.isna(v):
        return ""
    return re.sub(r"\s+", " ", str(v).strip().lower())


def norm_doi(v: object) -> str:
    s = norm_text(v)
    s = re.sub(r"^doi\s*:\s*", "", s)
    s = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", s)
    s = re.sub(r"^doi\.org/", "", s)
    return s


def safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]\:\*\?\/\\]", "_", name).strip()
    if not base:
        base = "unknown_doi"
    base = base[:31]
    out = base
    i = 1
    while out in used:
        suffix = f"_{i}"
        out = f"{base[: max(0, 31 - len(suffix))]}{suffix}"
        i += 1
    used.add(out)
    return out


def load_dev15_dois() -> List[str]:
    dev = pd.read_csv(DEV15_MANIFEST, sep="\t", dtype=str).fillna("")
    dois = dev["doi"].map(norm_doi)
    dois = dois[dois != ""].drop_duplicates().tolist()
    if len(dois) != 15:
        raise RuntimeError(f"Expected 15 DEV DOIs, got {len(dois)} from {DEV15_MANIFEST}")
    return dois


def load_curated_counts(dev15_dois: List[str]) -> pd.DataFrame:
    curated = pd.read_csv(CURATED_CSV, dtype=str).fillna("")
    doi_col = "reference"
    if doi_col not in curated.columns:
        raise RuntimeError(f"Expected DOI column '{doi_col}' in {CURATED_CSV}")
    curated["doi"] = curated[doi_col].map(norm_doi)
    curated = curated[curated["doi"].isin(set(dev15_dois))].copy()
    counts = curated.groupby("doi", as_index=False).size().rename(columns={"size": "curated_n"})
    return counts


def load_extracted_view_long() -> pd.DataFrame:
    xls = pd.ExcelFile(EXTRACTED_VIEW_XLSX)
    frames = []
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet, dtype=str).fillna("")
        if "doi" not in df.columns:
            continue
        df["doi"] = df["doi"].map(norm_doi)
        frames.append(df)
    if not frames:
        raise RuntimeError(f"No DOI sheets found in {EXTRACTED_VIEW_XLSX}")
    return pd.concat(frames, ignore_index=True)


def compute_extracted_counts(view_df: pd.DataFrame, dev15_dois: List[str]) -> pd.DataFrame:
    v = view_df[view_df["doi"].isin(set(dev15_dois))].copy()
    raw = v.groupby("doi", as_index=False).size().rename(columns={"size": "extracted_n_raw"})
    group_col = "grouping_id" if "grouping_id" in v.columns else "formulation_id"
    v["_group_key"] = v[group_col].astype(str).str.strip()
    if "formulation_id" in v.columns:
        v.loc[v["_group_key"] == "", "_group_key"] = v["formulation_id"].astype(str).str.strip()
    grouped = (
        v[v["_group_key"] != ""].groupby("doi")["_group_key"].nunique().reset_index(name="extracted_n_after_grouping")
    )
    return raw.merge(grouped, on="doi", how="outer")


def add_ratios_and_risk(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in ["curated_n", "extracted_n_raw", "extracted_n_after_grouping"]:
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).astype(int)

    out["ratio_raw"] = out.apply(
        lambda r: (r["extracted_n_raw"] / r["curated_n"]) if r["curated_n"] > 0 else float("nan"),
        axis=1,
    )
    out["ratio_grouped"] = out.apply(
        lambda r: (r["extracted_n_after_grouping"] / r["curated_n"]) if r["curated_n"] > 0 else float("nan"),
        axis=1,
    )

    def risk_flag(row: pd.Series) -> str:
        c = int(row["curated_n"])
        e = int(row["extracted_n_after_grouping"])
        if e >= c * 2 or e >= c + 10:
            return "HIGH"
        if e <= c * 1.2 and e >= c * 0.8:
            return "LOW"
        return "MED"

    out["risk_flag"] = out.apply(risk_flag, axis=1)
    return out


def make_signature_key_short(df: pd.DataFrame) -> pd.Series:
    fields = [
        "drug_name",
        "plga_mw",
        "la_ga_ratio",
        "drug_polymer_ratio",
        "surfactant_type",
        "organic_solvent",
    ]
    for c in fields:
        if c not in df.columns:
            df[c] = ""
    return df.apply(lambda r: "|".join(norm_text(r[c]) for c in fields), axis=1)


def write_audit_pack(view_df: pd.DataFrame, top5_high: List[str]) -> None:
    focused_cols = [
        "doi",
        "grouping_id",
        "formulation_id",
        "signature_hash",
        "drug_name",
        "plga_mw",
        "la_ga_ratio",
        "drug_polymer_ratio",
        "polymer_mass",
        "drug_mass",
        "surfactant_type",
        "surfactant_conc",
        "organic_solvent",
        "emulsion_type",
        "emulsion_method",
        "ee_value",
        "ee_evidence_source",
        "ee_evidence_pointer",
        "size_nm",
        "pdi",
    ]
    used: set[str] = set()
    OUT_AUDIT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUT_AUDIT_XLSX, engine="openpyxl") as writer:
        for doi in top5_high:
            d = view_df[view_df["doi"] == doi].copy()
            for c in focused_cols:
                if c not in d.columns:
                    d[c] = ""
            d = d[focused_cols].copy()
            d["signature_key_short"] = make_signature_key_short(d)
            d["grouping_id"] = d["grouping_id"].astype(str)
            d.loc[d["grouping_id"].str.strip() == "", "grouping_id"] = d["formulation_id"].astype(str)
            d = d.sort_values(["signature_key_short", "grouping_id", "formulation_id"], kind="stable")
            sheet = safe_sheet_name(doi, used)
            d.to_excel(writer, sheet_name=sheet, index=False)

    # Post-format workbook for readability.
    from openpyxl import load_workbook

    wb = load_workbook(OUT_AUDIT_XLSX)
    wrap_cols = {"ee_evidence_pointer", "signature_key_short"}
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        header = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        for idx, col_name in enumerate(header, start=1):
            col_name = "" if col_name is None else str(col_name)
            max_len = len(col_name)
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=idx).value
                s = "" if v is None else str(v)
                max_len = max(max_len, min(len(s), 120))
                if col_name in wrap_cols:
                    ws.cell(row=r, column=idx).alignment = Alignment(wrap_text=True, vertical="top")
            width = 60 if col_name in wrap_cols else max(12, min(max_len + 2, 36))
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
            ws.cell(row=1, column=idx).alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(OUT_AUDIT_XLSX)


def main() -> None:
    dev15_dois = load_dev15_dois()
    curated_counts = load_curated_counts(dev15_dois)
    print(f"TaskA_curated_source: {CURATED_CSV}")
    print("TaskA_curated_doi_column: reference")

    extracted_view = load_extracted_view_long()
    extracted_counts = compute_extracted_counts(extracted_view, dev15_dois)

    dashboard = pd.DataFrame({"doi": dev15_dois})
    dashboard = dashboard.merge(curated_counts, on="doi", how="left")
    dashboard = dashboard.merge(extracted_counts, on="doi", how="left")
    dashboard = add_ratios_and_risk(dashboard)
    dashboard = dashboard.sort_values(["risk_flag", "ratio_grouped", "doi"], ascending=[True, False, True])

    OUT_DASHBOARD_TSV.parent.mkdir(parents=True, exist_ok=True)
    dashboard.to_csv(OUT_DASHBOARD_TSV, sep="\t", index=False)

    high = dashboard[dashboard["risk_flag"] == "HIGH"].copy()
    high = high.sort_values("ratio_grouped", ascending=False)
    print("TaskA_top10_high_risk:")
    if high.empty:
        print("- <none>")
    else:
        for _, r in high.head(10).iterrows():
            print(
                f"- {r['doi']}: curated_n={int(r['curated_n'])}, extracted_n_raw={int(r['extracted_n_raw'])}, "
                f"extracted_n_after_grouping={int(r['extracted_n_after_grouping'])}, "
                f"ratio_raw={r['ratio_raw']:.3f}, ratio_grouped={r['ratio_grouped']:.3f}"
            )

    top5_high = high["doi"].head(5).tolist()
    write_audit_pack(extracted_view, top5_high)

    print(f"Saved dashboard TSV: {OUT_DASHBOARD_TSV}")
    print(f"Saved audit pack XLSX: {OUT_AUDIT_XLSX}")
    print("Selected top 5 HIGH risk DOIs:")
    if top5_high:
        for d in top5_high:
            print(f"- {d}")
    else:
        print("- <none>")


if __name__ == "__main__":
    main()
