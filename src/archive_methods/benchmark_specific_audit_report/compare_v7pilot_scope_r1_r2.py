#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


FIELD_MAP: List[Tuple[str, str]] = [
    ("polymer_name_raw", "plga_mw_kDa"),
    ("plga_mw_kDa", "plga_mw_kDa"),
    ("la_ga_ratio", "la_ga_ratio"),
    ("organic_solvent", "organic_solvent"),
    ("surfactant_type", "surfactant_name"),
    ("surfactant_conc", "surfactant_concentration_text"),
    ("encapsulation_efficiency_percent", "encapsulation_efficiency_percent"),
    ("size_nm", "size_nm"),
]

SHARED_FIELDS = ["plga_mw_kDa", "la_ga_ratio", "organic_solvent", "surfactant_type", "surfactant_conc"]


def normalize_doi(v: Any) -> str:
    s = "" if v is None else str(v)
    s = s.strip().lower()
    s = re.sub(r"^doi\s*:\s*", "", s)
    s = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", s)
    s = re.sub(r"^doi\.org/", "", s)
    return s.strip()


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Compare v7pilot1 vs v7pilot_r2 scope typing and export XLSX pack.")
    p.add_argument(
        "--manifest-tsv",
        default="data/cleaned/goren_2025/index/splits/dev_manifest_v7pilot3_2026-03-06.tsv",
    )
    p.add_argument(
        "--pilot1-tsv",
        default="data/results/run_20260306_1321_v7pilot3_dev/weak_labels_v7pilot/weak_labels__v7pilot.tsv",
    )
    p.add_argument(
        "--pilot-r2-tsv",
        default="data/results/run_20260306_1358_v7pilot3r2_dev/weak_labels_v7pilot_r2/weak_labels__v7pilot_r2.tsv",
    )
    p.add_argument(
        "--tag2",
        default="v7pilot_r2",
        help="Label for second pilot in output columns (e.g., v7pilot_r2 or v7pilot_r3).",
    )
    p.add_argument(
        "--out-xlsx",
        default="data/cleaned/labels/manual/dev_v7pilot3_scope_comparison_r2.xlsx",
    )
    p.add_argument(
        "--out-md",
        default="docs/methods/weak_labels_v7pilot3_r2_eval_2026-03-06.md",
    )
    return p.parse_args()


def to_long(df: pd.DataFrame, tag: str) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for _, r in df.iterrows():
        doi = str(r.get("doi", "")).strip()
        doi_norm = normalize_doi(doi)
        fid = str(r.get("formulation_id", "")).strip()
        role = str(r.get("formulation_role", "")).strip()
        inst_conf = str(r.get("instance_confidence", "")).strip()
        for display_name, src in FIELD_MAP:
            vtxt = str(r.get(f"{src}_value_text", "")).strip()
            vval = str(r.get(f"{src}_value", "")).strip()
            value_raw = vtxt if vtxt else vval
            if display_name == "polymer_name_raw":
                if not value_raw:
                    value_raw = ""
            scope = str(r.get(f"{src}_scope", "")).strip().lower()
            mc = str(r.get(f"{src}_membership_confidence", "")).strip().lower()
            er = str(r.get(f"{src}_evidence_region_type", "")).strip().lower()
            rows.append(
                {
                    "doi": doi,
                    "doi_norm": doi_norm,
                    "formulation_id": fid,
                    "formulation_role": role,
                    "instance_confidence": inst_conf,
                    "field_name": display_name,
                    "value_raw": value_raw,
                    f"{tag}_scope": scope,
                    f"{tag}_membership_confidence": mc,
                    f"{tag}_evidence_region_type": er,
                }
            )
    return pd.DataFrame(rows)


def format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        1: 24,
        2: 16,
        3: 18,
        4: 18,
        5: 24,
        6: 26,
        7: 16,
        8: 16,
        9: 24,
        10: 24,
        11: 26,
        12: 26,
    }
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def unknown_rate_for(df_long: pd.DataFrame, tag: str, field_name: str) -> float:
    sub = df_long[df_long["field_name"] == field_name].copy()
    if sub.empty:
        return float("nan")
    s = sub[f"{tag}_scope"].astype(str).str.strip().str.lower()
    return float((s.eq("") | s.eq("unknown")).mean())


def main() -> None:
    args = parse_args()
    manifest = pd.read_csv(args.manifest_tsv, sep="\t", dtype=str).fillna("")
    pilot1 = pd.read_csv(args.pilot1_tsv, sep="\t", dtype=str).fillna("")
    pilot2 = pd.read_csv(args.pilot_r2_tsv, sep="\t", dtype=str).fillna("")
    manifest["doi_norm"] = manifest["doi"].map(normalize_doi)

    p1_long = to_long(pilot1, "v7pilot1")
    p2_long = to_long(pilot2, args.tag2)
    merged = p1_long.merge(
        p2_long,
        on=["doi_norm", "formulation_id", "field_name"],
        how="outer",
        suffixes=("_p1", "_p2"),
    )
    merged["doi"] = merged["doi_p2"].where(merged["doi_p2"].astype(str).str.strip() != "", merged["doi_p1"])
    merged["formulation_role"] = merged["formulation_role_p2"].where(
        merged["formulation_role_p2"].astype(str).str.strip() != "", merged["formulation_role_p1"]
    )
    merged["instance_confidence"] = merged["instance_confidence_p2"].where(
        merged["instance_confidence_p2"].astype(str).str.strip() != "", merged["instance_confidence_p1"]
    )
    merged["value_raw"] = merged["value_raw_p2"].where(
        merged["value_raw_p2"].astype(str).str.strip() != "", merged["value_raw_p1"]
    )

    out_cols = [
        "doi",
        "formulation_id",
        "formulation_role",
        "instance_confidence",
        "field_name",
        "value_raw",
        "v7pilot1_scope",
        f"{args.tag2}_scope",
        "v7pilot1_membership_confidence",
        f"{args.tag2}_membership_confidence",
        "v7pilot1_evidence_region_type",
        f"{args.tag2}_evidence_region_type",
    ]
    for c in out_cols:
        if c not in merged.columns:
            merged[c] = ""
    merged = merged[out_cols + ["doi_norm"]].sort_values(["doi_norm", "formulation_id", "field_name"])

    out_xlsx = Path(args.out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        for _, mrow in manifest.iterrows():
            doi = str(mrow["doi"])
            doi_norm = str(mrow["doi_norm"])
            sub = merged[merged["doi_norm"] == doi_norm].copy()
            sheet_name = f"doi_{doi_norm[:28]}".replace("/", "_").replace(".", "_")
            sub[out_cols].to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(out_xlsx)
    for ws in wb.worksheets:
        format_sheet(ws)
    wb.save(out_xlsx)

    # Summary metrics.
    c1 = pilot1.assign(doi_norm=lambda d: d["doi"].map(normalize_doi)).groupby("doi_norm").size().rename("v7pilot1_count")
    c2 = pilot2.assign(doi_norm=lambda d: d["doi"].map(normalize_doi)).groupby("doi_norm").size().rename(f"{args.tag2}_count")
    counts = pd.concat([c1, c2], axis=1).fillna(0).reset_index()
    counts["v7pilot1_count"] = counts["v7pilot1_count"].astype(int)
    counts[f"{args.tag2}_count"] = counts[f"{args.tag2}_count"].astype(int)
    counts["delta_tag2_minus_v1"] = counts[f"{args.tag2}_count"] - counts["v7pilot1_count"]

    scope_dist_1 = (
        p1_long["v7pilot1_scope"].astype(str).str.strip().str.lower().replace("", "unknown").value_counts().to_dict()
    )
    scope_dist_2 = p2_long[f"{args.tag2}_scope"].astype(str).str.strip().str.lower().replace("", "unknown").value_counts().to_dict()

    unknown_rows = []
    for fn in SHARED_FIELDS:
        u1 = unknown_rate_for(p1_long, "v7pilot1", fn)
        u2 = unknown_rate_for(p2_long, args.tag2, fn)
        unknown_rows.append({"field": fn, "v7pilot1_unknown_rate": u1, f"{args.tag2}_unknown_rate": u2, "delta_tag2_minus_v1": u2 - u1})
    unknown_df = pd.DataFrame(unknown_rows).sort_values("delta_tag2_minus_v1")

    # Recommendation.
    avg_u1 = float(pd.to_numeric(unknown_df["v7pilot1_unknown_rate"], errors="coerce").fillna(1).mean())
    avg_u2 = float(pd.to_numeric(unknown_df[f"{args.tag2}_unknown_rate"], errors="coerce").fillna(1).mean())
    if avg_u2 <= 0.20 and avg_u2 < avg_u1:
        rec = "scale to 15 now"
    elif avg_u2 <= avg_u1:
        rec = "scale to 15 with caution"
    else:
        rec = "refine prompt again first"

    out_md = Path(args.out_md)
    out_md.parent.mkdir(parents=True, exist_ok=True)
    lines: List[str] = []
    lines.append("# weak_labels_v7pilot3 r2 Evaluation (2026-03-06)")
    lines.append("")
    lines.append("## Same 3 DOIs")
    for _, r in manifest.iterrows():
        lines.append(f"- `{r['doi']}` | key `{r['key']}`")
    lines.append("")
    lines.append("## Formulation counts per DOI")
    for _, r in counts.sort_values("doi_norm").iterrows():
        lines.append(
            f"- `{r['doi_norm']}`: v7pilot1={int(r['v7pilot1_count'])}, {args.tag2}={int(r[f'{args.tag2}_count'])}, delta={int(r['delta_tag2_minus_v1'])}"
        )
    lines.append("")
    lines.append(f"## Scope distribution comparison (v7pilot1 vs {args.tag2})")
    lines.append("- v7pilot1: " + json.dumps(scope_dist_1, ensure_ascii=False, sort_keys=True))
    lines.append(f"- {args.tag2}: " + json.dumps(scope_dist_2, ensure_ascii=False, sort_keys=True))
    lines.append("")
    lines.append("## Unknown rate comparison for shared-condition fields")
    for _, r in unknown_df.iterrows():
        lines.append(
            f"- `{r['field']}`: v1={r['v7pilot1_unknown_rate']:.1%}, {args.tag2}={r[f'{args.tag2}_unknown_rate']:.1%}, delta={r['delta_tag2_minus_v1']:+.1%}"
        )
    lines.append("")
    lines.append("## Recommendation")
    lines.append(f"- **{rec}**")
    out_md.write_text("\n".join(lines), encoding="utf-8")

    out = {
        "manifest_path": str(Path(args.manifest_tsv).resolve()),
        "pilot1_tsv": str(Path(args.pilot1_tsv).resolve()),
        "pilot_r2_tsv": str(Path(args.pilot_r2_tsv).resolve()),
        "comparison_xlsx": str(out_xlsx.resolve()),
        "summary_md": str(out_md.resolve()),
        "counts_per_doi": counts.sort_values("doi_norm").to_dict(orient="records"),
        "scope_dist_v7pilot1": scope_dist_1,
        f"scope_dist_{args.tag2}": scope_dist_2,
        "unknown_rate_shared_fields": unknown_df.to_dict(orient="records"),
        "recommendation": rec,
    }
    print(json.dumps(out, indent=2))


if __name__ == "__main__":
    main()
