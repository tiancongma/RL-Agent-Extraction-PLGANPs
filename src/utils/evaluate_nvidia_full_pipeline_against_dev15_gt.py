#!/usr/bin/env python3
"""
Diagnostic-only DEV15 GT evaluation for a completed NVIDIA full-pipeline Stage5 output.

This script compares only the completed Stage5 final table to GT.
It does not compare intermediate Stage2 probe or parsing surfaces.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from src.utils.active_data_source import resolve_artifact_path, resolve_run_context
    from src.utils.paths import DATA_RESULTS_DIR
except ModuleNotFoundError:
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from src.utils.active_data_source import resolve_artifact_path, resolve_run_context
    from src.utils.paths import DATA_RESULTS_DIR


@dataclass
class FinalRow:
    key: str
    nvidia_row_id: str
    raw_formulation_label: str
    polymer: str
    polymer_mw: str
    la_ga_ratio: str
    surfactant: str
    surfactant_concentration: str
    pva_concentration: str
    solvent: str
    drug: str
    drug_amount: str
    ee: str
    evidence_text: str


@dataclass
class GTRow:
    key: str
    formulation_id: str
    doi: str
    formulation_label_raw: str
    notes: str
    source_type: str
    source_locator: str
    extra_alignment_text: str


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_token(value: Any) -> str:
    text = normalize_text(value).lower()
    replacements = {
        "poly(ethylene glycol)": "peg",
        "polyethylene glycol": "peg",
        "poly(lactic acid-co-glycolic acid)": "plga",
        "poly(d,l-lactic-co-glycolic acid)": "plga",
        "poly-d,l-lactic-co-glycolic acid": "plga",
        "polysorbate 80": "tween80",
        "tween 80": "tween80",
        "pluronic f68": "f68",
        "lutrol f68": "f68",
        "dichloromethane": "dcm",
        "ethyl acetate": "ea",
        "polyvinyl alcohol": "pva",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    text = re.sub(r"[^a-z0-9%./+-]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def tokenize(value: Any) -> list[str]:
    return [token for token in normalize_token(value).split(" ") if token]


def meaningful_tokens(tokens: list[str]) -> list[str]:
    stop = {
        "nanoparticles",
        "nanoparticle",
        "loaded",
        "formulation",
        "formulations",
        "variant",
        "family",
        "plga",
    }
    return [token for token in tokens if token and token not in stop]


def parse_numeric_tokens(value: Any) -> list[str]:
    text = normalize_text(value).replace(",", ".")
    return re.findall(r"\d+(?:\.\d+)?", text)


def read_tsv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def make_run_dir() -> Path:
    today = datetime.now().strftime("%Y-%m-%d")
    candidate = DATA_RESULTS_DIR / f"run_{today}_dev15_nvidia_full_pipeline_gt_eval_v1"
    if not candidate.exists():
        return candidate
    version = 2
    while True:
        candidate = DATA_RESULTS_DIR / f"run_{today}_dev15_nvidia_full_pipeline_gt_eval_v{version}"
        if not candidate.exists():
            return candidate
        version += 1


def require_exists(path: Path, label: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Missing required {label}: {path}")


def load_gt_rows(workbook_path: Path, alignment_rows: list[dict[str, str]]) -> list[GTRow]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook["review_formulations"]
    raw_rows = worksheet.iter_rows(values_only=True)
    header = [str(value) if value is not None else "" for value in next(raw_rows)]

    alignment_text_by_gt_id: dict[str, list[str]] = defaultdict(list)
    for row in alignment_rows:
        gt_id = normalize_text(row.get("gt_formulation_id"))
        if not gt_id:
            continue
        parts = [
            row.get("gt_evidence_note", ""),
            row.get("canonical_signature_gt", ""),
            row.get("alignment_notes", ""),
            row.get("key_field_diff_summary", ""),
        ]
        joined = " ".join(normalize_text(part) for part in parts if normalize_text(part))
        if joined:
            alignment_text_by_gt_id[gt_id].append(joined)

    out: list[GTRow] = []
    for values in raw_rows:
        row = {
            header[idx]: "" if idx >= len(values) or values[idx] is None else str(values[idx])
            for idx in range(len(header))
        }
        if normalize_text(row.get("formulation_exists_gt")).lower() != "yes":
            continue
        if normalize_text(row.get("review_status")).lower() != "reviewed":
            continue
        formulation_id = normalize_text(row.get("formulation_id"))
        out.append(
            GTRow(
                key=normalize_text(row.get("paper_key")),
                formulation_id=formulation_id,
                doi=normalize_text(row.get("doi")),
                formulation_label_raw=normalize_text(row.get("formulation_label_raw")),
                notes=normalize_text(row.get("notes")),
                source_type=normalize_text(row.get("source_type")),
                source_locator=normalize_text(row.get("source_locator")),
                extra_alignment_text=" ".join(alignment_text_by_gt_id.get(formulation_id, [])),
            )
        )
    return out


def load_final_rows(final_table_tsv: Path) -> list[FinalRow]:
    rows = read_tsv_rows(final_table_tsv)
    out: list[FinalRow] = []
    for row in rows:
        key = normalize_text(row.get("key"))
        out.append(
            FinalRow(
                key=key,
                nvidia_row_id=normalize_text(row.get("final_formulation_id")),
                raw_formulation_label=normalize_text(row.get("raw_formulation_label") or row.get("representative_source_raw_formulation_label")),
                polymer=normalize_text(row.get("polymer_identity_final") or row.get("polymer_identity") or row.get("polymer_name_raw")),
                polymer_mw=normalize_text(row.get("polymer_mw_kDa_value_text") or row.get("polymer_mw_kDa_value")),
                la_ga_ratio=normalize_text(row.get("la_ga_ratio_value_text") or row.get("la_ga_ratio_value")),
                surfactant=normalize_text(row.get("surfactant_name_value_text") or row.get("surfactant_name_value")),
                surfactant_concentration=normalize_text(row.get("surfactant_concentration_text_value_text") or row.get("surfactant_concentration_text_value")),
                pva_concentration=normalize_text(row.get("pva_conc_percent_value_text") or row.get("pva_conc_percent_value")),
                solvent=normalize_text(row.get("organic_solvent_value_text") or row.get("organic_solvent_value")),
                drug=normalize_text(row.get("drug_name_value_text") or row.get("drug_name_value")),
                drug_amount=normalize_text(row.get("drug_feed_amount_text_value_text") or row.get("drug_feed_amount_text_value")),
                ee=normalize_text(row.get("encapsulation_efficiency_percent_value_text") or row.get("encapsulation_efficiency_percent_value")),
                evidence_text=normalize_text(row.get("evidence_span_text")),
            )
        )
    return out


def group_by_key(items: list[Any], key_attr: str) -> dict[str, list[Any]]:
    grouped: dict[str, list[Any]] = defaultdict(list)
    for item in items:
        grouped[getattr(item, key_attr)].append(item)
    return dict(grouped)


def build_gt_counts(gt_rows: list[GTRow]) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for row in gt_rows:
        counts[row.key] += 1
    return dict(counts)


def final_row_signal_tokens(row: FinalRow) -> dict[str, list[str]]:
    return {
        "label": meaningful_tokens(tokenize(row.raw_formulation_label)),
        "polymer": meaningful_tokens(tokenize(row.polymer)),
        "polymer_mw": parse_numeric_tokens(row.polymer_mw),
        "la_ga": parse_numeric_tokens(row.la_ga_ratio),
        "surfactant": meaningful_tokens(tokenize(row.surfactant)),
        "surfactant_concentration": parse_numeric_tokens(row.surfactant_concentration),
        "pva": parse_numeric_tokens(row.pva_concentration),
        "solvent": meaningful_tokens(tokenize(row.solvent)),
        "drug": meaningful_tokens(tokenize(row.drug)),
        "drug_amount": parse_numeric_tokens(row.drug_amount),
        "ee": parse_numeric_tokens(row.ee),
    }


def gt_search_text(row: GTRow) -> str:
    return normalize_token(
        " ".join(
            part
            for part in [
                row.formulation_label_raw,
                row.notes,
                row.source_type,
                row.source_locator,
                row.extra_alignment_text,
            ]
            if part
        )
    )


def score_pair(final_row: FinalRow, gt_row: GTRow) -> tuple[int, list[str]]:
    search_text = gt_search_text(gt_row)
    signals = final_row_signal_tokens(final_row)
    score = 0
    reasons: list[str] = []

    for field, weight in [
        ("drug", 5),
        ("polymer", 4),
        ("surfactant", 3),
        ("solvent", 2),
        ("label", 2),
    ]:
        matches = [token for token in signals[field] if token in search_text]
        if matches:
            score += min(weight, len(matches) * max(1, weight // 2))
            reasons.append(f"{field}={','.join(matches[:4])}")

    for field, weight in [
        ("polymer_mw", 2),
        ("la_ga", 2),
        ("surfactant_concentration", 1),
        ("pva", 1),
        ("drug_amount", 1),
        ("ee", 1),
    ]:
        matches = [token for token in signals[field] if token in search_text]
        if matches:
            score += weight
            reasons.append(f"{field}={','.join(matches[:3])}")

    return score, reasons


def compare_status_from_score(score: int) -> str:
    if score >= 6:
        return "matched"
    if score > 0:
        return "needs_manual_review"
    return "unresolved_boundary"


def build_layer2_rows_for_key(key: str, final_rows: list[FinalRow], gt_rows: list[GTRow]) -> list[dict[str, Any]]:
    candidate_pairs: list[dict[str, Any]] = []
    for final_row in final_rows:
        for gt_row in gt_rows:
            score, reasons = score_pair(final_row, gt_row)
            candidate_pairs.append(
                {
                    "key": key,
                    "nvidia_row_id": final_row.nvidia_row_id,
                    "matched_gt_row_id": gt_row.formulation_id,
                    "score": score,
                    "reasons": reasons,
                }
            )

    assigned_nvidia: set[str] = set()
    assigned_gt: set[str] = set()
    output_rows: list[dict[str, Any]] = []

    sorted_pairs = sorted(
        [row for row in candidate_pairs if row["score"] > 0],
        key=lambda row: (-int(row["score"]), row["nvidia_row_id"], row["matched_gt_row_id"]),
    )
    for pair in sorted_pairs:
        if pair["nvidia_row_id"] in assigned_nvidia or pair["matched_gt_row_id"] in assigned_gt:
            continue
        status = compare_status_from_score(int(pair["score"]))
        output_rows.append(
            {
                "key": key,
                "nvidia_row_id": pair["nvidia_row_id"],
                "matched_gt_row_id": pair["matched_gt_row_id"],
                "comparison_status": status,
                "notes": (
                    "token-based best-effort match"
                    + f"; score={pair['score']}"
                    + (f"; reasons={','.join(pair['reasons'])}" if pair["reasons"] else "")
                ),
            }
        )
        assigned_nvidia.add(pair["nvidia_row_id"])
        assigned_gt.add(pair["matched_gt_row_id"])

    remaining_nvidia = [row for row in final_rows if row.nvidia_row_id not in assigned_nvidia]
    remaining_gt = [row for row in gt_rows if row.formulation_id not in assigned_gt]
    fallback_pairs = min(len(remaining_nvidia), len(remaining_gt))
    for index in range(fallback_pairs):
        final_row = remaining_nvidia[index]
        gt_row = remaining_gt[index]
        output_rows.append(
            {
                "key": key,
                "nvidia_row_id": final_row.nvidia_row_id,
                "matched_gt_row_id": gt_row.formulation_id,
                "comparison_status": "unresolved_boundary",
                "notes": "count-aligned order fallback with no strong token match; manual review recommended",
            }
        )
        assigned_nvidia.add(final_row.nvidia_row_id)
        assigned_gt.add(gt_row.formulation_id)

    for final_row in final_rows:
        if final_row.nvidia_row_id in assigned_nvidia:
            continue
        output_rows.append(
            {
                "key": key,
                "nvidia_row_id": final_row.nvidia_row_id,
                "matched_gt_row_id": "",
                "comparison_status": "extra_candidate",
                "notes": "unmatched NVIDIA final-table row after best-effort token and order comparison",
            }
        )

    for gt_row in gt_rows:
        if gt_row.formulation_id in assigned_gt:
            continue
        output_rows.append(
            {
                "key": key,
                "nvidia_row_id": "",
                "matched_gt_row_id": gt_row.formulation_id,
                "comparison_status": "missing_vs_gt",
                "notes": "GT authority row remained unmatched after best-effort token and order comparison",
            }
        )

    return sorted(
        output_rows,
        key=lambda row: (
            row["key"],
            row["comparison_status"],
            row["nvidia_row_id"] or "ZZZ",
            row["matched_gt_row_id"] or "ZZZ",
        ),
    )


def priority_sort_key(row: dict[str, Any]) -> tuple[int, int, int, str]:
    return (
        -int(row["_abs_delta"]),
        -int(row["n_needs_manual_review"]),
        -int(row["n_layer2_extra"]),
        str(row["key"]),
    )


def build_run_context(
    *,
    run_dir: Path,
    full_run_dir: Path,
    source_run_context: dict[str, Any],
    final_table_tsv: Path,
    scope_manifest_tsv: Path,
    gt_workbook_xlsx: Path,
    gt_skeleton_tsv: Path | None,
    alignment_scaffold_tsv: Path | None,
) -> str:
    gt_skeleton_line = (
        f"- gt_skeleton_tsv: `{gt_skeleton_tsv}`" if gt_skeleton_tsv is not None else "- gt_skeleton_tsv: `not resolved`"
    )
    alignment_line = (
        f"- alignment_scaffold_tsv: `{alignment_scaffold_tsv}`" if alignment_scaffold_tsv is not None else "- alignment_scaffold_tsv: `not resolved`"
    )
    return "\n".join(
        [
            "# RUN_CONTEXT",
            "",
            "## 1. Run Type",
            "",
            "- `diagnostic-only`",
            "",
            "## 2. Purpose",
            "",
            "- Diagnostic-only GT evaluation of the NVIDIA-backed full pipeline final output on DEV15.",
            "- The evaluation object is the completed Stage5 final table only.",
            "",
            "## 3. Source Full Pipeline Run",
            "",
            f"- full_run_dir: `{full_run_dir}`",
            f"- final_table_tsv: `{final_table_tsv}`",
            "",
            "## 4. GT Authority Surfaces Used",
            "",
            f"- source_resolution: `{source_run_context['resolution_source']}`",
            f"- source_run_id: `{source_run_context['run_id']}`",
            f"- scope_manifest_tsv: `{scope_manifest_tsv}`",
            f"- gt_workbook_xlsx: `{gt_workbook_xlsx}`",
            gt_skeleton_line,
            alignment_line,
            "",
            "## 5. Comparison Limitations",
            "",
            "- Layer1 uses the completed final-table counts against the reviewed GT workbook.",
            "- Layer2 is best-effort diagnostic only when exact final-row identity alignment is not provable from the available final-table fields.",
            "- No official benchmark status is claimed here.",
            "",
            "## 6. Final Outputs",
            "",
            f"- `{run_dir / 'RUN_CONTEXT.md'}`",
            f"- `{run_dir / 'layer1_count_comparison.tsv'}`",
            f"- `{run_dir / 'layer2_identity_comparison.tsv'}`",
            f"- `{run_dir / 'per_paper_summary.tsv'}`",
            f"- `{run_dir / 'audit_priority.tsv'}`",
            f"- `{run_dir / 'evaluation_summary.md'}`",
        ]
    ) + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Evaluate a completed NVIDIA full-pipeline Stage5 final table against DEV15 GT."
    )
    parser.add_argument("--full-run-dir", required=True, type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    full_run_dir = args.full_run_dir.resolve()
    require_exists(full_run_dir, "full pipeline run directory")
    final_table_tsv = full_run_dir / "final_formulation_table_v1.tsv"
    require_exists(final_table_tsv, "final formulation table")

    source_run_context = resolve_run_context()
    scope_manifest_tsv = resolve_artifact_path(
        explicit_path=None,
        run_context=source_run_context,
        pointer_key="scope_manifest_tsv",
        required=True,
    )
    gt_workbook_xlsx = resolve_artifact_path(
        explicit_path=None,
        run_context=source_run_context,
        pointer_key="gt_workbook_xlsx",
        required=True,
    )
    gt_skeleton_tsv = resolve_artifact_path(
        explicit_path=None,
        run_context=source_run_context,
        pointer_key="gt_skeleton_tsv",
        required=False,
    )
    alignment_scaffold_tsv = resolve_artifact_path(
        explicit_path=None,
        run_context=source_run_context,
        pointer_key="alignment_scaffold_tsv",
        required=False,
    )
    assert scope_manifest_tsv is not None
    assert gt_workbook_xlsx is not None

    require_exists(scope_manifest_tsv, "scope manifest TSV")
    require_exists(gt_workbook_xlsx, "GT workbook XLSX")
    if alignment_scaffold_tsv is not None:
        require_exists(alignment_scaffold_tsv, "alignment scaffold TSV")

    run_dir = make_run_dir()
    run_dir.mkdir(parents=True, exist_ok=False)

    print(f"Resolved full pipeline run directory: {full_run_dir}")
    print(f"Resolved final table: {final_table_tsv}")
    print(f"Resolved scope manifest: {scope_manifest_tsv}")
    print(f"Resolved GT workbook: {gt_workbook_xlsx}")
    if alignment_scaffold_tsv is not None:
        print(f"Resolved alignment scaffold: {alignment_scaffold_tsv}")
    print(f"Output evaluation directory: {run_dir}")

    manifest_rows = read_tsv_rows(scope_manifest_tsv)
    manifest_keys = [
        normalize_text(row.get("key") or row.get("zotero_key") or row.get("paper_key"))
        for row in manifest_rows
        if normalize_text(row.get("key") or row.get("zotero_key") or row.get("paper_key"))
    ]
    alignment_rows = read_tsv_rows(alignment_scaffold_tsv) if alignment_scaffold_tsv is not None else []
    gt_rows = load_gt_rows(gt_workbook_xlsx, alignment_rows)
    final_rows = load_final_rows(final_table_tsv)

    gt_rows_by_key = group_by_key(gt_rows, "key")
    final_rows_by_key = group_by_key(final_rows, "key")
    gt_counts = build_gt_counts(gt_rows)

    layer1_rows: list[dict[str, Any]] = []
    layer2_rows: list[dict[str, Any]] = []
    per_paper_rows: list[dict[str, Any]] = []

    for key in manifest_keys:
        gt_rows_for_key = gt_rows_by_key.get(key, [])
        final_rows_for_key = final_rows_by_key.get(key, [])
        gt_count = int(gt_counts.get(key, 0))
        nvidia_count = len(final_rows_for_key)
        delta = nvidia_count - gt_count
        abs_delta = abs(delta)
        exact_match = "yes" if delta == 0 else "no"

        layer1_rows.append(
            {
                "key": key,
                "gt_count": gt_count,
                "nvidia_count": nvidia_count,
                "delta": delta,
                "abs_delta": abs_delta,
                "exact_match": exact_match,
            }
        )

        per_key_layer2_rows = build_layer2_rows_for_key(key, final_rows_for_key, gt_rows_for_key)
        layer2_rows.extend(per_key_layer2_rows)

        n_layer2_matched = sum(1 for row in per_key_layer2_rows if row["comparison_status"] == "matched")
        n_layer2_extra = sum(1 for row in per_key_layer2_rows if row["comparison_status"] == "extra_candidate")
        n_layer2_missing = sum(1 for row in per_key_layer2_rows if row["comparison_status"] == "missing_vs_gt")
        n_needs_manual_review = sum(
            1
            for row in per_key_layer2_rows
            if row["comparison_status"] in {"needs_manual_review", "unresolved_boundary"}
        )

        per_paper_rows.append(
            {
                "key": key,
                "gt_count": gt_count,
                "nvidia_count": nvidia_count,
                "exact_match": exact_match,
                "n_layer2_matched": n_layer2_matched,
                "n_layer2_extra": n_layer2_extra,
                "n_layer2_missing": n_layer2_missing,
                "n_needs_manual_review": n_needs_manual_review,
                "priority_rank": 0,
                "_abs_delta": abs_delta,
            }
        )

        print(
            f"{key}: gt_count={gt_count} nvidia_count={nvidia_count} "
            f"exact_match={exact_match} layer2_matched={n_layer2_matched} "
            f"extra={n_layer2_extra} missing={n_layer2_missing} manual_review={n_needs_manual_review}"
        )

    sorted_per_paper = sorted(per_paper_rows, key=priority_sort_key)
    for index, row in enumerate(sorted_per_paper, start=1):
        row["priority_rank"] = index

    audit_priority_rows = [
        {
            "priority_rank": row["priority_rank"],
            "key": row["key"],
            "gt_count": row["gt_count"],
            "nvidia_count": row["nvidia_count"],
            "abs_delta": row["_abs_delta"],
            "n_needs_manual_review": row["n_needs_manual_review"],
            "n_layer2_extra": row["n_layer2_extra"],
            "n_layer2_missing": row["n_layer2_missing"],
            "exact_match": row["exact_match"],
        }
        for row in sorted_per_paper
    ]

    total_gt_rows = sum(int(row["gt_count"]) for row in layer1_rows)
    total_nvidia_rows = sum(int(row["nvidia_count"]) for row in layer1_rows)
    layer1_exact_matches = sum(1 for row in layer1_rows if row["exact_match"] == "yes")
    top_priority = [row["key"] for row in audit_priority_rows[:5]]

    summary_lines = [
        "# NVIDIA Full Pipeline GT Evaluation Summary",
        "",
        "- Status: `diagnostic-only, not benchmark-valid final output`.",
        f"- Source final table: `{final_table_tsv}`",
        f"- GT workbook: `{gt_workbook_xlsx}`",
        "",
        "## Layer1",
        "",
        f"- papers with exact Layer1 count match: `{layer1_exact_matches}/{len(layer1_rows)}`",
        f"- total GT rows: `{total_gt_rows}`",
        f"- total NVIDIA final rows: `{total_nvidia_rows}`",
        "",
        "## Highest Priority Papers",
        "",
    ]
    for row in audit_priority_rows[:5]:
        summary_lines.append(
            "- `{key}`: abs_delta=`{abs_delta}` manual_review=`{n_needs_manual_review}` extra=`{n_layer2_extra}` missing=`{n_layer2_missing}`".format(
                **row
            )
        )
    summary_lines.extend(
        [
            "",
            "## Caveat",
            "",
            "- Layer2 identity comparison is best-effort diagnostic only when final-table fields do not fully prove GT row identity.",
            "- This evaluation compares only the completed Stage5 final output to GT and does not use intermediate Stage2 surfaces as the benchmark object.",
        ]
    )

    for row in sorted_per_paper:
        row.pop("_abs_delta", None)

    write_tsv(
        run_dir / "layer1_count_comparison.tsv",
        ["key", "gt_count", "nvidia_count", "delta", "abs_delta", "exact_match"],
        layer1_rows,
    )
    write_tsv(
        run_dir / "layer2_identity_comparison.tsv",
        ["key", "nvidia_row_id", "matched_gt_row_id", "comparison_status", "notes"],
        layer2_rows,
    )
    write_tsv(
        run_dir / "per_paper_summary.tsv",
        [
            "key",
            "gt_count",
            "nvidia_count",
            "exact_match",
            "n_layer2_matched",
            "n_layer2_extra",
            "n_layer2_missing",
            "n_needs_manual_review",
            "priority_rank",
        ],
        sorted_per_paper,
    )
    write_tsv(
        run_dir / "audit_priority.tsv",
        [
            "priority_rank",
            "key",
            "gt_count",
            "nvidia_count",
            "abs_delta",
            "n_needs_manual_review",
            "n_layer2_extra",
            "n_layer2_missing",
            "exact_match",
        ],
        audit_priority_rows,
    )
    (run_dir / "evaluation_summary.md").write_text("\n".join(summary_lines) + "\n", encoding="utf-8")
    (run_dir / "RUN_CONTEXT.md").write_text(
        build_run_context(
            run_dir=run_dir,
            full_run_dir=full_run_dir,
            source_run_context=source_run_context,
            final_table_tsv=final_table_tsv,
            scope_manifest_tsv=scope_manifest_tsv,
            gt_workbook_xlsx=gt_workbook_xlsx,
            gt_skeleton_tsv=gt_skeleton_tsv,
            alignment_scaffold_tsv=alignment_scaffold_tsv,
        ),
        encoding="utf-8",
    )

    print(
        f"Summary metrics: layer1_exact_matches={layer1_exact_matches}/{len(layer1_rows)} "
        f"total_gt_rows={total_gt_rows} total_nvidia_rows={total_nvidia_rows}"
    )
    print(f"Highest priority papers: {', '.join(top_priority)}")
    print(f"Outputs written to {run_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
