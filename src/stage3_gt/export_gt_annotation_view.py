#!/usr/bin/env python3
"""
export_gt_annotation_view.py

One-way export:
    authoritative GT decision TSV  ->  annotation-friendly Excel (XLSX)

Design goals
- NEVER edit the authoritative TSV by hand.
- Excel file is a UI for human annotation only.
- Only three fields are intended for editing:
    gt_decision, gt_value_text, gt_notes

Typical usage (from repo root):
    python src/stage3_gt/export_gt_annotation_view.py \
        --input-tsv data/cleaned/labels/manual/gt_field_decisions__run_XXXX.tsv \
        --out-xlsx  data/cleaned/labels/manual/gt_annotation_view__run_XXXX.xlsx

Notes
- Robust TSV parsing: supports quoted fields containing newlines.
- Adds Excel dropdown validation for gt_decision and highlights override rows.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import List, Dict, Tuple, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# Prefer centralized paths API when running inside the repo
try:
    from src.utils import paths as project_paths  # type: ignore
except Exception:
    project_paths = None


DEFAULT_DECISIONS = ["accept_model1", "accept_model2", "override", "unclear"]


def _read_tsv_robust(tsv_path: Path) -> pd.DataFrame:
    """
    Read TSV using Python's csv module to correctly handle multiline quoted fields.
    Returns a DataFrame with all columns as strings (no NA coercion).
    """
    rows: List[Dict[str, str]] = []
    with tsv_path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.DictReader(
            f,
            delimiter="\t",
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL,
            doublequote=True,
            escapechar=None,
        )
        if reader.fieldnames is None:
            raise ValueError(f"TSV has no header: {tsv_path}")

        for r in reader:
            # Normalize None -> "" to avoid pandas NaN surprises
            rows.append({k: (v if v is not None else "") for k, v in r.items()})

    df = pd.DataFrame(rows)
    # Ensure deterministic column order as in the file
    df = df[[c for c in reader.fieldnames]]  # type: ignore[arg-type]
    return df


def _pick_columns(df: pd.DataFrame) -> Tuple[List[str], List[str], List[str]]:
    """
    Returns (readonly_cols, editable_cols, missing_cols)
    We try to be flexible, but keep a stable view schema when possible.
    """
    # Candidate ID/join keys
    candidate_keys = ["key", "formulation_id", "field_name"]

    # Candidate model outputs and evidence
    candidate_readonly = [
        "model1",
        "model2",
        "preferred_model",
        "preferred_model_value",
        "preferred_value",
        "value_model1",
        "value_model2",
        "evidence_section_main",
        "evidence_span_text_main",
    ]

    editable = ["gt_decision", "gt_value_text", "gt_notes"]

    readonly: List[str] = []
    missing: List[str] = []

    for c in candidate_keys:
        if c in df.columns:
            readonly.append(c)
        else:
            missing.append(c)

    for c in candidate_readonly:
        if c in df.columns and c not in readonly:
            readonly.append(c)

    for c in editable:
        if c not in df.columns:
            df[c] = ""
        # editable cols always present
    return readonly, editable, missing


def _apply_excel_ui(xlsx_path: Path, sheet_name: str, editable_cols: List[str], decisions: List[str]) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    # Header row styling
    header_fill = PatternFill("solid", fgColor="1F2937")  # dark gray
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Column index mapping (1-based)
    col_index = {ws.cell(row=1, column=j).value: j for j in range(1, ws.max_column + 1)}

    # Freeze panes: freeze all readonly columns (left side) + header row
    # i.e., first editable column should be visible as the first scrollable column.
    first_editable = min(col_index[c] for c in editable_cols if c in col_index)
    ws.freeze_panes = ws.cell(row=2, column=first_editable)

    # Set reasonable widths (avoid auto-fit complexity; keep stable)
    for name, j in col_index.items():
        if name in editable_cols:
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 18
        elif name in ("evidence_span_text_main",):
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 70
        else:
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 22

    # Wrap long text
    wrap_cols = ["evidence_span_text_main", "gt_notes", "gt_value_text"]
    for name in wrap_cols:
        if name in col_index:
            j = col_index[name]
            for i in range(2, ws.max_row + 1):
                ws.cell(row=i, column=j).alignment = Alignment(wrap_text=True, vertical="top")

    # Shade readonly vs editable columns
    readonly_fill = PatternFill("solid", fgColor="F3F4F6")  # light gray
    editable_fill = PatternFill("solid", fgColor="FFF7ED")  # light orange
    for name, j in col_index.items():
        fill = editable_fill if name in editable_cols else readonly_fill
        for i in range(2, ws.max_row + 1):
            ws.cell(row=i, column=j).fill = fill

    # Dropdown validation for gt_decision
    if "gt_decision" in col_index:
        j = col_index["gt_decision"]
        formula = '"' + ",".join(decisions) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=True)
        dv.error = "Invalid gt_decision. Please choose from the dropdown."
        dv.errorTitle = "Invalid value"
        dv.prompt = "Choose one decision value."
        dv.promptTitle = "gt_decision"
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=j).coordinate}:{ws.cell(row=ws.max_row, column=j).coordinate}")

    # Highlight override rows (entire row)
    if "gt_decision" in col_index:
        j = col_index["gt_decision"]
        # Apply a pale yellow fill to the full row when gt_decision == "override"
        highlight_fill = PatternFill("solid", fgColor="FEF9C3")  # pale yellow
        # Formula: $<col_letter>2="override"
        col_letter = ws.cell(row=1, column=j).column_letter
        rule = FormulaRule(formula=[f'${col_letter}2="override"'], fill=highlight_fill)
        ws.conditional_formatting.add(f"A2:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}", rule)

    wb.save(xlsx_path)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-tsv", type=str, required=True, help="Authoritative TSV (read-only).")
    ap.add_argument("--out-xlsx", type=str, required=True, help="Annotation view XLSX to create/overwrite.")
    ap.add_argument("--sheet", type=str, default="annotation", help="Worksheet name.")
    ap.add_argument(
        "--decisions",
        type=str,
        default=",".join(DEFAULT_DECISIONS),
        help="Comma-separated allowed gt_decision values.",
    )
    ap.add_argument("--overwrite", action="store_true", help="Allow overwriting out-xlsx.")
    args = ap.parse_args()

    in_path = Path(args.input_tsv)
    out_path = Path(args.out_xlsx)

    if not in_path.exists():
        raise FileNotFoundError(f"input TSV not found: {in_path}")

    if out_path.exists() and not args.overwrite:
        raise FileExistsError(f"out-xlsx exists (use --overwrite): {out_path}")

    decisions = [s.strip() for s in args.decisions.split(",") if s.strip()]
    if not decisions:
        raise ValueError("No decisions provided.")

    df = _read_tsv_robust(in_path)
    readonly_cols, editable_cols, _missing = _pick_columns(df)

    view_cols = readonly_cols + editable_cols
    df_view = df[view_cols].copy()

    # Write with openpyxl engine to allow later styling
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_view.to_excel(writer, sheet_name=args.sheet, index=False)

    _apply_excel_ui(out_path, args.sheet, editable_cols=editable_cols, decisions=decisions)

    print(f"[OK] wrote annotation view: {out_path}")
    print(f"[info] editable columns: {editable_cols}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
