#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from src.utils import paths as P
from src.utils.run_id import is_valid_run_id
from src.utils.run_latest import inputs_fingerprint, write_latest


MAX_CELL_TEXT = 60000


def _sanitize_out_subdir(s: str) -> str:
    v = str(s or "").strip().replace("\\", "/")
    if not v:
        raise ValueError(
            "ERROR: --out-subdir is required when reusing a run_id. Use a stage/variant folder name, e.g. stage2_validation or stage5_signature_iter001."
        )
    if Path(v).is_absolute():
        raise ValueError("ERROR: --out-subdir must be a relative path.")
    parts = [p for p in v.split("/") if p]
    if not parts or any(p == ".." for p in parts):
        raise ValueError("ERROR: --out-subdir cannot contain path traversal ('..').")
    return "/".join(parts)


def _read_jsonl(path: Path) -> List[dict]:
    rows: List[dict] = []
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            s = line.strip()
            if not s:
                continue
            rows.append(json.loads(s))
    return rows


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = name.replace("\ufeff", "")
    raw = "".join(ch for ch in cleaned if ch not in r'[]:*?/\\')
    base = raw[:31] if raw else "KEY"
    candidate = base
    idx = 1
    while candidate in used:
        suffix = f"_{idx}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        idx += 1
    used.add(candidate)
    return candidate


def _join_list(v: object) -> str:
    if isinstance(v, list):
        return "; ".join(str(x) for x in v if str(x).strip())
    return ""


def _write_summary(ws, rows: List[dict]) -> None:
    cols = [
        "key",
        "doi",
        "has_text",
        "text_path",
        "preferred_table_source",
        "n_selected_tables",
        "selected_table_files",
        "notes",
    ]
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for r in rows:
        text = str(r.get("source_text", "") or "")
        ws.append(
            [
                str(r.get("key", "")),
                str(r.get("doi", "")),
                bool(text.strip()),
                str(r.get("source_text_path", "")),
                str(r.get("preferred_table_source", "none") or "none"),
                len(r.get("selected_table_files", []) or []),
                _join_list(r.get("selected_table_files", [])),
                _join_list(r.get("notes", [])),
            ]
        )

    widths = {
        "A": 14,
        "B": 22,
        "C": 10,
        "D": 54,
        "E": 20,
        "F": 18,
        "G": 60,
        "H": 60,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_key_sheet(ws, row: dict) -> None:
    key = str(row.get("key", ""))
    doi = str(row.get("doi", ""))
    pref = str(row.get("preferred_table_source", "none") or "none")
    text_path = str(row.get("source_text_path", ""))
    files = row.get("selected_table_files", [])
    if not isinstance(files, list):
        files = []
    notes = row.get("notes", [])
    if not isinstance(notes, list):
        notes = [str(notes)]
    src_text = str(row.get("source_text", "") or "")

    ws["A1"] = "Key"
    ws["B1"] = key
    ws["A2"] = "DOI"
    ws["B2"] = doi
    ws["A3"] = "preferred_table_source"
    ws["B3"] = pref
    ws["A4"] = "source_text_path"
    ws["B4"] = text_path
    ws["A5"] = "selected_table_files"
    ws["B5"] = _join_list(files)
    ws["A6"] = "notes"
    ws["B6"] = _join_list(notes)

    for r in range(1, 7):
        ws[f"A{r}"].font = Font(bold=True)

    ws["A8"] = "Cleaned Text"
    ws["A8"].font = Font(bold=True)
    truncated = False
    if len(src_text) > MAX_CELL_TEXT:
        src_text = src_text[:MAX_CELL_TEXT]
        truncated = True
    ws["A9"] = src_text
    ws["A9"].alignment = Alignment(wrap_text=True, vertical="top")
    if truncated:
        ws["A10"] = f"[TRUNCATED] source_text clipped to first {MAX_CELL_TEXT} chars."

    ws["A12"] = "Selected Tables"
    ws["A12"].font = Font(bold=True)

    tables = row.get("selected_tables_tsv", [])
    if not isinstance(tables, list):
        tables = []

    cur = 13
    if not files:
        ws[f"A{cur}"] = "(no selected tables)"
        cur += 2
    else:
        for i, fname in enumerate(files):
            ws[f"A{cur}"] = f"TABLE {i+1}: {fname}"
            ws[f"A{cur}"].font = Font(bold=True)
            cur += 1

            tsv = str(tables[i] if i < len(tables) else "")
            lines = [ln for ln in tsv.splitlines() if ln is not None]
            for ln in lines:
                cells = ln.split("\t")
                for cidx, val in enumerate(cells, start=1):
                    ws.cell(row=cur, column=cidx, value=val)
                cur += 1
            cur += 1

    ws.freeze_panes = "A13"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 120
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 24


def export_xlsx(jsonl_path: Path, out_xlsx: Path) -> Dict[str, object]:
    rows = _read_jsonl(jsonl_path)
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Summary"
    _write_summary(ws0, rows)

    used_names = {"Summary"}
    for rec in rows:
        key = str(rec.get("key", "")).strip() or "UNKNOWN"
        sheet_name = _safe_sheet_name(f"KEY_{key}", used_names)
        ws = wb.create_sheet(title=sheet_name)
        _write_key_sheet(ws, rec)

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    return {
        "out_xlsx": str(out_xlsx),
        "sheet_names": wb.sheetnames,
        "n_records": len(rows),
        "summary_rows": ws0.max_row - 1,
    }


def main() -> None:
    p = argparse.ArgumentParser(
        description="Export evidence_bundles_v1 JSONL to human-auditable XLSX."
    )
    p.add_argument("--run-id", default="")
    p.add_argument("--in-jsonl", default="")
    p.add_argument("--in-summary-tsv", default="")
    p.add_argument("--out-xlsx", default="")
    p.add_argument(
        "--out-subdir",
        default="",
        help="Optional subdirectory under data/results/<run_id>/ for run variants (e.g., iter_001).",
    )
    args = p.parse_args()
    rid = str(args.run_id).strip()
    if not rid:
        raise ValueError(
            "ERROR: --run-id is required. Generate/reuse a run_id via: python -m src.utils.run_preflight ..."
        )
    if not is_valid_run_id(rid):
        raise ValueError(f"Invalid --run-id (must match required regex): {rid}")
    out_subdir = _sanitize_out_subdir(args.out_subdir)

    base = P.DATA_RESULTS_DIR / rid
    base = base / out_subdir
    base = base / "stage2_validation"
    in_jsonl = Path(args.in_jsonl) if args.in_jsonl.strip() else (base / "evidence_bundles_v1.jsonl")
    _ = Path(args.in_summary_tsv) if args.in_summary_tsv.strip() else (base / "evidence_bundles_v1_summary.tsv")
    out_xlsx = Path(args.out_xlsx) if args.out_xlsx.strip() else (base / "evidence_audit_v1.xlsx")

    if not in_jsonl.exists():
        raise FileNotFoundError(f"evidence JSONL not found: {in_jsonl}")
    if str(args.out_xlsx).strip():
        try:
            out_xlsx.resolve().relative_to(base.resolve())
        except Exception:
            raise ValueError(
                f"ERROR: --out-xlsx must be under data/results/<run_id>/<out-subdir>/. Got: {out_xlsx}"
            )

    latest_path = write_latest(
        run_id=rid,
        meta={
            "subset": "goren2025",
            "stage": "stage2_validation",
            "inputs_fingerprint": inputs_fingerprint([in_jsonl]),
            "note": "export_evidence_bundle_audit_xlsx_v1",
        },
    )

    meta = export_xlsx(in_jsonl, out_xlsx)
    sheet_names = [str(x).replace("\ufeff", "") for x in meta["sheet_names"]]
    print(f"run_id\t{rid}")
    print(f"latest_pointer\t{latest_path}")
    print(f"output_xlsx\t{meta['out_xlsx']}")
    print(f"sheet_names\t{'; '.join(sheet_names)}")
    print(f"n_records\t{meta['n_records']}")
    print(f"summary_rows\t{meta['summary_rows']}")


if __name__ == "__main__":
    main()
