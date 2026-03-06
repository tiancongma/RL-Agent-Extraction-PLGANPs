#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from src.stage2_sampling_labels.auto_extract_weak_labels_v7pilot_r3_fixflat import (
    CORE_FIELDS,
    UNNAMED_LIST_FALLBACK_ORDER,
    _infer_field_name,
    normalize_doi,
)


TARGET_FIELDS = {
    "polymer_name_raw",
    "plga_mw_kDa",
    "la_ga_ratio",
    "encapsulation_efficiency_percent",
    "surfactant_type",
    "surfactant_concentration_text",
}

ALIAS_TO_CORE = {
    "surfactant_type": "surfactant_name",
    "surfactant": "surfactant_name",
    "polymer_name_raw": "plga_mw_kDa",
    "polymer_name": "plga_mw_kDa",
    "polymer_mw": "plga_mw_kDa",
}


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return " ".join(str(v).strip().split())


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="DOI-focused raw field -> flatten mapping audit for v7pilot fixflat run.")
    p.add_argument(
        "--run-dir",
        default="data/results/run_20260306_1552_v7pilot3r3fixflat_dev/weak_labels_v7pilot_r3_fixflat",
    )
    p.add_argument(
        "--doi",
        default="10.1016/j.ejpb.2004.09.002",
    )
    p.add_argument(
        "--out-xlsx",
        default="data/cleaned/labels/manual/doi_10.1016.j.ejpb.2004.09.002_raw_field_mapping_audit.xlsx",
    )
    return p.parse_args()


def resolve_io(run_dir: Path) -> Tuple[Path, Path]:
    json_candidates = [
        run_dir / "weak_labels__v7pilot.jsonl",
        run_dir / "weak_labels__v7pilot_r3_fixflat.jsonl",
        run_dir / "weak_labels__v7pilot_r3.jsonl",
    ]
    tsv_candidates = [
        run_dir / "weak_labels__v7pilot.tsv",
        run_dir / "weak_labels__v7pilot_r3_fixflat.tsv",
        run_dir / "weak_labels__v7pilot_r3.tsv",
    ]
    jsonl = next((p for p in json_candidates if p.exists()), None)
    tsv = next((p for p in tsv_candidates if p.exists()), None)
    if not jsonl or not tsv:
        raise FileNotFoundError(f"Could not resolve jsonl/tsv under: {run_dir}")
    return jsonl, tsv


def map_raw_item(
    item: Dict[str, Any],
    used: set[str],
    fallback_queue: List[str],
) -> Tuple[str, str]:
    raw_key = _norm(item.get("field_name") or item.get("key") or item.get("name") or item.get("field"))
    if raw_key in CORE_FIELDS:
        return raw_key, "exact_key"
    if raw_key and raw_key in ALIAS_TO_CORE:
        core = ALIAS_TO_CORE[raw_key]
        if core not in used:
            return core, "alias_key"
    guessed = _infer_field_name(item)
    if guessed and guessed not in used:
        return guessed, "inferred_from_label"
    while fallback_queue:
        nxt = fallback_queue.pop(0)
        if nxt not in used:
            return nxt, "fallback_order"
    return "", "unknown"


def target_tsv_cols(field: str) -> Tuple[str, str]:
    if field == "surfactant_type":
        return "surfactant_name_value", "surfactant_name_scope"
    if field == "polymer_name_raw":
        return "plga_mw_kDa_value_text", "plga_mw_kDa_scope"
    return f"{field}_value", f"{field}_scope"


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 60)


def main() -> None:
    args = parse_args()
    run_dir = Path(args.run_dir)
    out_xlsx = Path(args.out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    target_doi = normalize_doi(args.doi)

    jsonl_path, tsv_path = resolve_io(run_dir)
    tsv_df = pd.read_csv(tsv_path, sep="\t", dtype=str).fillna("")
    tsv_df["doi_norm"] = tsv_df["doi"].map(normalize_doi)
    tsv_df = tsv_df[tsv_df["doi_norm"] == target_doi].copy()

    raw_rows: List[Dict[str, Any]] = []
    trace_rows: List[Dict[str, Any]] = []

    with jsonl_path.open("r", encoding="utf-8", errors="replace") as f:
        for line in f:
            if not line.strip():
                continue
            obj = json.loads(line)
            doi_norm = normalize_doi(obj.get("doi", ""))
            if doi_norm != target_doi:
                continue
            forms = obj.get("formulations", [])
            for fm in forms:
                if not isinstance(fm, dict):
                    continue
                fid = _norm(fm.get("formulation_id", fm.get("id", "")))
                role = _norm(fm.get("formulation_role", ""))
                fields = fm.get("fields", [])

                used: set[str] = set()
                fallback_queue = list(UNNAMED_LIST_FALLBACK_ORDER)
                if isinstance(fields, dict):
                    ordered = []
                    for idx, (k, v) in enumerate(fields.items(), start=1):
                        if isinstance(v, dict):
                            vv = dict(v)
                        else:
                            vv = {"value": v}
                        vv["field_name"] = k
                        vv["_idx"] = idx
                        ordered.append(vv)
                    fields_list = ordered
                elif isinstance(fields, list):
                    fields_list = fields
                else:
                    fields_list = []

                tsv_match = tsv_df[tsv_df["formulation_id"].astype(str) == fid]
                tsv_row = tsv_match.iloc[0].to_dict() if len(tsv_match) > 0 else {}

                for idx, item in enumerate(fields_list, start=1):
                    if not isinstance(item, dict):
                        continue
                    raw_key = _norm(item.get("field_name") or item.get("key") or item.get("name") or item.get("field"))
                    raw_label = _norm(item.get("field_label") or item.get("label"))
                    raw_val = _norm(item.get("value"))
                    raw_scope = _norm(item.get("scope"))
                    raw_membership = _norm(item.get("membership_confidence"))
                    raw_evidence = _norm(item.get("evidence_region_type"))
                    raw_quote = _norm(item.get("evidence_quote") or item.get("value_text"))

                    mapped_target, method = map_raw_item(item, used, fallback_queue)
                    if mapped_target:
                        used.add(mapped_target)
                    mapped_tsv_value = ""
                    mapped_tsv_scope = ""
                    if mapped_target:
                        vcol, scol = target_tsv_cols(mapped_target)
                        mapped_tsv_value = _norm(tsv_row.get(vcol, ""))
                        mapped_tsv_scope = _norm(tsv_row.get(scol, ""))

                    raw_rows.append(
                        {
                            "doi": target_doi,
                            "formulation_id": fid,
                            "formulation_role": role,
                            "raw_field_index": idx,
                            "raw_field_key": raw_key,
                            "raw_field_label": raw_label,
                            "raw_value_raw": raw_val,
                            "raw_scope": raw_scope,
                            "raw_membership_confidence": raw_membership,
                            "raw_evidence_region_type": raw_evidence,
                            "raw_evidence_quote": raw_quote,
                        }
                    )
                    trace_rows.append(
                        {
                            "doi": target_doi,
                            "formulation_id": fid,
                            "raw_field_index": idx,
                            "raw_field_key": raw_key,
                            "raw_field_label": raw_label,
                            "raw_value_raw": raw_val,
                            "mapped_target_field": mapped_target if mapped_target else "",
                            "mapping_method": method,
                            "mapped_tsv_value": mapped_tsv_value,
                            "mapped_tsv_scope": mapped_tsv_scope,
                        }
                    )

    raw_df = pd.DataFrame(raw_rows)
    trace_df = pd.DataFrame(trace_rows)

    no_stable_key = int((trace_df["raw_field_key"].fillna("").str.strip() == "").sum()) if not trace_df.empty else 0
    method_counts = trace_df["mapping_method"].value_counts().to_dict() if not trace_df.empty else {}

    suspicious = trace_df[
        (trace_df["mapped_target_field"].isin(TARGET_FIELDS))
        & (
            trace_df["mapping_method"].isin(["fallback_order", "unknown"])
            | (trace_df["mapped_tsv_value"].fillna("").str.strip() == "")
        )
    ].copy()
    suspicious = suspicious.sort_values(["formulation_id", "raw_field_index"]).head(20)

    summary_rows: List[Dict[str, Any]] = [
        {"metric": "doi", "value": target_doi},
        {"metric": "jsonl_path", "value": str(jsonl_path.resolve())},
        {"metric": "tsv_path", "value": str(tsv_path.resolve())},
        {"metric": "formulation_count", "value": int(raw_df["formulation_id"].nunique()) if not raw_df.empty else 0},
        {"metric": "raw_field_object_count", "value": int(len(raw_df))},
        {"metric": "raw_field_objects_with_no_stable_key", "value": no_stable_key},
        {"metric": "mapped_by_exact_key", "value": int(method_counts.get("exact_key", 0))},
        {"metric": "mapped_by_alias_key", "value": int(method_counts.get("alias_key", 0))},
        {"metric": "mapped_by_inferred_from_label", "value": int(method_counts.get("inferred_from_label", 0))},
        {"metric": "mapped_by_fallback_order", "value": int(method_counts.get("fallback_order", 0))},
        {"metric": "mapped_by_unknown", "value": int(method_counts.get("unknown", 0))},
    ]
    summary_df = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xw:
        raw_df.to_excel(xw, sheet_name="RAW_FIELD_OBJECTS", index=False)
        trace_df.to_excel(xw, sheet_name="MAPPING_TRACE", index=False)
        summary_df.to_excel(xw, sheet_name="SUMMARY", index=False)
        if not suspicious.empty:
            suspicious.to_excel(xw, sheet_name="SUMMARY", index=False, startrow=len(summary_df) + 3)
            ws = xw.book["SUMMARY"]
            ws.cell(row=len(summary_df) + 3, column=1, value="SUSPICIOUS_MAPPINGS")

    wb = load_workbook(out_xlsx)
    for name in ["RAW_FIELD_OBJECTS", "MAPPING_TRACE", "SUMMARY"]:
        style_sheet(wb[name])
    wb.save(out_xlsx)

    top_susp = suspicious.head(10).to_dict(orient="records") if not suspicious.empty else []
    print(f"workbook_path={out_xlsx.resolve()}")
    print(f"formulations={int(raw_df['formulation_id'].nunique()) if not raw_df.empty else 0}")
    print(f"raw_field_objects={len(raw_df)}")
    print(f"fallback_order_mappings={int(method_counts.get('fallback_order', 0))}")
    print(f"top_suspicious_mappings={json.dumps(top_susp, ensure_ascii=False)}")


if __name__ == "__main__":
    main()
