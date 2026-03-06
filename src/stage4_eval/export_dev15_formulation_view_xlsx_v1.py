#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment


REPO_ROOT = Path.cwd()
OUT_XLSX = REPO_ROOT / "data/cleaned/labels/manual/dev15_extracted_formulation_view.xlsx"


def norm_text(v: object) -> str:
    if v is None or pd.isna(v):
        return ""
    return re.sub(r"\s+", " ", str(v).strip().lower())


def norm_doi(v: object) -> str:
    s = norm_text(v)
    s = re.sub(r"^doi\s*:\s*", "", s)
    s = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", s)
    s = re.sub(r"^doi\.org/", "", s)
    return s.strip()


def safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]\:\*\?\/\\]", "_", name).strip()
    if not base:
        base = "unknown_doi"
    base = base[:31]
    out = base
    i = 1
    while out in used:
        suffix = f"_{i}"
        out = f"{base[: max(0, 31 - len(suffix))]}{suffix}"
        i += 1
    used.add(out)
    return out


def print_header_and_first3(df: pd.DataFrame) -> None:
    print("header:")
    print("\t".join(df.columns.tolist()))
    print("first_3_rows:")
    if df.empty:
        print("<empty>")
        return
    for _, row in df.head(3).iterrows():
        vals = ["" if pd.isna(x) else str(x) for x in row.tolist()]
        print("\t".join(vals))


def discover_dev15_definition() -> Tuple[Path, pd.DataFrame, List[Path]]:
    refs: List[Path] = []
    search_roots = [
        REPO_ROOT / "docs/snapshots",
        REPO_ROOT / "project",
        REPO_ROOT / "data/results",
    ]
    patt = re.compile(r"(dev\s*15|dev15|dev_manifest_v1|15\s*doi|diagnostic subset)", re.IGNORECASE)
    for root in search_roots:
        if not root.exists():
            continue
        for p in root.rglob("*"):
            if not p.is_file():
                continue
            if p.suffix.lower() not in {".md", ".tsv", ".txt", ".jsonl"}:
                continue
            try:
                txt = p.read_text(encoding="utf-8", errors="replace")
            except Exception:
                continue
            if patt.search(txt):
                refs.append(p)

    manifest_candidates = [
        REPO_ROOT / "data/cleaned/goren_2025/index/splits/dev_manifest_v1.tsv",
        *sorted((REPO_ROOT / "data/results").glob("run_*/**/dev_manifest_v1.tsv")),
    ]
    for p in manifest_candidates:
        if not p.exists():
            continue
        df = pd.read_csv(p, sep="\t", dtype=str).fillna("")
        doi_col = "doi" if "doi" in df.columns else None
        if doi_col is None:
            continue
        n_doi = df[doi_col].map(norm_doi).replace("", pd.NA).dropna().nunique()
        if n_doi == 15:
            df["doi"] = df["doi"].map(norm_doi)
            return p, df, refs
    raise RuntimeError("Could not find a dev manifest with 15 unique DOIs.")


def candidate_source_paths() -> List[Path]:
    roots = [REPO_ROOT / "data/results", REPO_ROOT / "data/benchmark/goren_2025"]
    patterns = [
        "**/extracted_formulation_level*.tsv",
        "**/formulation_core.tsv",
        "**/measurements.tsv",
        "**/doi_level_ee_scaffold*.tsv",
        "**/modeling_ready*.tsv",
        "**/projected_to_curated*.tsv",
        "**/weak_labels__gemini.tsv",
    ]
    out: List[Path] = []
    seen: set[str] = set()
    for root in roots:
        if not root.exists():
            continue
        for pat in patterns:
            for p in root.glob(pat):
                k = str(p.resolve()).lower()
                if k in seen:
                    continue
                seen.add(k)
                out.append(p.resolve())
    return sorted(out)


def pick_best_source(dev_manifest: pd.DataFrame) -> Tuple[Path, pd.DataFrame]:
    dev_keys = set(dev_manifest.get("zotero_key", pd.Series(dtype=str)).astype(str).str.strip().tolist())
    best_path: Optional[Path] = None
    best_df: Optional[pd.DataFrame] = None
    best_score: float = -1.0

    for p in candidate_source_paths():
        try:
            df = pd.read_csv(p, sep="\t", dtype=str).fillna("")
        except Exception:
            continue
        if df.empty:
            continue
        cols = set(df.columns)
        key_col = "key" if "key" in cols else "zotero_key" if "zotero_key" in cols else "doc_key" if "doc_key" in cols else None
        overlap = 0
        if key_col:
            overlap = len(set(df[key_col].astype(str).str.strip().tolist()) & dev_keys)

        core_cols = [
            "drug_name",
            "plga_mw_kDa",
            "la_ga_ratio",
            "drug_feed_amount_text",
            "plga_mass_mg",
            "surfactant_name",
            "surfactant_concentration_text",
            "organic_solvent",
            "emul_type",
            "emul_method",
            "size_nm",
            "pdi",
        ]
        core_present = sum(1 for c in core_cols if c in cols)
        ee_present = 1 if ("encapsulation_efficiency_percent" in cols or "EE" in cols) else 0
        formulation_id_present = 1 if "formulation_id" in cols else 0
        source_bonus = 2 if "weak_labels__gemini.tsv" in p.name else 0

        score = 4 * formulation_id_present + 3 * ee_present + 0.8 * core_present + 0.2 * overlap + source_bonus
        if score > best_score:
            best_score = score
            best_path = p
            best_df = df

    if best_path is None or best_df is None:
        raise RuntimeError("No suitable formulation-level source TSV found.")
    return best_path, best_df


def find_instance_assignment(source_path: Path) -> Optional[Path]:
    direct = source_path.parent / "formulation_core_signature_v1" / "instance_assignment_v1.tsv"
    if direct.exists():
        return direct
    run_root = None
    for anc in source_path.parents:
        if anc.name.startswith("run_"):
            run_root = anc
            break
    if run_root:
        matches = sorted(run_root.glob("**/formulation_core_signature_v1/instance_assignment_v1.tsv"))
        if matches:
            return matches[0].resolve()
    all_matches = sorted((REPO_ROOT / "data/results").glob("run_*/**/formulation_core_signature_v1/instance_assignment_v1.tsv"))
    return all_matches[0].resolve() if all_matches else None


def build_signature_hash_from_core(row: pd.Series) -> str:
    parts = [
        norm_text(row.get("drug_name", "")),
        norm_text(row.get("plga_mw", "")),
        norm_text(row.get("la_ga_ratio", "")),
        norm_text(row.get("drug_polymer_ratio", "")),
        norm_text(row.get("surfactant_type", "")),
        norm_text(row.get("organic_solvent", "")),
    ]
    sig = "|".join(parts)
    return hashlib.sha1(sig.encode("utf-8")).hexdigest()


def build_view_dataframe(source: pd.DataFrame, dev_manifest: pd.DataFrame, inst_assign: Optional[pd.DataFrame]) -> pd.DataFrame:
    df = source.copy()
    rename_key = None
    if "key" in df.columns:
        rename_key = "key"
    elif "doc_key" in df.columns:
        rename_key = "doc_key"
    elif "zotero_key" in df.columns:
        rename_key = "zotero_key"
    if rename_key is None:
        raise RuntimeError("Source table has no key/doc_key/zotero_key column.")
    if rename_key != "zotero_key":
        df = df.rename(columns={rename_key: "zotero_key"})

    for col in ["zotero_key", "formulation_id"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
        else:
            df[col] = ""

    manifest_use = dev_manifest.copy()
    for col in ["zotero_key", "doi"]:
        manifest_use[col] = manifest_use[col].astype(str).str.strip()

    if "doi" not in df.columns:
        df = df.merge(
            manifest_use[["zotero_key", "doi"]],
            on="zotero_key",
            how="left",
        )
    else:
        df["doi"] = df["doi"].map(norm_doi)
        m = manifest_use.set_index("zotero_key")["doi"].to_dict()
        df["doi"] = df.apply(lambda r: r["doi"] if r["doi"] else m.get(r["zotero_key"], ""), axis=1)

    dev_dois = set(manifest_use["doi"].map(norm_doi).tolist())
    df["doi"] = df["doi"].map(norm_doi)
    df = df[df["doi"].isin(dev_dois)].copy()

    if inst_assign is not None and not inst_assign.empty:
        ia = inst_assign.copy()
        ia = ia.rename(columns={"doc_key": "zotero_key"})
        for col in ["zotero_key", "formulation_id"]:
            if col in ia.columns:
                ia[col] = ia[col].astype(str).str.strip()
        keep_cols = [c for c in ["zotero_key", "formulation_id", "formulation_core_id", "signature_hash", "critical_missing_json"] if c in ia.columns]
        ia = ia[keep_cols].drop_duplicates(subset=["zotero_key", "formulation_id"], keep="first")
        df = df.merge(ia, on=["zotero_key", "formulation_id"], how="left")

    df["ee_value"] = pd.to_numeric(df.get("encapsulation_efficiency_percent", ""), errors="coerce")
    df["ee_evidence_source"] = df.get("evidence_section", "").astype(str).map(
        lambda x: "table" if "table" in x.lower() else ("text" if x.strip() else "")
    )
    df["ee_evidence_pointer"] = df.apply(
        lambda r: "|".join(
            [
                str(r.get("table_filename", "") or ""),
                str(r.get("evidence_section", "") or ""),
                str(r.get("evidence_span_start", "") or ""),
                str(r.get("evidence_span_end", "") or ""),
                str(r.get("evidence_method", "") or ""),
            ]
        ).strip("|"),
        axis=1,
    )

    q_map = {"A": 3, "B": 2, "C": 1, "D": 0}
    df["evidence_support_score"] = df.get("evidence_quality", "").astype(str).str.upper().map(q_map)

    def extract_ratio(v: object) -> str:
        s = "" if pd.isna(v) else str(v)
        m = re.search(r"(\d+(?:\.\d+)?)\s*[:/]\s*(\d+(?:\.\d+)?)", s)
        return f"{m.group(1)}:{m.group(2)}" if m else ""

    df["drug_polymer_ratio"] = df.get("drug_feed_amount_text", "").map(extract_ratio)
    df["drug_mass"] = df.get("drug_feed_amount_text", "")

    out = pd.DataFrame(
        {
            "doi": df.get("doi", ""),
            "zotero_key": df.get("zotero_key", ""),
            "formulation_id": df.get("formulation_id", ""),
            "grouping_id": df.get("formulation_core_id", ""),
            "signature_hash": df.get("signature_hash", ""),
            "derived_flag_summary": "",
            "evidence_support_score": df.get("evidence_support_score", ""),
            "ee_value": df.get("ee_value", ""),
            "ee_evidence_source": df.get("ee_evidence_source", ""),
            "ee_evidence_pointer": df.get("ee_evidence_pointer", ""),
            "drug_name": df.get("drug_name", ""),
            "plga_mw": df.get("plga_mw_kDa", ""),
            "la_ga_ratio": df.get("la_ga_ratio", ""),
            "drug_polymer_ratio": df.get("drug_polymer_ratio", ""),
            "polymer_mass": df.get("plga_mass_mg", ""),
            "drug_mass": df.get("drug_mass", ""),
            "surfactant_type": df.get("surfactant_name", ""),
            "surfactant_conc": df.get("surfactant_concentration_text", ""),
            "organic_solvent": df.get("organic_solvent", ""),
            "emulsion_type": df.get("emul_type", ""),
            "emulsion_method": df.get("emul_method", ""),
            "size_nm": df.get("size_nm", ""),
            "pdi": df.get("pdi", ""),
        }
    )

    out["grouping_id"] = out["grouping_id"].astype(str).str.strip()
    out.loc[out["grouping_id"] == "", "grouping_id"] = out["formulation_id"].astype(str)

    out["signature_hash"] = out["signature_hash"].astype(str).str.strip()
    missing_sig = out["signature_hash"] == ""
    out.loc[missing_sig, "signature_hash"] = out[missing_sig].apply(build_signature_hash_from_core, axis=1)

    return out


def apply_excel_formatting(out_xlsx: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(out_xlsx)
    long_cols = {"ee_evidence_pointer", "derived_flag_summary"}
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row > 1 and ws.max_column > 0:
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            colname_to_idx = {str(v): i + 1 for i, v in enumerate(header) if v is not None}
            for col_name, idx in colname_to_idx.items():
                max_len = len(col_name)
                for r in range(2, ws.max_row + 1):
                    val = ws.cell(row=r, column=idx).value
                    sval = "" if val is None else str(val)
                    max_len = max(max_len, min(len(sval), 100))
                    if col_name in long_cols:
                        ws.cell(row=r, column=idx).alignment = Alignment(wrap_text=True, vertical="top")
                width = 18
                if col_name in long_cols:
                    width = 60
                else:
                    width = max(12, min(max_len + 2, 36))
                ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
            for c in range(1, ws.max_column + 1):
                ws.cell(row=1, column=c).alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(out_xlsx)


def print_step0(step0_manifest: Path, dev_manifest: pd.DataFrame, refs: Iterable[Path]) -> None:
    print("STEP0_DEV15_DEFINITION")
    print(f"manifest_file: {step0_manifest}")
    refs_list = [str(p) for p in refs if "dev_manifest_v1.tsv" in p.name or "dev18" in p.name.lower()]
    if not refs_list:
        refs_list = [str(step0_manifest)]
    print("reference_files:")
    for r in sorted(set(refs_list)):
        print(f"- {r}")
    first5 = dev_manifest["doi"].dropna().astype(str).head(5).tolist()
    print("first_5_dois:")
    for d in first5:
        print(f"- {d}")


def print_step1(source_path: Path, source_df: pd.DataFrame) -> None:
    print("STEP1_CHOSEN_SOURCE")
    print(f"source_file: {source_path}")
    print_header_and_first3(source_df)


def print_verification(view_df: pd.DataFrame) -> None:
    print("STEP5_VERIFICATION_REPORT")
    n_dois = int(view_df["doi"].nunique())
    total_rows = int(len(view_df))
    print(f"n_dois: {n_dois}")
    print(f"total_rows: {total_rows}")
    counts = view_df.groupby("doi", dropna=False).size().reset_index(name="rows").sort_values("rows", ascending=False)
    print("rows_per_doi_top10:")
    for _, r in counts.head(10).iterrows():
        print(f"- {r['doi']}: {int(r['rows'])}")
    if not counts.empty:
        preview_doi = str(counts.iloc[0]["doi"])
        cols = ["doi", "zotero_key", "formulation_id", "grouping_id", "ee_value", "drug_name", "plga_mw", "la_ga_ratio", "size_nm", "pdi"]
        cols = [c for c in cols if c in view_df.columns]
        print(f"preview_doi: {preview_doi}")
        print("preview_first_5_rows:")
        print_header_and_first3(view_df[view_df["doi"] == preview_doi][cols].head(5))


def main() -> None:
    step0_manifest, dev_manifest, refs = discover_dev15_definition()
    print_step0(step0_manifest, dev_manifest, refs)

    source_path, source_df = pick_best_source(dev_manifest)
    print_step1(source_path, source_df)

    inst_path = find_instance_assignment(source_path)
    inst_df = pd.DataFrame()
    if inst_path and inst_path.exists():
        inst_df = pd.read_csv(inst_path, sep="\t", dtype=str).fillna("")
        print(f"instance_assignment_file: {inst_path}")
    else:
        print("instance_assignment_file: <not found>")

    view_df = build_view_dataframe(source_df, dev_manifest, inst_df)
    view_df = view_df.sort_values(["doi", "grouping_id", "formulation_id"], kind="stable").reset_index(drop=True)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    used_sheet_names: set[str] = set()
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        for doi, doi_df in view_df.groupby("doi", dropna=False):
            sheet_name = safe_sheet_name(str(doi), used_sheet_names)
            doi_df.to_excel(writer, sheet_name=sheet_name, index=False)

    apply_excel_formatting(OUT_XLSX)
    print_verification(view_df)
    print(f"output_xlsx: {OUT_XLSX}")


if __name__ == "__main__":
    main()
