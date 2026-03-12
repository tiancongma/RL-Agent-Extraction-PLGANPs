#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

try:
    from src.utils.paths import DATA_CLEANED_DIR, DATA_RESULTS_DIR, DOCS_DIR
except ModuleNotFoundError:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from src.utils.paths import DATA_CLEANED_DIR, DATA_RESULTS_DIR, DOCS_DIR


GT_WORKBOOK = (
    DATA_CLEANED_DIR
    / "labels"
    / "manual"
    / "dev15_formulation_skeleton"
    / "dev15_formulation_skeleton_review_v1_fixed.xlsx"
)
COMBINED_EVAL_TSV = (
    DATA_CLEANED_DIR
    / "labels"
    / "manual"
    / "formulation_instance_dev15_combined_eval_2026-03-10_reconciled.tsv"
)
REMAINING12_EVAL_TSV = (
    DATA_CLEANED_DIR
    / "labels"
    / "manual"
    / "formulation_instance_remaining12_eval_2026-03-10"
    / "per_doi_formulation_instance_summary.tsv"
)
TUNED3_EVAL_TSV = (
    DATA_CLEANED_DIR
    / "labels"
    / "manual"
    / "formulation_instance_pilot3_eval_synthmethod_2026-03-10"
    / "per_doi_formulation_instance_summary.tsv"
)
OUT_DIR = DATA_RESULTS_DIR / "dev15_review"
OUT_XLSX = OUT_DIR / "dev15_instance_review_v1.xlsx"


PAPER_SUMMARY_COLUMNS = [
    "zotero_key",
    "doi",
    "GT_count",
    "predicted_count",
    "count_diff",
    "error_type",
    "notes",
]

PREDICTED_INSTANCE_COLUMNS = [
    "zotero_key",
    "doi",
    "instance_id",
    "instance_kind",
    "parent_instance_id",
    "evidence_block_id",
    "evidence_snippet",
    "reviewer_decision",
]

REVIEWER_DECISION_OPTIONS = [
    "",
    "valid_formulation",
    "not_a_formulation",
    "needs_second_pass",
]


def norm_text(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def norm_doi(value: object) -> str:
    doi = norm_text(value).lower()
    doi = re.sub(r"^doi\s*:\s*", "", doi)
    doi = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", doi)
    doi = re.sub(r"^doi\.org/", "", doi)
    return doi.strip()


def read_tsv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, sep="\t", dtype=str).fillna("")


def read_gt_counts(gt_workbook: Path) -> pd.DataFrame:
    source_summary = pd.read_excel(
        gt_workbook,
        sheet_name="source_summary",
        dtype=str,
        engine="openpyxl",
    ).fillna("")
    source_summary["paper_key"] = source_summary["paper_key"].map(norm_text)
    source_summary["doi"] = source_summary["doi"].map(norm_doi)
    source_summary["GT_count"] = pd.to_numeric(source_summary["GT_rows"], errors="coerce").fillna(0).astype(int)
    return source_summary[["paper_key", "doi", "paper_title", "GT_count"]].copy()


def discover_latest_weak_label_tsv(expected_keys: Iterable[str]) -> Path:
    expected = set(expected_keys)
    candidates: List[Tuple[float, Path]] = []
    for path in DATA_RESULTS_DIR.rglob("weak_labels__v7pilot_r3_fixparse.tsv"):
        try:
            df = read_tsv(path)
        except Exception:
            continue
        if "key" not in df.columns:
            continue
        keys = {norm_text(v) for v in df["key"].tolist() if norm_text(v)}
        if keys == expected:
            candidates.append((path.stat().st_mtime, path.resolve()))
    if not candidates:
        raise FileNotFoundError(f"No weak-label TSV found for key set: {sorted(expected)}")
    candidates.sort(key=lambda item: (item[0], str(item[1])))
    return candidates[-1][1]


def infer_notes(row: pd.Series) -> str:
    parts: List[str] = []
    source_group = norm_text(row.get("source_group", ""))
    paper_pattern = norm_text(row.get("paper_pattern", ""))
    notes = norm_text(row.get("notes", ""))
    if source_group:
        parts.append(source_group)
    if paper_pattern:
        parts.append(paper_pattern)
    if notes:
        parts.append(notes)
    return " | ".join(parts)


def build_paper_summary(gt_counts: pd.DataFrame, combined_eval: pd.DataFrame) -> pd.DataFrame:
    combined = combined_eval.copy()
    combined["paper_key"] = combined["paper_key"].map(norm_text)
    combined["doi"] = combined["doi"].map(norm_doi)
    combined["predicted_formulation_count"] = pd.to_numeric(
        combined["predicted_formulation_count"], errors="coerce"
    ).fillna(0).astype(int)
    combined["gt_formulation_count"] = pd.to_numeric(
        combined["gt_formulation_count"], errors="coerce"
    ).fillna(0).astype(int)

    merged = gt_counts.merge(
        combined,
        on=["paper_key", "doi"],
        how="left",
        suffixes=("_gt", "_eval"),
    )
    merged["GT_count"] = merged["GT_count"].fillna(merged["gt_formulation_count"]).astype(int)
    merged["predicted_count"] = merged["predicted_formulation_count"].fillna(0).astype(int)
    merged["count_diff"] = merged["predicted_count"] - merged["GT_count"]
    merged["error_type"] = merged["count_diff"].map(
        lambda diff: "exact" if diff == 0 else ("under_segmentation" if diff < 0 else "over_segmentation")
    )
    merged["notes"] = merged.apply(infer_notes, axis=1)
    paper_summary = merged.rename(columns={"paper_key": "zotero_key"})[PAPER_SUMMARY_COLUMNS].copy()
    paper_summary = paper_summary.sort_values(["error_type", "zotero_key"], kind="stable").reset_index(drop=True)
    return paper_summary


def is_formulation_instance(instance_kind: str) -> bool:
    return norm_text(instance_kind) in {"new_formulation", "variant_formulation"}


def parse_supporting_evidence_refs(raw: str) -> List[dict]:
    text = norm_text(raw)
    if not text:
        return []
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        return []
    if isinstance(parsed, list):
        return [item for item in parsed if isinstance(item, dict)]
    return []


def build_evidence_block_id(row: pd.Series) -> str:
    refs = parse_supporting_evidence_refs(row.get("supporting_evidence_refs", ""))
    if refs:
        first = refs[0]
        explicit = norm_text(first.get("block_id", "")) or norm_text(first.get("evidence_block_id", ""))
        if explicit:
            return explicit
        parts = [
            norm_text(first.get("region_type", "")),
            norm_text(first.get("section", "")),
        ]
        span_start = norm_text(row.get("evidence_span_start", ""))
        span_end = norm_text(row.get("evidence_span_end", ""))
        if span_start or span_end:
            parts.append(f"{span_start}-{span_end}".strip("-"))
        block_id = "|".join(part for part in parts if part)
        if block_id:
            return block_id
    section = norm_text(row.get("evidence_section", ""))
    region = norm_text(row.get("instance_evidence_region_type", "")) or norm_text(row.get("evidence_region_type", ""))
    span_start = norm_text(row.get("evidence_span_start", ""))
    span_end = norm_text(row.get("evidence_span_end", ""))
    fallback_parts = [region, section]
    if span_start or span_end:
        fallback_parts.append(f"{span_start}-{span_end}".strip("-"))
    return "|".join(part for part in fallback_parts if part)


def build_evidence_snippet(row: pd.Series, max_chars: int = 400) -> str:
    snippet = norm_text(row.get("evidence_span_text", ""))
    if not snippet:
        refs = parse_supporting_evidence_refs(row.get("supporting_evidence_refs", ""))
        if refs:
            snippet = norm_text(refs[0].get("span_text", ""))
    if len(snippet) > max_chars:
        return snippet[: max_chars - 3].rstrip() + "..."
    return snippet


def build_instance_id(row: pd.Series) -> str:
    for column in ["local_instance_id", "formulation_id", "raw_formulation_label"]:
        value = norm_text(row.get(column, ""))
        if value:
            return value
    return ""


def build_predicted_instances(weak_label_paths: List[Path], paper_summary: pd.DataFrame) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []
    for path in weak_label_paths:
        df = read_tsv(path)
        df["source_weak_label_path"] = str(path)
        frames.append(df)
    all_rows = pd.concat(frames, ignore_index=True)
    all_rows["key"] = all_rows["key"].map(norm_text)
    all_rows["doi"] = all_rows["doi"].map(norm_doi)

    paper_meta = paper_summary.rename(columns={"zotero_key": "key"})[["key", "doi"]].drop_duplicates()
    merged = all_rows.merge(paper_meta, on=["key", "doi"], how="inner")

    predicted = pd.DataFrame(
        {
            "zotero_key": merged["key"],
            "doi": merged["doi"],
            "instance_id": merged.apply(build_instance_id, axis=1),
            "instance_kind": merged["instance_kind"].map(norm_text),
            "parent_instance_id": merged["parent_instance_id"].map(norm_text),
            "evidence_block_id": merged.apply(build_evidence_block_id, axis=1),
            "evidence_snippet": merged.apply(build_evidence_snippet, axis=1),
            "reviewer_decision": "",
        }
    )
    predicted = predicted.sort_values(
        ["zotero_key", "instance_kind", "instance_id"],
        kind="stable",
    ).reset_index(drop=True)
    return predicted[PREDICTED_INSTANCE_COLUMNS]


def build_review_queue(paper_summary: pd.DataFrame) -> pd.DataFrame:
    return paper_summary[paper_summary["error_type"] != "exact"].copy().reset_index(drop=True)


def style_worksheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column_cells in ws.columns:
        max_len = max(len(norm_text(cell.value)) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def print_preview(label: str, df: pd.DataFrame, rows: int = 5) -> None:
    print(f"\n[{label}]")
    print("\t".join(df.columns.tolist()))
    if df.empty:
        print("<empty>")
        return
    preview = df.head(rows).fillna("")
    for _, row in preview.iterrows():
        values = [norm_text(v).encode("ascii", "replace").decode("ascii") for v in row.tolist()]
        print("\t".join(values))


def main() -> int:
    gt_counts = read_gt_counts(GT_WORKBOOK)
    combined_eval = read_tsv(COMBINED_EVAL_TSV)
    remaining12_eval = read_tsv(REMAINING12_EVAL_TSV)
    tuned3_eval = read_tsv(TUNED3_EVAL_TSV)

    tuned3_keys = sorted(combined_eval.loc[combined_eval["source_group"] == "tuned_3paper", "paper_key"].map(norm_text).tolist())
    remaining12_keys = sorted(
        combined_eval.loc[combined_eval["source_group"] == "remaining_12paper", "paper_key"].map(norm_text).tolist()
    )
    tuned3_weak_labels = discover_latest_weak_label_tsv(tuned3_keys)
    remaining12_weak_labels = discover_latest_weak_label_tsv(remaining12_keys)

    paper_summary = build_paper_summary(gt_counts, combined_eval)
    predicted_instances = build_predicted_instances([tuned3_weak_labels, remaining12_weak_labels], paper_summary)
    review_queue = build_review_queue(paper_summary)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        paper_summary.to_excel(writer, sheet_name="paper_summary", index=False)
        predicted_instances.to_excel(writer, sheet_name="predicted_instances", index=False)
        review_queue.to_excel(writer, sheet_name="review_queue", index=False)

        wb = writer.book
        for sheet_name in ["paper_summary", "predicted_instances", "review_queue"]:
            style_worksheet(wb[sheet_name])

    exact = int((paper_summary["error_type"] == "exact").sum())
    under = int((paper_summary["error_type"] == "under_segmentation").sum())
    over = int((paper_summary["error_type"] == "over_segmentation").sum())

    print(f"number_of_papers={len(paper_summary)}")
    print(f"exact={exact}")
    print(f"under_segmentation={under}")
    print(f"over_segmentation={over}")
    print(f"generated_excel={OUT_XLSX}")
    print(f"input_gt_workbook={GT_WORKBOOK}")
    print(f"input_combined_eval={COMBINED_EVAL_TSV}")
    print(f"input_remaining12_eval={REMAINING12_EVAL_TSV}")
    print(f"input_tuned3_eval={TUNED3_EVAL_TSV}")
    print(f"input_tuned3_weak_labels={tuned3_weak_labels}")
    print(f"input_remaining12_weak_labels={remaining12_weak_labels}")

    print_preview("paper_summary", paper_summary)
    print_preview("predicted_instances", predicted_instances)
    print_preview("review_queue", review_queue)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
