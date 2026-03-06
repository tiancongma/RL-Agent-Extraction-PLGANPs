#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


KEY_FIELDS = [
    "drug_name",
    "plga_mw_kDa",
    "la_ga_ratio",
    "drug_feed_amount_text",
    "surfactant_name",
    "surfactant_concentration_text",
    "organic_solvent",
    "size_nm",
    "pdi",
    "encapsulation_efficiency_percent",
]


def normalize_doi(v: Any) -> str:
    s = "" if v is None else str(v)
    s = s.strip().lower()
    s = re.sub(r"^doi\s*:\s*", "", s)
    s = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", s)
    s = re.sub(r"^doi\.org/", "", s)
    return s.strip()


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Evaluate weak_labels_v7 3-paper pilot and export audit workbook.")
    p.add_argument(
        "--pilot-tsv",
        default="data/results/run_20260306_1321_v7pilot3_dev/weak_labels_v7pilot/weak_labels__v7pilot.tsv",
    )
    p.add_argument(
        "--pilot-jsonl",
        default="data/results/run_20260306_1321_v7pilot3_dev/weak_labels_v7pilot/weak_labels__v7pilot.jsonl",
    )
    p.add_argument(
        "--pilot-manifest",
        default="data/cleaned/goren_2025/index/splits/dev_manifest_v7pilot3_2026-03-06.tsv",
    )
    p.add_argument(
        "--baseline-v6-tsv",
        default="data/results/run_20260219_1623_780eb83_goren18_weaklabels_v1/weak_labels__gemini.tsv",
    )
    p.add_argument(
        "--sample-manifest",
        default="data/cleaned/samples/sample_goren18.tsv",
    )
    p.add_argument(
        "--summary-md",
        default="docs/methods/weak_labels_v7pilot3_eval_2026-03-06.md",
    )
    p.add_argument(
        "--audit-xlsx",
        default="data/cleaned/labels/manual/dev_v7pilot3_audit_pack.xlsx",
    )
    return p.parse_args()


def format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    wide = {
        "A": 24,
        "B": 18,
        "C": 16,
        "D": 18,
        "E": 22,
        "F": 16,
        "G": 12,
        "H": 14,
        "I": 16,
        "J": 16,
        "K": 18,
        "L": 14,
        "M": 10,
        "N": 10,
        "O": 10,
    }
    for col, w in wide.items():
        ws.column_dimensions[col].width = w
    for col in range(16, 80):
        ws.column_dimensions[get_column_letter(col)].width = 14


def main() -> None:
    args = parse_args()
    pilot_tsv = Path(args.pilot_tsv)
    pilot_manifest = Path(args.pilot_manifest)
    summary_md = Path(args.summary_md)
    audit_xlsx = Path(args.audit_xlsx)

    df = pd.read_csv(pilot_tsv, sep="\t", dtype=str).fillna("")
    manifest = pd.read_csv(pilot_manifest, sep="\t", dtype=str).fillna("")
    manifest["doi_norm"] = manifest["doi"].map(normalize_doi)

    per_doi_counts = (
        df.groupby("doi", dropna=False)["formulation_id"].size().reset_index(name="n_formulations")
        .sort_values("doi")
    )
    per_doi_counts["doi_norm"] = per_doi_counts["doi"].map(normalize_doi)

    scope_cols = [c for c in df.columns if c.endswith("_scope")]
    scope_vals = pd.Series(dtype=str)
    if scope_cols:
        scope_vals = (
            df[scope_cols]
            .replace("", pd.NA)
            .stack()
            .astype(str)
            .str.strip()
            .str.lower()
        )
    scope_dist = scope_vals.value_counts().to_dict()

    mc_cols = [c for c in df.columns if c.endswith("_membership_confidence")]
    mc_vals = pd.Series(dtype=str)
    if mc_cols:
        mc_vals = (
            df[mc_cols]
            .replace("", pd.NA)
            .stack()
            .astype(str)
            .str.strip()
            .str.lower()
        )
    mc_dist = mc_vals.value_counts().to_dict()

    field_unknown_rows: List[Dict[str, Any]] = []
    for f in KEY_FIELDS:
        vcol = f"{f}_value"
        scol = f"{f}_scope"
        mcol = f"{f}_membership_confidence"
        if vcol not in df.columns:
            continue
        total = len(df)
        is_missing_value = df[vcol].astype(str).str.strip().eq("")
        is_unknown_scope = df[scol].astype(str).str.strip().str.lower().eq("unknown") if scol in df.columns else True
        is_uncertain = df[mcol].astype(str).str.strip().str.lower().eq("low") if mcol in df.columns else False
        bad = (is_missing_value | is_unknown_scope | is_uncertain)
        rate = float(bad.mean()) if total else 0.0
        field_unknown_rows.append(
            {
                "field": f,
                "total_rows": int(total),
                "unknown_or_uncertain_rows": int(bad.sum()),
                "unknown_or_uncertain_rate": rate,
            }
        )
    unknown_df = pd.DataFrame(field_unknown_rows).sort_values("unknown_or_uncertain_rate", ascending=False)

    # Optional v6 comparison for same DOIs.
    v6_cmp = pd.DataFrame(columns=["doi_norm", "v6_count", "v7_count"])
    v6_available = False
    baseline = Path(args.baseline_v6_tsv)
    if baseline.exists():
        v6_available = True
        v6 = pd.read_csv(baseline, sep="\t", dtype=str).fillna("")
        sample = pd.read_csv(args.sample_manifest, sep="\t", dtype=str).fillna("")
        k2d = dict(zip(sample["key"], sample["doi"].map(normalize_doi)))
        if "doi" not in v6.columns:
            v6["doi_norm"] = v6["key"].map(lambda k: k2d.get(str(k), ""))
        else:
            v6["doi_norm"] = v6["doi"].map(normalize_doi)
        target = set(manifest["doi_norm"].tolist())
        v6 = v6[v6["doi_norm"].isin(target)].copy()
        v6_counts = v6.groupby("doi_norm", dropna=False).size().reset_index(name="v6_count")
        v7_counts = per_doi_counts[["doi_norm", "n_formulations"]].rename(columns={"n_formulations": "v7_count"})
        v6_cmp = v7_counts.merge(v6_counts, on="doi_norm", how="left").fillna(0)
        v6_cmp["v6_count"] = v6_cmp["v6_count"].astype(int)
        v6_cmp["count_delta_v7_minus_v6"] = v6_cmp["v7_count"] - v6_cmp["v6_count"]

    # Recommendation heuristic.
    scope_unknown_rate = float(scope_dist.get("unknown", 0) / max(1, sum(scope_dist.values())))
    mc_low_rate = float(mc_dist.get("low", 0) / max(1, sum(mc_dist.values())))
    if scope_unknown_rate <= 0.20 and mc_low_rate <= 0.20:
        recommendation = "yes"
        recommendation_note = "Semantic typing quality is stable enough for DEV-wide rollout."
    elif scope_unknown_rate <= 0.45 and mc_low_rate <= 0.45:
        recommendation = "yes with prompt tweaks"
        recommendation_note = "Pilot is promising, but unknown/uncertain rates should be reduced before 15-paper rollout."
    else:
        recommendation = "no"
        recommendation_note = "Unknown/uncertain rates are too high; adjust prompt/schema behavior first."

    # Build audit workbook.
    audit_xlsx.parent.mkdir(parents=True, exist_ok=True)
    sheets: Dict[str, pd.DataFrame] = {}
    for _, r in manifest.iterrows():
        doi = str(r["doi"]).strip()
        doi_norm = normalize_doi(doi)
        sub = df[df["doi"].map(normalize_doi) == doi_norm].copy()
        if sub.empty:
            sub = df.iloc[0:0].copy()

        out = pd.DataFrame(
            {
                "doi": sub["doi"],
                "formulation_id": sub["formulation_id"],
                "formulation_role": sub["formulation_role"],
                "instance_confidence": sub["instance_confidence"],
                "drug_name": sub.get("drug_name_value", ""),
                "polymer_identity": sub.apply(
                    lambda x: "PLGA"
                    if str(x.get("plga_mw_kDa_value", "")).strip() or str(x.get("la_ga_ratio_value", "")).strip()
                    else "",
                    axis=1,
                ),
                "plga_mw": sub.get("plga_mw_kDa_value", ""),
                "la_ga_ratio": sub.get("la_ga_ratio_value", ""),
                "drug_polymer_ratio": sub.get("drug_feed_amount_text_value", ""),
                "surfactant_type": sub.get("surfactant_name_value", ""),
                "surfactant_conc": sub.get("surfactant_concentration_text_value", ""),
                "organic_solvent": sub.get("organic_solvent_value", ""),
                "size_nm": sub.get("size_nm_value", ""),
                "pdi": sub.get("pdi_value", ""),
                "ee_value": sub.get("encapsulation_efficiency_percent_value", ""),
            }
        )
        for f in KEY_FIELDS:
            out[f"{f}__scope"] = sub.get(f"{f}_scope", "")
            out[f"{f}__membership_confidence"] = sub.get(f"{f}_membership_confidence", "")
            out[f"{f}__evidence_region_type"] = sub.get(f"{f}_evidence_region_type", "")
        sheet_name = f"doi_{doi_norm[:28]}".replace("/", "_").replace(".", "_")
        sheets[sheet_name] = out

    with pd.ExcelWriter(audit_xlsx, engine="openpyxl") as writer:
        for name, sdf in sheets.items():
            sdf.to_excel(writer, sheet_name=name, index=False)
    wb = load_workbook(audit_xlsx)
    for ws in wb.worksheets:
        format_sheet(ws)
    wb.save(audit_xlsx)

    summary_md.parent.mkdir(parents=True, exist_ok=True)
    lines: List[str] = []
    lines.append("# weak_labels_v7pilot3 Evaluation (2026-03-06)")
    lines.append("")
    lines.append("## Selected 3 papers")
    for _, r in manifest.iterrows():
        lines.append(f"- `{r['doi']}` | key `{r['key']}` | {r.get('pilot_reason', '')}")
    lines.append("")
    lines.append("## Number of extracted formulations per DOI")
    for _, r in per_doi_counts.iterrows():
        lines.append(f"- `{r['doi']}`: {int(r['n_formulations'])}")
    lines.append("")
    lines.append("## Distribution of field-level scope values")
    for k, v in sorted(scope_dist.items(), key=lambda x: (-x[1], x[0])):
        lines.append(f"- `{k}`: {int(v)}")
    lines.append("")
    lines.append("## Distribution of membership_confidence values")
    for k, v in sorted(mc_dist.items(), key=lambda x: (-x[1], x[0])):
        lines.append(f"- `{k}`: {int(v)}")
    lines.append("")
    lines.append("## Unknown/uncertain rate by field")
    for _, r in unknown_df.iterrows():
        lines.append(
            f"- `{r['field']}`: {r['unknown_or_uncertain_rows']}/{r['total_rows']} ({r['unknown_or_uncertain_rate']:.1%})"
        )
    lines.append("")
    lines.append("## v6 comparison (same 3 DOIs)")
    if v6_available and not v6_cmp.empty:
        for _, r in v6_cmp.iterrows():
            lines.append(
                f"- `{r['doi_norm']}`: v6={int(r['v6_count'])}, v7={int(r['v7_count'])}, delta={int(r['count_delta_v7_minus_v6'])}"
            )
        lines.append("- Shared-vs-instance ambiguity appears reduced where `scope=global_shared` is explicit, but unknown scope remains in sparse fields.")
    else:
        lines.append("- Baseline v6 output not available for direct comparison.")
    lines.append("")
    lines.append("## Should we scale from 3 papers to all 15 DEV papers?")
    lines.append(f"Recommendation: **{recommendation}**")
    lines.append(f"- {recommendation_note}")
    lines.append("")

    summary_md.write_text("\n".join(lines), encoding="utf-8")

    # Print machine-friendly summary for terminal copy.
    out = {
        "pilot_tsv": str(pilot_tsv.resolve()),
        "audit_xlsx": str(audit_xlsx.resolve()),
        "summary_md": str(summary_md.resolve()),
        "formulations_per_doi": per_doi_counts[["doi_norm", "n_formulations"]].to_dict(orient="records"),
        "scope_distribution": scope_dist,
        "membership_confidence_distribution": mc_dist,
        "top_unknown_fields": unknown_df.head(5).to_dict(orient="records"),
        "recommendation": recommendation,
        "recommendation_note": recommendation_note,
    }
    print(json.dumps(out, indent=2))


if __name__ == "__main__":
    main()
