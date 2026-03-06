#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


TARGET_FIELDS = [
    "polymer_name_raw",
    "plga_mw_kDa",
    "la_ga_ratio",
    "organic_solvent",
    "surfactant_type",
    "surfactant_concentration_text",
    "encapsulation_efficiency_percent",
    "size_nm",
]


def normalize_doi(v: Any) -> str:
    s = "" if v is None else str(v)
    s = s.strip().lower()
    s = re.sub(r"^doi\s*:\s*", "", s)
    s = re.sub(r"^https?://(?:dx\.)?doi\.org/", "", s)
    s = re.sub(r"^doi\.org/", "", s)
    return s.strip()


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build JSON->TSV field mapping audit workbook for v7pilot3 run.")
    p.add_argument(
        "--run-dir",
        default="data/results/run_20260306_1513_v7pilot3r3_dev/weak_labels_v7pilot_r3",
    )
    p.add_argument(
        "--jsonl",
        default="",
        help="Optional explicit JSONL path. If empty, auto-detect under run-dir.",
    )
    p.add_argument(
        "--tsv",
        default="",
        help="Optional explicit TSV path. If empty, auto-detect under run-dir.",
    )
    p.add_argument(
        "--out-xlsx",
        default="data/cleaned/labels/manual/dev_v7pilot3_field_mapping_audit.xlsx",
    )
    return p.parse_args()


def _norm_text(v: Any) -> str:
    s = "" if v is None else str(v)
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _to_num(v: Any) -> float | None:
    try:
        s = _norm_text(v)
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def _bool_match(a: Any, b: Any) -> bool:
    sa = _norm_text(a).lower()
    sb = _norm_text(b).lower()
    if sa == sb:
        return True
    na = _to_num(sa)
    nb = _to_num(sb)
    if na is not None and nb is not None:
        return abs(na - nb) <= 1e-9
    return False


def _field_rec(value_raw: Any = "", scope: Any = "", membership: Any = "", evidence_type: Any = "", quote: Any = "") -> Dict[str, str]:
    return {
        "value_raw": _norm_text(value_raw),
        "scope": _norm_text(scope).lower(),
        "membership_confidence": _norm_text(membership).lower(),
        "evidence_region_type": _norm_text(evidence_type).lower(),
        "evidence_quote": _norm_text(quote),
    }


def extract_json_fields_from_dict(fields: Dict[str, Any]) -> Dict[str, Dict[str, str]]:
    out = {f: _field_rec() for f in TARGET_FIELDS}
    mapping = {
        "plga_mw_kDa": "plga_mw_kDa",
        "la_ga_ratio": "la_ga_ratio",
        "organic_solvent": "organic_solvent",
        "surfactant_name": "surfactant_type",
        "surfactant_concentration_text": "surfactant_concentration_text",
        "encapsulation_efficiency_percent": "encapsulation_efficiency_percent",
        "size_nm": "size_nm",
    }
    for jk, tf in mapping.items():
        if jk not in fields:
            continue
        obj = fields.get(jk)
        if isinstance(obj, dict):
            out[tf] = _field_rec(
                obj.get("value"),
                obj.get("scope", ""),
                obj.get("membership_confidence", ""),
                obj.get("evidence_region_type", ""),
                obj.get("value_text", ""),
            )
        else:
            out[tf] = _field_rec(obj, "", "", "", "")

    # polymer_name_raw comes from plga_mw_kDa textual descriptor in this pilot schema.
    pmw = fields.get("plga_mw_kDa")
    if isinstance(pmw, dict):
        out["polymer_name_raw"] = _field_rec(
            pmw.get("value_text", ""),
            pmw.get("scope", ""),
            pmw.get("membership_confidence", ""),
            pmw.get("evidence_region_type", ""),
            pmw.get("value_text", ""),
        )
    return out


def extract_json_fields_from_list(fields_list: List[Any]) -> Dict[str, Dict[str, str]]:
    # Heuristic parser for malformed list-style "fields" outputs.
    out = {f: _field_rec() for f in TARGET_FIELDS}
    for it in fields_list:
        if not isinstance(it, dict):
            continue
        v = _norm_text(it.get("value", ""))
        vt = _norm_text(it.get("value_text", ""))
        scope = _norm_text(it.get("scope", "")).lower()
        membership = _norm_text(it.get("membership_confidence", "")).lower()
        e = _norm_text(it.get("evidence_region_type", "")).lower()
        blob = (vt + " " + v).lower()

        rec = _field_rec(v if v else vt, scope, membership, e, vt)

        if re.search(r"\bresomer\b|\brg\s*50[234]\b|plga grade", blob):
            out["polymer_name_raw"] = rec
            # MW is intentionally unknown for product code-only mentions.
            out["plga_mw_kDa"] = _field_rec("", scope, membership, e, vt)
            continue
        if re.search(r"\b\d+\s*:\s*\d+\b", blob):
            out["la_ga_ratio"] = rec
            continue
        if re.search(r"\bacetone\b|\bdcm\b|dichloromethane|ethyl acetate|chloroform|methanol|acetonitrile", blob):
            out["organic_solvent"] = rec
            continue
        if re.search(r"\bnm\b", blob):
            out["size_nm"] = rec
            continue
        if re.search(r"encapsulation|entrapment|\bee\b", blob):
            out["encapsulation_efficiency_percent"] = rec
            continue
        if re.search(r"polysorbate|labrafil|surfactant|pva|polyvinyl alcohol", blob):
            if re.search(r"\b\d+(\.\d+)?\s*(%|mg/ml|mg)\b", blob):
                out["surfactant_concentration_text"] = rec
            else:
                out["surfactant_type"] = rec
            continue
    return out


def parse_jsonl_to_long(jsonl_path: Path) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    with jsonl_path.open("r", encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            doi = normalize_doi(obj.get("doi", ""))
            forms = obj.get("formulations", [])
            if not isinstance(forms, list):
                continue
            for fm in forms:
                if not isinstance(fm, dict):
                    continue
                fid = _norm_text(fm.get("formulation_id", fm.get("id", "")))
                role = _norm_text(fm.get("formulation_role", ""))
                fobj = fm.get("fields", {})
                if isinstance(fobj, dict):
                    fld = extract_json_fields_from_dict(fobj)
                elif isinstance(fobj, list):
                    fld = extract_json_fields_from_list(fobj)
                else:
                    fld = {f: _field_rec() for f in TARGET_FIELDS}

                for fn in TARGET_FIELDS:
                    rr = fld.get(fn, _field_rec())
                    rows.append(
                        {
                            "doi": doi,
                            "formulation_id": fid,
                            "formulation_role": role,
                            "field_name": fn,
                            "json_value_raw": rr["value_raw"],
                            "json_scope": rr["scope"],
                            "json_membership": rr["membership_confidence"],
                            "json_evidence_type": rr["evidence_region_type"],
                            "json_evidence_quote": rr["evidence_quote"],
                        }
                    )
    return pd.DataFrame(rows)


def parse_tsv_to_long(tsv_path: Path) -> pd.DataFrame:
    df = pd.read_csv(tsv_path, sep="\t", dtype=str).fillna("")
    rows: List[Dict[str, Any]] = []
    for _, r in df.iterrows():
        doi = normalize_doi(r.get("doi", ""))
        fid = _norm_text(r.get("formulation_id", ""))
        role = _norm_text(r.get("formulation_role", ""))

        tsv_map = {
            "polymer_name_raw": ("plga_mw_kDa_value_text", "plga_mw_kDa_scope", "plga_mw_kDa_membership_confidence", "plga_mw_kDa_evidence_region_type"),
            "plga_mw_kDa": ("plga_mw_kDa_value", "plga_mw_kDa_scope", "plga_mw_kDa_membership_confidence", "plga_mw_kDa_evidence_region_type"),
            "la_ga_ratio": ("la_ga_ratio_value", "la_ga_ratio_scope", "la_ga_ratio_membership_confidence", "la_ga_ratio_evidence_region_type"),
            "organic_solvent": ("organic_solvent_value", "organic_solvent_scope", "organic_solvent_membership_confidence", "organic_solvent_evidence_region_type"),
            "surfactant_type": ("surfactant_name_value", "surfactant_name_scope", "surfactant_name_membership_confidence", "surfactant_name_evidence_region_type"),
            "surfactant_concentration_text": ("surfactant_concentration_text_value", "surfactant_concentration_text_scope", "surfactant_concentration_text_membership_confidence", "surfactant_concentration_text_evidence_region_type"),
            "encapsulation_efficiency_percent": ("encapsulation_efficiency_percent_value", "encapsulation_efficiency_percent_scope", "encapsulation_efficiency_percent_membership_confidence", "encapsulation_efficiency_percent_evidence_region_type"),
            "size_nm": ("size_nm_value", "size_nm_scope", "size_nm_membership_confidence", "size_nm_evidence_region_type"),
        }
        for fn, cols in tsv_map.items():
            vcol, scol, mcol, ecol = cols
            rows.append(
                {
                    "doi": doi,
                    "formulation_id": fid,
                    "formulation_role": role,
                    "field_name": fn,
                    "tsv_value": _norm_text(r.get(vcol, "")),
                    "tsv_scope": _norm_text(r.get(scol, "")).lower(),
                    "tsv_membership": _norm_text(r.get(mcol, "")).lower(),
                    "tsv_evidence_type": _norm_text(r.get(ecol, "")).lower(),
                }
            )
    return pd.DataFrame(rows)


def format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        width = min(max(10, max_len + 2), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main() -> None:
    args = parse_args()
    run_dir = Path(args.run_dir)
    jsonl_path = Path(args.jsonl) if args.jsonl else next(iter(sorted(run_dir.glob("weak_labels__*.jsonl"))))
    tsv_path = Path(args.tsv) if args.tsv else next(iter(sorted(run_dir.glob("weak_labels__*.tsv"))))

    json_long = parse_jsonl_to_long(jsonl_path)
    tsv_long = parse_tsv_to_long(tsv_path)

    merged = json_long.merge(
        tsv_long,
        on=["doi", "formulation_id", "formulation_role", "field_name"],
        how="outer",
    )
    for c in [
        "json_value_raw",
        "json_scope",
        "json_membership",
        "json_evidence_type",
        "json_evidence_quote",
        "tsv_value",
        "tsv_scope",
        "tsv_membership",
        "tsv_evidence_type",
    ]:
        if c not in merged.columns:
            merged[c] = ""
        merged[c] = merged[c].fillna("")

    merged["value_match"] = merged.apply(lambda r: _bool_match(r["json_value_raw"], r["tsv_value"]), axis=1)
    merged["scope_match"] = merged.apply(lambda r: _bool_match(r["json_scope"], r["tsv_scope"]), axis=1)
    merged["membership_match"] = merged.apply(lambda r: _bool_match(r["json_membership"], r["tsv_membership"]), axis=1)
    merged["evidence_match"] = merged.apply(lambda r: _bool_match(r["json_evidence_type"], r["tsv_evidence_type"]), axis=1)

    out_cols = [
        "doi",
        "formulation_id",
        "formulation_role",
        "field_name",
        "json_value_raw",
        "tsv_value",
        "value_match",
        "json_scope",
        "tsv_scope",
        "scope_match",
        "json_membership",
        "tsv_membership",
        "membership_match",
        "json_evidence_type",
        "tsv_evidence_type",
        "evidence_match",
    ]
    merged = merged[out_cols].copy()
    merged = merged.sort_values(["doi", "formulation_id", "field_name"]).reset_index(drop=True)

    # Summary metrics
    total_rows = int(len(merged))
    value_mismatches = int((merged["value_match"] == False).sum())  # noqa: E712
    scope_mismatches = int((merged["scope_match"] == False).sum())  # noqa: E712
    membership_mismatches = int((merged["membership_match"] == False).sum())  # noqa: E712
    evidence_mismatches = int((merged["evidence_match"] == False).sum())  # noqa: E712

    field_counts = (
        merged.assign(
            any_mismatch=lambda d: (~d["value_match"]) | (~d["scope_match"]) | (~d["membership_match"]) | (~d["evidence_match"])
        )
        .groupby("field_name", dropna=False)["any_mismatch"]
        .sum()
        .reset_index(name="mismatch_count")
        .sort_values("mismatch_count", ascending=False)
    )

    out_xlsx = Path(args.out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        # SUMMARY sheet
        summary_rows = [
            {"metric": "total_rows", "value": total_rows},
            {"metric": "value_mismatches", "value": value_mismatches},
            {"metric": "scope_mismatches", "value": scope_mismatches},
            {"metric": "membership_mismatches", "value": membership_mismatches},
            {"metric": "evidence_mismatches", "value": evidence_mismatches},
            {"metric": "jsonl_path", "value": str(jsonl_path)},
            {"metric": "tsv_path", "value": str(tsv_path)},
        ]
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="SUMMARY", index=False, startrow=0)
        field_counts.to_excel(writer, sheet_name="SUMMARY", index=False, startrow=len(summary_rows) + 3)

        # One sheet per DOI
        for doi, g in merged.groupby("doi", dropna=False):
            d = str(doi).strip() or "unknown_doi"
            sheet_name = f"doi_{d[:26]}".replace("/", "_").replace(".", "_")
            g.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(out_xlsx)
    for ws in wb.worksheets:
        format_sheet(ws)
    wb.save(out_xlsx)

    top5 = field_counts.head(5).to_dict(orient="records")
    print(f"audit_workbook={out_xlsx.resolve()}")
    print(f"rows_analyzed={total_rows}")
    print(f"value_mismatch_count={value_mismatches}")
    print(f"scope_mismatch_count={scope_mismatches}")
    print("top_5_problematic_fields=" + json.dumps(top5, ensure_ascii=False))


if __name__ == "__main__":
    main()
